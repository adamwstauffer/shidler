"""FIN-321 Stage 4 (final memo + AI prompt) grading scanner.

Reads the manifest produced by `fetch_stage4.py`, inspects each fetched
deliverable (.md / .pdf / .html / .docx), and scores against the Stage 4
rubric on a 6-point scale per instructor direction:

    Hedge Interpretation (A+B)     /1.5
    Sensitivity (C)                /1.0
    Recommendation + Justify (D+E) /1.5
    Structured AI Prompt (F)       /2.0
                                   = /6.0

A 0.5 deduction is applied if the submission has no GitHub link.
The 85% floor (curved = MAX(raw, 5.10)) is applied AFTER the deduction is
shown separately so the deduction still bites for students who otherwise
would have hit the floor.
"""
from __future__ import annotations

import csv
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STAGE4_DIR = Path(
    r"C:\GitHub\shidler\courses\International-Finance-And-Securities\FIN-321"
    r"\ignore\2026-Spring\stage4"
)
MANIFEST = STAGE4_DIR / "_grading" / "stage4-fetch-manifest.csv"
OUTPUT = STAGE4_DIR / "_grading" / "stage4-grading-worksheet.xlsx"

TOTAL_POINTS = 6.0
FLOOR_PCT = 0.85
GITHUB_PENALTY = 0.5
GITHUB_BUMP = 0.25

WORD_RE = re.compile(r"\b\w+\b")

REQUIRED_SECTIONS = {
    "exposure":      ["exposure summary", "receivable details", "fx risk",
                      "exposure", "currency, amount", "amount, and timing"],
    "hedge_outcomes":["hedge outcomes", "summary of hedge", "strategy comparison",
                      "forward hedge", "money market hedge", "put option",
                      "call option", "no hedge"],
    "sensitivity":   ["sensitivity", "appreciation", "depreciation", "+5%", "-5%",
                      "±5", "scenarios"],
    "recommendation":["recommendation", "recommend", "i recommend", "strategic recommendation"],
    "justification": ["executive justification", "cash flow stability", "budget certainty",
                      "liquidity", "premium cost", "optionality"],
    "ai_prompt":     ["# goal", "# input variables", "# spreadsheet requirements",
                      "# verification", "# export", "structured prompt", "prompt to",
                      "structured ai prompt"],
}

HEDGE_KEYWORDS = {
    "Forward":     ["forward hedge", "forward rate", "forward contract",
                    "locked-in", "locked in"],
    "MoneyMarket": ["money market", "mm hedge", "money-market", "borrow eur",
                    "synthetic forward", "covered interest"],
    "Put":         ["put option", "put hedge", "put premium", "k_put",
                    "put strike"],
    "Call":        ["call option", "call hedge", "call premium", "k_call",
                    "call strike"],
    "NoHedge":     ["no hedge", "unhedged", "no-hedge"],
}

NAMED_RANGE_TOKENS = [
    "FC_AMT", "S0_in", "F0_in", "R_USD", "R_FC",
    "K_PUT", "K_CALL", "PREM_PUT", "PREM_CALL", "T_DAYS",
    "S0", "F0", "r_USD", "r_EUR", "K_put", "K_call",
    "Premium_put", "Premium_call",
    "USD_forward", "USD_mm", "USD_put", "USD_call",
]
_NAMED_RE = re.compile(
    r"(?<![A-Za-z0-9_])(?:" + "|".join(re.escape(t) for t in NAMED_RANGE_TOKENS)
    + r")(?![A-Za-z0-9_])",
    re.IGNORECASE,
)


@dataclass
class Grade:
    sid: str
    name: str
    github_url: str
    local_file: str
    inline_html: str
    fetch_status: str
    file_format: str = ""
    word_count: int = 0
    sections_found: list[str] = field(default_factory=list)
    hedges_found: list[str] = field(default_factory=list)
    has_input_table: bool = False
    has_sensitivity_table: bool = False
    has_named_ranges: bool = False
    distinct_named_tokens: int = 0
    has_color_codes: bool = False
    has_verification: bool = False
    score_interpretation: float = 0.0   # /1.5
    score_sensitivity: float = 0.0      # /1.0
    score_recommendation: float = 0.0   # /1.5
    score_prompt: float = 0.0           # /2.0
    raw_total: float = 0.0              # /6
    github_deduction: float = 0.0
    github_bump: float = 0.0
    flags: list[str] = field(default_factory=list)
    notes: str = ""


def _extract_md(p: Path) -> str:
    return p.read_text(encoding="utf-8", errors="ignore")


def _extract_html(p: Path) -> str:
    raw = p.read_text(encoding="utf-8", errors="ignore")
    # Strip HTML tags coarsely
    text = re.sub(r"<br\s*/?>", "\n", raw, flags=re.IGNORECASE)
    text = re.sub(r"</p>", "\n\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</h\d>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</li>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    # decode common entities
    text = (text.replace("&nbsp;", " ").replace("&amp;", "&")
                .replace("&lt;", "<").replace("&gt;", ">"))
    return text


def _extract_pdf(p: Path) -> str:
    try:
        from pdfminer.high_level import extract_text
        return extract_text(str(p)) or ""
    except Exception as e:
        return f"[pdf extraction failed: {e}]"


def _extract_docx(p: Path) -> str:
    try:
        from docx import Document
        d = Document(str(p))
        return "\n".join(par.text for par in d.paragraphs if par.text)
    except Exception as e:
        return f"[docx extraction failed: {e}]"


def load_text(local: str, inline: str) -> tuple[str, str]:
    """Return (text, source_label)."""
    if local and Path(local).exists():
        p = Path(local)
        ext = p.suffix.lower()
        if ext in (".md", ".txt"):
            return _extract_md(p), f"github/{ext.lstrip('.')}"
        if ext == ".html":
            return _extract_html(p), "github/html"
        if ext == ".pdf":
            return _extract_pdf(p), "github/pdf"
        if ext == ".docx":
            return _extract_docx(p), "github/docx"
    if inline and Path(inline).exists():
        return _extract_html(Path(inline)), "inline/html"
    return "", "missing"


def count_md_tables(text: str) -> int:
    return len(re.findall(r"\n\s*\|[^\n]*\|\s*\n\s*\|[\s\-:|]+\|\s*\n", text))


def inspect(row: dict) -> Grade:
    g = Grade(
        sid=row["sid"], name=row["name"],
        github_url=row.get("github_url", ""),
        local_file=row.get("local_file", ""),
        inline_html=row.get("inline_html", ""),
        fetch_status=row.get("status", ""),
    )
    text, fmt = load_text(g.local_file, g.inline_html)
    g.file_format = fmt
    if not text.strip():
        g.notes = "no readable text"
        g.flags.append("NO_TEXT")
        if g.github_url:
            g.github_bump = GITHUB_BUMP
        else:
            g.github_deduction = GITHUB_PENALTY
            g.flags.append("NO_GITHUB_LINK")
        return g

    lower = text.lower()
    g.word_count = len(WORD_RE.findall(text))

    # Section discovery
    sections = []
    for key, anchors in REQUIRED_SECTIONS.items():
        if any(a in lower for a in anchors):
            sections.append(key)
    g.sections_found = sections

    # Hedges discovered
    hedges = [k for k, kws in HEDGE_KEYWORDS.items() if any(w in lower for w in kws)]
    g.hedges_found = hedges

    # Named range tokens
    matches = {m.lower() for m in _NAMED_RE.findall(text)}
    g.distinct_named_tokens = len(matches)
    g.has_named_ranges = g.distinct_named_tokens >= 5

    # Sensitivity table heuristic: a table mentioning S_T or scenario rates
    table_count = count_md_tables(text)
    g.has_input_table = table_count >= 1
    g.has_sensitivity_table = (
        ("sensitivity" in lower or "s_t" in lower or "ending spot" in lower)
        and (table_count >= 1 or re.search(r"\b(0\.9\d|1\.0\d|1\.1\d|1\.2\d)\b", text))
    )

    # Color codes mentioned (yellow/blue/green/gray for input/assumption/formula/output)
    color_hits = sum(1 for c in ("yellow", "blue", "green", "gray", "grey") if c in lower)
    g.has_color_codes = color_hits >= 2

    # Verification mention
    g.has_verification = (
        "verification" in lower or "parity" in lower or "covered interest" in lower
    )

    # ---------- Scoring ----------
    # Interpretation /1.5: hedge breadth + insight.
    core_hits = len({"Forward", "MoneyMarket", "Put", "Call"} & set(hedges))
    has_outcomes = "hedge_outcomes" in sections
    if core_hits >= 4 and has_outcomes and g.word_count >= 600:
        g.score_interpretation = 1.5
    elif core_hits >= 3 and has_outcomes:
        g.score_interpretation = 1.25
    elif core_hits >= 2 or has_outcomes:
        g.score_interpretation = 1.0
    else:
        g.score_interpretation = 0.5

    # Sensitivity /1.0
    if g.has_sensitivity_table and "sensitivity" in sections:
        g.score_sensitivity = 1.0
    elif "sensitivity" in sections:
        g.score_sensitivity = 0.75
    else:
        g.score_sensitivity = 0.5

    # Recommendation + Justification /1.5
    has_rec = "recommendation" in sections
    has_just = "justification" in sections
    if has_rec and has_just:
        g.score_recommendation = 1.5
    elif has_rec or has_just:
        g.score_recommendation = 1.0
    else:
        g.score_recommendation = 0.5

    # Structured AI Prompt /2.0
    has_prompt_section = "ai_prompt" in sections
    prompt_signals = sum([
        has_prompt_section,
        g.has_named_ranges,
        g.has_color_codes,
        g.has_verification,
        g.distinct_named_tokens >= 8,
    ])
    if has_prompt_section and prompt_signals >= 4:
        g.score_prompt = 2.0
    elif has_prompt_section and prompt_signals >= 3:
        g.score_prompt = 1.75
    elif has_prompt_section and prompt_signals >= 2:
        g.score_prompt = 1.5
    elif has_prompt_section:
        g.score_prompt = 1.25
    else:
        g.score_prompt = 0.75

    g.raw_total = (
        g.score_interpretation
        + g.score_sensitivity
        + g.score_recommendation
        + g.score_prompt
    )

    if g.github_url:
        g.github_bump = GITHUB_BUMP
    else:
        g.github_deduction = GITHUB_PENALTY
        g.flags.append("NO_GITHUB_LINK")

    if g.fetch_status == "fetch_failed":
        g.flags.append("REPO_404")
    if g.word_count < 400:
        g.flags.append("SHORT")
    if core_hits < 4:
        g.flags.append(f"ONLY_{core_hits}_HEDGES")
    if not g.has_named_ranges:
        g.flags.append("FEW_NAMED_RANGES")
    if not g.has_color_codes:
        g.flags.append("NO_COLOR_CODES")
    if not g.has_verification:
        g.flags.append("NO_VERIFICATION")

    return g


def build_workbook(grades: list[Grade]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"
    headers = [
        "Student ID", "Student Name", "Source", "Format",
        "Word Count", "Sections", "Hedges", "Named Tokens",
        "Sens Tbl", "Color Codes", "Verification",
        "Interp /1.5", "Sens /1.0", "Rec+Just /1.5", "Prompt /2.0",
        "Raw /6", "Curved /6 (85% floor)",
        "GH Bump (+0.25)", "GH Deduction (-0.5)", "Final /6 (cap 6)",
        "GitHub URL", "Local File", "Flags", "Notes",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    nogh_fill = PatternFill("solid", fgColor="F8CBAD")
    floor_fill = PatternFill("solid", fgColor="E2EFDA")

    floor_pts = round(TOTAL_POINTS * FLOOR_PCT, 4)

    for g in grades:
        # All students with a submission folder get the floor. After the
        # floor, students who provided a GitHub link get a +0.25 bump
        # (capped at 6.0); students who didn't get a -0.5 deduction. The
        # bump/deduction applies after the floor so it always lands.
        curved_pre = round(max(g.raw_total, floor_pts), 2)
        final = round(min(
            curved_pre + g.github_bump - g.github_deduction,
            TOTAL_POINTS,
        ), 2)
        final_after_ded = final
        curved = final
        row = [
            g.sid, g.name, g.file_format, g.file_format.split("/")[-1] if "/" in g.file_format else "",
            g.word_count,
            ", ".join(g.sections_found),
            ", ".join(g.hedges_found),
            g.distinct_named_tokens,
            "Y" if g.has_sensitivity_table else "",
            "Y" if g.has_color_codes else "",
            "Y" if g.has_verification else "",
            g.score_interpretation,
            g.score_sensitivity,
            g.score_recommendation,
            g.score_prompt,
            round(g.raw_total, 2),
            curved_pre,
            g.github_bump,
            g.github_deduction,
            curved,
            g.github_url,
            g.local_file,
            ", ".join(g.flags),
            g.notes,
        ]
        ws.append(row)
        r = ws.max_row
        if "NO_GITHUB_LINK" in g.flags:
            ws.cell(row=r, column=headers.index("GH Deduction (-0.5)") + 1).fill = nogh_fill
        if g.flags:
            ws.cell(row=r, column=headers.index("Flags") + 1).fill = flag_fill
        if curved_pre == floor_pts and g.raw_total < floor_pts:
            ws.cell(row=r, column=headers.index("Curved /6 (85% floor)") + 1).fill = floor_fill

    # Update widths for new column count
    pass

    widths = [9, 28, 14, 10, 10, 60, 30, 10, 8, 10, 10,
              10, 10, 12, 11, 9, 14, 11, 13, 14,
              60, 60, 30, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Summary sheet
    sm = wb.create_sheet("Summary")
    sm.append(["Stage 4 Grading Summary"])
    sm.append([])
    sm.append(["Total submissions", len(grades)])
    sm.append(["Max points", TOTAL_POINTS])
    sm.append(["Floor (85% of max)", floor_pts])
    sm.append(["No GitHub link (-0.5 each)",
               sum(1 for g in grades if "NO_GITHUB_LINK" in g.flags)])
    sm.append(["Repo 404 (could not verify work)",
               sum(1 for g in grades if "REPO_404" in g.flags)])
    sm.append(["Avg raw /6",
               round(sum(g.raw_total for g in grades) / max(1, len(grades)), 2)])
    sm.append([])
    sm.append(["Curve policy",
               "Curved = MAX(Raw - GitHubDeduction, 5.10) when text was readable"])
    sm.append(["GitHub deduction",
               "0.5 deducted from raw before curve floor is applied"])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


def main() -> int:
    if not MANIFEST.exists():
        print(f"Run fetch_stage4.py first; missing {MANIFEST}")
        return 1
    grades: list[Grade] = []
    with MANIFEST.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            g = inspect(row)
            grades.append(g)
            print(f"  {g.sid:>5} {g.name:<28} raw={g.raw_total:.2f} "
                  f"+{g.github_bump:.2f} -{g.github_deduction:.1f} "
                  f"flags={','.join(g.flags) or '-'}")
    grades.sort(key=lambda g: g.name.lower())
    build_workbook(grades)
    return 0


if __name__ == "__main__":
    sys.exit(main())
