"""Build BUS-629 team roster in Vietnamese: Team, Name. Optional note in col C."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# (Team, Vietnamese Name, Note)
ROWS = [
    (1, "Markeiz Ryan",            ""),
    (1, "Trương Hoàng Phúc",       ""),
    (1, "Nguyễn Bùi Ngọc Linh",    ""),
    (1, "Nguyễn Mạnh Thái",        ""),
    (1, "Trần Trí Hiệp",           ""),
    (1, "Phan Thị Xuân Mai",       ""),
    (2, "Nguyễn Đức Duy",          ""),
    (2, "Hà Lê Bích Thuỳ",         ""),
    (2, "Cầm Thị Thu Hiền",        ""),
    (2, "Hà Tuấn Nghiệp",          ""),
    (2, "Chu Cảnh Chiêu",          ""),
    (2, "Phạm Thông",              ""),
    (3, "Taylor Blasco",           ""),
    (3, "Lương Duy Phương",        ""),
    (3, "Phạm Nguyễn Thanh Trúc",  ""),
    (3, "Đặng Thị Hải",            ""),
    (3, "Cao Bội Ngọc",            ""),
    (3, "Hoàng Thị Anh Minh",      "not enrolled"),
    (4, "Phan Trần Hoài Khanh",    ""),
    (4, "Trần Thanh Tân",          ""),
    (4, "Nguyễn Phượng Anh",       ""),
    (4, "Nguyễn Khổng Thanh Thảo", ""),
    (4, "Nguyễn Lan Anh",          ""),
]

wb = Workbook()
ws = wb.active
ws.title = "Danh sách Nhóm"

uh_green = "024731"
header_font = Font(name="Open Sans", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color=uh_green)
body_font = Font(name="Open Sans", size=11)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["Nhóm", "Họ và Tên", "Ghi chú"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for i, (team, name, note) in enumerate(ROWS, start=2):
    ws.cell(row=i, column=1, value=team).alignment = center
    ws.cell(row=i, column=2, value=name).alignment = left
    ws.cell(row=i, column=3, value=note).alignment = left
    for col in range(1, 4):
        cell = ws.cell(row=i, column=col)
        cell.font = body_font
        cell.border = border

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 24
ws.row_dimensions[1].height = 22

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:C{len(ROWS)+1}"

out = r"C:\GitHub\shidler\courses\International-Corporate-Finance\BUS-629-VEMBA\BUS-629 Danh sách Nhóm.xlsx"
wb.save(out)
print(f"Saved: {out}")
