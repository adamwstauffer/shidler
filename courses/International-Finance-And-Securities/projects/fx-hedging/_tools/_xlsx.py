"""Shared workbook inspection for the FIN-321 fx-hedging graders.

Used by Stage 3 (AI build + audit) and Stage 4 (live-data population, which
re-audits the same workbook). The headline check is mechanical and central to
v2: **every calculated cell must be a formula referencing named ranges** — a
pasted number where a formula belongs earns nothing.

`audit_workbook(bytes)` loads the workbook with formulas visible
(data_only=False) and reports:
  - which of the 10 contract named ranges are attached,
  - the formula ratio over non-input numeric cells on calculation tabs,
  - how many formulas actually reference a named range,
  - hedge-family presence, sensitivity + chart, cover/legend/notes tabs, and
    the count of distinct fill colors (evidence of the color convention).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import load_workbook

NAMED_RANGE_CONTRACT = [
    "FC_AMT", "S0_in", "F0_in", "R_USD", "R_FC",
    "K_PUT", "K_CALL", "PREM_PUT", "PREM_CALL", "T_DAYS",
]

# Tabs where a bare number is a legitimate INPUT, not a missing formula.
INPUT_TAB_HINTS = ("input", "cover", "legend", "key", "note", "assumption",
                   "readme", "instruction", "provenance", "source")

HEDGE_KEYWORDS = {
    "Forward": ["forward hedge", "forward rate", "forward contract", "locked-in", "locked in", "f0_in"],
    "MoneyMarket": ["money market", "money-market", "mm hedge", "borrow", "synthetic forward", "covered interest"],
    "Put": ["put option", "put hedge", "put premium", "k_put", "put strike", "prem_put"],
    "Call": ["call option", "call hedge", "call premium", "k_call", "call strike", "prem_call"],
    "Sensitivity": ["sensitivity", "±5", "+/-5", "scenario", "s_t", "ending spot", "s_t_grid"],
}

FINDING_LINE_RE = re.compile(r"^\s*(?:\d+\.|[-*])\s+\S")
FINDING_VERB_RE = re.compile(
    r"\b(found|fixed|checked|corrected|confirmed|verified|caught|missing|hardcoded|wrong)\b",
    re.IGNORECASE,
)


@dataclass
class WorkbookAudit:
    opened: bool = False
    error: str = ""
    named_ranges_present: list[str] = field(default_factory=list)
    formula_cells: int = 0
    hardcoded_numeric_cells: int = 0
    formulas_using_named_ranges: int = 0
    hedges_found: list[str] = field(default_factory=list)
    sensitivity_detected: bool = False
    chart_count: int = 0
    has_cover_tab: bool = False
    has_legend_tab: bool = False
    distinct_fill_colors: int = 0
    has_notes_tab: bool = False

    @property
    def formula_ratio(self) -> float:
        denom = self.formula_cells + self.hardcoded_numeric_cells
        return self.formula_cells / denom if denom else 0.0

    @property
    def has_option_hedge(self) -> bool:
        return bool({"Put", "Call"} & set(self.hedges_found))


def _input_cell_coords(wb) -> set[tuple[str, str]]:
    """Cells targeted by the contract named ranges — allowed to be constants."""
    coords: set[tuple[str, str]] = set()
    for name in wb.defined_names:
        if name not in NAMED_RANGE_CONTRACT:
            continue
        try:
            for title, coord in wb.defined_names[name].destinations:
                coords.add((title, coord.replace("$", "")))
        except Exception:
            pass
    return coords


def audit_workbook(xbytes: bytes) -> WorkbookAudit:
    a = WorkbookAudit()
    try:
        wb = load_workbook(BytesIO(xbytes), data_only=False)
    except Exception as e:  # corrupt / not a workbook
        a.error = f"open failed: {type(e).__name__}: {e}"
        return a
    a.opened = True

    names = list(wb.defined_names)
    a.named_ranges_present = [n for n in NAMED_RANGE_CONTRACT if n in names]
    input_coords = _input_cell_coords(wb)

    lowered = {s: s.lower() for s in wb.sheetnames}
    a.has_cover_tab = any("cover" in s for s in lowered.values())
    a.has_legend_tab = any(("legend" in s or "key" in s) for s in lowered.values())
    a.has_notes_tab = any(("note" in s or "assumption" in s) for s in lowered.values())

    text_blob = ""
    fills: set[str] = set()
    for sname in wb.sheetnames:
        ws = wb[sname]
        a.chart_count += len(getattr(ws, "_charts", []))
        is_input_tab = any(h in lowered[sname] for h in INPUT_TAB_HINTS)
        max_r = min(ws.max_row or 1, 1500)
        max_c = min(ws.max_column or 1, 50)
        for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    text_blob += v.lower() + " "
                try:
                    fill = cell.fill
                    if fill and fill.fgColor and fill.fgColor.rgb:
                        rgb = str(fill.fgColor.rgb)
                        if rgb not in ("00000000", "FFFFFFFF"):
                            fills.add(rgb)
                except Exception:
                    pass
                if is_input_tab:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    a.formula_cells += 1
                    if any(n in v for n in NAMED_RANGE_CONTRACT):
                        a.formulas_using_named_ranges += 1
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    if (sname, cell.coordinate) not in input_coords:
                        a.hardcoded_numeric_cells += 1

    a.distinct_fill_colors = len(fills)
    a.hedges_found = [k for k, kws in HEDGE_KEYWORDS.items()
                      if k != "Sensitivity" and any(w in text_blob for w in kws)]
    a.sensitivity_detected = (
        any(w in text_blob for w in HEDGE_KEYWORDS["Sensitivity"]) or a.chart_count > 0
    )
    return a


def count_audit_findings(text: str) -> int:
    """Approximate substantive-finding count in a build-audit / validation note."""
    if not text:
        return 0
    list_items = sum(1 for ln in text.splitlines() if FINDING_LINE_RE.match(ln))
    verbs = len(FINDING_VERB_RE.findall(text))
    return max(min(list_items, verbs), list_items if verbs else 0)
