"""FIN-321 (fx-hedging v2) Stage 3 grader — AI-Assisted Build + Audit.

Stage 3's deliverables live in the student's GitHub repo:
  - workbook:  models/builds/YYYY-MM-DD-{lastname}-{scenario}-model.xlsx
  - audit note: analysis/YYYY-MM-DD-{lastname}-build-audit.md

The graded skill is *specifying precisely and auditing ruthlessly* — so the
headline check here is mechanical and new to v2: **every calculated cell must
be a formula referencing named ranges.** A pasted number where a formula
belongs earns nothing for that element. We approximate this with a formula
ratio over non-input numeric cells in the calculation tabs.

Rubric (criterion weights = % of the stage, from _weights.CRITERIA_WEIGHTS[3]):
    Contract compliance        50   (named ranges + formulas-only + hedges + sensitivity)
    Structure & presentation   25   (cover, legend/key, color convention, layout)
    Audit note                 25   (>=3 substantive findings, committed)

Scores are on a 0-100 "% of the stage" scale; the stage's project weight (17%)
is applied later in the gradebook. Generosity-only curve floor from _curve.

Outputs (all under <stage3>/graded/, mirroring BUS-629):
  - STAGE3_GRADES.md                           internal — HAS scores
  - _pr_feedback/{lastname}/feedback-file.md   score-free, PR-ready

CLI:
    python grade_stage3.py <export.zip|dir> [--floor N] [--prior-stage2 PATH]
                           [--out-dir DIR] [--today YYYY-MM-DD] [--no-move]

Score privacy (Adam's policy): score numbers live ONLY in STAGE3_GRADES.md and
instructor email — never in the PR feedback pushed to a student's public repo.
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

import _repo
from _repo import RepoState, Submission
from _weights import CRITERIA_WEIGHTS, STAGE_FLOOR_PCT, stage_pct
from _curve import curved_score, floor_applied
from _grading_comments import core, backward, next_stage_pointer, render_suggestions

STAGE_N = 3
STAGE_LABEL = "AI-Assisted Build + Audit"
CRIT = CRITERIA_WEIGHTS[STAGE_N]           # {'contract_compliance':50, ...}
DEFAULT_FLOOR_PCT = STAGE_FLOOR_PCT[STAGE_N]

NAMED_RANGE_CONTRACT = [
    "FC_AMT", "S0_in", "F0_in", "R_USD", "R_FC",
    "K_PUT", "K_CALL", "PREM_PUT", "PREM_CALL", "T_DAYS",
]

# Tabs where a bare number is a legitimate INPUT, not a missing formula.
INPUT_TAB_HINTS = ("input", "cover", "legend", "key", "note", "assumption",
                   "readme", "instruction", "provenance", "source")

HEDGE_KEYWORDS = {
    "Forward": ["forward hedge", "forward rate", "forward contract", "locked-in", "locked in", "f0_in"],
    "MoneyMarket": ["money market", "money-market", "mm hedge", "borrow", "synthetic forward", "covered interest"],
    "Put": ["put option", "put hedge", "put premium", "k_put", "put strike", "prem_put"],
    "Call": ["call option", "call hedge", "call premium", "k_call", "call strike", "prem_call"],
    "Sensitivity": ["sensitivity", "±5", "+/-5", "scenario", "s_t", "ending spot", "s_t_grid"],
}

WORKBOOK_RE = re.compile(r"models/builds/.*\.xlsx?$", re.IGNORECASE)
AUDIT_NOTE_RE = re.compile(r"analysis/.*(?:audit|build-audit).*\.md$", re.IGNORECASE)
FINDING_LINE_RE = re.compile(r"^\s*(?:\d+\.|[-*])\s+\S")
FINDING_VERB_RE = re.compile(
    r"\b(found|fixed|checked|corrected|confirmed|verified|caught|missing|hardcoded|wrong)\b",
    re.IGNORECASE,
)

LETTER_SCALE = [
    ("A", 93), ("A-", 90), ("B+", 87), ("B", 83), ("B-", 80),
    ("C+", 77), ("C", 73), ("C-", 70), ("D+", 67), ("D", 63), ("D-", 60), ("F", 0),
]


def letter(pct: float) -> str:
    for name, lo in LETTER_SCALE:
        if pct >= lo:
            return name
    return "F"


# ---------------------------------------------------------------- workbook audit
@dataclass
class WorkbookAudit:
    opened: bool = False
    error: str = ""
    named_ranges_present: list[str] = field(default_factory=list)
    formula_cells: int = 0
    hardcoded_numeric_cells: int = 0
    formulas_using_named_ranges: int = 0
    hedges_found: list[str] = field(default_factory=list)
    sensitivity_detected: bool = False
    chart_count: int = 0
    has_cover_tab: bool = False
    has_legend_tab: bool = False
    distinct_fill_colors: int = 0
    has_notes_tab: bool = False

    @property
    def formula_ratio(self) -> float:
        denom = self.formula_cells + self.hardcoded_numeric_cells
        return self.formula_cells / denom if denom else 0.0


def _input_cell_coords(wb) -> set[tuple[str, str]]:
    """Cells targeted by the contract named ranges — allowed to be constants."""
    coords: set[tuple[str, str]] = set()
    for name in wb.defined_names:
        if name not in NAMED_RANGE_CONTRACT:
            continue
        dn = wb.defined_names[name]
        try:
            for title, coord in dn.destinations:
                coords.add((title, coord.replace("$", "")))
        except Exception:
            pass
    return coords


def audit_workbook(xbytes: bytes) -> WorkbookAudit:
    a = WorkbookAudit()
    try:
        wb = load_workbook(BytesIO(xbytes), data_only=False)
    except Exception as e:  # corrupt / not a workbook
        a.error = f"open failed: {type(e).__name__}: {e}"
        return a
    a.opened = True

    names = list(wb.defined_names)
    a.named_ranges_present = [n for n in NAMED_RANGE_CONTRACT if n in names]
    input_coords = _input_cell_coords(wb)

    lowered_sheets = {s: s.lower() for s in wb.sheetnames}
    a.has_cover_tab = any("cover" in s for s in lowered_sheets.values())
    a.has_legend_tab = any(("legend" in s or "key" in s) for s in lowered_sheets.values())
    a.has_notes_tab = any(("note" in s or "assumption" in s) for s in lowered_sheets.values())

    text_blob = ""
    fills: set[str] = set()
    for sname in wb.sheetnames:
        ws = wb[sname]
        a.chart_count += len(getattr(ws, "_charts", []))
        is_input_tab = any(h in lowered_sheets[sname] for h in INPUT_TAB_HINTS)
        max_r = min(ws.max_row or 1, 1500)
        max_c = min(ws.max_column or 1, 50)
        for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    text_blob += v.lower() + " "
                try:
                    fill = cell.fill
                    if fill and fill.fgColor and fill.fgColor.rgb:
                        rgb = str(fill.fgColor.rgb)
                        if rgb not in ("00000000", "FFFFFFFF"):
                            fills.add(rgb)
                except Exception:
                    pass
                if is_input_tab:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    a.formula_cells += 1
                    if any(n in v for n in NAMED_RANGE_CONTRACT):
                        a.formulas_using_named_ranges += 1
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    if (sname, cell.coordinate) not in input_coords:
                        a.hardcoded_numeric_cells += 1

    a.distinct_fill_colors = len(fills)
    a.hedges_found = [k for k, kws in HEDGE_KEYWORDS.items()
                      if k != "Sensitivity" and any(w in text_blob for w in kws)]
    a.sensitivity_detected = (
        any(w in text_blob for w in HEDGE_KEYWORDS["Sensitivity"]) or a.chart_count > 0
    )
    return a


def count_audit_findings(text: str) -> int:
    """Approximate substantive-finding count in the build-audit note."""
    if not text:
        return 0
    list_items = sum(1 for ln in text.splitlines() if FINDING_LINE_RE.match(ln))
    verbs = len(FINDING_VERB_RE.findall(text))
    return max(min(list_items, verbs), list_items if verbs else 0)


# ---------------------------------------------------------------- grading
@dataclass
class Grade:
    sid: str
    name: str
    submitted_at: datetime | None
    github_url: str
    repo: RepoState | None = None
    workbook_path: str = ""
    audit_path: str = ""
    audit: WorkbookAudit | None = None
    note_present: bool = False
    note_findings: int = 0
    cc: float = 0.0            # contract-compliance points earned (/50)
    sp: float = 0.0            # structure points earned (/25)
    an: float = 0.0            # audit-note points earned (/25)
    prior_weak: bool = False
    flags: list[str] = field(default_factory=list)
    error: str = ""

    @property
    def raw_pct(self) -> float:
        return round(self.cc + self.sp + self.an, 1)


def _score_contract(a: WorkbookAudit, g: Grade) -> float:
    nr_frac = len(a.named_ranges_present) / len(NAMED_RANGE_CONTRACT)
    if len(a.named_ranges_present) < 8:
        g.flags.append("FEW_NAMED_RANGES")
    formula_score = min(1.0, a.formula_ratio / 0.9) if a.formula_ratio else 0.0
    if a.formula_cells and a.formulas_using_named_ranges == 0:
        formula_score = min(formula_score, 0.5)
        g.flags.append("FORMULAS_NOT_USING_NAMES")
    if a.formula_ratio < 0.8:
        g.flags.append("HARDCODED_OUTPUTS")
    have = set(a.hedges_found)
    hedge_frac = (
        0.5 * (len({"Forward", "MoneyMarket"} & have) / 2)
        + 0.5 * (1.0 if ({"Put", "Call"} & have) else 0.0)
    )
    for h in ("Forward", "MoneyMarket"):
        if h not in have:
            g.flags.append(f"MISSING_{h.upper()}")
    if not ({"Put", "Call"} & have):
        g.flags.append("MISSING_OPTION")
    sens = 1.0 if (a.sensitivity_detected and a.chart_count > 0) else (
        0.5 if a.sensitivity_detected else 0.0)
    if not a.sensitivity_detected:
        g.flags.append("NO_SENSITIVITY")
    if a.chart_count == 0:
        g.flags.append("NO_CHART")
    frac = 0.40 * nr_frac + 0.35 * formula_score + 0.15 * hedge_frac + 0.10 * sens
    return round(frac * CRIT["contract_compliance"], 1)


def _score_structure(a: WorkbookAudit, g: Grade) -> float:
    if not a.has_cover_tab:
        g.flags.append("NO_COVER")
    if not a.has_legend_tab:
        g.flags.append("NO_LEGEND")
    color_ok = a.distinct_fill_colors >= 3
    if not color_ok:
        g.flags.append("WEAK_COLOR_CONVENTION")
    frac = (0.35 * a.has_cover_tab + 0.35 * a.has_legend_tab
            + 0.20 * color_ok + 0.10 * a.has_notes_tab)
    return round(frac * CRIT["structure_presentation"], 1)


def _score_audit_note(g: Grade) -> float:
    if not g.note_present:
        g.flags.append("AUDIT_NOTE_MISSING")
        return 0.0
    if g.note_findings >= 3:
        return float(CRIT["audit_note"])
    g.flags.append("AUDIT_NOTE_THIN")
    return round(0.5 * CRIT["audit_note"], 1)


def grade_submission(sub: Submission) -> Grade:
    g = Grade(sub.student_id, sub.name, sub.submitted_at, sub.github_url)
    parsed = sub.repo
    if not parsed:
        g.flags.append("NO_GITHUB_LINK")
        g.error = "no github url in submission"
        return g
    owner, repo = parsed
    g.repo = _repo.repo_state(owner, repo)
    if not g.repo.accessible:
        g.flags.append("REPO_404")
        g.error = "repo not accessible via gh"
        return g
    if not g.repo.public:
        g.flags.append("NOT_PUBLIC")
    if not g.repo.instructor_is_collaborator:
        g.flags.append("INSTRUCTOR_NOT_COLLABORATOR")

    branch = g.repo.default_branch
    wb_paths = [p for p in g.repo.tree if WORKBOOK_RE.search(p)]
    note_paths = [p for p in g.repo.tree if AUDIT_NOTE_RE.search(p)]

    if wb_paths:
        g.workbook_path = sorted(wb_paths, key=len)[0]
        raw = _repo.download_bytes(owner, repo, g.workbook_path, branch)
        g.audit = audit_workbook(raw) if raw else WorkbookAudit(error="download failed")
    else:
        g.flags.append("NO_WORKBOOK")
        g.audit = WorkbookAudit(error="no workbook in models/builds/")

    if note_paths:
        g.audit_path = sorted(note_paths, key=len)[0]
        text = _repo.download_text(owner, repo, g.audit_path, branch) or ""
        g.note_present = bool(text.strip())
        g.note_findings = count_audit_findings(text)

    a = g.audit
    if a and a.opened:
        g.cc = _score_contract(a, g)
        g.sp = _score_structure(a, g)
    elif a and a.error.startswith("open failed"):
        g.flags.append("WORKBOOK_OPEN_FAILED")
    g.an = _score_audit_note(g)

    if g.raw_pct >= 92 and not g.flags:
        g.flags.append("STRONG")
    return g


# ---------------------------------------------------------------- suggestions
def _suggestions_for(g: Grade):
    s = []
    f = set(g.flags)
    if "NO_WORKBOOK" in f:
        s.append(core("No workbook found under `models/builds/`. Commit the generated "
                      "`.xlsx` there with the convention-compliant filename."))
    if "WORKBOOK_OPEN_FAILED" in f:
        s.append(core("The workbook couldn't be opened — re-export it as a valid .xlsx "
                      "and re-commit."))
    if "FEW_NAMED_RANGES" in f:
        s.append(core("Fewer than 8 of the 10 contract named ranges were found. Attach "
                      "every one (FC_AMT, S0_in, F0_in, R_USD, R_FC, K_PUT, K_CALL, "
                      "PREM_PUT, PREM_CALL, T_DAYS) to its input cell."))
    if "HARDCODED_OUTPUTS" in f:
        s.append(core("Several calculated cells hold typed numbers rather than formulas. "
                      "Every output must be a formula referencing named ranges — pasted "
                      "results don't earn contract credit."))
    if "FORMULAS_NOT_USING_NAMES" in f:
        s.append(core("Your formulas use cell addresses (e.g. `$F$7`) instead of named "
                      "ranges. Swap in the contract names so spec, workbook, and prompt "
                      "share one vocabulary."))
    for h, msg in (("MISSING_FORWARD", "forward hedge"),
                   ("MISSING_MONEYMARKET", "money-market hedge"),
                   ("MISSING_OPTION", "option (put/call) hedge")):
        if h in f:
            s.append(core(f"The {msg} isn't detectable — make sure it's built and labelled."))
    if "NO_CHART" in f:
        s.append(core("Add the sensitivity line chart (USD outcome vs. ending spot, one "
                      "series per strategy)."))
    if "NO_LEGEND" in f:
        s.append(core("Add a Legend/Key tab documenting the color convention (yellow "
                      "inputs, blue assumptions, green formulas, gray outputs)."))
    if "AUDIT_NOTE_MISSING" in f:
        s.append(core("The build-audit note is missing. Add `analysis/…-build-audit.md` "
                      "with at least three findings you checked or fixed."))
    if "AUDIT_NOTE_THIN" in f:
        s.append(core("The audit note has fewer than three substantive findings. For "
                      "each: what you checked, what you found, what you did."))
    if "INSTRUCTOR_NOT_COLLABORATOR" in f:
        s.append(core("I'm not a collaborator on the repo yet — add `adamwstauffer` so I "
                      "can leave inline review comments."))
    if "STRONG" in f:
        s.append(core("Clean build — named ranges complete, outputs formula-driven, hedges "
                      "and sensitivity all present. Nicely audited."))
    if g.prior_weak:
        s.append(backward("Your Stage 2 spec scored below the floor — the clearer the "
                          "spec's named-range contract and calculation flow, the less the "
                          "build has to guess. Tightening it can still lift the Stage 2 "
                          "score at the revision sweep."))
    nxt = next_stage_pointer(STAGE_N)
    if nxt:
        s.append(nxt)
    return s


# ---------------------------------------------------------------- prior lookup
_PRIOR_HEADER_RE = re.compile(
    r"^##\s+\d+\.\s+(?P<name>.+?)\s+—\s+.*?\*\*(?P<score>\d+(?:\.\d+)?)\s*/\s*100\*\*",
    re.MULTILINE,
)


def parse_prior_report(path: Path | None) -> dict[str, float]:
    """Map normalized student name -> Stage 2 final score (0-100)."""
    out: dict[str, float] = {}
    if not path or not Path(path).exists():
        return out
    text = Path(path).read_text(encoding="utf-8", errors="ignore")
    for m in _PRIOR_HEADER_RE.finditer(text):
        out[_repo.normalize_name(m.group("name"))] = float(m.group("score"))
    return out


# ---------------------------------------------------------------- report writers
def _criterion_rows(g: Grade) -> list[tuple[str, str, str]]:
    a = g.audit or WorkbookAudit()
    return [
        ("Contract compliance", f"{g.cc:g} / {CRIT['contract_compliance']}",
         f"{len(a.named_ranges_present)}/10 named ranges; "
         f"{a.formula_ratio * 100:.0f}% of calc cells are formulas; "
         f"hedges: {', '.join(a.hedges_found) or 'none'}; "
         f"{'chart + ' if a.chart_count else 'no chart, '}sensitivity "
         f"{'yes' if a.sensitivity_detected else 'no'}."),
        ("Structure & presentation", f"{g.sp:g} / {CRIT['structure_presentation']}",
         f"cover {'Y' if a.has_cover_tab else 'N'}, legend {'Y' if a.has_legend_tab else 'N'}, "
         f"{a.distinct_fill_colors} fill colors, notes {'Y' if a.has_notes_tab else 'N'}."),
        ("Audit note", f"{g.an:g} / {CRIT['audit_note']}",
         f"{'present' if g.note_present else 'missing'}, ~{g.note_findings} findings."),
    ]


def _student_section(n: int, g: Grade, floor_pct: int) -> list[str]:
    accessible = bool(g.repo and g.repo.accessible)
    final = curved_score(g.raw_pct, STAGE_N, accessible=accessible)
    floored = floor_applied(g.raw_pct, STAGE_N, accessible=accessible)

    if final == 0:
        tag = " (no gradable submission)"
    elif floored:
        tag = f" ({letter(final)}, floor applied)"
    else:
        tag = f" ({letter(final)})"

    lines = [f"## {n}. {g.name} — **{final:g} / 100**{tag}", ""]
    if g.github_url:
        lines.append(f"**Repo:** {g.github_url}")
    if g.workbook_path:
        lines.append(f"**Workbook:** `{g.workbook_path}`")
    if g.audit_path:
        lines.append(f"**Audit note:** `{g.audit_path}`")
    if g.submitted_at:
        lines.append(f"**Submitted:** {g.submitted_at:%Y-%m-%d %H:%M}")
    lines.append("")

    lines.append("| Criterion | Earned | Notes |")
    lines.append("|-----------|--------|-------|")
    for label, earned, note in _criterion_rows(g):
        lines.append(f"| {label} | {earned} | {note} |")
    if floored:
        lines.append(f"| **Raw total** | **{g.raw_pct:g} / 100** | — |")
        lines.append(f"| **Floor adjustment** | **+{final - g.raw_pct:g}** | lifted to {floor_pct}% floor |")
    final_note = ("no gradable submission" if final == 0
                  else "floor applied" if floored else "earned on merit")
    lines.append(f"| **Final** | **{final:g} / 100** | {final_note} |")
    lines.append("")
    if g.flags:
        lines.append(f"*Flags: {', '.join(g.flags)}*")
        lines.append("")
    lines.extend(render_suggestions(_suggestions_for(g), stage_n=STAGE_N))
    lines.append("---")
    return lines


def build_report(grades: list[Grade], floor_pct: int, today: date) -> str:
    lines = [
        "# FIN-321 Stage 3 — Grade Report",
        "",
        f"**Stage:** {STAGE_LABEL} ({stage_pct(STAGE_N)}% of project score)",
        f"**Graded:** {today:%Y-%m-%d}",
        f"**Submissions reviewed:** {len(grades)}",
        f"**Floor policy:** {floor_pct}% floor for any accessible repo with a workbook present.",
        "**Score privacy:** scores live in this internal file only — never in the "
        "PR feedback pushed to student repos.",
        "",
        "---",
        "## Rubric (recap)",
        "",
        "| Criterion | Weight |",
        "|-----------|--------|",
        f"| Contract compliance (named ranges, formulas-only, hedges, sensitivity) | {CRIT['contract_compliance']}% |",
        f"| Structure & presentation (cover, legend/key, color convention) | {CRIT['structure_presentation']}% |",
        f"| Audit note (>=3 findings) | {CRIT['audit_note']}% |",
        "",
        "---",
    ]
    ordered = sorted(grades, key=lambda g: _repo.normalize_name(g.name))
    summary: list[tuple[str, float, str]] = []
    for i, g in enumerate(ordered, 1):
        lines.extend(_student_section(i, g, floor_pct))
        accessible = bool(g.repo and g.repo.accessible)
        final = curved_score(g.raw_pct, STAGE_N, accessible=accessible)
        note = "no submission" if final == 0 else (
            "floor applied" if floor_applied(g.raw_pct, STAGE_N, accessible=accessible)
            else "earned")
        summary.append((g.name, final, note))

    lines += ["## Class summary", "", "| Student | Score | Notes |", "|---------|-------|-------|"]
    for name, sc, note in summary:
        lines.append(f"| {name} | {sc:g} / 100 | {note} |")
    submitted = [s for _, s, n in summary if n != "no submission"]
    if submitted:
        floored_n = sum(1 for _, _, n in summary if n == "floor applied")
        lines += [
            "",
            f"**Mean (submissions only):** {sum(submitted) / len(submitted):.1f}",
            f"**Submission rate:** {len(submitted)} of {len(summary)}",
            f"**Floor applied:** {floored_n} of {len(submitted)} submissions",
        ]
    lines.append("")
    return "\n".join(lines)


def build_pr_feedback(g: Grade, today: date) -> str:
    a = g.audit or WorkbookAudit()
    opt = "✓" if ({"Put", "Call"} & set(a.hedges_found)) else "—"
    lines = [
        f"# Stage 3 review — {today:%Y-%m-%d}",
        "",
        "## Build contract checklist",
        "",
        "| Check | Status |",
        "|-------|--------|",
        f"| Workbook committed (`models/builds/`) | {'✓' if g.workbook_path else '—'} |",
        f"| Named ranges attached | {len(a.named_ranges_present)}/10 |",
        f"| Calculated cells are formulas | {a.formula_ratio * 100:.0f}% |",
        f"| Formulas use named ranges | {'✓' if a.formulas_using_named_ranges else '—'} |",
        f"| Forward hedge | {'✓' if 'Forward' in a.hedges_found else '—'} |",
        f"| Money-market hedge | {'✓' if 'MoneyMarket' in a.hedges_found else '—'} |",
        f"| Option hedge (put/call) | {opt} |",
        f"| Sensitivity table | {'✓' if a.sensitivity_detected else '—'} |",
        f"| Sensitivity chart | {'✓' if a.chart_count else '—'} |",
        "",
        "## Presentation",
        "",
        "| Element | Status |",
        "|---------|--------|",
        f"| Cover tab | {'✓' if a.has_cover_tab else '—'} |",
        f"| Legend/Key tab | {'✓' if a.has_legend_tab else '—'} |",
        f"| Color convention (distinct fills) | {a.distinct_fill_colors} |",
        f"| Notes/Assumptions tab | {'✓' if a.has_notes_tab else '—'} |",
        "",
        "## Audit note",
        "",
        f"- Present: **{'yes' if g.note_present else 'no'}**",
        f"- Substantive findings counted: **{g.note_findings}** (brief asks for ≥3)",
        "",
    ]
    lines.extend(render_suggestions(_suggestions_for(g), stage_n=STAGE_N))
    lines += [
        "",
        "*This review is feedback-only — no scores included. Score numbers live in the "
        "internal grade report and your instructor email.*",
    ]
    return "\n".join(lines)


# ---------------------------------------------------------------- driver
def _resolve_out_dir(export: Path, out_dir: str | None) -> Path:
    if out_dir:
        return Path(out_dir)
    if export.parent.name.lower() == "ungraded":
        return export.parent.parent / "graded"
    base = export.parent if export.is_file() else export
    return base / "graded"


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="FIN-321 Stage 3 grader (AI build + audit).")
    ap.add_argument("export", type=Path, help="Lamaku export .zip or extracted dir")
    ap.add_argument("--floor", type=int, default=DEFAULT_FLOOR_PCT)
    ap.add_argument("--prior-stage2", type=Path, default=None,
                    help="STAGE2_GRADES.md for carry-forward recognition")
    ap.add_argument("--out-dir", default=None)
    ap.add_argument("--today", default=None, help="YYYY-MM-DD (defaults to today)")
    ap.add_argument("--no-move", action="store_true",
                    help="(reserved) keep the source export in place")
    args = ap.parse_args(argv)

    if not args.export.exists():
        print(f"error: export not found: {args.export}", file=sys.stderr)
        return 1
    today = (datetime.strptime(args.today, "%Y-%m-%d").date()
             if args.today else datetime.now().date())

    subs = _repo.discover_submissions(args.export)
    print(f"Discovered {len(subs)} submissions.")
    prior = parse_prior_report(args.prior_stage2)

    grades: list[Grade] = []
    for sub in subs:
        print(f"  grading {sub.student_id} {sub.name} ...", end=" ", flush=True)
        g = grade_submission(sub)
        key = _repo.normalize_name(g.name)
        g.prior_weak = key in prior and prior[key] < STAGE_FLOOR_PCT[2]
        grades.append(g)
        print(f"raw={g.raw_pct:g}/100 flags={','.join(g.flags) or '-'}")

    out_dir = _resolve_out_dir(args.export, args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / "STAGE3_GRADES.md"
    report_path.write_text(build_report(grades, args.floor, today), encoding="utf-8")
    print(f"Wrote {report_path}")

    fb_root = out_dir / "_pr_feedback"
    for g in grades:
        d = fb_root / _repo.lastname_slug(g.name)
        d.mkdir(parents=True, exist_ok=True)
        (d / "feedback-file.md").write_text(build_pr_feedback(g, today), encoding="utf-8")
    print(f"Wrote {len(grades)} PR-feedback files under {fb_root}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
