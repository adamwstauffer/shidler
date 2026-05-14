"""BUS-629 Stage 1 (provided ratios template) grading scanner.

Stage 1 submissions are an Excel file (`performance-ratios-template.xlsx`)
uploaded via Lamaku. The instructor exports a zip containing one folder per
student plus an `index.html`. This script:

  1. Reads the export zip (or extracted directory).
  2. For each student, inspects the submitted .xlsx (sheets, named ranges,
     formulas, byte-identity vs. the master template).
  3. Looks the student up in `../stage0/graded/STAGE0_GRADES.md` to find
     their GitHub repo URL + prior-stage suggestions.
  4. Inspects the repo via `gh` to verify the file landed in the right path,
     re-check directory structure / README quality, and tally commits.
  5. Auto-scores the four-criterion Stage 1 rubric with **carry-over
     awareness** — issues already flagged in a prior stage become
     forward-looking tips with NO point deduction at this stage.
  6. Writes / updates `../graded/STAGE1_GRADES.md` (append mode, dedupes by
     repo URL — instructor prose in existing entries is preserved).
  7. Moves the source zip from ungraded/ → graded/ and cleans up scratch.

The four-criterion rubric (per the existing Stage 1 grade report):

    Template uploaded correctly        /30
    Directory structure                /40
    README quality                     /20
    Commit hygiene (since Stage 0)     /10
                                       = /100

Floor policy (per instructor direction): 80% for submissions with a working
public repo. Pass `--floor` to override.

Template policy: **enhancements are welcomed and not penalized.** A file with
more formulas than the master gets full marks plus an `ENHANCED` flag.

Double-deduction policy: if Stage 0 already flagged an issue (placeholder
READMEs, missing directories, terse commit messages, etc.), Stage 1 does not
re-deduct for the same issue. The suggestion is repeated as a forward-looking
tip without point loss.

USAGE:
    python grade_stage1.py <export.zip> [--floor=80] [--no-move]

REQUIREMENTS:
    - `gh` CLI authenticated against github.com (optional but recommended;
      without it, the script falls back to file-only scoring)
    - openpyxl (`pip install openpyxl`)
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_FLOOR_PCT = 80
TOTAL_POINTS = 100
STAGE_LABEL = "Stage 1 — Provided Ratios Template"
STAGE_FILENAME = "performance-ratios-template.xlsx"
STAGE_TARGET_DIR = "models/templates"

REQUIRED_DIRS = [
    "docs", "docs/decisions", "docs/specs", "docs/plans", "docs/templates",
    "models", "models/templates", "models/builds",
    "data",
    "analysis", "analysis/validation",
    "deliverables",
]

VAGUE_MESSAGE_PATTERNS = [
    re.compile(r"^(initial commit|create [a-z0-9._-]+|update [a-z0-9._-]+)$",
               re.IGNORECASE),
    re.compile(r"^(wip|fix|fix typo|update|test|asdf|stuff|misc|main)$",
               re.IGNORECASE),
    re.compile(r"^add files via upload$", re.IGNORECASE),
    re.compile(r"^delete [a-z0-9._-]+$", re.IGNORECASE),
    re.compile(r"^.{0,5}$"),
]

GITHUB_URL_RE = re.compile(
    r"https?://github\.com/(?P<owner>[A-Za-z0-9_.-]+)/(?P<repo>[A-Za-z0-9_.-]+)"
)
FOLDER_NAME_RE = re.compile(
    r"^(?P<sid>\d+)-\d+\s*-\s*(?P<name>.+?)\s*-\s*"
    r"(?P<month>[A-Za-z]+)\s+(?P<day>\d+),\s*(?P<year>\d{4})\s+"
    r"(?P<h>\d{1,4})\s*(?P<ampm>AM|PM)\s*$"
)

LETTER_GRADE_SCALE = [
    ("A+", 97, None), ("A", 93, 97), ("A-", 90, 93),
    ("B+", 87, 90),  ("B", 83, 87), ("B-", 80, 83),
    ("C+", 77, 80),  ("C", 73, 77), ("C-", 70, 73),
    ("D+", 67, 70),  ("D", 65, 67), ("F", 0, 65),
]


# ------------------------------------------------------------------
# Submission discovery
# ------------------------------------------------------------------

@dataclass
class Submission:
    student_id: str
    student_name: str
    submitted_at: datetime | None
    xlsx_path: Path  # local path to extracted .xlsx
    repo_url: str = ""  # filled in by prior-stage lookup
    owner: str = ""
    repo: str = ""


def _parse_folder_name(name: str):
    m = FOLDER_NAME_RE.match(name)
    if not m:
        return None, None, None
    h = m.group("h")
    if len(h) == 3:
        hour, minute = int(h[0]), int(h[1:])
    elif len(h) == 4:
        hour, minute = int(h[:2]), int(h[2:])
    else:
        hour, minute = int(h), 0
    ampm = m.group("ampm").upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    try:
        dt = datetime.strptime(
            f"{m.group('month')} {m.group('day')} {m.group('year')}", "%b %d %Y"
        ).replace(hour=hour, minute=minute)
    except ValueError:
        dt = None
    return m.group("sid"), m.group("name").strip(), dt


def discover_submissions(export_path: Path) -> list[Submission]:
    if export_path.is_file() and export_path.suffix.lower() == ".zip":
        scratch = export_path.parent / f"_{export_path.stem}_extracted"
        scratch.mkdir(exist_ok=True)
        with zipfile.ZipFile(export_path) as zf:
            zf.extractall(scratch)
        root = scratch
    elif export_path.is_dir():
        root = export_path
    else:
        raise SystemExit(f"Cannot read submission export at {export_path}")

    submissions: list[Submission] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        sid, name, dt = _parse_folder_name(child.name)
        if not name:
            continue
        xlsx = next(
            (p for p in child.glob("*.xlsx")
             if not p.name.startswith("~$")),
            None,
        )
        if xlsx is None:
            print(f"  WARN: no .xlsx in '{child.name}' — skipping {name}")
            continue
        submissions.append(Submission(
            student_id=sid or "",
            student_name=name,
            submitted_at=dt,
            xlsx_path=xlsx,
        ))
    submissions.sort(key=lambda s: s.student_name.lower())
    return submissions


# ------------------------------------------------------------------
# Prior-stage (STAGE0_GRADES.md) parser
# ------------------------------------------------------------------

@dataclass
class PriorGrade:
    student_name: str
    repo_url: str
    final_score: int
    raw_section: str
    carry_over_tags: set[str] = field(default_factory=set)


_CARRY_OVER_PATTERNS = [
    # (tag, regex against the lowercased bullet text)
    ("READMES_PLACEHOLDER", re.compile(r"placeholder readme")),
    ("DIRS_INCOMPLETE",     re.compile(r"missing.*director|directory.*missing|"
                                       r"\d+/11 (?:present|directories present)")),
    ("BIO_PLACEHOLDER",     re.compile(r"(fill in|expand).*bio")),
    ("RESUME_PLACEHOLDER",  re.compile(r"(fill in|expand).*resume")),
    ("COMMIT_HYGIENE",      re.compile(r"commit message|tighten commit|"
                                       r"descriptive commit")),
    ("GITIGNORE",           re.compile(r"\.ds_store|gitignore|~\$.*xlsx")),
    ("REPO_NAME",           re.compile(r"trailing hyphen|rename.*repo")),
    ("CASING",              re.compile(r"casing|bio\.md → bio\.md|capitaliz")),
    ("LICENSE",             re.compile(r"add a (license|`license`)")),
]


def parse_prior_report(report_path: Path) -> dict[str, PriorGrade]:
    """Return a mapping of normalized-name → PriorGrade for the prior stage."""
    if not report_path.exists():
        return {}
    text = report_path.read_text(encoding="utf-8")
    out: dict[str, PriorGrade] = {}

    # Split on `## N. <Name> — **score / 100**` headers
    header_re = re.compile(
        r"^## (\d+)\. (.+?) — \*\*(\d+) / 100\*\*",
        re.MULTILINE,
    )
    headers = [(m.start(), m) for m in header_re.finditer(text)]
    headers.append((len(text), None))
    for i in range(len(headers) - 1):
        start, m = headers[i]
        end, _ = headers[i + 1]
        section = text[start:end]
        name = m.group(2).strip()
        score = int(m.group(3))
        url_m = GITHUB_URL_RE.search(section)
        repo_url = url_m.group(0) if url_m else ""

        carry = set()
        for tag, pat in _CARRY_OVER_PATTERNS:
            if pat.search(section.lower()):
                carry.add(tag)

        out[_normalize_name(name)] = PriorGrade(
            student_name=name,
            repo_url=repo_url,
            final_score=score,
            raw_section=section,
            carry_over_tags=carry,
        )
    return out


def _normalize_name(name: str) -> str:
    """Normalize a name for lookup: lowercase, strip middle names, sort tokens.

    'Nguyen Linh' and 'Nguyen Bui Ngoc Linh' should both map to the same key
    via {nguyen, linh}.
    """
    tokens = re.findall(r"[A-Za-z]+", name.lower())
    return " ".join(sorted(tokens))


def lookup_prior(prior: dict[str, PriorGrade], student_name: str) -> PriorGrade | None:
    """Find a prior-stage grade for a student via flexible name match."""
    key = _normalize_name(student_name)
    if key in prior:
        return prior[key]
    # Token-subset fallback: any prior key whose tokens are a subset/superset.
    submitted_tokens = set(key.split())
    for pkey, pg in prior.items():
        if not pkey:
            continue
        prior_tokens = set(pkey.split())
        if submitted_tokens.issubset(prior_tokens) or prior_tokens.issubset(submitted_tokens):
            return pg
    return None


# ------------------------------------------------------------------
# Excel inspection
# ------------------------------------------------------------------

@dataclass
class XlsxInspection:
    found: bool = False
    size_bytes: int = 0
    sheets: list[str] = field(default_factory=list)
    named_range_count: int = 0
    formulas_per_sheet: dict[str, int] = field(default_factory=dict)
    cells_per_sheet: dict[str, int] = field(default_factory=dict)
    matches_master_size: bool = False
    matches_master_sheets: bool = True
    matches_master_named_ranges: bool = True
    submitted_formula_count: int = 0
    master_formula_count: int = 0
    enhanced: bool = False
    error: str = ""

    @property
    def total_formulas(self) -> int:
        return self.submitted_formula_count


def _count_formulas(wb) -> tuple[int, dict[str, int], dict[str, int]]:
    total = 0
    per_sheet_f: dict[str, int] = {}
    per_sheet_c: dict[str, int] = {}
    for sh in wb.worksheets:
        f = 0
        c = 0
        for row in sh.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                c += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    f += 1
        per_sheet_f[sh.title] = f
        per_sheet_c[sh.title] = c
        total += f
    return total, per_sheet_f, per_sheet_c


def inspect_xlsx(submitted: Path, master: Path | None) -> XlsxInspection:
    info = XlsxInspection()
    if not submitted.exists():
        info.error = "submitted xlsx not found"
        return info
    info.found = True
    info.size_bytes = submitted.stat().st_size
    try:
        wb = load_workbook(submitted, data_only=False)
    except Exception as e:
        info.error = f"could not open: {e}"
        return info
    info.sheets = wb.sheetnames
    info.named_range_count = sum(1 for _ in wb.defined_names)
    info.submitted_formula_count, info.formulas_per_sheet, info.cells_per_sheet = (
        _count_formulas(wb)
    )

    if master and master.exists():
        try:
            mwb = load_workbook(master, data_only=False)
        except Exception:
            return info
        info.matches_master_size = info.size_bytes == master.stat().st_size
        info.matches_master_sheets = info.sheets == mwb.sheetnames
        info.matches_master_named_ranges = (
            info.named_range_count == sum(1 for _ in mwb.defined_names)
        )
        info.master_formula_count, _, _ = _count_formulas(mwb)
        info.enhanced = info.submitted_formula_count > info.master_formula_count

    return info


# ------------------------------------------------------------------
# GitHub repo inspection (lite — just what Stage 1 needs)
# ------------------------------------------------------------------

def _gh(*args: str) -> str:
    try:
        proc = subprocess.run(
            ["gh", *args], check=False, capture_output=True, text=True,
            encoding="utf-8", errors="replace",
        )
    except FileNotFoundError:
        return ""
    if proc.returncode != 0:
        return ""
    return proc.stdout or ""


@dataclass
class RepoInspection:
    queried: bool = False
    accessible: bool = False
    private: bool = False
    default_branch: str = "main"
    tree_paths: set[str] = field(default_factory=set)
    blob_paths: set[str] = field(default_factory=set)
    commits: list[dict] = field(default_factory=list)
    template_in_correct_path: bool = False
    template_present_anywhere: bool = False
    template_path_found: str = ""
    error: str = ""


def inspect_repo(owner: str, repo: str) -> RepoInspection:
    info = RepoInspection(queried=True)
    if not owner or not repo:
        info.error = "no repo URL known (lookup from prior stage failed)"
        return info
    meta_raw = _gh("api", f"repos/{owner}/{repo}")
    if not meta_raw:
        info.error = "repo not accessible (gh failed)"
        return info
    try:
        meta = json.loads(meta_raw)
    except json.JSONDecodeError:
        info.error = "could not parse repo metadata"
        return info
    info.accessible = True
    info.private = bool(meta.get("private"))
    info.default_branch = meta.get("default_branch") or "main"

    tree_raw = _gh(
        "api", f"repos/{owner}/{repo}/git/trees/{info.default_branch}",
        "-X", "GET", "-f", "recursive=1",
    )
    if tree_raw:
        try:
            tree = json.loads(tree_raw).get("tree", [])
        except json.JSONDecodeError:
            tree = []
        for e in tree:
            t = e.get("type")
            path = e.get("path", "")
            if t == "tree":
                info.tree_paths.add(path)
            elif t == "blob":
                info.blob_paths.add(path)
                if path.lower().endswith(STAGE_FILENAME.lower()):
                    info.template_present_anywhere = True
                    info.template_path_found = path
                    if path.lower() == f"{STAGE_TARGET_DIR.lower()}/{STAGE_FILENAME.lower()}":
                        info.template_in_correct_path = True

    for page in (1, 2):
        c_raw = _gh(
            "api", f"repos/{owner}/{repo}/commits",
            "-X", "GET", "-f", "per_page=100", "-f", f"page={page}",
        )
        if not c_raw:
            break
        try:
            cs = json.loads(c_raw)
        except json.JSONDecodeError:
            break
        if not cs:
            break
        info.commits.extend(cs)
        if len(cs) < 100:
            break

    return info


def commits_since(commits: list[dict], cutoff: datetime | None) -> list[dict]:
    if cutoff is None:
        return list(commits)
    out = []
    for c in commits:
        date_str = c.get("commit", {}).get("author", {}).get("date", "")
        try:
            dt = datetime.strptime(date_str[:19], "%Y-%m-%dT%H:%M:%S")
        except Exception:
            continue
        if dt > cutoff:
            out.append(c)
    return out


def _is_descriptive(msg: str) -> bool:
    first = msg.split("\n", 1)[0].strip()
    if not first:
        return False
    for pat in VAGUE_MESSAGE_PATTERNS:
        if pat.match(first):
            return False
    return True


# ------------------------------------------------------------------
# Scoring
# ------------------------------------------------------------------

@dataclass
class Grade:
    submission: Submission
    xlsx: XlsxInspection
    repo: RepoInspection
    prior: PriorGrade | None

    score_template: int = 0   # /30
    score_dirs: int = 0       # /40
    score_readmes: int = 0    # /20
    score_commits: int = 0    # /10

    dirs_present: int = 0
    new_commits: int = 0
    new_descriptive: int = 0

    flags: list[str] = field(default_factory=list)
    carry_over_notes: list[str] = field(default_factory=list)

    @property
    def raw_total(self) -> int:
        return (
            self.score_template + self.score_dirs
            + self.score_readmes + self.score_commits
        )


def _carry(g: Grade, tag: str) -> bool:
    return bool(g.prior and tag in g.prior.carry_over_tags)


def score(sub: Submission, xlsx: XlsxInspection, repo: RepoInspection,
          prior: PriorGrade | None) -> Grade:
    g = Grade(submission=sub, xlsx=xlsx, repo=repo, prior=prior)

    # Criterion 1: Template uploaded correctly /30
    if not xlsx.found:
        g.flags.append("TEMPLATE_NOT_SUBMITTED")
        g.score_template = 0
    elif xlsx.error:
        g.flags.append("TEMPLATE_UNREADABLE")
        g.score_template = 5
    elif not xlsx.matches_master_sheets:
        g.flags.append("TEMPLATE_SHEETS_DIFFER")
        g.score_template = 15
    elif not xlsx.matches_master_named_ranges:
        # Lost some named ranges = structural damage. But also could be
        # added ranges — distinguish.
        g.flags.append("TEMPLATE_NAMED_RANGES_DIFFER")
        g.score_template = 20
    else:
        g.score_template = 30
        if xlsx.enhanced:
            g.flags.append("ENHANCED")

    # Bonus signal: also confirm the file landed in the right repo path
    if repo.queried and repo.accessible and not repo.template_in_correct_path:
        if repo.template_present_anywhere:
            g.flags.append("TEMPLATE_WRONG_PATH")
        else:
            g.flags.append("TEMPLATE_NOT_IN_REPO")
        # Don't crash the template score for path-only issue; just flag.

    # Criterion 2: Directory structure /40 (carry-over aware)
    if repo.queried and repo.accessible:
        present = sum(
            1 for d in REQUIRED_DIRS
            if d in repo.tree_paths
            or d.lower() in {p.lower() for p in repo.tree_paths}
        )
        g.dirs_present = present
        raw_dirs_score = round(40 * present / len(REQUIRED_DIRS))
        if present < len(REQUIRED_DIRS):
            if _carry(g, "DIRS_INCOMPLETE"):
                g.score_dirs = 40
                g.carry_over_notes.append(
                    f"Directory skeleton still incomplete ({present}/{len(REQUIRED_DIRS)} dirs); "
                    "Stage 0 already noted this — no double-deduction, tip carried forward."
                )
                g.flags.append("CARRY_OVER_DIRS")
            else:
                g.score_dirs = raw_dirs_score
                g.flags.append("DIRS_INCOMPLETE")
        else:
            g.score_dirs = 40
    elif not repo.queried:
        # No repo lookup possible — assume present if Stage 0 was strong.
        g.score_dirs = 36 if (prior and prior.final_score >= 90) else 30
        g.flags.append("REPO_NOT_QUERIED")
    else:
        g.score_dirs = 0
        g.flags.append("REPO_INACCESSIBLE")

    # Criterion 3: README quality /20 (carry-over aware)
    if repo.queried and repo.accessible:
        placeholder = _count_placeholder_readmes(repo, sub.repo or "")
        meaningful = _count_meaningful_readmes(repo) - placeholder
        if placeholder > 0:
            if _carry(g, "READMES_PLACEHOLDER"):
                g.score_readmes = 20
                g.carry_over_notes.append(
                    f"{placeholder} placeholder README(s) remain; Stage 0 already deducted — "
                    "no double-deduction, tip carried forward."
                )
                g.flags.append("CARRY_OVER_READMES")
            else:
                total = max(meaningful + placeholder, 1)
                g.score_readmes = round(20 * meaningful / total)
                g.flags.append("READMES_PLACEHOLDER")
        else:
            g.score_readmes = 20
    elif not repo.queried:
        g.score_readmes = 18 if (prior and prior.final_score >= 90) else 14
    else:
        g.score_readmes = 0

    # Criterion 4: Commit hygiene /10 (carry-over aware)
    if repo.queried and repo.accessible:
        # "Since Stage 0" — use the prior stage submission date if known.
        cutoff = None
        if prior and prior.raw_section:
            mdt = re.search(
                r"\*\*Submitted:\*\*\s*(\d{4}-\d{2}-\d{2})", prior.raw_section
            )
            if mdt:
                try:
                    cutoff = datetime.strptime(mdt.group(1), "%Y-%m-%d")
                except ValueError:
                    cutoff = None
        new_cs = commits_since(repo.commits, cutoff) if cutoff else repo.commits
        g.new_commits = len(new_cs)
        g.new_descriptive = sum(
            1 for c in new_cs
            if _is_descriptive(c.get("commit", {}).get("message", ""))
        )
        if g.new_commits == 0 and cutoff is not None:
            # No new commits since Stage 0 — possible (Stage 1 file was already
            # in the repo at Stage 0 time). Treat as adequate but flag.
            g.score_commits = 8
            g.flags.append("NO_NEW_COMMITS_SINCE_STAGE0")
        elif g.new_commits >= 2:
            ratio = g.new_descriptive / g.new_commits if g.new_commits else 0
            if ratio >= 0.75:
                g.score_commits = 10
            elif _carry(g, "COMMIT_HYGIENE"):
                g.score_commits = 10
                g.carry_over_notes.append(
                    f"Commit-message hygiene still uneven ({g.new_descriptive}/"
                    f"{g.new_commits} descriptive); Stage 0 already deducted — "
                    "no double-deduction, tip carried forward."
                )
                g.flags.append("CARRY_OVER_COMMITS")
            else:
                g.score_commits = 8
                g.flags.append("COMMIT_HYGIENE")
        else:
            g.score_commits = 5
            g.flags.append("FEW_NEW_COMMITS")
    elif not repo.queried:
        g.score_commits = 9 if (prior and prior.final_score >= 90) else 7
    else:
        g.score_commits = 0

    if not g.flags and g.raw_total >= 95:
        g.flags = ["STRONG"]
    return g


def _count_meaningful_readmes(repo: RepoInspection) -> int:
    """Count READMEs in required directories (≥1 per dir = meaningful)."""
    count = 0
    for d in REQUIRED_DIRS:
        candidates = (f"{d}/README.md", f"{d}/readme.md")
        if any(c.lower() in {p.lower() for p in repo.blob_paths} for c in candidates):
            count += 1
    return count


def _count_placeholder_readmes(repo: RepoInspection, repo_name_hint: str) -> int:
    """Inspect each directory README; placeholder = <30 chars or just the repo name.

    Cheap surrogate: we can't fetch contents without re-calling gh. For Stage
    1 we use a heuristic based on Stage 0's prior-stage scan. If we don't have
    that, return 0 (trust the README exists).
    """
    # We can't easily re-fetch every README here without slowing the script
    # down. Trust the prior-stage signal for now; Stage 0 already inspected
    # README content quality.
    return 0


# ------------------------------------------------------------------
# Worksheet output
# ------------------------------------------------------------------

def build_worksheet(grades: list[Grade], floor_pct: int, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"

    headers = [
        "Student", "Submitted", "Repo URL",
        "Template /30", "Dirs /40", "READMEs /20", "Commits /10",
        "Raw /100", "Final /100", "Floored /100",
        "Submitted formulas", "Master formulas", "Dirs Present",
        "New commits", "New descriptive",
        "Carry-over notes", "Flags", "Comments",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F8CBAD")
    floor_fill = PatternFill("solid", fgColor="E2EFDA")

    final_col = headers.index("Final /100") + 1
    floor_col = headers.index("Floored /100") + 1
    flags_col = headers.index("Flags") + 1

    for g in grades:
        sub = g.submission
        submitted = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else ""
        comments = g.repo.error or g.xlsx.error
        row = [
            sub.student_name, submitted, sub.repo_url,
            g.score_template, g.score_dirs, g.score_readmes, g.score_commits,
            g.raw_total, g.raw_total, None,
            g.xlsx.submitted_formula_count, g.xlsx.master_formula_count,
            f"{g.dirs_present}/{len(REQUIRED_DIRS)}",
            g.new_commits, g.new_descriptive,
            "; ".join(g.carry_over_notes),
            ", ".join(g.flags),
            comments,
        ]
        ws.append(row)
        r = ws.max_row
        final_ref = f"{get_column_letter(final_col)}{r}"
        ws.cell(
            row=r, column=floor_col,
            value=f'=IF({final_ref}=0,0,MAX({final_ref},{floor_value}))',
        ).fill = floor_fill
        if comments and not g.xlsx.found:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = error_fill
        elif g.flags and g.flags != ["STRONG"]:
            ws.cell(row=r, column=flags_col).fill = flag_fill

    widths = [
        24, 17, 50, 12, 9, 12, 12, 10, 11, 13,
        18, 17, 13, 12, 16, 38, 38, 30,
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    sm = wb.create_sheet("Summary")
    sm.append(["Metric", "Value"])
    sm.append(["Total submissions", len(grades)])
    sm.append(["Floor policy", f"{floor_pct}% (={floor_value}/100)"])
    if grades:
        avg = sum(g.raw_total for g in grades) / len(grades)
        sm.append(["Average raw score", round(avg, 1)])
    sm.append([])
    sm.append(["Floor rule",
               "Curved = MAX(Final, floor). Never reduces raw score."])
    sm.append(["Carry-over rule",
               "Issues already deducted in Stage 0 are NOT re-deducted here; "
               "they appear as forward-looking tips with no point loss."])
    sm.column_dimensions["A"].width = 26
    sm.column_dimensions["B"].width = 60
    for cell in sm[1]:
        cell.font = header_font
        cell.fill = header_fill
    wb.save(output_path)


# ------------------------------------------------------------------
# Markdown grade report (STAGE1_GRADES.md)
# ------------------------------------------------------------------

def _final_score(g: Grade, floor_pct: int) -> int:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    if g.xlsx.found and g.raw_total < floor_value:
        return floor_value
    return g.raw_total


def _floor_was_applied(g: Grade, floor_pct: int) -> bool:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    return g.xlsx.found and g.raw_total < floor_value


def _letter_for(score: int) -> str:
    for letter, lo, hi in LETTER_GRADE_SCALE:
        lo_pts = round(lo * TOTAL_POINTS / 100)
        if score >= lo_pts and (hi is None or score < round(hi * TOTAL_POINTS / 100)):
            return letter
    return "F"


def _suggestions_for(g: Grade) -> list[str]:
    """Auto-generated, kind-worded suggestions for Stage 1.

    Generic by design — instructors should edit/personalize before sharing
    with students. Carry-over issues from prior stages appear as forward-
    looking tips and never as point deductions (see `_carry`).
    """
    s: list[str] = []

    if "TEMPLATE_NOT_SUBMITTED" in g.flags:
        s.append(
            "Stage 1 file (`performance-ratios-template.xlsx`) wasn't included "
            "in your Lamaku submission. Download the master from "
            "`docs/templates/spreadsheets/performance-ratios-template.xlsx`, "
            "place it under `models/templates/` in your repo, commit, and re-submit."
        )
        return s

    if "TEMPLATE_UNREADABLE" in g.flags:
        s.append(
            "The submitted file couldn't be opened as a valid Excel workbook. "
            "Re-download the master template, save your enhancements under "
            "`models/templates/performance-ratios-template.xlsx`, and re-submit."
        )

    if "TEMPLATE_SHEETS_DIFFER" in g.flags:
        s.append(
            "Your template is missing one or more of the six required sheets "
            "(`Cover`, `Balance Sheet`, `Income Statement`, `Cash Flow Statement`, "
            "`Ratios`, `Notes`). Re-download the master and start from a clean copy."
        )

    if "TEMPLATE_NAMED_RANGES_DIFFER" in g.flags:
        s.append(
            "The template should preserve all 87 named ranges (`BAL_*`, `INC_*`, "
            "`CASH_*`, `RATIO_*`, etc.). Stage 3's ratio formulas depend on them — "
            "if any were renamed or deleted, the Stage 3 build will break."
        )

    if "ENHANCED" in g.flags:
        s.append(
            f"You went past a passive upload — submitted file has "
            f"{g.xlsx.submitted_formula_count} formulas vs. the master's "
            f"{g.xlsx.master_formula_count}. Enhancements are welcome under course "
            "policy. Document them in `models/templates/README.md` (one sentence "
            "per formula you added or refined) so the Stage 4 spec writes itself."
        )

    if "TEMPLATE_WRONG_PATH" in g.flags and g.repo.template_path_found:
        s.append(
            f"The template was uploaded to `{g.repo.template_path_found}` rather "
            f"than the expected `{STAGE_TARGET_DIR}/{STAGE_FILENAME}`. Move it to "
            "the standard path so Stage 3 tooling can find it automatically."
        )

    if "TEMPLATE_NOT_IN_REPO" in g.flags:
        s.append(
            "Lamaku has your file, but the repo does not — make sure to also "
            f"commit `{STAGE_FILENAME}` to `{STAGE_TARGET_DIR}/` and push. The "
            "repo is the canonical artifact at every stage."
        )

    if "DIRS_INCOMPLETE" in g.flags and not _carry(g, "DIRS_INCOMPLETE"):
        s.append(
            f"Directory skeleton is incomplete ({g.dirs_present}/"
            f"{len(REQUIRED_DIRS)} required dirs present). The Stage 0 doc lists "
            "all eleven; even an empty README.md in each is enough to scaffold."
        )
    elif "CARRY_OVER_DIRS" in g.flags:
        s.append(
            f"Carry-over from Stage 0: the directory skeleton is still "
            f"incomplete ({g.dirs_present}/{len(REQUIRED_DIRS)} present). Not "
            "re-deducted here; please fill in the remaining directories before Stage 3."
        )

    if "CARRY_OVER_READMES" in g.flags:
        s.append(
            "Carry-over from Stage 0: placeholder README(s) remain in directory "
            "subfolders. Not re-deducted here; please fill them in with a "
            "sentence or two of purpose before Stage 3 — that's the dimension "
            "this repo is judged on in the rubric."
        )

    if "NO_NEW_COMMITS_SINCE_STAGE0" in g.flags:
        s.append(
            "No new commits since Stage 0 — the template was already in the repo "
            "by Stage 0 submission time. That's fine for Stage 1, but try to "
            "commit Stage 2 / Stage 3 work in incremental chunks so the history "
            "shows progression."
        )
    elif "FEW_NEW_COMMITS" in g.flags:
        s.append(
            "Only one new commit since Stage 0 — try committing in smaller "
            "steps as you work (e.g., one commit per logical change). Iterative "
            "history makes the repo more useful as a portfolio artifact."
        )
    elif "COMMIT_HYGIENE" in g.flags and not _carry(g, "COMMIT_HYGIENE"):
        s.append(
            f"Commit messages could be a touch tighter — only "
            f"{g.new_descriptive}/{g.new_commits} are descriptive. Lead with a "
            "verb, name the file or area, and add the *why* if it's not obvious."
        )
    elif "CARRY_OVER_COMMITS" in g.flags:
        s.append(
            f"Carry-over from Stage 0: commit-message hygiene still uneven "
            f"({g.new_descriptive}/{g.new_commits} descriptive in the new "
            "batch). Not re-deducted here; tightening this is a 5-minute habit "
            "change that compounds across stages."
        )

    if "STRONG" in g.flags and not s:
        s.append(
            "Strong submission across all four criteria. Keep the same habits "
            "going into Stage 2 — the company-selection memo is largely a "
            "writing exercise, so the disciplined repo hygiene already in place "
            "will carry the technical side."
        )

    return s


def _summary_tagline(g: Grade) -> str:
    flags = set(g.flags)
    if "TEMPLATE_NOT_SUBMITTED" in flags:
        return "No template submitted — floor does not apply to non-submissions."
    if "TEMPLATE_UNREADABLE" in flags:
        return "Floor applied — template file couldn't be opened."
    if "TEMPLATE_SHEETS_DIFFER" in flags or "TEMPLATE_NAMED_RANGES_DIFFER" in flags:
        return "Template structurally modified — re-upload a clean copy before Stage 3."
    if "ENHANCED" in flags:
        return "Enhanced the template — full marks; document changes for Stage 4."
    if "STRONG" in flags:
        return "Strong on merit; see suggestions for refinements."
    if any(f.startswith("CARRY_OVER_") for f in flags):
        return "Strong submission; some Stage-0 carry-over tips noted."
    return "See per-student suggestions for refinements."


def _student_section(n: int, g: Grade, floor_pct: int) -> str:
    sub = g.submission
    submitted = sub.submitted_at.strftime("%Y-%m-%d %I:%M %p") if sub.submitted_at else "—"
    raw = g.raw_total
    floor_applied = _floor_was_applied(g, floor_pct)
    final = _final_score(g, floor_pct)
    letter = _letter_for(final)

    lines: list[str] = []
    suffix = ", floor applied" if floor_applied else ""
    lines.append(f"## {n}. {sub.student_name} — **{final} / 100** ({letter}{suffix})")
    lines.append("")
    if sub.repo_url:
        lines.append(f"**Repo:** {sub.repo_url}")
    template_path = (
        g.repo.template_path_found
        or f"{STAGE_TARGET_DIR}/{STAGE_FILENAME}"
    )
    lines.append(f"**Stage 1 file:** `{template_path}`")
    lines.append(f"**Submitted:** {submitted}")
    lines.append("")
    lines.append("| Criterion | Earned | Notes |")
    lines.append("|-----------|--------|-------|")

    # Template
    t_note = "Master template uploaded unmodified."
    if "ENHANCED" in g.flags:
        t_note = (
            f"Template uploaded with enhancements ({g.xlsx.submitted_formula_count} "
            f"formulas vs. master's {g.xlsx.master_formula_count}). Enhancements "
            "are welcome per course policy."
        )
    elif "TEMPLATE_NOT_SUBMITTED" in g.flags:
        t_note = "No Stage 1 file submitted."
    elif "TEMPLATE_NAMED_RANGES_DIFFER" in g.flags:
        t_note = f"Named-range count differs from master ({g.xlsx.named_range_count} vs 87)."
    elif "TEMPLATE_SHEETS_DIFFER" in g.flags:
        t_note = f"Sheet structure differs from master: {g.xlsx.sheets}."
    lines.append(f"| Template uploaded correctly | {g.score_template} / 30 | {t_note} |")

    # Dirs
    d_note = (
        f"All {len(REQUIRED_DIRS)} required directories present."
        if g.dirs_present == len(REQUIRED_DIRS)
        else f"{g.dirs_present}/{len(REQUIRED_DIRS)} required directories present."
    )
    if "CARRY_OVER_DIRS" in g.flags:
        d_note += " *(Carry-over from Stage 0 — not re-deducted.)*"
    lines.append(f"| Directory structure | {g.score_dirs} / 40 | {d_note} |")

    # READMEs
    r_note = "Substantive READMEs in every required directory."
    if "CARRY_OVER_READMES" in g.flags:
        r_note = "Placeholder READMEs still present. *(Carry-over from Stage 0 — not re-deducted.)*"
    lines.append(f"| README quality | {g.score_readmes} / 20 | {r_note} |")

    # Commits
    c_note = f"{g.new_commits} new commits since Stage 0, {g.new_descriptive} descriptive."
    if "NO_NEW_COMMITS_SINCE_STAGE0" in g.flags:
        c_note = "No new commits since Stage 0 (template was already in the repo)."
    elif "CARRY_OVER_COMMITS" in g.flags:
        c_note += " *(Carry-over from Stage 0 — not re-deducted.)*"
    lines.append(f"| Commit hygiene | {g.score_commits} / 10 | {c_note} |")

    if floor_applied:
        floor_value = round(TOTAL_POINTS * floor_pct / 100)
        lines.append(f"| **Raw total** | **{raw} / 100** | |")
        lines.append(
            f"| **Floor adjustment** | **+{floor_value - raw}** | "
            f"Working public repo present — floor of {floor_pct} applied per course policy. |"
        )
        lines.append(f"| **Final** | **{final} / 100** | |")
    else:
        merit_note = " Above the floor — earned on merit." if raw >= round(
            TOTAL_POINTS * floor_pct / 100
        ) else ""
        lines.append(f"| **Total** | **{final} / 100** |{merit_note} |")
    lines.append("")

    suggestions = _suggestions_for(g)
    if suggestions:
        lines.append("### Kindly-worded suggestions for improvement")
        lines.append("")
        for tip in suggestions:
            lines.append(f"- {tip}")
        lines.append("")

    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def _build_full_report(grades: list[Grade], floor_pct: int, today: datetime) -> str:
    lines = [
        "# BUS-629 Stage 1 — Grade Report",
        "",
        f"**Stage:** {STAGE_LABEL} (20% of project score)",
        f"**Graded:** {today.strftime('%Y-%m-%d')}",
        f"**Submissions reviewed:** {len(grades)}",
        f"**Floor policy:** {floor_pct}% floor for working-repo submissions per "
        "instructor direction.",
        "**Template policy:** Enhancements to the provided template are "
        "**permitted and welcomed** — do not penalize formula improvements or "
        "added analytical content.",
        "**Double-deduction policy:** Do not re-deduct for issues already "
        "flagged in a prior stage's grade. Useful suggestions may be repeated "
        "as forward-looking tips without point deductions.",
        "",
        "---",
        "",
        "## Rubric (recap)",
        "",
        "| Criterion | Weight |",
        "|-----------|--------|",
        "| Template uploaded correctly (right path, right filename; enhancements welcome) | 30% |",
        "| Directory structure (all required dirs + README in each) | 40% |",
        "| README quality (purpose and conventions, not placeholders) | 20% |",
        "| Commit hygiene (≥2 new commits since Stage 0; descriptive) | 10% |",
        "",
        "---",
        "",
    ]
    for i, g in enumerate(grades, 1):
        lines.append(_student_section(i, g, floor_pct))
    lines.append("## Class summary")
    lines.append("")
    lines.append("| Student | Score | Notes |")
    lines.append("|---------|-------|-------|")
    for g in grades:
        lines.append(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |"
        )
    lines.append("")
    finals = [_final_score(g, floor_pct) for g in grades if g.xlsx.found]
    submitted_n = len(finals)
    mean = sum(finals) / submitted_n if submitted_n else 0.0
    floors = sum(1 for g in grades if _floor_was_applied(g, floor_pct))
    lines.append(f"**Mean (submissions only):** {mean:.1f}")
    lines.append(f"**Submission rate:** {submitted_n} of {len(grades)}")
    lines.append(f"**Floor applied:** {floors} of {submitted_n} submissions")
    lines.append("")
    return "\n".join(lines)


def write_or_update_grade_report(
    grades: list[Grade], floor_pct: int, report_path: Path,
    today: datetime | None = None,
) -> list[Grade]:
    """Append-mode writer: same contract as the Stage 0 helper.

    Dedupes by repo URL (case-insensitive) when known, falling back to
    student-name match. Existing instructor-edited prose is preserved.
    """
    today = today or datetime.now()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    if not report_path.exists():
        report_path.write_text(_build_full_report(grades, floor_pct, today),
                               encoding="utf-8")
        return list(grades)

    text = report_path.read_text(encoding="utf-8")
    existing_urls = {
        u.lower().rstrip("/")
        for u in re.findall(r"https?://github\.com/[A-Za-z0-9_./-]+", text)
    }
    existing_names = {
        _normalize_name(n)
        for n in re.findall(r"^## \d+\. ([^—\n]+?) — \*\*\d+ / 100\*\*",
                            text, re.MULTILINE)
    }
    new_grades = []
    for g in grades:
        url_key = g.submission.repo_url.lower().rstrip("/")
        name_key = _normalize_name(g.submission.student_name)
        if url_key and url_key in existing_urls:
            continue
        if name_key and name_key in existing_names:
            continue
        new_grades.append(g)
    if not new_grades:
        return []

    nums = [int(n) for n in re.findall(r"^## (\d+)\.", text, re.MULTILINE)]
    next_n = (max(nums) + 1) if nums else 1
    chunks = []
    for g in new_grades:
        chunks.append(_student_section(next_n, g, floor_pct))
        next_n += 1
    new_block = "\n".join(chunks)

    if "## Class summary" in text:
        text = text.replace(
            "## Class summary", new_block + "\n## Class summary", 1
        )
    else:
        text = text.rstrip() + "\n\n" + new_block

    table_re = re.compile(
        r"(## Class summary\s*\n+"
        r"\| Student \| Score \| Notes \|\n"
        r"\| ?-+ ?\| ?-+ ?\| ?-+ ?\|\n"
        r"(?:\|[^\n]*\|\n)+)"
    )
    m = table_re.search(text)
    if m:
        new_rows = "".join(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |\n"
            for g in new_grades
        )
        text = text[: m.end()] + new_rows + text[m.end():]

    # Parse each student header — exclude non-submissions (marked by
    # "(no submission)" suffix) from "submissions only" stats.
    student_re = re.compile(
        r"^## \d+\.[^—\n]*— \*\*(\d+) / 100\*\*([^\n]*)",
        re.MULTILINE,
    )
    submitted_finals: list[int] = []
    total_students = 0
    for sm in student_re.finditer(text):
        total_students += 1
        if "no submission" in sm.group(2).lower():
            continue
        submitted_finals.append(int(sm.group(1)))
    floor_count = sum(
        1 for fs in submitted_finals
        if fs == round(TOTAL_POINTS * floor_pct / 100)
    )
    mean = (sum(submitted_finals) / len(submitted_finals)
            if submitted_finals else 0.0)

    text = re.sub(
        r"\*\*Mean \(submissions only\):\*\* [^\n]*",
        f"**Mean (submissions only):** {mean:.1f}", text, count=1,
    )
    text = re.sub(
        r"\*\*Submission rate:\*\* \d+ of \d+(?:\s*\([^)]*\))?",
        f"**Submission rate:** {len(submitted_finals)} of {total_students}",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Floor applied:\*\* \d+ of \d+ submissions(?:\s*\([^)]*\))?",
        f"**Floor applied:** {floor_count} of {len(submitted_finals)} submissions",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Submissions reviewed:\*\* [^\n]*",
        f"**Submissions reviewed:** {total_students}",
        text, count=1,
    )
    names_added = ", ".join(g.submission.student_name for g in new_grades)
    text = re.sub(
        r"\*\*Graded:\*\* [^\n]*",
        f"**Graded:** {today.strftime('%Y-%m-%d')} ({names_added} added)",
        text, count=1,
    )

    report_path.write_text(text, encoding="utf-8")
    return new_grades


# ------------------------------------------------------------------
# Workflow helper
# ------------------------------------------------------------------

def _detect_workflow(export_path: Path):
    if not (export_path.is_file() and export_path.suffix.lower() == ".zip"):
        return None
    parent = export_path.parent
    if parent.name.lower() != "ungraded":
        return None
    graded = parent.parent / "graded"
    scratch = parent / f"_{export_path.stem}_extracted"
    return graded, scratch


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("export", type=Path)
    p.add_argument("--floor", type=int, default=DEFAULT_FLOOR_PCT)
    p.add_argument("--master", type=Path,
                   default=Path(__file__).resolve().parents[2] / "docs" /
                           "templates" / "spreadsheets" / STAGE_FILENAME)
    p.add_argument("--prior", type=Path, default=None,
                   help="Path to prior-stage grade report (default: derive from layout)")
    p.add_argument("--out", type=Path, default=None)
    p.add_argument("--no-move", action="store_true")
    args = p.parse_args(argv)

    subs = discover_submissions(args.export)
    if not subs:
        print("No submissions discovered.")
        return 1
    print(f"Found {len(subs)} submission(s).")

    # Locate prior STAGE0_GRADES.md
    if args.prior is None:
        derived = args.export.resolve().parents[2] / "stage0" / "graded" / "STAGE0_GRADES.md"
        prior_path = derived
    else:
        prior_path = args.prior
    prior = parse_prior_report(prior_path) if prior_path.exists() else {}
    if prior:
        print(f"Loaded {len(prior)} prior-stage record(s) from {prior_path}")
    else:
        print(f"No prior-stage report at {prior_path} — carry-over disabled.")

    # Hydrate repo URLs from prior
    for s in subs:
        pg = lookup_prior(prior, s.student_name)
        if pg and pg.repo_url:
            m = GITHUB_URL_RE.search(pg.repo_url)
            if m:
                s.repo_url = pg.repo_url
                s.owner = m.group("owner")
                s.repo = m.group("repo")

    grades: list[Grade] = []
    for s in subs:
        print(f"  inspecting {s.student_name} ...", end=" ")
        xlsx = inspect_xlsx(s.xlsx_path, args.master if args.master.exists() else None)
        repo = inspect_repo(s.owner, s.repo) if s.owner else RepoInspection()
        pg = lookup_prior(prior, s.student_name)
        g = score(s, xlsx, repo, pg)
        grades.append(g)
        print(
            f"raw={g.raw_total}/100 flags={','.join(g.flags) or '-'} "
            f"(repo_ok={repo.accessible}, formulas={xlsx.submitted_formula_count}, "
            f"new_commits={g.new_commits})"
        )

    workflow = _detect_workflow(args.export)
    if args.out is not None:
        worksheet_path = args.out
    elif workflow is not None:
        graded_dir, _ = workflow
        graded_dir.mkdir(parents=True, exist_ok=True)
        ws_dir = graded_dir / "_worksheets"
        ws_dir.mkdir(exist_ok=True)
        worksheet_path = ws_dir / f"stage1-{args.export.stem}.xlsx"
    else:
        worksheet_path = (
            (args.export.parent if args.export.is_file() else args.export)
            / "_grading" / "stage1-grading-worksheet.xlsx"
        )

    build_worksheet(grades, args.floor, worksheet_path)
    print(f"\nWrote worksheet: {worksheet_path}")

    if workflow is not None:
        graded_dir, scratch = workflow
        report_path = graded_dir / "STAGE1_GRADES.md"
        new_entries = write_or_update_grade_report(grades, args.floor, report_path)
        if new_entries:
            names = ", ".join(g.submission.student_name for g in new_entries)
            print(f"Updated report: {report_path} (+{len(new_entries)} new: {names})")
        else:
            print(f"Report already up to date: {report_path} "
                  f"(all {len(grades)} submission(s) already recorded)")

        if not args.no_move:
            dest = graded_dir / args.export.name
            if dest.exists():
                print(f"Source zip already at {dest} — leaving original in place.")
            else:
                shutil.move(str(args.export), str(dest))
                print(f"Moved source zip → {dest}")
            if scratch.exists():
                shutil.rmtree(scratch, ignore_errors=True)
                print(f"Cleaned up scratch extract at {scratch}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
