"""BUS-629 Stage 5 (LLM analysis + executive evaluation + repo polish) grading scanner.

Stage 5's deliverable is the repo URL itself — students submit a Lamaku
upload pointing at their public GitHub repo. The graded artifact is the
repo's final state across six required files plus the overall polish pass.

This script:

  1. Reads the Lamaku export zip (or extracted directory). The zip is
     typically URL-only (HTML pointer); the actual artifacts live in the
     student's repo.
  2. For each student, inspects the repo via `gh`:
       - Required files (raw LLM output, verification table, final
         analysis, spec retrospective, prompt log, optional Stage 2
         feedback-response memo)
       - Final analysis structure (required sections, length, ratio
         citations, Du Pont commentary)
       - Verification table structure (≥5 rows, manual vs LLM columns,
         named-range formulas, discrepancy notes)
       - Spec retrospective structure (section verdicts, top-three gaps,
         revisions, effectiveness rating)
       - Stage 2 feedback incorporation (follow-up memo OR commits on
         the Stage 2 memo file after the Stage 2 deadline)
       - Repo polish (LICENSE, .gitignore, repo description, per-
         directory READMEs, filename convention, commit hygiene,
         public visibility)
  3. Looks the student up in prior-stage reports (Stage 0..4) for repo
     URL and carry-over flags.
  4. Scores the 6-criterion rubric (25/10/25/20/5/15), 70% floor.
  5. Writes/updates `../graded/STAGE5_GRADES.md` (internal — has scores).
  6. Writes per-student PR-ready feedback files under
     `../graded/_pr_feedback/{lastname}/feedback-file.md` (NO scores —
     these are intended to be opened as PRs on student repos).
  7. Moves the source zip to `graded/` and cleans scratch.

The six-criterion rubric (per `stage5-llm-analysis-evaluation.md`):

    Analytical correctness                       /25
    Manual verification artifact                 /10
    LLM evaluation + spec retrospective          /25
    Strategic recommendations + executive voice  /20
    Stage 2 feedback incorporation               /5
    Repo polish                                  /15
                                                = /100

Floor policy: 70% for any submission with a working repo and a final-
analysis file present. Pass `--floor` to override.

Score-privacy: scores live ONLY in the internal `STAGE5_GRADES.md` and
instructor email. The per-student feedback files written under
`_pr_feedback/` omit scores so they can be pushed to public student repos.

USAGE:
    python grade_stage5.py <export.zip> [--floor=70] [--no-move]
        [--prior-stage0=path] [--prior-stage2=path] [--prior-stage3=path]
        [--prior-stage4=path]

REQUIREMENTS:
    - `gh` CLI authenticated against github.com
"""
from __future__ import annotations

import argparse
import base64
import json
import re
import shutil
import subprocess
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _grading_comments import (
    Suggestion,
    backward,
    core,
    render_suggestions,
)
from _safe_zip import safe_extractall

STAGE_N = 5
DEFAULT_FLOOR_PCT = 70
TOTAL_POINTS = 100
STAGE_LABEL = "Stage 5 — LLM Analysis, Executive Evaluation, and Repo Polish"
INSTRUCTOR_GITHUB_HANDLE = "adamwstauffer"

# Validation regexes for subprocess argv safety
_SAFE_OWNER_RE = re.compile(r"^[A-Za-z0-9](?:[A-Za-z0-9_.-]{0,38})$")
_SAFE_REPO_RE = re.compile(r"^[A-Za-z0-9_.][A-Za-z0-9_.-]{0,99}$")
_SAFE_PATH_RE = re.compile(r"^[A-Za-z0-9._/-]{1,200}$")
_SAFE_BRANCH_RE = re.compile(r"^[A-Za-z0-9_][A-Za-z0-9._/-]{0,99}$")
_SAFE_HANDLE_RE = re.compile(r"^[A-Za-z0-9](?:[A-Za-z0-9-]{0,38})$")

# Naming convention: YYYY-MM-DD-{lastname}-{company}-{kind}.{ext}
CANONICAL_FILENAME_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2}-[a-z]+(?:-[a-z]+)*-[a-z0-9]+(?:-[a-z0-9]+)*-"
    r"[a-z]+(?:-[a-z]+)*\.[a-z]+$"
)
# Loose pattern — date prefix is enough to call the file "stage-tagged"
DATED_FILENAME_RE = re.compile(r"^\d{4}-\d{2}-\d{2}-.+\.[a-z]+$")

# Path heuristics for the six required Stage 5 artifacts
ARTIFACT_PATTERNS = {
    "llm_raw":        [r"deliverables/.*llm[- ]raw.*\.md$"],
    "verification":   [r"analysis/validation/.*stage5[- ]verification.*\.md$",
                       r"analysis/validation/.*verification.*\.md$"],
    "final_analysis": [r"deliverables/.*final[- ]analysis.*\.md$"],
    "retrospective":  [r"deliverables/.*spec[- ]retrospective.*\.md$",
                       r"deliverables/.*retrospective.*\.md$"],
    "prompt_log":     [r"deliverables/prompt[- ]log\.md$"],
    "s2_feedback":    [r"docs/decisions/.*stage2[- ]feedback[- ]response.*\.md$",
                       r"docs/decisions/.*feedback[- ]response.*\.md$"],
}

# Final-analysis required section headings (loose match)
FINAL_ANALYSIS_SECTIONS = {
    "company_summary": [r"company\s*(&|and)?\s*data\s+summary", r"company\s+summary"],
    "ratio_results":   [r"ratio\s+results?", r"results?\s*(&|and)?\s*interpretation"],
    "du_pont":         [r"du\s*pont", r"roe\s+decomposition"],
    "strategic_recs":  [r"strategic\s+recommendation", r"recommendation"],
    "llm_evaluation":  [r"llm\s+evaluation", r"evaluation\s*(&|and)?\s*annotation",
                        r"ai\s+evaluation"],
    "executive_voice": [r"executive\s+justification", r"executive\s+thesis",
                        r"executive\s+summary", r"final\s+thesis"],
}

# Spec retrospective required components
RETRO_SIGNALS = {
    "verdicts":      [r"clear\s*/\s*vague\s*/\s*missing", r"section.by.section\s+verdict",
                      r"\bverdict\b"],
    "gaps":          [r"top[- ]three\s+gaps", r"top\s+3\s+gaps", r"\bgap\s*\d", r"gap\s+#?\d"],
    "revisions":     [r"\bthree\s+revisions?\b", r"revisions?\s+I\s+would", r"\brevision\s*\d"],
    "rating":        [r"effectiveness\s+rating", r"rating.*[1-5]\s*/\s*5", r"\b[1-5]\s*/\s*5\b"],
    "process_fb":    [r"process\s+feedback", r"retrospective\s+process",
                      r"template\s+suggestion"],
}

GITHUB_URL_RE = re.compile(
    r"https?://github\.com/(?P<owner>[A-Za-z0-9_.-]+)/(?P<repo>[A-Za-z0-9_.-]+)"
)
FOLDER_NAME_RE = re.compile(
    r"^(?P<sid>\d+)-\d+\s*-\s*(?P<name>.+?)\s*-\s*"
    r"(?P<month>[A-Za-z]+)\s+(?P<day>\d+),\s*(?P<year>\d{4})\s+"
    r"(?P<h>\d{1,4})\s*(?P<ampm>AM|PM)\s*$"
)

LETTER_GRADE_SCALE = [
    ("A+", 97, None), ("A", 93, 97), ("A-", 90, 93),
    ("B+", 87, 90), ("B", 83, 87), ("B-", 80, 83),
    ("C+", 77, 80), ("C", 73, 77), ("C-", 70, 73),
    ("D+", 67, 70), ("D", 65, 67), ("F", 0, 65),
]


# ------------------------------------------------------------------
# Submission discovery
# ------------------------------------------------------------------

@dataclass
class Submission:
    student_id: str
    student_name: str
    submitted_at: datetime | None
    repo_url: str = ""
    owner: str = ""
    repo: str = ""


def _parse_folder_name(name: str):
    m = FOLDER_NAME_RE.match(name)
    if not m:
        return None, None, None
    h = m.group("h")
    if len(h) == 3:
        hour, minute = int(h[0]), int(h[1:])
    elif len(h) == 4:
        hour, minute = int(h[:2]), int(h[2:])
    else:
        hour, minute = int(h), 0
    ampm = m.group("ampm").upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    try:
        dt = datetime.strptime(
            f"{m.group('month')} {m.group('day')} {m.group('year')}", "%b %d %Y"
        ).replace(hour=hour, minute=minute)
    except ValueError:
        dt = None
    return m.group("sid"), m.group("name").strip(), dt


def _scan_html_for_url(folder: Path) -> str | None:
    for html in folder.glob("*.html"):
        text = html.read_text(encoding="utf-8", errors="ignore")
        m = GITHUB_URL_RE.search(text)
        if m:
            return m.group(0)
    return None


def discover_submissions(export_path: Path) -> list[Submission]:
    if export_path.is_file() and export_path.suffix.lower() == ".zip":
        scratch = export_path.parent / f"_{export_path.stem}_extracted"
        scratch.mkdir(exist_ok=True)
        with zipfile.ZipFile(export_path) as zf:
            safe_extractall(zf, scratch)
        root = scratch
    elif export_path.is_dir():
        root = export_path
    else:
        raise SystemExit(f"Cannot read submission export at {export_path}")

    subs: list[Submission] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        sid, name, dt = _parse_folder_name(child.name)
        if not name:
            continue
        repo_url = _scan_html_for_url(child) or ""
        owner, repo = "", ""
        if repo_url:
            m = GITHUB_URL_RE.search(repo_url)
            if m:
                owner, repo = m.group("owner"), m.group("repo")
        subs.append(Submission(
            student_id=sid or "", student_name=name, submitted_at=dt,
            repo_url=repo_url, owner=owner, repo=repo,
        ))
    subs.sort(key=lambda s: s.student_name.lower())
    return subs


# ------------------------------------------------------------------
# Prior-stage parsing
# ------------------------------------------------------------------

@dataclass
class PriorGrade:
    student_name: str
    repo_url: str
    final_score: int
    raw_section: str
    carry_over_tags: set[str] = field(default_factory=set)


def _normalize_name(name: str) -> str:
    tokens = re.findall(r"[A-Za-z]+", name.lower())
    return " ".join(sorted(tokens))


def parse_prior_report(report_path: Path) -> dict[str, PriorGrade]:
    if not report_path.exists():
        return {}
    text = report_path.read_text(encoding="utf-8")
    out: dict[str, PriorGrade] = {}
    header_re = re.compile(r"^## (\d+)\. (.+?) — \*\*(\d+) / 100\*\*", re.MULTILINE)
    headers = [(m.start(), m) for m in header_re.finditer(text)]
    headers.append((len(text), None))
    for i in range(len(headers) - 1):
        start, m = headers[i]
        end, _ = headers[i + 1]
        section = text[start:end]
        name = m.group(2).strip()
        score = int(m.group(3))
        url_m = GITHUB_URL_RE.search(section)
        repo_url = url_m.group(0) if url_m else ""
        carry = set()
        if re.search(r"not a write collaborator|instructor_not_collaborator",
                     section, re.IGNORECASE):
            carry.add("INSTRUCTOR_NOT_COLLAB")
        if re.search(r"filename convention", section, re.IGNORECASE):
            carry.add("FILENAME_CONVENTION")
        out[_normalize_name(name)] = PriorGrade(
            student_name=name, repo_url=repo_url, final_score=score,
            raw_section=section, carry_over_tags=carry,
        )
    return out


def lookup_prior(prior: dict[str, PriorGrade], student_name: str):
    key = _normalize_name(student_name)
    if key in prior:
        return prior[key]
    submitted_tokens = set(key.split())
    for pkey, pg in prior.items():
        if not pkey:
            continue
        prior_tokens = set(pkey.split())
        if submitted_tokens.issubset(prior_tokens) or prior_tokens.issubset(submitted_tokens):
            return pg
    return None


def merge_prior(*priors: dict[str, PriorGrade]) -> dict[str, PriorGrade]:
    out: dict[str, PriorGrade] = {}
    for prior in priors:
        for key, pg in prior.items():
            if key not in out:
                out[key] = PriorGrade(
                    student_name=pg.student_name, repo_url=pg.repo_url,
                    final_score=pg.final_score, raw_section=pg.raw_section,
                    carry_over_tags=set(pg.carry_over_tags),
                )
            else:
                out[key].carry_over_tags |= pg.carry_over_tags
                if not out[key].repo_url and pg.repo_url:
                    out[key].repo_url = pg.repo_url
    return out


# ------------------------------------------------------------------
# gh helpers
# ------------------------------------------------------------------

def _gh(*args: str) -> str:
    try:
        proc = subprocess.run(
            ["gh", *args], check=False, capture_output=True, text=True,
            encoding="utf-8", errors="replace",
        )
    except FileNotFoundError:
        return ""
    if proc.returncode != 0:
        return ""
    return proc.stdout or ""


def _safe_inputs(owner: str, repo: str, branch: str | None = None) -> bool:
    if not _SAFE_OWNER_RE.match(owner or ""):
        return False
    if not _SAFE_REPO_RE.match(repo or ""):
        return False
    if branch is not None and not _SAFE_BRANCH_RE.match(branch):
        return False
    return True


def _download_blob(owner: str, repo: str, path: str, branch: str) -> str | None:
    if not _safe_inputs(owner, repo, branch):
        return None
    if not _SAFE_PATH_RE.match(path):
        return None
    raw = _gh(
        "api", f"repos/{owner}/{repo}/contents/{path}",
        "-X", "GET", "-f", f"ref={branch}",
    )
    if not raw:
        return None
    try:
        meta = json.loads(raw)
    except json.JSONDecodeError:
        return None
    content = meta.get("content")
    if not content:
        return None
    try:
        return base64.b64decode(content).decode("utf-8", errors="replace")
    except Exception:
        return None


# ------------------------------------------------------------------
# Repo inspection
# ------------------------------------------------------------------

@dataclass
class RepoInspection:
    queried: bool = False
    accessible: bool = False
    private: bool = False
    default_branch: str = "main"
    description: str = ""
    license_name: str = ""
    has_gitignore: bool = False
    has_license_file: bool = False

    tree_paths: list[str] = field(default_factory=list)
    artifact_paths: dict[str, str] = field(default_factory=dict)  # key -> path
    dir_readmes: dict[str, bool] = field(default_factory=dict)    # dir -> has README
    dated_files_total: int = 0
    dated_files_canonical: int = 0

    commit_count: int = 0
    descriptive_commit_count: int = 0
    stage2_memo_path: str = ""
    stage2_memo_recent_commit_messages: list[str] = field(default_factory=list)

    instructor_is_collaborator: bool = False
    collaborator_permission: str = ""
    collaborator_check_ran: bool = False
    error: str = ""


VAGUE_MESSAGE_PATTERNS = [
    re.compile(r"^(initial commit|create [a-z0-9._-]+|update [a-z0-9._-]+)$",
               re.IGNORECASE),
    re.compile(r"^(wip|fix|fix typo|update|test|asdf|stuff|misc)$",
               re.IGNORECASE),
    re.compile(r"^.{0,5}$"),
]


def _is_descriptive(msg: str) -> bool:
    first = (msg or "").split("\n", 1)[0].strip()
    if not first:
        return False
    for pat in VAGUE_MESSAGE_PATTERNS:
        if pat.match(first):
            return False
    return True


def _match_artifact(paths: list[str], patterns: list[str]) -> str:
    for p in paths:
        low = p.lower()
        for pat in patterns:
            if re.search(pat, low):
                return p
    return ""


def inspect_repo(owner: str, repo: str) -> RepoInspection:
    info = RepoInspection(queried=True)
    if not owner or not repo:
        info.error = "no repo URL known"
        return info
    if not _safe_inputs(owner, repo):
        info.error = "unsafe owner/repo characters"
        return info
    meta_raw = _gh("api", f"repos/{owner}/{repo}")
    if not meta_raw:
        info.error = "repo not accessible"
        return info
    try:
        meta = json.loads(meta_raw)
    except json.JSONDecodeError:
        info.error = "could not parse repo metadata"
        return info
    info.accessible = True
    info.private = bool(meta.get("private"))
    info.description = (meta.get("description") or "").strip()
    info.default_branch = meta.get("default_branch") or "main"
    if not _SAFE_BRANCH_RE.match(info.default_branch):
        info.default_branch = "main"
    if meta.get("license"):
        info.license_name = meta["license"].get("spdx_id") or meta["license"].get("name") or ""

    tree_raw = _gh(
        "api", f"repos/{owner}/{repo}/git/trees/{info.default_branch}",
        "-X", "GET", "-f", "recursive=1",
    )
    if tree_raw:
        try:
            tree = json.loads(tree_raw).get("tree", [])
        except json.JSONDecodeError:
            tree = []
        blob_paths = [e.get("path", "") for e in tree if e.get("type") == "blob"]
        tree_dirs = [e.get("path", "") for e in tree if e.get("type") == "tree"]
        info.tree_paths = blob_paths

        # License + .gitignore
        for p in blob_paths:
            low = p.lower()
            if low in ("license", "license.md", "license.txt"):
                info.has_license_file = True
            if low == ".gitignore":
                info.has_gitignore = True

        # Required artifacts
        for key, patterns in ARTIFACT_PATTERNS.items():
            match = _match_artifact(blob_paths, patterns)
            if match:
                info.artifact_paths[key] = match

        # Per-directory README presence
        for d in tree_dirs:
            if not d or d.startswith("."):
                continue
            has_readme = any(
                p.lower() == f"{d}/readme.md" for p in blob_paths
            )
            info.dir_readmes[d] = has_readme

        # Dated-filename / canonical-filename tallies (root-level metadata
        # files like LICENSE / README.md are exempted; we only care about
        # stage artifacts dropped into deliverables/, analysis/, docs/,
        # models/, data/).
        considered_prefixes = (
            "deliverables/", "analysis/", "docs/decisions/", "docs/specs/",
            "models/", "data/",
        )
        for p in blob_paths:
            if not any(p.startswith(pre) for pre in considered_prefixes):
                continue
            name = p.rsplit("/", 1)[-1]
            if name.lower() == "readme.md":
                continue
            if name.lower().startswith("."):
                continue
            if DATED_FILENAME_RE.match(name):
                info.dated_files_total += 1
                if CANONICAL_FILENAME_RE.match(name):
                    info.dated_files_canonical += 1

        # Stage 2 memo path (used to detect feedback incorporation commits)
        for p in blob_paths:
            if re.search(r"docs/decisions/.*selection\.md$", p, re.IGNORECASE):
                info.stage2_memo_path = p
                break

    # Commits — paginate up to ~300
    all_commits: list[dict] = []
    for page in (1, 2, 3):
        c_raw = _gh(
            "api", f"repos/{owner}/{repo}/commits",
            "-X", "GET", "-f", "per_page=100", "-f", f"page={page}",
        )
        if not c_raw:
            break
        try:
            page_commits = json.loads(c_raw)
        except json.JSONDecodeError:
            break
        if not page_commits:
            break
        all_commits.extend(page_commits)
        if len(page_commits) < 100:
            break

    info.commit_count = len(all_commits)
    info.descriptive_commit_count = sum(
        1 for c in all_commits
        if _is_descriptive(c.get("commit", {}).get("message", ""))
    )

    # Recent-ish messages on the Stage 2 memo (signal of feedback incorporation)
    if info.stage2_memo_path:
        if _SAFE_PATH_RE.match(info.stage2_memo_path):
            mc_raw = _gh(
                "api", f"repos/{owner}/{repo}/commits",
                "-X", "GET", "-f", f"path={info.stage2_memo_path}",
                "-f", "per_page=20",
            )
            if mc_raw:
                try:
                    mc = json.loads(mc_raw)
                except json.JSONDecodeError:
                    mc = []
                info.stage2_memo_recent_commit_messages = [
                    (c.get("commit", {}).get("message") or "")[:200]
                    for c in mc
                ]

    # Collaborator check
    if _SAFE_HANDLE_RE.match(INSTRUCTOR_GITHUB_HANDLE):
        perm_raw = _gh(
            "api",
            f"repos/{owner}/{repo}/collaborators/{INSTRUCTOR_GITHUB_HANDLE}/permission",
        )
        if perm_raw:
            info.collaborator_check_ran = True
            try:
                perm = json.loads(perm_raw).get("permission", "")
            except json.JSONDecodeError:
                perm = ""
            info.collaborator_permission = perm
            info.instructor_is_collaborator = perm in {"admin", "write", "maintain"}
    return info


# ------------------------------------------------------------------
# Per-artifact inspection
# ------------------------------------------------------------------

@dataclass
class ArtifactInspection:
    final_analysis_words: int = 0
    final_analysis_sections: dict[str, bool] = field(default_factory=dict)
    final_analysis_ratio_citations: int = 0    # numeric ratios cited (%, x, days)
    final_analysis_recs_count: int = 0          # detected H1/H2/numbered recs

    verification_rows: int = 0                  # data rows in verification table
    verification_match_columns: bool = False    # has Match? column
    verification_distinct_categories: int = 0   # distinct ratio categories covered

    retro_signals: dict[str, bool] = field(default_factory=dict)
    retro_words: int = 0

    prompt_log_words: int = 0
    prompt_log_blocks: int = 0
    has_s5_session: bool = False                # prompt-log mentions Stage 5

    s2_feedback_response_words: int = 0
    s2_feedback_commit_signals: int = 0         # commits referencing PR/feedback


def _word_count(text: str) -> int:
    return len(re.findall(r"\b\w+\b", text or ""))


def _has_heading(text: str, patterns: list[str]) -> bool:
    for pat in patterns:
        if re.search(rf"^\s*#{{1,4}}[^\n]*{pat}", text, re.MULTILINE | re.IGNORECASE):
            return True
    return False


def inspect_artifacts(owner: str, repo: str, branch: str,
                      paths: dict[str, str]) -> ArtifactInspection:
    a = ArtifactInspection()

    # Final analysis
    fp = paths.get("final_analysis")
    if fp:
        text = _download_blob(owner, repo, fp, branch) or ""
        a.final_analysis_words = _word_count(text)
        for key, pats in FINAL_ANALYSIS_SECTIONS.items():
            a.final_analysis_sections[key] = _has_heading(text, pats)
        # Numeric citations: %, x-multipliers, days, currency totals
        a.final_analysis_ratio_citations = (
            len(re.findall(r"\b\d+(?:\.\d+)?\s*%", text))
            + len(re.findall(r"\b\d+(?:\.\d+)?\s*x\b", text, re.IGNORECASE))
            + len(re.findall(r"\b\d+(?:\.\d+)?\s*days?\b", text, re.IGNORECASE))
        )
        # Recommendations count — sub-headings under Strategic Recommendations
        # OR enumerated items in 1./2./3./4./5. form near the recs section
        m = re.search(r"strategic\s+recommendations?", text, re.IGNORECASE)
        if m:
            after = text[m.end():m.end() + 6000]
            nums = re.findall(
                r"^\s*(?:#{2,4}\s*Recommendation\s*\d+|"
                r"^\s*\*\*Recommendation\s*\d+|"
                r"^\s*\d+\.\s+\*\*[A-Z])",
                after, re.MULTILINE,
            )
            a.final_analysis_recs_count = len(nums)

    # Verification table
    vp = paths.get("verification")
    if vp:
        text = _download_blob(owner, repo, vp, branch) or ""
        rows = 0
        for line in text.splitlines():
            stripped = line.strip()
            if not stripped.startswith("|") or "---" in stripped:
                continue
            if stripped.count("|") < 4:  # need ≥4 columns
                continue
            # Skip header rows (first occurrence after a separator)
            rows += 1
        # Subtract 1 header row per table (rough; assumes 1 table)
        a.verification_rows = max(0, rows - 1)
        a.verification_match_columns = (
            "match?" in text.lower() or "match " in text.lower()
        )
        cats_present = set()
        low = text.lower()
        for cat in ("roa", "roe", "current ratio", "quick ratio", "asset turnover",
                    "debt", "interest", "turnover", "margin", "eva", "mva"):
            if cat in low:
                cats_present.add(cat)
        a.verification_distinct_categories = len(cats_present)

    # Spec retrospective
    rp = paths.get("retrospective")
    if rp:
        text = _download_blob(owner, repo, rp, branch) or ""
        a.retro_words = _word_count(text)
        for key, pats in RETRO_SIGNALS.items():
            present = any(re.search(p, text, re.IGNORECASE) for p in pats)
            a.retro_signals[key] = present

    # Prompt log
    pp = paths.get("prompt_log")
    if pp:
        text = _download_blob(owner, repo, pp, branch) or ""
        a.prompt_log_words = _word_count(text)
        blocks = re.findall(
            r"^(?:#{2,4}\s*Prompt\s*\d+|^\*\*Prompt\s*\d+[\*:])",
            text, re.MULTILINE | re.IGNORECASE,
        )
        a.prompt_log_blocks = len(blocks)
        a.has_s5_session = bool(
            re.search(r"stage\s*5|final\s+analysis|verification\s+table|"
                      r"spec[- ]retrospective", text, re.IGNORECASE)
        )

    # Stage 2 feedback response
    s2p = paths.get("s2_feedback")
    if s2p:
        text = _download_blob(owner, repo, s2p, branch) or ""
        a.s2_feedback_response_words = _word_count(text)

    return a


def _count_s2_feedback_commit_signals(messages: list[str]) -> int:
    """Count commits whose messages reference PR review / feedback / instructor."""
    signals = 0
    pat = re.compile(
        r"\b(pr\s+(?:comment|feedback|review)|instructor\s+pr|"
        r"review\s+(?:comment|feedback)|feedback\s+incorporation|"
        r"address.*feedback|respond.*pr|stage\s*2\s+feedback)",
        re.IGNORECASE,
    )
    for m in messages:
        if pat.search(m or ""):
            signals += 1
    return signals


# ------------------------------------------------------------------
# Scoring
# ------------------------------------------------------------------

@dataclass
class Grade:
    submission: Submission
    repo: RepoInspection
    artifacts: ArtifactInspection
    prior: PriorGrade | None

    score_analytical: int = 0   # /25
    score_verification: int = 0 # /10
    score_evaluation: int = 0   # /25 (LLM eval + spec retrospective)
    score_strategic: int = 0    # /20 (recs + executive voice)
    score_s2_feedback: int = 0  # /5
    score_polish: int = 0       # /15

    flags: list[str] = field(default_factory=list)

    @property
    def raw_total(self) -> int:
        return (self.score_analytical + self.score_verification
                + self.score_evaluation + self.score_strategic
                + self.score_s2_feedback + self.score_polish)


def score(sub: Submission, repo: RepoInspection,
          artifacts: ArtifactInspection, prior: PriorGrade | None) -> Grade:
    g = Grade(submission=sub, repo=repo, artifacts=artifacts, prior=prior)
    if not repo.accessible:
        g.flags.append("REPO_INACCESSIBLE")
        return g
    paths = repo.artifact_paths

    # ----- Criterion 1: Analytical correctness, 25 pts -----
    # Necessarily approximate. Signals:
    #   - final analysis present (5)
    #   - required sections present (up to 10 — 1.66/section, 6 sections)
    #   - length in band (1,200–1,800 words excl appendix) (5)
    #   - ratio citations present (5)
    c1 = 0
    if paths.get("final_analysis"):
        c1 += 5
        present = sum(1 for v in artifacts.final_analysis_sections.values() if v)
        c1 += round(10 * present / max(1, len(artifacts.final_analysis_sections)))
        wc = artifacts.final_analysis_words
        if 1000 <= wc <= 2400:
            c1 += 5
        elif 600 <= wc <= 3500:
            c1 += 3
        elif wc > 0:
            c1 += 1
        if artifacts.final_analysis_ratio_citations >= 20:
            c1 += 5
        elif artifacts.final_analysis_ratio_citations >= 10:
            c1 += 3
        elif artifacts.final_analysis_ratio_citations >= 3:
            c1 += 1
    else:
        g.flags.append("FINAL_ANALYSIS_MISSING")
    g.score_analytical = min(25, c1)

    # ----- Criterion 2: Manual verification artifact, 10 pts -----
    # 10 if file present AND ≥5 rows AND has Match? column AND covers multiple cats
    # 7 if ≥3 rows, has comparison
    # 3 if file present but thin
    # 0 if missing
    c2 = 0
    if paths.get("verification"):
        rows = artifacts.verification_rows
        if rows >= 5 and artifacts.verification_match_columns:
            c2 = 10
        elif rows >= 5:
            c2 = 8
        elif rows >= 3:
            c2 = 7
        elif rows >= 1:
            c2 = 4
        else:
            c2 = 2
            g.flags.append("VERIFICATION_EMPTY")
        if artifacts.verification_distinct_categories < 3 and c2 >= 5:
            c2 -= 1  # nudge for category-coverage thinness
    else:
        g.flags.append("VERIFICATION_MISSING")
    g.score_verification = min(10, max(0, c2))

    # ----- Criterion 3: LLM evaluation + spec retrospective, 25 pts -----
    # LLM evaluation: section in final analysis (up to 8)
    # Spec retrospective file: presence (5) + signal coverage (up to 12)
    c3 = 0
    if artifacts.final_analysis_sections.get("llm_evaluation"):
        c3 += 8
    else:
        c3 += 0
        g.flags.append("LLM_EVAL_SECTION_MISSING")
    if paths.get("retrospective"):
        c3 += 5
        signal_hits = sum(1 for v in artifacts.retro_signals.values() if v)
        c3 += round(12 * signal_hits / max(1, len(artifacts.retro_signals)))
        if artifacts.retro_words < 400:
            g.flags.append("RETRO_THIN")
            c3 -= 2
    else:
        g.flags.append("RETROSPECTIVE_MISSING")
    g.score_evaluation = min(25, max(0, c3))

    # ----- Criterion 4: Strategic recs + executive voice, 20 pts -----
    # Strategic Recs section (8) + Executive Justification (6) + recs count (6)
    c4 = 0
    if artifacts.final_analysis_sections.get("strategic_recs"):
        c4 += 8
    if artifacts.final_analysis_sections.get("executive_voice"):
        c4 += 6
    else:
        g.flags.append("EXEC_VOICE_MISSING")
    recs = artifacts.final_analysis_recs_count
    if recs >= 3:
        c4 += 6
    elif recs >= 1:
        c4 += 3
    else:
        # Brief allows recs to live as numbered list; soft-fail
        pass
    g.score_strategic = min(20, c4)

    # ----- Criterion 5: Stage 2 feedback incorporation, 5 pts -----
    # Either a follow-up memo or commits referencing PR feedback
    c5 = 0
    if paths.get("s2_feedback") and artifacts.s2_feedback_response_words >= 100:
        c5 = 5
    elif paths.get("s2_feedback"):
        c5 = 3
    sig = _count_s2_feedback_commit_signals(repo.stage2_memo_recent_commit_messages)
    if sig and c5 < 5:
        c5 = max(c5, 4)
    if c5 == 0:
        # Soft signal: did the Stage 2 memo see ANY commits after its initial add?
        if len(repo.stage2_memo_recent_commit_messages) >= 2:
            c5 = 2
        else:
            g.flags.append("S2_FEEDBACK_MISSING")
    g.score_s2_feedback = min(5, c5)

    # ----- Criterion 6: Repo polish, 15 pts -----
    # LICENSE (2) + .gitignore (1) + description (2) + dir READMEs (3) +
    # canonical filenames (3) + commit hygiene (3) + public (1)
    c6 = 0
    if repo.has_license_file or repo.license_name:
        c6 += 2
    else:
        g.flags.append("LICENSE_MISSING")
    if repo.has_gitignore:
        c6 += 1
    else:
        g.flags.append("GITIGNORE_MISSING")
    if repo.description:
        c6 += 2
    else:
        g.flags.append("DESCRIPTION_MISSING")
    if repo.dir_readmes:
        with_readme = sum(1 for v in repo.dir_readmes.values() if v)
        c6 += round(3 * with_readme / len(repo.dir_readmes))
    if repo.dated_files_total > 0:
        canonical_share = repo.dated_files_canonical / repo.dated_files_total
        if canonical_share >= 0.8:
            c6 += 3
        elif canonical_share >= 0.5:
            c6 += 2
        elif canonical_share >= 0.25:
            c6 += 1
        else:
            g.flags.append("FILENAME_CONVENTION")
    if repo.commit_count >= 10:
        ratio = repo.descriptive_commit_count / max(1, repo.commit_count)
        if ratio >= 0.75:
            c6 += 3
        elif ratio >= 0.5:
            c6 += 2
        else:
            c6 += 1
            g.flags.append("COMMIT_HYGIENE")
    elif repo.commit_count >= 5:
        c6 += 1
    if not repo.private:
        c6 += 1
    else:
        g.flags.append("PRIVATE_REPO")
    g.score_polish = min(15, c6)

    # Required-artifact summary flag (used by tagline & PR feedback)
    missing = [k for k in ("llm_raw", "verification", "final_analysis",
                            "retrospective", "prompt_log") if not paths.get(k)]
    if missing:
        g.flags.append("ARTIFACTS_INCOMPLETE")

    if repo.collaborator_check_ran and not repo.instructor_is_collaborator:
        g.flags.append("INSTRUCTOR_NOT_COLLABORATOR")

    if not g.flags and g.raw_total >= 95:
        g.flags = ["STRONG"]
    return g


# ------------------------------------------------------------------
# Suggestions
# ------------------------------------------------------------------

def _suggestions_for(g: Grade) -> list[Suggestion]:
    s: list[Suggestion] = []
    repo = g.repo
    art = g.artifacts
    paths = repo.artifact_paths

    if "REPO_INACCESSIBLE" in g.flags:
        s.append(core(
            "We weren't able to reach your repo via the GitHub API. Confirm "
            "the URL on Lamaku and that the repo is **public** (Settings → "
            "Danger Zone → Change visibility). Stage 5's deliverable is the "
            "repo URL — without access we cannot grade the artifact."
        ))
        return s

    # Missing required artifacts
    if "FINAL_ANALYSIS_MISSING" in g.flags:
        s.append(core(
            "**Final analysis file not found.** The Stage 5 brief asks for "
            "`deliverables/YYYY-MM-DD-{lastname}-{company}-final-analysis.md` "
            "with six required sections (Company & Data Summary, Ratio Results, "
            "Du Pont, Strategic Recommendations, LLM Evaluation & Annotations, "
            "Executive Justification). Without this file, most of the 25-pt "
            "analytical-correctness rubric line cannot be earned."
        ))
    if "VERIFICATION_MISSING" in g.flags:
        s.append(core(
            "**Manual verification table not found** at "
            "`analysis/validation/YYYY-MM-DD-{lastname}-{company}-stage5-verification.md`. "
            "The brief asks for at least five ratios recomputed by hand from "
            "your Stage 3 financials, compared against the LLM's output, with "
            "a Match? column and a one-line note on each discrepancy. This is "
            "10% of the Stage 5 rubric — purely additive if you build it."
        ))
    elif "VERIFICATION_EMPTY" in g.flags:
        s.append(core(
            f"The verification table at `{paths.get('verification')}` was "
            f"found but has zero data rows. Recompute at least five ratios "
            f"from across categories (don't pick five liquidity ratios — "
            f"show coverage), with formula in named-range notation, manual "
            f"value, LLM value, Match? indicator, and a one-line note."
        ))
    elif art.verification_rows < 5:
        s.append(core(
            f"The verification table at `{paths.get('verification')}` has "
            f"{art.verification_rows} data row(s) — the brief asks for ≥5 "
            f"ratios. Strong work picks ratios the LLM is most likely to get "
            f"wrong (averages, start-of-year values, unit conversions)."
        ))

    if "RETROSPECTIVE_MISSING" in g.flags:
        s.append(core(
            "**Spec retrospective not found.** Copy "
            "`docs/templates/spec-retrospective-template.md`, rename it per "
            "`YYYY-MM-DD-{lastname}-{company}-spec-retrospective.md`, and "
            "drop it under `deliverables/`. The template requires: section-"
            "by-section Clear/Vague/Missing verdicts, top three gaps with "
            "evidence, three revisions, an effectiveness rating, a forward "
            "link, and a ≤150-word process-feedback note. \"My spec was "
            "perfect\" earns nothing; specificity is the rubric."
        ))
    elif "RETRO_THIN" in g.flags:
        missing_signals = [k for k, v in art.retro_signals.items() if not v]
        s.append(core(
            f"The spec retrospective at `{paths.get('retrospective')}` is "
            f"on the short side ({art.retro_words} words). Sections the "
            f"detector didn't see clear signals for: "
            f"{', '.join(missing_signals) or 'none'}. Each part of the "
            f"template — verdict rows, top three gaps with evidence, three "
            f"revisions, 1–5 effectiveness rating, process feedback — should "
            f"be visibly tagged in the markdown so a reviewer can find it."
        ))

    if "LLM_EVAL_SECTION_MISSING" in g.flags:
        s.append(core(
            "**LLM Evaluation & Annotations section missing** from the final "
            "analysis. This is where you tell the reader what the LLM got "
            "right, where it deviated/hallucinated/oversimplified, and "
            "whether each issue was a spec gap or an LLM limitation. The "
            "Stage 5 grade leans heavily on this section because it's the "
            "evidence of judgment."
        ))
    if "EXEC_VOICE_MISSING" in g.flags:
        s.append(core(
            "**Executive Justification section missing** from the final "
            "analysis. Close with the investment/strategic thesis in your "
            "own voice — the \"so what?\" that only a human with judgment "
            "can provide. The brief calls this out explicitly because it's "
            "the differentiator between an annotated LLM dump and a "
            "professional deliverable."
        ))

    # Stage 2 feedback
    if "S2_FEEDBACK_MISSING" in g.flags:
        s.append(core(
            "**No visible response to the instructor's Stage 2 PR feedback.** "
            "Two acceptable forms (either counts): (a) commits on your Stage 2 "
            "memo that reference specific PR comments (e.g., \"Tighten "
            "hypothesis 2 per instructor PR comment #3\"), or (b) a follow-up "
            "memo at `docs/decisions/YYYY-MM-DD-{lastname}-stage2-feedback-"
            "response.md`. This is 5% — modest, but easy to earn."
        ))

    # Repo polish
    if "LICENSE_MISSING" in g.flags:
        s.append(core(
            "Add a `LICENSE` at the repo root (MIT or Apache-2.0 — pick one). "
            "Signals \"this is intended as a portfolio piece you can "
            "reference.\" Stage 5 polish checklist."
        ))
    if "GITIGNORE_MISSING" in g.flags:
        s.append(core(
            "Add a `.gitignore` excluding common scratch (`.DS_Store`, "
            "`~$*.xlsx`, `*.tmp`). Keeps the commit history focused on real "
            "work — checklist item."
        ))
    if "DESCRIPTION_MISSING" in g.flags:
        s.append(core(
            "Set the one-line **repo description** on the GitHub repo page "
            "(the field at the top — not the README). One sentence "
            "summarizing the project. The repo description is what shows up "
            "in LinkedIn link previews and search results."
        ))
    if "FILENAME_CONVENTION" in g.flags:
        s.append(core(
            f"Only {g.repo.dated_files_canonical}/{g.repo.dated_files_total} "
            f"dated stage files match the canonical "
            f"`YYYY-MM-DD-{{lastname}}-{{company}}-{{kind}}.{{ext}}` "
            f"pattern. Renames are cheap; the convention pays off when "
            f"someone (or a future you) scans the repo six months from now."
        ))
    if "COMMIT_HYGIENE" in g.flags:
        ratio = g.repo.descriptive_commit_count / max(1, g.repo.commit_count)
        s.append(core(
            f"Commit-message hygiene is thin "
            f"({g.repo.descriptive_commit_count}/{g.repo.commit_count} "
            f"descriptive, {ratio*100:.0f}%). Even a quick pass — squashing "
            f"`Update X` messages into descriptive ones — improves the "
            f"history a recruiter would actually read."
        ))

    if "INSTRUCTOR_NOT_COLLABORATOR" in g.flags:
        s.append(backward(
            f"**Still open from Stage 2:** `@{INSTRUCTOR_GITHUB_HANDLE}` is "
            "not a Write collaborator on your repo. Without it, the Stage 2 "
            "feedback-PR workflow couldn't run — which directly affects the "
            "5-pt feedback-incorporation rubric line at this stage. "
            "Settings → Collaborators → Add people → **Write**."
        ))

    if "ARTIFACTS_INCOMPLETE" in g.flags and "FINAL_ANALYSIS_MISSING" not in g.flags:
        present = ", ".join(k for k in paths) or "none"
        s.append(core(
            f"Artifact checklist: detected ({present}). Stage 5 asks for six "
            f"files (raw LLM output, verification table, final analysis, spec "
            f"retrospective, prompt log, plus the optional Stage 2 feedback-"
            f"response). Each missing file is unscored credit — even a sparse "
            f"draft of the missing artifacts moves the needle."
        ))

    if "STRONG" in g.flags:
        s.append(core(
            "Strong submission across all six rubric lines — artifacts complete, "
            "named-range notation visible, repo polish in place. Consider a "
            "LinkedIn post once the deadline lifts; this is a portfolio-shaped "
            "project and the repo can carry weight in a job search."
        ))

    # Stage 5 is the final stage — no forward pointer to a Stage 6.
    return s


def _summary_tagline(g: Grade) -> str:
    flags = set(g.flags)
    if "REPO_INACCESSIBLE" in flags:
        return "Repo not accessible — confirm URL and public visibility."
    if "STRONG" in flags:
        return "Strong on merit; see suggestions for refinements."
    if "FINAL_ANALYSIS_MISSING" in flags:
        return "Final analysis file missing — most of the 25-pt analytical rubric unearned."
    if "RETROSPECTIVE_MISSING" in flags and "VERIFICATION_MISSING" in flags:
        return "Retrospective + verification missing — both are quick wins worth pursuing."
    if "VERIFICATION_MISSING" in flags:
        return "Verification table missing — purely additive 10 pts to recover."
    if "RETROSPECTIVE_MISSING" in flags:
        return "Spec retrospective missing — half of the 25-pt evaluation rubric depends on it."
    if "S2_FEEDBACK_MISSING" in flags:
        return "Stage 2 feedback incorporation not visible — quickest 5 pts to recover."
    return "See per-student suggestions for refinements."


# ------------------------------------------------------------------
# Report writers
# ------------------------------------------------------------------

def _letter_for(score_pts: int) -> str:
    for letter, lo, hi in LETTER_GRADE_SCALE:
        lo_pts = round(lo * TOTAL_POINTS / 100)
        if score_pts >= lo_pts and (hi is None or score_pts < round(hi * TOTAL_POINTS / 100)):
            return letter
    return "F"


def _final_score(g: Grade, floor_pct: int) -> int:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    if g.repo.accessible and g.raw_total < floor_value:
        return floor_value
    return g.raw_total


def _floor_was_applied(g: Grade, floor_pct: int) -> bool:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    return g.repo.accessible and g.raw_total < floor_value


def _student_section(n: int, g: Grade, floor_pct: int) -> str:
    sub = g.submission
    repo = g.repo
    art = g.artifacts
    submitted = sub.submitted_at.strftime("%Y-%m-%d %I:%M %p") if sub.submitted_at else "—"
    raw = g.raw_total
    floor_applied = _floor_was_applied(g, floor_pct)
    final = _final_score(g, floor_pct)
    letter = _letter_for(final)

    lines: list[str] = []
    suffix = ", floor applied" if floor_applied else ""
    lines.append(f"## {n}. {sub.student_name} — **{final} / 100** ({letter}{suffix})")
    lines.append("")
    lines.append(f"**Repo:** {sub.repo_url or '(not provided)'}")
    lines.append(f"**Submitted:** {submitted}")
    lines.append("")
    lines.append("| Criterion | Earned | Notes |")
    lines.append("|-----------|--------|-------|")
    fa_secs = sum(1 for v in art.final_analysis_sections.values() if v)
    fa_total = max(1, len(art.final_analysis_sections))
    lines.append(
        f"| Analytical correctness | {g.score_analytical} / 25 | "
        f"final analysis {art.final_analysis_words}w, "
        f"{fa_secs}/{fa_total} required sections, "
        f"{art.final_analysis_ratio_citations} ratio citations. |"
    )
    lines.append(
        f"| Manual verification | {g.score_verification} / 10 | "
        f"{art.verification_rows} data row(s); Match? col "
        f"{'✓' if art.verification_match_columns else '—'}. |"
    )
    retro_hits = sum(1 for v in art.retro_signals.values() if v)
    retro_total = max(1, len(art.retro_signals))
    lines.append(
        f"| LLM eval + spec retrospective | {g.score_evaluation} / 25 | "
        f"LLM-eval section "
        f"{'✓' if art.final_analysis_sections.get('llm_evaluation') else '—'}; "
        f"retro {art.retro_words}w, {retro_hits}/{retro_total} template signals. |"
    )
    lines.append(
        f"| Strategic recs + exec voice | {g.score_strategic} / 20 | "
        f"recs detected: {art.final_analysis_recs_count}; exec-voice section "
        f"{'✓' if art.final_analysis_sections.get('executive_voice') else '—'}. |"
    )
    lines.append(
        f"| Stage 2 feedback incorporation | {g.score_s2_feedback} / 5 | "
        f"response memo "
        f"{'✓ (' + str(art.s2_feedback_response_words) + 'w)' if art.s2_feedback_response_words else '—'}; "
        f"S2-memo commits: {len(repo.stage2_memo_recent_commit_messages)}. |"
    )
    polish_bits = []
    polish_bits.append("LIC ✓" if (repo.has_license_file or repo.license_name) else "LIC —")
    polish_bits.append(".gi ✓" if repo.has_gitignore else ".gi —")
    polish_bits.append("desc ✓" if repo.description else "desc —")
    polish_bits.append(f"dirREADMEs {sum(1 for v in repo.dir_readmes.values() if v)}/{len(repo.dir_readmes)}")
    polish_bits.append(f"canon {repo.dated_files_canonical}/{repo.dated_files_total}")
    polish_bits.append(f"commits {repo.descriptive_commit_count}/{repo.commit_count}")
    lines.append(
        f"| Repo polish | {g.score_polish} / 15 | "
        f"{'; '.join(polish_bits)}. |"
    )

    if floor_applied:
        floor_value = round(TOTAL_POINTS * floor_pct / 100)
        lines.append(f"| **Raw total** | **{raw} / 100** | |")
        lines.append(
            f"| **Floor adjustment** | **+{floor_value - raw}** | "
            f"Working repo present — floor of {floor_pct} applied. |"
        )
        lines.append(f"| **Final** | **{final} / 100** | |")
    else:
        merit_note = " Above the floor — earned on merit." if raw >= round(
            TOTAL_POINTS * floor_pct / 100
        ) else ""
        lines.append(f"| **Total** | **{final} / 100** |{merit_note} |")
    lines.append("")

    suggestions = _suggestions_for(g)
    lines.extend(render_suggestions(suggestions, stage_n=STAGE_N))
    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def _build_full_report(grades: list[Grade], floor_pct: int, today: datetime) -> str:
    lines = [
        "# BUS-629 Stage 5 — Grade Report",
        "",
        f"**Stage:** {STAGE_LABEL} (25% of project score)",
        f"**Graded:** {today.strftime('%Y-%m-%d')}",
        f"**Submissions reviewed:** {len(grades)}",
        f"**Floor policy:** {floor_pct}% floor for any working repo with a "
        "final analysis file present.",
        "**Score privacy:** scores live in this internal file only. The "
        "per-student suggestion blocks are also rendered under "
        "`_pr_feedback/<lastname>/feedback-file.md` *without scores* for "
        "PR delivery to student repos.",
        "",
        "---",
        "",
        "## Rubric (recap)",
        "",
        "| Criterion | Weight |",
        "|-----------|--------|",
        "| Analytical correctness | 25% |",
        "| Manual verification artifact | 10% |",
        "| LLM evaluation + spec retrospective | 25% |",
        "| Strategic recommendations + executive voice | 20% |",
        "| Stage 2 feedback incorporation | 5% |",
        "| Repo polish | 15% |",
        "",
        "---",
        "",
    ]
    for i, g in enumerate(grades, 1):
        lines.append(_student_section(i, g, floor_pct))
    lines.append("## Class summary")
    lines.append("")
    lines.append("| Student | Score | Notes |")
    lines.append("|---------|-------|-------|")
    for g in grades:
        lines.append(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |"
        )
    lines.append("")
    finals = [_final_score(g, floor_pct) for g in grades if g.repo.accessible]
    submitted_n = len(finals)
    mean = sum(finals) / submitted_n if submitted_n else 0.0
    floors = sum(1 for g in grades if _floor_was_applied(g, floor_pct))
    lines.append(f"**Mean (submissions only):** {mean:.1f}")
    lines.append(f"**Submission rate:** {submitted_n} of {len(grades)}")
    lines.append(f"**Floor applied:** {floors} of {submitted_n} submissions")
    lines.append("")
    return "\n".join(lines)


def write_or_update_grade_report(grades: list[Grade], floor_pct: int,
                                   report_path: Path,
                                   today: datetime | None = None) -> list[Grade]:
    today = today or datetime.now()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    if not report_path.exists():
        report_path.write_text(_build_full_report(grades, floor_pct, today),
                               encoding="utf-8")
        return list(grades)

    text = report_path.read_text(encoding="utf-8")
    existing_urls = {
        u.lower().rstrip("/")
        for u in re.findall(r"https?://github\.com/[A-Za-z0-9_./-]+", text)
    }
    existing_names = {
        _normalize_name(n)
        for n in re.findall(r"^## \d+\. ([^—\n]+?) — \*\*\d+ / 100\*\*",
                            text, re.MULTILINE)
    }
    new_grades = []
    for g in grades:
        url_key = (g.submission.repo_url or "").lower().rstrip("/")
        name_key = _normalize_name(g.submission.student_name)
        if url_key and url_key in existing_urls:
            continue
        if name_key and name_key in existing_names:
            continue
        new_grades.append(g)
    if not new_grades:
        return []

    nums = [int(n) for n in re.findall(r"^## (\d+)\.", text, re.MULTILINE)]
    next_n = (max(nums) + 1) if nums else 1
    chunks = []
    for g in new_grades:
        chunks.append(_student_section(next_n, g, floor_pct))
        next_n += 1
    new_block = "\n".join(chunks)

    if "## Class summary" in text:
        text = text.replace("## Class summary", new_block + "\n## Class summary", 1)
    else:
        text = text.rstrip() + "\n\n" + new_block

    table_re = re.compile(
        r"(## Class summary\s*\n+"
        r"\| Student \| Score \| Notes \|\n"
        r"\| ?-+ ?\| ?-+ ?\| ?-+ ?\|\n"
        r"(?:\|[^\n]*\|\n)+)"
    )
    m = table_re.search(text)
    if m:
        new_rows = "".join(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |\n"
            for g in new_grades
        )
        text = text[: m.end()] + new_rows + text[m.end():]

    text = re.sub(
        r"\*\*Graded:\*\* [^\n]*",
        f"**Graded:** {today.strftime('%Y-%m-%d')} "
        f"({', '.join(g.submission.student_name for g in new_grades)} added)",
        text, count=1,
    )

    report_path.write_text(text, encoding="utf-8")
    return new_grades


# ------------------------------------------------------------------
# PR-ready feedback file (no scores)
# ------------------------------------------------------------------

def _lastname_slug(name: str) -> str:
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9'-]+", name)
    if not tokens:
        return "student"
    last = tokens[-1].lower()
    return re.sub(r"[^a-z0-9-]", "", last) or "student"


def write_pr_feedback(g: Grade, today: datetime, feedback_dir: Path) -> Path:
    slug = _lastname_slug(g.submission.student_name)
    out_dir = feedback_dir / slug
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "feedback-file.md"

    date_str = today.strftime("%Y-%m-%d")
    repo = g.repo
    art = g.artifacts
    paths = repo.artifact_paths

    lines: list[str] = [f"# Stage 5 review — {date_str}", ""]

    lines.append("## Artifact checklist")
    lines.append("")
    lines.append("| Artifact | Status | Path |")
    lines.append("|---|---|---|")
    art_labels = [
        ("Raw LLM output", "llm_raw"),
        ("Manual verification table", "verification"),
        ("Final analysis", "final_analysis"),
        ("Spec retrospective", "retrospective"),
        ("Prompt log", "prompt_log"),
        ("Stage 2 feedback response (optional)", "s2_feedback"),
    ]
    for label, key in art_labels:
        if paths.get(key):
            lines.append(f"| {label} | ✓ | `{paths[key]}` |")
        else:
            lines.append(f"| {label} | — | *(not detected)* |")
    lines.append("")

    lines.append("## Final analysis structure")
    lines.append("")
    lines.append(f"- Length: **{art.final_analysis_words} words** "
                 "(brief targets 1,200–1,800 words excluding appendix)")
    lines.append(f"- Ratio citations counted: **{art.final_analysis_ratio_citations}**")
    lines.append(f"- Recommendations detected: **{art.final_analysis_recs_count}**")
    lines.append("")
    section_labels = {
        "company_summary": "Company & Data Summary",
        "ratio_results": "Ratio Results & Interpretation",
        "du_pont": "Du Pont Analysis",
        "strategic_recs": "Strategic Recommendations",
        "llm_evaluation": "LLM Evaluation & Annotations",
        "executive_voice": "Executive Justification",
    }
    lines.append("| Required section | Detected? |")
    lines.append("|---|---|")
    for key, label in section_labels.items():
        present = "✓" if art.final_analysis_sections.get(key) else "—"
        lines.append(f"| {label} | {present} |")
    lines.append("")

    lines.append("## Verification table")
    lines.append("")
    lines.append(f"- Data rows counted: **{art.verification_rows}** "
                 "(brief asks for ≥5)")
    lines.append(f"- Match? column present: "
                 f"**{'yes' if art.verification_match_columns else 'no'}**")
    lines.append(f"- Distinct ratio types referenced: "
                 f"**{art.verification_distinct_categories}**")
    lines.append("")

    lines.append("## Spec retrospective")
    lines.append("")
    lines.append(f"- Length: **{art.retro_words} words**")
    lines.append("")
    retro_label = {
        "verdicts": "Section-by-section verdicts (Clear/Vague/Missing)",
        "gaps": "Top three gaps",
        "revisions": "Three revisions",
        "rating": "Effectiveness rating (1–5)",
        "process_fb": "Process feedback note",
    }
    lines.append("| Template signal | Detected? |")
    lines.append("|---|---|")
    for key, label in retro_label.items():
        present = "✓" if art.retro_signals.get(key) else "—"
        lines.append(f"| {label} | {present} |")
    lines.append("")

    lines.append("## Repo polish snapshot")
    lines.append("")
    license_state = repo.license_name or ("present" if repo.has_license_file else "missing")
    lines.append(f"- LICENSE: **{license_state}**")
    lines.append(f"- .gitignore: **{'present' if repo.has_gitignore else 'missing'}**")
    lines.append(f"- Repo description set: **{'yes' if repo.description else 'no'}**")
    if repo.dir_readmes:
        with_readme = sum(1 for v in repo.dir_readmes.values() if v)
        lines.append(f"- Per-directory READMEs: **{with_readme}/{len(repo.dir_readmes)}**")
    lines.append(f"- Filename convention: "
                 f"**{repo.dated_files_canonical}/{repo.dated_files_total}** "
                 "dated files match canonical pattern")
    lines.append(f"- Commit hygiene: **{repo.descriptive_commit_count}/{repo.commit_count}** "
                 "commits descriptive")
    lines.append(f"- Public visibility: **{'yes' if not repo.private else 'no (private)'}**")
    lines.append("")

    suggestions = _suggestions_for(g)
    lines.extend(render_suggestions(suggestions, stage_n=STAGE_N))
    lines.append("")
    lines.append(
        "*This review is feedback-only — no scores included.* Score numbers "
        "live in the internal grade report and the instructor's email; this "
        "file is intended for review against your repo state."
    )
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


# ------------------------------------------------------------------
# Worksheet
# ------------------------------------------------------------------

def build_worksheet(grades: list[Grade], floor_pct: int, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"
    headers = [
        "Student", "Submitted", "Repo URL",
        "Analytical /25", "Verification /10", "Eval+Retro /25",
        "Strategic+Voice /20", "S2 Feedback /5", "Polish /15",
        "Raw /100", "Final /100", "Floored /100",
        "LLM-raw", "Verification path", "Final-analysis path",
        "Retrospective path", "Prompt-log path", "S2-feedback path",
        "FA words", "FA sections", "FA cites", "FA recs",
        "Verif rows", "Retro words", "Retro signals",
        "LIC", ".gitignore", "desc", "dir-READMEs", "canonical",
        "commits/desc", "Flags",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F8CBAD")
    floor_fill = PatternFill("solid", fgColor="E2EFDA")
    final_col = headers.index("Final /100") + 1
    floor_col = headers.index("Floored /100") + 1
    flags_col = headers.index("Flags") + 1

    for g in grades:
        sub = g.submission
        repo = g.repo
        art = g.artifacts
        paths = repo.artifact_paths
        submitted = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else ""
        fa_secs = sum(1 for v in art.final_analysis_sections.values() if v)
        fa_total = len(art.final_analysis_sections) or 1
        retro_hits = sum(1 for v in art.retro_signals.values() if v)
        retro_total = len(art.retro_signals) or 1
        row = [
            sub.student_name, submitted, sub.repo_url,
            g.score_analytical, g.score_verification, g.score_evaluation,
            g.score_strategic, g.score_s2_feedback, g.score_polish,
            g.raw_total, g.raw_total, None,
            paths.get("llm_raw", ""), paths.get("verification", ""),
            paths.get("final_analysis", ""), paths.get("retrospective", ""),
            paths.get("prompt_log", ""), paths.get("s2_feedback", ""),
            art.final_analysis_words, f"{fa_secs}/{fa_total}",
            art.final_analysis_ratio_citations, art.final_analysis_recs_count,
            art.verification_rows, art.retro_words, f"{retro_hits}/{retro_total}",
            "Y" if (repo.has_license_file or repo.license_name) else "N",
            "Y" if repo.has_gitignore else "N",
            "Y" if repo.description else "N",
            f"{sum(1 for v in repo.dir_readmes.values() if v)}/{len(repo.dir_readmes)}",
            f"{repo.dated_files_canonical}/{repo.dated_files_total}",
            f"{repo.descriptive_commit_count}/{repo.commit_count}",
            ", ".join(g.flags),
        ]
        ws.append(row)
        r = ws.max_row
        final_ref = f"{get_column_letter(final_col)}{r}"
        ws.cell(
            row=r, column=floor_col,
            value=f'=IF({final_ref}=0,0,MAX({final_ref},{floor_value}))',
        ).fill = floor_fill
        if not repo.accessible:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = error_fill
        elif g.flags and g.flags != ["STRONG"]:
            ws.cell(row=r, column=flags_col).fill = flag_fill

    widths = [24, 17, 50,
              12, 13, 14, 16, 13, 11,
              10, 11, 13,
              30, 32, 32, 32, 28, 32,
              10, 12, 10, 10,
              11, 12, 14,
              5, 11, 6, 14, 12, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(output_path)


# ------------------------------------------------------------------
# Sweep entry point
# ------------------------------------------------------------------

def rescore_from_repo(
    student_name: str,
    repo_url: str,
    submitted_at: datetime | None = None,
    prior: PriorGrade | None = None,
    student_id: str = "",
) -> Grade | None:
    m = GITHUB_URL_RE.search(repo_url)
    if not m:
        return None
    owner, repo = m.group("owner"), m.group("repo")
    repo_info = inspect_repo(owner, repo)
    if repo_info.accessible:
        art = inspect_artifacts(owner, repo, repo_info.default_branch,
                                 repo_info.artifact_paths)
    else:
        art = ArtifactInspection()
    sub = Submission(
        student_id=student_id, student_name=student_name,
        submitted_at=submitted_at, repo_url=repo_url, owner=owner, repo=repo,
    )
    return score(sub, repo_info, art, prior)


# ------------------------------------------------------------------
# Workflow helper
# ------------------------------------------------------------------

def _detect_workflow(export_path: Path):
    if not (export_path.is_file() and export_path.suffix.lower() == ".zip"):
        return None
    parent = export_path.parent
    if parent.name.lower() != "ungraded":
        return None
    graded = parent.parent / "graded"
    scratch = parent / f"_{export_path.stem}_extracted"
    return graded, scratch


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("export", type=Path)
    p.add_argument("--floor", type=int, default=DEFAULT_FLOOR_PCT)
    p.add_argument("--prior-stage0", type=Path, default=None)
    p.add_argument("--prior-stage2", type=Path, default=None)
    p.add_argument("--prior-stage3", type=Path, default=None)
    p.add_argument("--prior-stage4", type=Path, default=None)
    p.add_argument("--out", type=Path, default=None)
    p.add_argument("--no-move", action="store_true")
    args = p.parse_args(argv)

    subs = discover_submissions(args.export)
    if not subs:
        print("No submissions discovered.")
        return 1
    print(f"Found {len(subs)} submission(s).")

    base_root = args.export.resolve().parents[2]
    s0 = args.prior_stage0 or (base_root / "stage0" / "graded" / "STAGE0_GRADES.md")
    s2 = args.prior_stage2 or (base_root / "stage2" / "graded" / "STAGE2_GRADES.md")
    s3 = args.prior_stage3 or (base_root / "stage3" / "graded" / "STAGE3_GRADES.md")
    s4 = args.prior_stage4 or (base_root / "stage4" / "graded" / "STAGE4_GRADES.md")
    priors = []
    for label, p_path in (("Stage 0", s0), ("Stage 2", s2),
                           ("Stage 3", s3), ("Stage 4", s4)):
        pm = parse_prior_report(p_path) if p_path.exists() else {}
        if pm:
            print(f"Loaded {label} records: {len(pm)} from {p_path}")
        priors.append(pm)
    prior = merge_prior(*priors)

    # Repo-URL fill-in from priors.
    for s in subs:
        if s.repo_url:
            continue
        pg = lookup_prior(prior, s.student_name)
        if pg and pg.repo_url:
            m = GITHUB_URL_RE.search(pg.repo_url)
            if m:
                s.repo_url = pg.repo_url
                s.owner = m.group("owner")
                s.repo = m.group("repo")

    grades: list[Grade] = []
    for s in subs:
        print(f"  inspecting {s.student_name} ...", end=" ", flush=True)
        if not s.owner:
            grades.append(score(s, RepoInspection(error="no repo URL"),
                                 ArtifactInspection(), lookup_prior(prior, s.student_name)))
            print("NO REPO URL")
            continue
        repo_info = inspect_repo(s.owner, s.repo)
        if repo_info.accessible:
            art = inspect_artifacts(s.owner, s.repo, repo_info.default_branch,
                                     repo_info.artifact_paths)
        else:
            art = ArtifactInspection()
        pg = lookup_prior(prior, s.student_name)
        g = score(s, repo_info, art, pg)
        grades.append(g)
        if not repo_info.accessible:
            print(f"REPO INACCESSIBLE: {repo_info.error}")
        else:
            paths = repo_info.artifact_paths
            artifact_count = sum(1 for k in ("llm_raw", "verification",
                                              "final_analysis", "retrospective",
                                              "prompt_log") if paths.get(k))
            collab = (
                f"collab=Y({repo_info.collaborator_permission})"
                if repo_info.collaborator_check_ran and repo_info.instructor_is_collaborator
                else (f"collab=N({repo_info.collaborator_permission or 'none'})"
                       if repo_info.collaborator_check_ran
                       else "collab=?")
            )
            print(
                f"raw={g.raw_total}/100 "
                f"art={artifact_count}/5 "
                f"a={g.score_analytical} v={g.score_verification} "
                f"e={g.score_evaluation} s={g.score_strategic} "
                f"s2={g.score_s2_feedback} p={g.score_polish} "
                f"{collab} flags={','.join(g.flags) or '-'}"
            )

    workflow = _detect_workflow(args.export)
    if args.out is not None:
        worksheet_path = args.out
    elif workflow is not None:
        graded_dir, _ = workflow
        graded_dir.mkdir(parents=True, exist_ok=True)
        ws_dir = graded_dir / "_worksheets"
        ws_dir.mkdir(exist_ok=True)
        worksheet_path = ws_dir / f"stage5-{args.export.stem}.xlsx"
    else:
        worksheet_path = (
            (args.export.parent if args.export.is_file() else args.export)
            / "_grading" / "stage5-grading-worksheet.xlsx"
        )

    build_worksheet(grades, args.floor, worksheet_path)
    print(f"\nWrote worksheet: {worksheet_path}")

    today = datetime.now()
    if workflow is not None:
        graded_dir, scratch = workflow
        report_path = graded_dir / "STAGE5_GRADES.md"
        new_entries = write_or_update_grade_report(grades, args.floor,
                                                    report_path, today)
        if new_entries:
            names = ", ".join(g.submission.student_name for g in new_entries)
            print(f"Updated report: {report_path} (+{len(new_entries)} new: {names})")
        else:
            print(f"Report already up to date: {report_path}")

        pr_dir = graded_dir / "_pr_feedback"
        pr_dir.mkdir(exist_ok=True)
        for g in grades:
            if not g.repo.accessible:
                continue
            out = write_pr_feedback(g, today, pr_dir)
            print(f"Wrote PR feedback: {out}")

        if not args.no_move:
            dest = graded_dir / args.export.name
            if dest.exists():
                print(f"Source zip already exists at {dest} — leaving original in place.")
            else:
                shutil.move(str(args.export), str(dest))
                print(f"Moved source zip → {dest}")
            if scratch.exists():
                shutil.rmtree(scratch, ignore_errors=True)
                print(f"Cleaned up scratch extract at {scratch}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
