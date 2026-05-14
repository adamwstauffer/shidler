"""BUS-629 Stage 0 (portfolio repo setup) grading scanner.

Stage 0 submissions are URL-only: students upload an HTML pointer to their
public GitHub repo via Lamaku. The instructor exports a zip containing one
`index.html` and per-student folders. This script:

  1. Reads the export zip (or a directory of extracted HTML files).
  2. Parses each submission for (student name, GitHub URL, submitted_at).
  3. Inspects each repo via the `gh` CLI:
       - public visibility
       - directory skeleton presence and per-directory README size
       - README / BIO / RESUME word counts
       - commit count + descriptive-message ratio
  4. Auto-scores the five-criterion Stage 0 rubric.
  5. Writes an Excel worksheet with rubric signals, tentative scores, and a
     curving block (80% / 90% / 100% floor options).

The five-criterion rubric (per `stage0-repo-setup.md`):

    Repo public + accessible                  /15
    Directory skeleton + READMEs              /20
    Bio quality (150–200 words; structured)   /25
    Resume quality (Penn-style; quantified)   /25
    Commit hygiene (≥2 commits; descriptive)  /15
                                              = /100

Floor policy is configurable: pass --floor=90 (default), --floor=80, or any
other percentage. The floor applies only to submissions with a *working*
public repo.

USAGE:
    python grade_stage0.py <export.zip> [--floor=90] [--out=path.xlsx]

REQUIREMENTS:
    - `gh` CLI authenticated against github.com
    - openpyxl (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_FLOOR_PCT = 90
TOTAL_POINTS = 100

REQUIRED_DIRS = [
    "docs",
    "docs/decisions",
    "docs/specs",
    "docs/plans",
    "models",
    "models/templates",
    "models/builds",
    "data",
    "analysis",
    "analysis/validation",
    "deliverables",
]
REQUIRED_TOP_FILES = ["README.md", "RESUME.md"]
OPTIONAL_TOP_FILES = ["BIO.md"]

VAGUE_MESSAGE_PATTERNS = [
    re.compile(r"^(initial commit|create [a-z0-9._-]+|update [a-z0-9._-]+)$",
               re.IGNORECASE),
    re.compile(r"^(wip|fix|fix typo|update|test|asdf|stuff|misc)$",
               re.IGNORECASE),
    re.compile(r"^.{0,5}$"),  # 5 chars or fewer
]

GITHUB_URL_RE = re.compile(
    r"https?://github\.com/(?P<owner>[A-Za-z0-9_.-]+)/(?P<repo>[A-Za-z0-9_.-]+)"
)
FOLDER_NAME_RE = re.compile(
    r"^(?P<sid>\d+)-\d+\s*-\s*(?P<name>.+?)\s*-\s*"
    r"(?P<month>[A-Za-z]+)\s+(?P<day>\d+),\s*(?P<year>\d{4})\s+"
    r"(?P<h>\d{1,4})\s*(?P<ampm>AM|PM)\s*$"
)


# ------------------------------------------------------------------
# Submission discovery
# ------------------------------------------------------------------

@dataclass
class Submission:
    student_id: str
    student_name: str
    submitted_at: datetime | None
    repo_url: str
    owner: str
    repo: str


class _LinkExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        for k, v in attrs:
            if k.lower() == "href" and v:
                self.links.append(v)


def _parse_folder_name(name: str) -> tuple[str | None, str | None, datetime | None]:
    """Folder convention: '<sid>-<course> - <Name> - <Mon D, YYYY HHMM AM|PM>'."""
    m = FOLDER_NAME_RE.match(name)
    if not m:
        return None, None, None
    h = m.group("h")
    if len(h) == 3:
        hour, minute = int(h[0]), int(h[1:])
    elif len(h) == 4:
        hour, minute = int(h[:2]), int(h[2:])
    else:
        hour, minute = int(h), 0
    ampm = m.group("ampm").upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    try:
        dt = datetime.strptime(
            f"{m.group('month')} {m.group('day')} {m.group('year')}", "%b %d %Y"
        ).replace(hour=hour, minute=minute)
    except ValueError:
        dt = None
    return m.group("sid"), m.group("name").strip(), dt


def discover_submissions(export_path: Path) -> list[Submission]:
    """Walk the export zip or directory and yield one Submission per student."""
    if export_path.is_file() and export_path.suffix.lower() == ".zip":
        scratch = export_path.parent / f"_{export_path.stem}_extracted"
        scratch.mkdir(exist_ok=True)
        with zipfile.ZipFile(export_path) as zf:
            zf.extractall(scratch)
        root = scratch
    elif export_path.is_dir():
        root = export_path
    else:
        raise SystemExit(f"Cannot read submission export at {export_path}")

    # Key by repo URL (lowercased) so the index.html and folder name don't
    # produce two entries for the same student. Prefer per-folder data when
    # available because the folder name carries the canonical "First Last".
    submissions: dict[str, Submission] = {}

    # Walk folders first — they have canonical name ordering.
    for child in root.iterdir():
        if not child.is_dir():
            continue
        sid, name, dt = _parse_folder_name(child.name)
        if not name:
            continue
        url = _scan_html_for_url(child)
        if not url:
            continue
        m = GITHUB_URL_RE.search(url)
        if not m:
            continue
        submissions[url.lower()] = Submission(
            student_id=sid or "",
            student_name=name,
            submitted_at=dt,
            repo_url=url,
            owner=m.group("owner"),
            repo=m.group("repo"),
        )

    # Fill in anyone the folder pass missed using index.html as a fallback.
    index_html = next(root.rglob("index.html"), None)
    if index_html is not None:
        text = index_html.read_text(encoding="utf-8", errors="ignore")
        for sub in _parse_index_html(text):
            submissions.setdefault(sub.repo_url.lower(), sub)

    return sorted(submissions.values(), key=lambda s: s.student_name.lower())


def _parse_index_html(text: str) -> list[Submission]:
    """Pull (name, github_url, submitted_at) tuples from the export index.

    The Lamaku index uses one row per student with a bold <b>Name</b>, an
    <a href> to the repo, and a `Submitted: <date>` cell.
    """
    rows: list[Submission] = []
    # Each student appears in a <tr><td colspan=2>...<b>NAME</b></td></tr>
    # followed by their content row. We split on the bold name marker.
    name_re = re.compile(r"<b>([^<]+)</b>")
    chunk_starts = [(m.start(), m.group(1).strip()) for m in name_re.finditer(text)]
    # Filter out non-name bolds like "Comments:" / "Submitted:" by requiring
    # a comma in the bolded text (the export uses "Last, First").
    chunk_starts = [(s, n) for s, n in chunk_starts if "," in n]
    chunk_starts.append((len(text), ""))
    for i in range(len(chunk_starts) - 1):
        start, name = chunk_starts[i]
        end, _ = chunk_starts[i + 1]
        chunk = text[start:end]
        url_match = GITHUB_URL_RE.search(chunk)
        if not url_match:
            continue
        url = url_match.group(0)
        sub_match = re.search(
            r"Submitted:.*?([A-Za-z]+ \d+,\s*\d{4}[^<]*)", chunk, re.DOTALL
        )
        dt: datetime | None = None
        if sub_match:
            raw = re.sub(r"\s+", " ", sub_match.group(1)).strip()
            for fmt in ("%b %d, %Y %I%M %p", "%B %d, %Y %I%M %p",
                        "%b %d, %Y", "%B %d, %Y"):
                try:
                    dt = datetime.strptime(raw, fmt)
                    break
                except ValueError:
                    continue
        # Normalize "Last, First" -> "First Last"
        if "," in name:
            last, first = (s.strip() for s in name.split(",", 1))
            display = f"{first} {last}"
        else:
            display = name
        rows.append(Submission(
            student_id="",
            student_name=display,
            submitted_at=dt,
            repo_url=url,
            owner=url_match.group("owner"),
            repo=url_match.group("repo"),
        ))
    return rows


def _scan_html_for_url(folder: Path) -> str | None:
    for html in folder.glob("*.html"):
        text = html.read_text(encoding="utf-8", errors="ignore")
        m = GITHUB_URL_RE.search(text)
        if m:
            return m.group(0)
    return None


# ------------------------------------------------------------------
# GitHub repo inspection
# ------------------------------------------------------------------

def _gh(*args: str) -> str:
    """Run `gh` and return stdout. Returns empty string on failure."""
    try:
        proc = subprocess.run(
            ["gh", *args], check=False, capture_output=True, text=True,
            encoding="utf-8", errors="replace",
        )
    except FileNotFoundError:
        raise SystemExit("`gh` CLI not found. Install GitHub CLI to grade Stage 0.")
    if proc.returncode != 0:
        return ""
    return proc.stdout or ""


@dataclass
class RepoInspection:
    accessible: bool = False
    private: bool = False
    default_branch: str = "main"
    tree: list[dict] = field(default_factory=list)
    commits: list[dict] = field(default_factory=list)
    readme: str = ""
    bio: str = ""
    resume: str = ""
    has_license: bool = False
    error: str = ""


def _fetch_file(owner: str, repo: str, path: str, branch: str) -> str:
    raw = _gh(
        "api", f"repos/{owner}/{repo}/contents/{path}?ref={branch}",
        "-H", "Accept: application/vnd.github.raw",
    )
    return raw


def inspect_repo(sub: Submission) -> RepoInspection:
    info = RepoInspection()
    meta_raw = _gh("api", f"repos/{sub.owner}/{sub.repo}")
    if not meta_raw:
        info.error = "repo not found or not accessible"
        return info
    try:
        meta = json.loads(meta_raw)
    except json.JSONDecodeError:
        info.error = "could not parse repo metadata"
        return info
    info.accessible = True
    info.private = bool(meta.get("private"))
    info.default_branch = meta.get("default_branch") or "main"
    info.has_license = meta.get("license") is not None

    tree_raw = _gh(
        "api",
        f"repos/{sub.owner}/{sub.repo}/git/trees/{info.default_branch}",
        "-X", "GET", "-f", "recursive=1",
    )
    if tree_raw:
        try:
            info.tree = json.loads(tree_raw).get("tree", [])
        except json.JSONDecodeError:
            pass

    # paginate up to ~100 commits — generous for Stage 0
    for page in (1, 2, 3):
        c_raw = _gh(
            "api", f"repos/{sub.owner}/{sub.repo}/commits",
            "-X", "GET", "-f", "per_page=100", "-f", f"page={page}",
        )
        if not c_raw:
            break
        try:
            page_commits = json.loads(c_raw)
        except json.JSONDecodeError:
            break
        if not page_commits:
            break
        info.commits.extend(page_commits)
        if len(page_commits) < 100:
            break

    # README / BIO / RESUME (case-insensitive)
    paths_by_lower: dict[str, str] = {}
    for entry in info.tree:
        if entry.get("type") != "blob":
            continue
        path = entry.get("path", "")
        if "/" in path:
            continue
        paths_by_lower[path.lower()] = path
    for key, attr in (("readme.md", "readme"), ("bio.md", "bio"),
                      ("resume.md", "resume")):
        actual = paths_by_lower.get(key)
        if actual:
            setattr(info, attr, _fetch_file(
                sub.owner, sub.repo, actual, info.default_branch
            ))
    return info


# ------------------------------------------------------------------
# Rubric scoring
# ------------------------------------------------------------------

@dataclass
class Grade:
    submission: Submission
    inspection: RepoInspection
    score_public: int = 0          # /15
    score_skeleton: int = 0        # /20
    score_bio: int = 0             # /25
    score_resume: int = 0          # /25
    score_commits: int = 0         # /15

    word_count_readme: int = 0
    word_count_bio: int = 0
    word_count_resume: int = 0
    dirs_present: int = 0
    dirs_total: int = len(REQUIRED_DIRS)
    placeholder_readme_count: int = 0
    meaningful_readme_count: int = 0
    commit_count: int = 0
    descriptive_commit_count: int = 0
    flags: list[str] = field(default_factory=list)

    @property
    def raw_total(self) -> int:
        return (
            self.score_public + self.score_skeleton + self.score_bio
            + self.score_resume + self.score_commits
        )


WORD_RE = re.compile(r"\b\w+\b")


def _word_count(text: str) -> int:
    return len(WORD_RE.findall(text))


def _is_placeholder(content: str, repo_name: str) -> bool:
    """A README is placeholder if it's <30 chars or just '# <reponame>'."""
    s = content.strip()
    if len(s) < 30:
        return True
    s_low = re.sub(r"[^a-z0-9]+", "", s.lower())
    rn_low = re.sub(r"[^a-z0-9]+", "", repo_name.lower())
    return s_low in {rn_low, "readme", "todo"}


def _is_descriptive(msg: str) -> bool:
    first = msg.split("\n", 1)[0].strip()
    if not first:
        return False
    for pat in VAGUE_MESSAGE_PATTERNS:
        if pat.match(first):
            return False
    return True


def score(sub: Submission, info: RepoInspection) -> Grade:
    g = Grade(submission=sub, inspection=info)

    # Criterion 1: Repo public + accessible /15
    if not info.accessible:
        g.flags.append("REPO_INACCESSIBLE")
        return g
    if info.private:
        g.flags.append("REPO_PRIVATE")
        g.score_public = 5
    else:
        g.score_public = 15

    # Build path index
    blob_paths = {
        e["path"] for e in info.tree if e.get("type") == "blob"
    }
    tree_paths = {
        e["path"] for e in info.tree if e.get("type") == "tree"
    }

    # Criterion 2: Directory skeleton + READMEs /20
    dirs_present = sum(1 for d in REQUIRED_DIRS
                       if d in tree_paths or d.lower() in {p.lower() for p in tree_paths})
    g.dirs_present = dirs_present
    skeleton_dir_pts = round(10 * dirs_present / len(REQUIRED_DIRS))

    # Per-directory README quality
    readme_per_dir: list[tuple[str, str]] = []
    for d in REQUIRED_DIRS:
        for candidate in (f"{d}/README.md", f"{d}/readme.md"):
            actual = next(
                (p for p in blob_paths if p.lower() == candidate.lower()), None
            )
            if actual:
                content = _fetch_file(
                    sub.owner, sub.repo, actual, info.default_branch
                )
                readme_per_dir.append((d, content))
                break
    meaningful = sum(
        1 for _, c in readme_per_dir if not _is_placeholder(c, sub.repo)
    )
    placeholders = len(readme_per_dir) - meaningful
    g.meaningful_readme_count = meaningful
    g.placeholder_readme_count = placeholders
    if readme_per_dir:
        readme_quality_pts = round(10 * meaningful / len(readme_per_dir))
    else:
        readme_quality_pts = 0
    g.score_skeleton = skeleton_dir_pts + readme_quality_pts

    # Criterion 3: Bio quality /25
    # Use BIO.md when present and substantive, else README.md.
    g.word_count_readme = _word_count(info.readme)
    g.word_count_bio = _word_count(info.bio)
    bio_text = info.bio if g.word_count_bio >= 80 else info.readme
    bio_wc = max(g.word_count_bio, g.word_count_readme) if not info.bio else g.word_count_bio
    if _is_placeholder(info.readme, sub.repo) and not info.bio:
        g.score_bio = 0
        g.flags.append("BIO_MISSING")
    elif bio_wc < 50:
        g.score_bio = 5
        g.flags.append("BIO_STUB")
    elif bio_wc < 100:
        g.score_bio = 12
    elif bio_wc < 150:
        g.score_bio = 18
    elif bio_wc <= 250:
        g.score_bio = 25
    else:
        g.score_bio = 22  # over the 200-word ceiling but still substantive

    # Criterion 4: Resume quality /25
    g.word_count_resume = _word_count(info.resume)
    if _is_placeholder(info.resume, sub.repo):
        g.score_resume = 0
        g.flags.append("RESUME_MISSING")
    elif g.word_count_resume < 100:
        g.score_resume = 5
        g.flags.append("RESUME_STUB")
    else:
        # Heuristics for Penn-style + quantification
        signals = 0
        lower = info.resume.lower()
        for kw in ("education", "experience", "skills"):
            if kw in lower:
                signals += 1
        # Quantification: count $, %, and bare numbers >= 3 digits.
        quant = (
            len(re.findall(r"\$[\d,]+", info.resume))
            + len(re.findall(r"\d+%", info.resume))
            + len(re.findall(r"\b\d{3,}\b", info.resume))
        )
        if signals >= 3 and quant >= 5 and g.word_count_resume >= 250:
            g.score_resume = 25
        elif signals >= 2 and quant >= 3:
            g.score_resume = 20
        elif signals >= 2:
            g.score_resume = 15
        else:
            g.score_resume = 10

    # Criterion 5: Commit hygiene /15
    g.commit_count = len(info.commits)
    g.descriptive_commit_count = sum(
        1 for c in info.commits
        if _is_descriptive(c.get("commit", {}).get("message", ""))
    )
    if g.commit_count < 2:
        g.score_commits = 0
        g.flags.append("FEW_COMMITS")
    else:
        ratio = (
            g.descriptive_commit_count / g.commit_count if g.commit_count else 0
        )
        if g.commit_count >= 5 and ratio >= 0.75:
            g.score_commits = 15
        elif g.commit_count >= 3 and ratio >= 0.5:
            g.score_commits = 13
        elif ratio >= 0.5:
            g.score_commits = 10
        else:
            g.score_commits = 8

    if not g.flags and g.raw_total >= 90:
        g.flags = ["STRONG"]
    return g


# ------------------------------------------------------------------
# Excel output
# ------------------------------------------------------------------

LETTER_GRADE_SCALE = [
    ("A+", 97, None), ("A", 93, 97), ("A-", 90, 93),
    ("B+", 87, 90),  ("B", 83, 87), ("B-", 80, 83),
    ("C+", 77, 80),  ("C", 73, 77), ("C-", 70, 73),
    ("D+", 67, 70),  ("D", 65, 67), ("F", 0, 65),
]


def build_worksheet(grades: list[Grade], floor_pct: int, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"

    headers = [
        "Student", "Submitted", "Repo URL",
        "Public /15", "Skeleton /20", "Bio /25", "Resume /25", "Commits /15",
        "Raw /100", "Final /100", "Floored /100",
        "Bio WC", "Resume WC", "Dirs Present", "Meaningful READMEs",
        "Commits", "Descriptive Commits",
        "Flags", "Comments",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    floor_value = round(TOTAL_POINTS * floor_pct / 100)

    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F8CBAD")
    floor_fill = PatternFill("solid", fgColor="E2EFDA")

    final_col = headers.index("Final /100") + 1
    floor_col = headers.index("Floored /100") + 1
    flags_col = headers.index("Flags") + 1

    first_data_row = 2
    for g in grades:
        sub = g.submission
        info = g.inspection
        submitted = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else ""
        row = [
            sub.student_name,
            submitted,
            sub.repo_url,
            g.score_public,
            g.score_skeleton,
            g.score_bio,
            g.score_resume,
            g.score_commits,
            g.raw_total,
            g.raw_total,  # editable final
            None,
            g.word_count_bio if info.bio else g.word_count_readme,
            g.word_count_resume,
            f"{g.dirs_present}/{g.dirs_total}",
            f"{g.meaningful_readme_count}/{g.meaningful_readme_count + g.placeholder_readme_count}",
            g.commit_count,
            g.descriptive_commit_count,
            ", ".join(g.flags),
            info.error,
        ]
        ws.append(row)
        r = ws.max_row
        # Floor formula: floor only applies if the repo was accessible.
        final_ref = f"{get_column_letter(final_col)}{r}"
        ws.cell(
            row=r, column=floor_col,
            value=(
                f'=IF({final_ref}=0,0,'
                f'MAX({final_ref},{floor_value}))'
            ),
        ).fill = floor_fill
        if info.error:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = error_fill
        elif g.flags and g.flags != ["STRONG"]:
            ws.cell(row=r, column=flags_col).fill = flag_fill

    widths = [
        24, 17, 50,
        11, 12, 9, 11, 11,
        10, 11, 13,
        9, 11, 13, 18,
        9, 18,
        38, 30,
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # ----- Summary sheet -----
    sm = wb.create_sheet("Summary")
    sm.append(["Metric", "Value"])
    sm.append(["Total submissions", len(grades)])
    sm.append(["Accessible repos", sum(1 for g in grades if g.inspection.accessible)])
    sm.append(["Floor policy", f"{floor_pct}% (={floor_value}/100)"])
    if grades:
        avg = sum(g.raw_total for g in grades) / len(grades)
        sm.append(["Average raw score", round(avg, 1)])
    sm.append([])
    sm.append(["Floor rule",
               "Curved = MAX(Final, floor). Never reduces raw score; "
               "only lifts working-repo submissions to the floor."])
    sm.append(["Non-submissions", "Final stays at 0 (floor not applied)."])

    cv_col_letter = get_column_letter(floor_col)
    cv_rng = f"Grading!${cv_col_letter}${first_data_row}:${cv_col_letter}${ws.max_row}"

    sm.append([])
    sm.append(["Letter", "Min %", "Min /100", "Count", "Histogram"])
    hr = sm.max_row
    for col in range(1, 6):
        c = sm.cell(row=hr, column=col)
        c.font = header_font
        c.fill = header_fill
    for letter, min_pct, max_pct in LETTER_GRADE_SCALE:
        min_pts = round(min_pct * TOTAL_POINTS / 100, 2)
        r = sm.max_row + 1
        if letter == "F":
            max_pts = round(max_pct * TOTAL_POINTS / 100, 2)
            cv_f = f'=COUNTIFS({cv_rng},">0",{cv_rng},"<"&{max_pts})'
        elif max_pct is None:
            cv_f = f'=COUNTIF({cv_rng},">="&{min_pts})'
        else:
            max_pts = round(max_pct * TOTAL_POINTS / 100, 2)
            cv_f = f'=COUNTIFS({cv_rng},">="&{min_pts},{cv_rng},"<"&{max_pts})'
        sm.cell(row=r, column=1, value=letter)
        sm.cell(row=r, column=2, value=min_pct)
        sm.cell(row=r, column=3, value=min_pts)
        sm.cell(row=r, column=4, value=cv_f)
        sm.cell(row=r, column=5, value=f'=REPT("█",D{r})')
    sm.column_dimensions["A"].width = 30
    sm.column_dimensions["B"].width = 30
    sm.column_dimensions["E"].width = 40
    for cell in sm[1]:
        cell.font = header_font
        cell.fill = header_fill

    wb.save(output_path)


# ------------------------------------------------------------------
# Markdown grade report (STAGE0_GRADES.md)
# ------------------------------------------------------------------

def _final_score(g: Grade, floor_pct: int) -> int:
    info = g.inspection
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    if info.accessible and not info.private and g.raw_total < floor_value:
        return floor_value
    return g.raw_total


def _floor_was_applied(g: Grade, floor_pct: int) -> bool:
    info = g.inspection
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    return info.accessible and not info.private and g.raw_total < floor_value


def _letter_for(score: int) -> str:
    for letter, lo, hi in LETTER_GRADE_SCALE:
        lo_pts = round(lo * TOTAL_POINTS / 100)
        if score >= lo_pts and (hi is None or score < round(hi * TOTAL_POINTS / 100)):
            return letter
    return "F"


def _repo_note(g: Grade) -> str:
    info = g.inspection
    if not info.accessible:
        return "Repo not accessible. " + (info.error or "")
    if info.private:
        return "Repo is private — flip to public so graders can see it."
    return "Public, accessible without login."


def _summary_tagline(g: Grade) -> str:
    """Short note for the class-summary table row."""
    flags = set(g.flags)
    if "REPO_INACCESSIBLE" in flags:
        return "Repo not accessible — confirm URL and visibility."
    if "REPO_PRIVATE" in flags:
        return "Repo is private — flip to public."
    if "BIO_MISSING" in flags and "RESUME_MISSING" in flags:
        return "Floor applied — bio and resume both need real content."
    if "BIO_MISSING" in flags or "BIO_STUB" in flags:
        return "Floor applied — bio needs to be fleshed out."
    if "RESUME_MISSING" in flags or "RESUME_STUB" in flags:
        return "Floor applied — resume needs to be fleshed out."
    if "FEW_COMMITS" in flags:
        return "Floor applied — needs more commits."
    if "STRONG" in flags:
        return "Strong on merit; see suggestions for refinements."
    return "See per-student suggestions for refinements."


def _suggestions_for(g: Grade) -> list[str]:
    """Kindly-worded, per-student improvement suggestions derived from flags + scores.

    Auto-generated suggestions are templated and generic; instructors are expected
    to edit/personalize them before sharing the report with students.
    """
    info = g.inspection
    s: list[str] = []

    if "REPO_INACCESSIBLE" in g.flags:
        s.append(
            "We weren't able to reach your repo. Double-check the URL on Lamaku "
            "and confirm the repo is set to **Public** under Settings → Danger Zone → "
            "Change visibility. If you intended to keep it private, share access with "
            "the instructor's GitHub account."
        )
        return s

    if "REPO_PRIVATE" in g.flags:
        s.append(
            "Flip the repo to **Public** in GitHub Settings → Danger Zone → Change "
            "visibility. Right now graders (and recruiters) can't see your work."
        )

    if g.dirs_present < g.dirs_total:
        s.append(
            f"Add the missing required directories ({g.dirs_present}/{g.dirs_total} "
            "present today). The Stage 0 doc lists all eleven; even creating an empty "
            "placeholder README.md in each is enough to scaffold the layout."
        )
    if g.placeholder_readme_count > 0:
        s.append(
            f"Replace the {g.placeholder_readme_count} placeholder README(s) with a "
            "sentence or two describing what belongs in each folder. Each directory's "
            "README is what helps a recruiter (or future you) navigate the repo six "
            "months from now."
        )

    if "BIO_MISSING" in g.flags:
        s.append(
            "Add `BIO.md` with a 150–200 word professional bio. The Stage 0 doc walks "
            "through using Claude or ChatGPT with the bio template — 30–45 minutes "
            "with an LLM gets you a strong first draft."
        )
    elif "BIO_STUB" in g.flags:
        bio_wc = max(g.word_count_bio, g.word_count_readme)
        s.append(
            f"Expand your bio — currently about {bio_wc} words; aim for the 150–200 "
            "word target. A complete bio covers role/company, expertise/achievements, "
            "education, and a forward-looking goal."
        )
    elif g.word_count_bio and g.word_count_bio > 250:
        s.append(
            f"`BIO.md` is on the long side ({g.word_count_bio} words vs. the 150–200 "
            "target). Trim a paragraph or merge background sections — easier to scan."
        )

    if "RESUME_MISSING" in g.flags:
        s.append(
            "Fill in `RESUME.md` with a Penn-style resume: Education → Experience → "
            "Skills, with quantified bullets (%, $, headcount). The Stage 0 doc links "
            "to a template."
        )
    elif "RESUME_STUB" in g.flags:
        s.append(
            f"Expand `RESUME.md` — currently {g.word_count_resume} words; aim for 250+ "
            "in Penn-style format with quantified bullets."
        )

    if "FEW_COMMITS" in g.flags:
        s.append(
            "Try committing in smaller steps as you work, rather than one big upload. "
            "Even 3–4 commits for Stage 0 (skeleton, README, BIO, RESUME) shows "
            "iterative work to anyone reading the history."
        )
    elif g.commit_count and (g.descriptive_commit_count / g.commit_count) < 0.75:
        s.append(
            f"A few commit messages could be tighter — {g.descriptive_commit_count}/"
            f"{g.commit_count} are descriptive. Rule of thumb: lead with a verb, name "
            "the file or area, and add the *why* if it's not obvious. Avoid `wip`, "
            "`update`, `fix typo`, or repeated `Update README.md`."
        )

    if info.accessible and not info.has_license:
        s.append(
            "Optional: add a `LICENSE` (MIT or CC-BY-4.0 are both fine). GitHub "
            "displays the license prominently on the repo landing page — a small "
            "recruiter-friendly touch."
        )

    if "STRONG" in g.flags and not s:
        s.append(
            "Strong submission across all five criteria — no specific suggestions "
            "from the rubric scan. Keep up the iterative habits (descriptive commits, "
            "meaningful READMEs) as the project gets more substantive."
        )

    return s


def _student_section(n: int, g: Grade, floor_pct: int) -> str:
    sub = g.submission
    info = g.inspection
    submitted = (
        sub.submitted_at.strftime("%Y-%m-%d %I:%M %p")
        if sub.submitted_at else "—"
    )
    raw = g.raw_total
    floor_applied = _floor_was_applied(g, floor_pct)
    final = _final_score(g, floor_pct)
    letter = _letter_for(final)

    bio_used = "BIO.md" if info.bio else "README.md"
    bio_wc_shown = g.word_count_bio if info.bio else g.word_count_readme

    lines: list[str] = []
    suffix = ", floor applied" if floor_applied else ""
    lines.append(f"## {n}. {sub.student_name} — **{final} / 100** ({letter}{suffix})")
    lines.append("")
    lines.append(f"**Repo:** {sub.repo_url}")
    lines.append(f"**Submitted:** {submitted}")
    lines.append("")
    lines.append("| Criterion | Earned | Notes |")
    lines.append("|-----------|--------|-------|")
    lines.append(
        f"| Repo public + accessible | {g.score_public} / 15 | {_repo_note(g)} |"
    )
    lines.append(
        f"| Directory skeleton + READMEs | {g.score_skeleton} / 20 | "
        f"{g.dirs_present}/{g.dirs_total} directories present; "
        f"{g.meaningful_readme_count}/"
        f"{g.meaningful_readme_count + g.placeholder_readme_count} READMEs meaningful. |"
    )
    lines.append(
        f"| Bio quality | {g.score_bio} / 25 | "
        f"`{bio_used}` is {bio_wc_shown} words. |"
    )
    lines.append(
        f"| Resume quality | {g.score_resume} / 25 | "
        f"`RESUME.md` is {g.word_count_resume} words. |"
    )
    lines.append(
        f"| Commit hygiene | {g.score_commits} / 15 | "
        f"{g.commit_count} commits, {g.descriptive_commit_count} descriptive. |"
    )
    if floor_applied:
        floor_value = round(TOTAL_POINTS * floor_pct / 100)
        lines.append(f"| **Raw total** | **{raw} / 100** | |")
        lines.append(
            f"| **Floor adjustment** | **+{floor_value - raw}** | "
            f"Working public repo present — floor of {floor_pct} applied per "
            "course policy. |"
        )
        lines.append(f"| **Final** | **{final} / 100** | |")
    else:
        merit_note = " Above the floor — earned on merit." if raw >= round(
            TOTAL_POINTS * floor_pct / 100
        ) else ""
        lines.append(f"| **Total** | **{final} / 100** |{merit_note} |")
    lines.append("")

    suggestions = _suggestions_for(g)
    if suggestions:
        lines.append("### Kindly-worded suggestions for improvement")
        lines.append("")
        for tip in suggestions:
            lines.append(f"- {tip}")
        lines.append("")

    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def _build_full_report(grades: list[Grade], floor_pct: int, today: datetime) -> str:
    lines = [
        "# BUS-629 Stage 0 — Grade Report",
        "",
        "**Stage:** Stage 0 — Personal Portfolio Repository (5% of project score)",
        f"**Graded:** {today.strftime('%Y-%m-%d')}",
        f"**Submissions reviewed:** {len(grades)}",
        f"**Floor policy:** Any submission with a working public GitHub repo "
        f"receives a floor of {floor_pct}%.",
        "",
        "---",
        "",
        "## Rubric (recap)",
        "",
        "| Criterion | Weight |",
        "|-----------|--------|",
        "| Repo public + accessible | 15% |",
        "| Directory skeleton + READMEs | 20% |",
        "| Bio quality (150–200 words; structured; iteratively revised) | 25% |",
        "| Resume quality (Penn-style; quantified; concise) | 25% |",
        "| Commit hygiene (≥2 commits; descriptive messages) | 15% |",
        "",
        "---",
        "",
    ]
    for i, g in enumerate(grades, 1):
        lines.append(_student_section(i, g, floor_pct))
    lines.append("## Class summary")
    lines.append("")
    lines.append("| Student | Score | Notes |")
    lines.append("|---------|-------|-------|")
    for g in grades:
        lines.append(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | "
            f"{_summary_tagline(g)} |"
        )
    lines.append("")
    finals = [_final_score(g, floor_pct) for g in grades]
    mean = sum(finals) / len(finals) if finals else 0.0
    floors = sum(1 for g in grades if _floor_was_applied(g, floor_pct))
    lines.append(f"**Mean:** {mean:.1f}  ")
    lines.append(f"**Floor applied:** {floors} of {len(grades)} submissions")
    lines.append("")
    return "\n".join(lines)


def write_or_update_grade_report(
    grades: list[Grade], floor_pct: int, report_path: Path,
    today: datetime | None = None,
) -> list[Grade]:
    """Write STAGE0_GRADES.md, appending new entries if the file already exists.

    Existing entries are detected by repo URL; we never overwrite them, so
    instructor-personalized suggestions are preserved across re-runs.

    Returns the list of newly-added grades.
    """
    today = today or datetime.now()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    if not report_path.exists():
        report_path.write_text(
            _build_full_report(grades, floor_pct, today), encoding="utf-8"
        )
        return list(grades)

    text = report_path.read_text(encoding="utf-8")
    existing_urls = {
        u.lower().rstrip("/")
        for u in re.findall(r"https?://github\.com/[A-Za-z0-9_./-]+", text)
    }
    new_grades = [
        g for g in grades
        if g.submission.repo_url.lower().rstrip("/") not in existing_urls
    ]
    if not new_grades:
        return []

    nums = [int(n) for n in re.findall(r"^## (\d+)\.", text, re.MULTILINE)]
    next_n = (max(nums) + 1) if nums else 1

    chunks = []
    for g in new_grades:
        chunks.append(_student_section(next_n, g, floor_pct))
        next_n += 1
    new_block = "\n".join(chunks)

    if "## Class summary" in text:
        text = text.replace(
            "## Class summary", new_block + "\n## Class summary", 1
        )
    else:
        text = text.rstrip() + "\n\n" + new_block

    # Append rows to the class-summary table (between the header separator and
    # the blank line that precedes **Mean:**).
    table_re = re.compile(
        r"(## Class summary\s*\n+"
        r"\| Student \| Score \| Notes \|\n"
        r"\| ?-+ ?\| ?-+ ?\| ?-+ ?\|\n"
        r"(?:\|[^\n]*\|\n)+)"
    )
    m = table_re.search(text)
    if m:
        new_rows = "".join(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | "
            f"{_summary_tagline(g)} |\n"
            for g in new_grades
        )
        text = text[: m.end()] + new_rows + text[m.end():]

    # Recompute Mean and Floor-applied counts by parsing the (now-updated) file.
    finals = [
        int(s) for s in re.findall(
            r"^## \d+\.[^—]*— \*\*(\d+) / 100\*\*", text, re.MULTILINE
        )
    ]
    floor_count = len(re.findall(r"floor applied", text, re.IGNORECASE))
    mean = sum(finals) / len(finals) if finals else 0.0

    text = re.sub(
        r"\*\*Mean:\*\* [^\n]*",
        f"**Mean:** {mean:.1f}  ",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Floor applied:\*\* \d+ of \d+ submissions",
        f"**Floor applied:** {floor_count} of {len(finals)} submissions",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Submissions reviewed:\*\* \d+",
        f"**Submissions reviewed:** {len(finals)}",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Graded:\*\* [^\n]*",
        f"**Graded:** {today.strftime('%Y-%m-%d')} "
        f"({', '.join(g.submission.student_name for g in new_grades)} added)",
        text, count=1,
    )

    report_path.write_text(text, encoding="utf-8")
    return new_grades


# ------------------------------------------------------------------
# Workflow helpers (ungraded/ → graded/ move)
# ------------------------------------------------------------------

def _detect_workflow(export_path: Path) -> tuple[Path, Path] | None:
    """Detect the ungraded/graded workflow.

    Returns (graded_dir, scratch_extract_dir) when:
      - export_path is a zip file
      - export_path's immediate parent folder is named ``ungraded`` (any case)
      - a sibling ``graded`` folder exists or can be created

    Returns None otherwise (fall back to the legacy _grading/ output path).
    """
    if not (export_path.is_file() and export_path.suffix.lower() == ".zip"):
        return None
    parent = export_path.parent
    if parent.name.lower() != "ungraded":
        return None
    graded = parent.parent / "graded"
    scratch = parent / f"_{export_path.stem}_extracted"
    return graded, scratch


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("export", type=Path,
                   help="Path to the Lamaku export zip or extracted directory")
    p.add_argument("--floor", type=int, default=DEFAULT_FLOOR_PCT,
                   help="Floor percentage for working-repo submissions (default 90)")
    p.add_argument("--out", type=Path, default=None,
                   help="Override worksheet output path. Default in workflow mode: "
                        "<root>/graded/_worksheets/stage0-<zipstem>.xlsx. Default "
                        "otherwise: <export-dir>/_grading/stage0-grading-worksheet.xlsx.")
    p.add_argument("--no-move", action="store_true",
                   help="Skip moving the source zip from ungraded/ to graded/ "
                        "(workflow mode only).")
    args = p.parse_args(argv)

    subs = discover_submissions(args.export)
    if not subs:
        print("No submissions discovered.")
        return 1
    print(f"Found {len(subs)} submission(s).")

    grades: list[Grade] = []
    for s in subs:
        print(f"  inspecting {s.student_name} ({s.repo_url}) ...", end=" ")
        info = inspect_repo(s)
        g = score(s, info)
        grades.append(g)
        if info.error:
            print(f"ERROR: {info.error}")
        else:
            print(f"raw={g.raw_total}/100 flags={','.join(g.flags) or '-'}")

    workflow = _detect_workflow(args.export)

    if args.out is not None:
        worksheet_path = args.out
    elif workflow is not None:
        graded_dir, _scratch = workflow
        graded_dir.mkdir(parents=True, exist_ok=True)
        ws_dir = graded_dir / "_worksheets"
        ws_dir.mkdir(exist_ok=True)
        worksheet_path = ws_dir / f"stage0-{args.export.stem}.xlsx"
    else:
        worksheet_path = (
            (args.export.parent if args.export.is_file() else args.export)
            / "_grading" / "stage0-grading-worksheet.xlsx"
        )

    build_worksheet(grades, args.floor, worksheet_path)
    print(f"\nWrote worksheet: {worksheet_path}")

    if workflow is not None:
        graded_dir, scratch = workflow
        report_path = graded_dir / "STAGE0_GRADES.md"
        new_entries = write_or_update_grade_report(grades, args.floor, report_path)
        if new_entries:
            names = ", ".join(g.submission.student_name for g in new_entries)
            print(f"Updated report: {report_path} (+{len(new_entries)} new: {names})")
        else:
            print(f"Report already up to date: {report_path} "
                  f"(all {len(grades)} submission(s) already recorded)")

        if not args.no_move:
            dest = graded_dir / args.export.name
            if dest.exists():
                print(f"Source zip already exists at {dest} — leaving original in place.")
            else:
                shutil.move(str(args.export), str(dest))
                print(f"Moved source zip → {dest}")
            if scratch.exists():
                shutil.rmtree(scratch, ignore_errors=True)
                print(f"Cleaned up scratch extract at {scratch}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
