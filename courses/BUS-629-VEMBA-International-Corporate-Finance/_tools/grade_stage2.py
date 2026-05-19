"""BUS-629 Stage 2 (company selection memo) grading scanner.

Stage 2 submissions are a Markdown memo (`.md`) named per the convention
`YYYY-MM-DD-{lastname}-{company-slug}-selection.md`, uploaded via Lamaku
and/or committed to `docs/decisions/` in the student's GitHub repo.

This script:

  1. Reads the Lamaku export zip (or extracted directory).
  2. For each student, parses the submitted memo:
       - YAML frontmatter sanity (required fields, valid stage / template)
       - Required-section presence (6 numbered sections per the spec)
       - Hypothesis pattern detection ("I expect X because Y" form)
       - Data-source naming (named, specific sources vs. vague ones)
       - Word count target (400–600 words of prose)
       - Filename convention check
  3. Looks the student up in `../stage0/graded/STAGE0_GRADES.md` AND
     `../stage1/graded/STAGE1_GRADES.md` to pick up the repo URL and
     prior-stage carry-over flags (no double-deduction across stages).
  4. Optional: verifies the memo also exists in the repo at the canonical
     path, and that the instructor (`adamwstauffer`) has been added as a
     Write collaborator (a Stage 2 submission-checklist requirement).
  5. Scores the 4-criterion rubric (25% each).
  6. Writes/updates `../graded/STAGE2_GRADES.md` (append mode, dedupes).
  7. Moves the source zip to `graded/` and cleans scratch.
  8. **Optional PR push** — when `--push-pr` is set and the instructor is
     a Write collaborator on the student's repo, clones the repo, creates
     an `instructor/stage2-feedback-{date}` branch, adds the per-student
     feedback file under `docs/decisions/`, and opens a PR. Off by default.

The four-criterion rubric (per `stage2-company-selection-memo.md`):

    Company Selection & Rationale        /25
    Analytical Framing & Hypotheses      /25
    Data Source Identification           /25
    Professionalism & Communication      /25
                                         = /100

Floor policy: 80% for working-repo submissions (instructor direction).

Double-deduction policy: issues already flagged in Stage 0 / Stage 1 are
**not** re-deducted here. They appear as forward-looking tips with no
point loss. See `feedback-no-double-deductions` in instructor memory and
the inline `_carry()` helper.

USAGE:
    python grade_stage2.py <export.zip>
        [--floor=80] [--no-move] [--push-pr] [--no-push-pr]
        [--prior-stage0=path/to/STAGE0_GRADES.md]
        [--prior-stage1=path/to/STAGE1_GRADES.md]

REQUIREMENTS:
    - `gh` CLI authenticated against github.com
    - `git` CLI (only used when `--push-pr` is on)
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _grading_comments import (
    Suggestion,
    backward,
    core,
    forward,
    next_stage_pointer,
    render_suggestions,
)
from _safe_zip import safe_extractall

STAGE_N = 2
DEFAULT_FLOOR_PCT = 80
TOTAL_POINTS = 100
STAGE_LABEL = "Stage 2 — Company Selection Memo"
STAGE_TARGET_DIR = "docs/decisions"
INSTRUCTOR_GITHUB_HANDLE = "adamwstauffer"

REQUIRED_FRONTMATTER_FIELDS_STAGE2 = [
    "template", "stage", "author", "date",
    "company", "ticker", "exchange",
]
# Generic memo template (docs/templates/memo-template.md) only marks itself as
# a memo — students using it identify company/ticker in the title/`re:` field.
REQUIRED_FRONTMATTER_FIELDS_GENERIC = ["template"]
# `purpose` and `audience` describe the *template* (per docs/templates/memo-template.md
# frontmatter convention), not an individual memo — they're metadata about the template
# itself. Memos written from the template don't need to carry them.
#
# Each Stage 2 section is satisfied by either the Stage 2 spec's canonical
# heading OR the generic memo-template equivalent — the Stage 2 brief tells
# students to use either template. Content quality is judged by the rubric
# criteria; this list just confirms structural coverage.
REQUIRED_SECTIONS = [
    ("Company Overview", [
        re.compile(r"^#+\s*\d?\.?\s*Company Overview", re.MULTILINE | re.IGNORECASE),
        re.compile(r"^#+\s*\d?\.?\s*Background", re.MULTILINE | re.IGNORECASE),
    ]),
    ("Selection Rationale", [
        re.compile(r"^#+\s*\d?\.?\s*Selection Rationale", re.MULTILINE | re.IGNORECASE),
        re.compile(r"^#+\s*\d?\.?\s*Executive Summary", re.MULTILINE | re.IGNORECASE),
    ]),
    ("Data Availability & Sources", [
        re.compile(r"^#+\s*\d?\.?\s*Data Availability", re.MULTILINE | re.IGNORECASE),
        re.compile(r"^#+\s*\d?\.?\s*Method", re.MULTILINE | re.IGNORECASE),
        re.compile(r"^#+\s*\d?\.?\s*References", re.MULTILINE | re.IGNORECASE),
    ]),
    ("Preliminary Observations", [
        re.compile(r"^#+\s*\d?\.?\s*Preliminary Observations", re.MULTILINE | re.IGNORECASE),
        re.compile(r"^#+\s*\d?\.?\s*Findings", re.MULTILINE | re.IGNORECASE),
    ]),
    ("Ratio Categories Preview", [
        re.compile(r"^#+\s*\d?\.?\s*Ratio Categories", re.MULTILINE | re.IGNORECASE),
        re.compile(r"^#+\s*\d?\.?\s*Implications", re.MULTILINE | re.IGNORECASE),
    ]),
    ("Data Collection Plan", [
        re.compile(r"^#+\s*\d?\.?\s*Data Collection Plan", re.MULTILINE | re.IGNORECASE),
        re.compile(r"^#+\s*\d?\.?\s*Limitations", re.MULTILINE | re.IGNORECASE),
        re.compile(r"^#+\s*\d?\.?\s*Next Steps", re.MULTILINE | re.IGNORECASE),
    ]),
]
# `(TICKER: Exchange)`, `(TICKER, Exchange)`, or `(Exchange: TICKER)` pattern
# in title / `re:` field — used to detect company identification for memos
# that put the ticker in the memo header rather than in dedicated frontmatter
# fields. Either side may be numeric (e.g., HKEX:1929 for Hong Kong tickers).
TITLE_TICKER_RE = re.compile(r"\(\s*[A-Z][A-Z0-9.\-]+\s*[:,]\s*[A-Za-z0-9][A-Za-z0-9 \-]+\)")
FILENAME_RE = re.compile(
    r"^(?P<date>\d{4}-\d{2}-\d{2})-"
    r"(?P<lastname>[a-z]+(?:-[a-z]+)*)-"
    r"(?P<company>[a-z0-9]+(?:-[a-z0-9]+)*)-"
    r"selection\.md$"
)
# A hypothesis is a directional claim + a reason. The Stage 2 spec uses
# "I expect X because Y" as the canonical form, but the substantive
# requirement is direction + cause. We accept either:
#   - explicit "I expect X (because|driven by|due to|...) Y", OR
#   - a directional verb (expect / will / should / anticipate / etc.) within
#     ~250 chars of a causal connector (because / driven by / reflecting /
#     as a / due to / given / attributable to)
_DIRECTIONAL = r"\b(?:I\s+expect|will|should|likely\s+to|anticipate|expect\s+to)\b"
_CAUSAL = (
    r"\b(?:because|driven\s+by|due\s+to|reflecting|reflect\s+the|"
    r"given\s+the|as\s+a\s+result|attributable\s+to|caused\s+by|"
    r"on\s+the\s+back\s+of|in\s+light\s+of|as\s+a\s+concentrate|"
    r"as\s+a\s+franchise|as\s+the)\b"
)
HYPOTHESIS_PATTERN = re.compile(
    rf"{_DIRECTIONAL}.{{1,250}}?{_CAUSAL}",
    re.IGNORECASE | re.DOTALL,
)
# Soft (hedged) claims — only flagged when they appear in a sentence that
# could otherwise have been a hypothesis (mentions a ratio/financial term).
SOFT_HYPOTHESIS_PATTERN = re.compile(
    r"\b(?:may|might|could|may\s+signal|appears?\s+to)\b",
    re.IGNORECASE,
)
_FIN_CONTEXT = re.compile(
    r"\b(ratio|leverage|margin|turnover|liquidity|coverage|profitability|"
    r"solvency|efficiency|valuation|revenue|earnings)\b",
    re.IGNORECASE,
)
NAMED_SOURCES = [
    "sec edgar", "edgar.sec.gov", "10-k", "annual report",
    "hose", "hnx", "sgx", "set", "psx", "bursa", "idx",
    "yahoo finance", "bloomberg", "mergent", "capital iq",
    "investor relations", "ir.", "company filings",
    "factset", "morningstar", "s&p capital", "refinitiv",
    "reuters", "investing.com", "stockanalysis.com",
]
VAGUE_SOURCES = ["the internet", "google", "google search", "wikipedia alone"]

GITHUB_URL_RE = re.compile(
    r"https?://github\.com/(?P<owner>[A-Za-z0-9_.-]+)/(?P<repo>[A-Za-z0-9_.-]+)"
)
FOLDER_NAME_RE = re.compile(
    r"^(?P<sid>\d+)-\d+\s*-\s*(?P<name>.+?)\s*-\s*"
    r"(?P<month>[A-Za-z]+)\s+(?P<day>\d+),\s*(?P<year>\d{4})\s+"
    r"(?P<h>\d{1,4})\s*(?P<ampm>AM|PM)\s*$"
)
LETTER_GRADE_SCALE = [
    ("A+", 97, None), ("A", 93, 97), ("A-", 90, 93),
    ("B+", 87, 90),  ("B", 83, 87), ("B-", 80, 83),
    ("C+", 77, 80),  ("C", 73, 77), ("C-", 70, 73),
    ("D+", 67, 70),  ("D", 65, 67), ("F", 0, 65),
]


# ------------------------------------------------------------------
# Submission discovery
# ------------------------------------------------------------------

@dataclass
class Submission:
    student_id: str
    student_name: str
    submitted_at: datetime | None
    memo_path: Path
    repo_url: str = ""
    owner: str = ""
    repo: str = ""


def _parse_folder_name(name: str):
    m = FOLDER_NAME_RE.match(name)
    if not m:
        return None, None, None
    h = m.group("h")
    if len(h) == 3:
        hour, minute = int(h[0]), int(h[1:])
    elif len(h) == 4:
        hour, minute = int(h[:2]), int(h[2:])
    else:
        hour, minute = int(h), 0
    ampm = m.group("ampm").upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    try:
        dt = datetime.strptime(
            f"{m.group('month')} {m.group('day')} {m.group('year')}", "%b %d %Y"
        ).replace(hour=hour, minute=minute)
    except ValueError:
        dt = None
    return m.group("sid"), m.group("name").strip(), dt


def discover_submissions(export_path: Path) -> list[Submission]:
    if export_path.is_file() and export_path.suffix.lower() == ".zip":
        scratch = export_path.parent / f"_{export_path.stem}_extracted"
        scratch.mkdir(exist_ok=True)
        with zipfile.ZipFile(export_path) as zf:
            safe_extractall(zf, scratch)
        root = scratch
    elif export_path.is_dir():
        root = export_path
    else:
        raise SystemExit(f"Cannot read submission export at {export_path}")

    subs: list[Submission] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        sid, name, dt = _parse_folder_name(child.name)
        if not name:
            continue
        memo = next(child.glob("*.md"), None)
        if memo is None:
            print(f"  WARN: no .md memo in '{child.name}' — skipping {name}")
            continue
        subs.append(Submission(
            student_id=sid or "",
            student_name=name,
            submitted_at=dt,
            memo_path=memo,
        ))
    subs.sort(key=lambda s: s.student_name.lower())
    return subs


# ------------------------------------------------------------------
# Prior-stage parsing (carry-over awareness)
# ------------------------------------------------------------------

@dataclass
class PriorGrade:
    student_name: str
    repo_url: str
    final_score: int
    raw_section: str
    carry_over_tags: set[str] = field(default_factory=set)


_CARRY_OVER_PATTERNS = [
    ("READMES_PLACEHOLDER", re.compile(r"placeholder readme")),
    ("DIRS_INCOMPLETE",     re.compile(r"missing.*director|directory.*missing|"
                                       r"\d+/1[12] (?:present|directories)")),
    ("BIO_PLACEHOLDER",     re.compile(r"(fill in|expand).*bio")),
    ("RESUME_PLACEHOLDER",  re.compile(r"(fill in|expand).*resume")),
    ("COMMIT_HYGIENE",      re.compile(r"commit message|tighten commit|"
                                       r"descriptive commit")),
    ("GITIGNORE",           re.compile(r"\.ds_store|gitignore|~\$.*xlsx")),
    ("REPO_NAME",           re.compile(r"trailing hyphen|rename.*repo")),
    ("CASING",              re.compile(r"casing|capitaliz")),
    ("LICENSE",             re.compile(r"add a (license|`license`)")),
    ("ROOT_CLUTTER",        re.compile(r"repo root.*crowded|cluttered.*root|"
                                       r"stage\[?0-5\]?-.*md.*root")),
]


def parse_prior_report(report_path: Path) -> dict[str, PriorGrade]:
    if not report_path.exists():
        return {}
    text = report_path.read_text(encoding="utf-8")
    out: dict[str, PriorGrade] = {}
    header_re = re.compile(
        r"^## (\d+)\. (.+?) — \*\*(\d+) / 100\*\*", re.MULTILINE
    )
    headers = [(m.start(), m) for m in header_re.finditer(text)]
    headers.append((len(text), None))
    for i in range(len(headers) - 1):
        start, m = headers[i]
        end, _ = headers[i + 1]
        section = text[start:end]
        name = m.group(2).strip()
        score = int(m.group(3))
        url_m = GITHUB_URL_RE.search(section)
        repo_url = url_m.group(0) if url_m else ""
        carry = set()
        for tag, pat in _CARRY_OVER_PATTERNS:
            if pat.search(section.lower()):
                carry.add(tag)
        out[_normalize_name(name)] = PriorGrade(
            student_name=name, repo_url=repo_url, final_score=score,
            raw_section=section, carry_over_tags=carry,
        )
    return out


def _normalize_name(name: str) -> str:
    tokens = re.findall(r"[A-Za-z]+", name.lower())
    return " ".join(sorted(tokens))


def lookup_prior(prior: dict[str, PriorGrade], student_name: str):
    key = _normalize_name(student_name)
    if key in prior:
        return prior[key]
    submitted_tokens = set(key.split())
    for pkey, pg in prior.items():
        if not pkey:
            continue
        prior_tokens = set(pkey.split())
        if submitted_tokens.issubset(prior_tokens) or prior_tokens.issubset(submitted_tokens):
            return pg
    return None


def merge_prior(*priors: dict[str, PriorGrade]) -> dict[str, PriorGrade]:
    """Combine carry-over tags across multiple prior-stage reports."""
    out: dict[str, PriorGrade] = {}
    for prior in priors:
        for key, pg in prior.items():
            if key not in out:
                out[key] = PriorGrade(
                    student_name=pg.student_name, repo_url=pg.repo_url,
                    final_score=pg.final_score, raw_section=pg.raw_section,
                    carry_over_tags=set(pg.carry_over_tags),
                )
            else:
                out[key].carry_over_tags |= pg.carry_over_tags
                if not out[key].repo_url and pg.repo_url:
                    out[key].repo_url = pg.repo_url
    return out


# ------------------------------------------------------------------
# Memo inspection
# ------------------------------------------------------------------

@dataclass
class MemoInspection:
    found: bool = False
    filename: str = ""
    filename_valid: bool = False
    filename_lastname: str = ""
    filename_company: str = ""
    frontmatter: dict = field(default_factory=dict)
    frontmatter_missing_fields: list[str] = field(default_factory=list)
    sections_present: list[str] = field(default_factory=list)
    sections_missing: list[str] = field(default_factory=list)
    word_count_prose: int = 0
    hypothesis_count: int = 0
    soft_hypothesis_count: int = 0
    sources_named: list[str] = field(default_factory=list)
    sources_vague: list[str] = field(default_factory=list)
    body_text: str = ""
    template_kind: str = "unknown"  # 'stage2', 'generic', or 'unknown'
    title_ticker_present: bool = False  # company/ticker named in title or `re:` field
    error: str = ""


def _detect_template_kind(fm: dict) -> str:
    """Classify the memo's template per its frontmatter `template:` field.

    Stage 2 explicitly permits either the Stage-2-specific template (with
    dedicated company/ticker/exchange fields) or the generic memo template
    (where the company is named in title/`re:`/H1 instead).
    """
    t = (fm.get("template") or "").lower().strip()
    if t in {"company-selection-memo", "stage2-company-selection",
             "selection-memo"}:
        return "stage2"
    if t in {"memo", "memo-template"}:
        return "generic"
    return "unknown"


def _parse_frontmatter(text: str) -> tuple[dict, str]:
    if not text.startswith("---"):
        return {}, text
    end = text.find("\n---", 3)
    if end == -1:
        return {}, text
    fm_text = text[3:end].strip()
    body = text[end + 4:].lstrip("\n")
    fm: dict[str, str] = {}
    for line in fm_text.splitlines():
        if ":" in line:
            k, _, v = line.partition(":")
            fm[k.strip()] = v.strip().strip('"').strip("'")
    return fm, body


def _word_count(text: str) -> int:
    # Strip tables, headings, list markers, frontmatter — count prose words.
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith(("#", "|", "---", ">", "```")):
            continue
        s = re.sub(r"^[-*+]\s+", "", s)
        s = re.sub(r"^\d+\.\s+", "", s)
        lines.append(s)
    prose = " ".join(lines)
    return len(re.findall(r"\b\w+\b", prose))


def inspect_memo(memo_path: Path) -> MemoInspection:
    info = MemoInspection(filename=memo_path.name)
    if not memo_path.exists():
        info.error = "memo file not found"
        return info
    info.found = True

    m = FILENAME_RE.match(memo_path.name)
    if m:
        info.filename_valid = True
        info.filename_lastname = m.group("lastname")
        info.filename_company = m.group("company")

    text = memo_path.read_text(encoding="utf-8", errors="replace")
    fm, body = _parse_frontmatter(text)
    info.frontmatter = fm
    info.body_text = body
    # Required frontmatter depends on which template the student used.
    # The Stage 2 brief points students to either the Stage 2 specific template
    # (which has dedicated company/ticker/exchange fields) or the generic memo
    # template (which puts those in title/re instead).
    template_kind = _detect_template_kind(fm)
    info.template_kind = template_kind
    required_fm = (
        REQUIRED_FRONTMATTER_FIELDS_GENERIC
        if template_kind == "generic"
        else REQUIRED_FRONTMATTER_FIELDS_STAGE2
    )
    # Accept `from` as a synonym for `author` (memo-routing convention).
    fm_with_synonyms = dict(fm)
    if fm.get("from") and not fm.get("author"):
        fm_with_synonyms["author"] = fm["from"]
    info.frontmatter_missing_fields = [
        k for k in required_fm if not fm_with_synonyms.get(k)
    ]

    for label, patterns in REQUIRED_SECTIONS:
        if any(pat.search(body) for pat in patterns):
            info.sections_present.append(label)
        else:
            info.sections_missing.append(label)

    info.word_count_prose = _word_count(body)

    # Hypotheses can land in any of three sections depending on which template
    # the student followed: Preliminary Observations (Stage 2 spec), Findings
    # (generic memo template), or Implications (some students put the
    # falsifiable predictions under "what this means"). Concatenate all three
    # candidate sections; fall back to whole body if none match.
    hyp_sections = re.findall(
        r"#+\s*\d?\.?\s*(?:Preliminary Observations|Findings|Implications)"
        r"[^\n]*\n(.+?)(?=^#+\s|\Z)",
        body, re.MULTILINE | re.DOTALL,
    )
    target_text = "\n\n".join(hyp_sections) if hyp_sections else body
    # Split on numbered list items (`1. `, `2. `), bulleted list items
    # (`* `, `- `, `• `), bolded "Hypothesis N" labels, or blank lines, to
    # count distinct claims. Students use a mix of conventions.
    items = re.split(
        r"\n\s*\n|^\s*\d+\.\s+|^\s*[*\-•]\s+|\*\*Hypothesis\s+\d+\b",
        target_text, flags=re.MULTILINE | re.IGNORECASE,
    )
    items = [it for it in items if it and len(it.strip()) > 20]

    hyp = 0
    soft = 0
    for it in items:
        if HYPOTHESIS_PATTERN.search(it):
            hyp += 1
        elif SOFT_HYPOTHESIS_PATTERN.search(it) and _FIN_CONTEXT.search(it):
            soft += 1
    info.hypothesis_count = hyp
    info.soft_hypothesis_count = soft

    body_lower = body.lower()
    info.sources_named = [s for s in NAMED_SOURCES if s in body_lower]
    info.sources_vague = [s for s in VAGUE_SOURCES if s in body_lower]

    # Title/`re:` ticker detection — e.g., "(MWG, HOSE)" or "(KO: NYSE)".
    # Memos using the generic template put company/ticker here rather than
    # in dedicated frontmatter fields.
    title_blob = " ".join([
        str(fm.get("title") or ""),
        str(fm.get("re") or ""),
        body[:600],  # first ~half a page also catches H1
    ])
    info.title_ticker_present = bool(TITLE_TICKER_RE.search(title_blob))

    return info


# ------------------------------------------------------------------
# GitHub repo inspection (memo in repo + collaborator check)
# ------------------------------------------------------------------

def _gh(*args: str) -> str:
    try:
        proc = subprocess.run(
            ["gh", *args], check=False, capture_output=True, text=True,
            encoding="utf-8", errors="replace",
        )
    except FileNotFoundError:
        return ""
    if proc.returncode != 0:
        return ""
    return proc.stdout or ""


@dataclass
class RepoInspection:
    queried: bool = False
    accessible: bool = False
    private: bool = False
    default_branch: str = "main"
    memo_in_repo: bool = False
    memo_repo_path: str = ""
    instructor_is_collaborator: bool = False
    collaborator_permission: str = ""  # raw permission value; "" if unknown
    collaborator_check_ran: bool = False  # False if gh missing or API didn't respond
    error: str = ""


def inspect_repo(owner: str, repo: str, expected_filename: str) -> RepoInspection:
    info = RepoInspection(queried=True)
    if not owner or not repo:
        info.error = "no repo URL known (lookup from prior stages failed)"
        return info
    meta_raw = _gh("api", f"repos/{owner}/{repo}")
    if not meta_raw:
        info.error = "repo not accessible"
        return info
    try:
        meta = json.loads(meta_raw)
    except json.JSONDecodeError:
        info.error = "could not parse repo metadata"
        return info
    info.accessible = True
    info.private = bool(meta.get("private"))
    info.default_branch = meta.get("default_branch") or "main"

    tree_raw = _gh(
        "api", f"repos/{owner}/{repo}/git/trees/{info.default_branch}",
        "-X", "GET", "-f", "recursive=1",
    )
    if tree_raw:
        try:
            tree = json.loads(tree_raw).get("tree", [])
        except json.JSONDecodeError:
            tree = []
        for e in tree:
            if e.get("type") != "blob":
                continue
            path = e.get("path", "")
            if path.lower().endswith(expected_filename.lower()):
                info.memo_in_repo = True
                info.memo_repo_path = path
                break

    # Collaborator check — gh returns 204 (empty stdout, returncode 0) for true,
    # 404 (empty stdout, returncode non-zero) for false. Use a richer endpoint.
    perm_raw = _gh(
        "api",
        f"repos/{owner}/{repo}/collaborators/{INSTRUCTOR_GITHUB_HANDLE}/permission",
    )
    if perm_raw:
        info.collaborator_check_ran = True
        try:
            perm = json.loads(perm_raw).get("permission", "")
        except json.JSONDecodeError:
            perm = ""
        info.collaborator_permission = perm
        info.instructor_is_collaborator = perm in {"admin", "write", "maintain"}

    return info


# ------------------------------------------------------------------
# Scoring
# ------------------------------------------------------------------

COLLAB_PENALTY = 5  # points off the raw total when instructor isn't a
                   # Write collaborator on the student's repo (instructor
                   # direction, 2026-05-13). Stage 2 submission checklist
                   # requires this; tracked feedback can't be delivered against
                   # the work without it.


@dataclass
class Grade:
    submission: Submission
    memo: MemoInspection
    repo: RepoInspection
    prior: PriorGrade | None

    score_selection: int = 0   # /25
    score_hypotheses: int = 0  # /25
    score_sources: int = 0     # /25
    score_professionalism: int = 0  # /25
    collab_penalty: int = 0    # subtracted from raw_total

    flags: list[str] = field(default_factory=list)
    carry_over_notes: list[str] = field(default_factory=list)

    @property
    def rubric_subtotal(self) -> int:
        return (
            self.score_selection + self.score_hypotheses
            + self.score_sources + self.score_professionalism
        )

    @property
    def raw_total(self) -> int:
        return max(0, self.rubric_subtotal - self.collab_penalty)


def _carry(g: Grade, tag: str) -> bool:
    return bool(g.prior and tag in g.prior.carry_over_tags)


def score(sub: Submission, memo: MemoInspection, repo: RepoInspection,
          prior: PriorGrade | None) -> Grade:
    g = Grade(submission=sub, memo=memo, repo=repo, prior=prior)

    if not memo.found:
        g.flags.append("MEMO_NOT_SUBMITTED")
        return g

    # Criterion 1: Company Selection & Rationale /25
    has_overview = "Company Overview" in memo.sections_present
    has_rationale = "Selection Rationale" in memo.sections_present
    company_named = bool(
        (memo.frontmatter.get("company") and memo.frontmatter.get("ticker"))
        or memo.title_ticker_present
    )
    if has_overview and has_rationale and company_named:
        # Rationale richness: word count of the section is a proxy. Accept
        # either the Stage 2 heading or the generic template's equivalent
        # (Executive Summary / Background carries the rationale narrative).
        rationale_match = re.search(
            r"#+\s*\d?\.?\s*(?:Selection Rationale|Executive Summary|Background)"
            r"(.+?)(?=^#+\s|\Z)",
            memo.body_text, re.MULTILINE | re.DOTALL,
        )
        rationale_text = rationale_match.group(1) if rationale_match else ""
        rationale_wc = len(re.findall(r"\b\w+\b", rationale_text))
        if rationale_wc < 40:
            g.score_selection = 18
            g.flags.append("RATIONALE_THIN")
        elif rationale_wc < 80:
            g.score_selection = 22
        else:
            g.score_selection = 25
    elif has_overview and has_rationale:
        g.score_selection = 18
        g.flags.append("COMPANY_INFO_INCOMPLETE")
    elif has_overview or has_rationale:
        g.score_selection = 12
        g.flags.append("SELECTION_SECTIONS_MISSING")
    else:
        g.score_selection = 5
        g.flags.append("SELECTION_SECTIONS_MISSING")

    # Criterion 2: Analytical Framing & Hypotheses /25
    h = memo.hypothesis_count
    if h >= 3:
        if memo.soft_hypothesis_count == 0:
            g.score_hypotheses = 25
        else:
            g.score_hypotheses = 23
            g.flags.append("HYPOTHESES_SOME_HEDGED")
    elif h == 2:
        if memo.soft_hypothesis_count >= 1:
            g.score_hypotheses = 22  # 2 strict + 1 soft = 3 total ideas, but one hedged
            g.flags.append("HYPOTHESES_SOME_HEDGED")
        else:
            g.score_hypotheses = 20
            g.flags.append("HYPOTHESES_UNDER_TARGET")
    elif h == 1:
        g.score_hypotheses = 15
        g.flags.append("HYPOTHESES_UNDER_TARGET")
    else:
        g.score_hypotheses = 8
        g.flags.append("NO_FALSIFIABLE_HYPOTHESES")

    # Criterion 3: Data Source Identification /25
    has_sources_section = "Data Availability & Sources" in memo.sections_present
    n_named = len(memo.sources_named)
    has_collection_plan = "Data Collection Plan" in memo.sections_present
    if has_sources_section and n_named >= 3 and has_collection_plan and not memo.sources_vague:
        g.score_sources = 25
    elif has_sources_section and n_named >= 2:
        g.score_sources = 22
    elif has_sources_section and n_named >= 1:
        g.score_sources = 18
        g.flags.append("SOURCES_THIN")
    elif memo.sources_vague:
        g.score_sources = 10
        g.flags.append("SOURCES_VAGUE")
    else:
        g.score_sources = 5
        g.flags.append("SOURCES_MISSING")

    # Criterion 4: Professionalism & Communication /25
    pro_score = 25
    if memo.frontmatter_missing_fields:
        pro_score -= min(8, 2 * len(memo.frontmatter_missing_fields))
        g.flags.append("FRONTMATTER_INCOMPLETE")
    if not memo.filename_valid:
        pro_score -= 4
        g.flags.append("FILENAME_NONSTANDARD")
    wc = memo.word_count_prose
    if wc < 300:
        pro_score -= 4
        g.flags.append("WORD_COUNT_LOW")
    elif wc < 400:
        pro_score -= 2
        g.flags.append("WORD_COUNT_BELOW_TARGET")
    elif wc > 700:
        pro_score -= 3
        g.flags.append("WORD_COUNT_HIGH")
    if memo.sections_missing:
        pro_score -= min(8, 2 * len(memo.sections_missing))
        g.flags.append("SECTIONS_MISSING")
    g.score_professionalism = max(pro_score, 0)

    # Collaborator + repo checks
    if repo.queried and repo.accessible:
        if not repo.memo_in_repo:
            g.flags.append("MEMO_NOT_IN_REPO")
        # Only flag (and penalize) when the collaborator check actually ran.
        # If `gh` is missing or the API didn't respond, we can't tell — don't
        # dock a student for an environment issue on the grader's side.
        if repo.collaborator_check_ran and not repo.instructor_is_collaborator:
            g.flags.append("INSTRUCTOR_NOT_COLLABORATOR")
            g.collab_penalty = COLLAB_PENALTY

    if (not g.flags or g.flags == ["INSTRUCTOR_NOT_COLLABORATOR"]) \
       and g.rubric_subtotal >= 95:
        # Strong rubric work; the collab flag (if any) is the only blemish.
        if "INSTRUCTOR_NOT_COLLABORATOR" not in g.flags:
            g.flags = ["STRONG"]
    return g


# ------------------------------------------------------------------
# Markdown grade report
# ------------------------------------------------------------------

def _final_score(g: Grade, floor_pct: int) -> int:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    if g.memo.found and g.raw_total < floor_value:
        return floor_value
    return g.raw_total


def _floor_was_applied(g: Grade, floor_pct: int) -> bool:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    return g.memo.found and g.raw_total < floor_value


def _letter_for(score: int) -> str:
    for letter, lo, hi in LETTER_GRADE_SCALE:
        lo_pts = round(lo * TOTAL_POINTS / 100)
        if score >= lo_pts and (hi is None or score < round(hi * TOTAL_POINTS / 100)):
            return letter
    return "F"


def _suggestions_for(g: Grade) -> list[Suggestion]:
    """Auto-generated, kind-worded suggestions for Stage 2.

    Returns a list of `Suggestion` objects tagged by bucket:
      - CORE     — observations about Stage 2's rubric performance
      - BACKWARD — carry-forwards from Stage 0 / Stage 1 (no points lost;
                   can bump the prior stage's score at the post-deadline
                   revision sweep)
      - FORWARD  — looking ahead to Stage 3
    """
    s: list[Suggestion] = []
    memo = g.memo

    if "MEMO_NOT_SUBMITTED" in g.flags:
        s.append(core(
            "No Stage 2 memo was included in your Lamaku submission. The "
            "deliverable is a Markdown memo named "
            "`YYYY-MM-DD-{lastname}-{company-slug}-selection.md` saved to "
            "`docs/decisions/` in your repo. The course memo template under "
            "`docs/templates/memo-template.md` is a good starting point."
        ))
        return s

    if "FILENAME_NONSTANDARD" in g.flags:
        s.append(core(
            f"The filename `{memo.filename}` doesn't match the convention "
            "`YYYY-MM-DD-{lastname}-{company-slug}-selection.md` (all lowercase, "
            "hyphen-separated, ISO date prefix). Rename in-repo and re-commit; "
            "Stage 4/5 tooling indexes files by this convention."
        ))

    if "FRONTMATTER_INCOMPLETE" in g.flags and memo.frontmatter_missing_fields:
        missing = ", ".join(f"`{f}`" for f in memo.frontmatter_missing_fields)
        if memo.template_kind == "generic":
            s.append(core(
                f"YAML frontmatter is missing {missing}. You're using the "
                "generic memo template, which is fine — but Stage 4 / Stage 5 "
                "tooling indexes by `template`, `date`, and the memo-routing "
                "fields (`to`/`from`/`re`). Add those if not already present."
            ))
        else:
            s.append(core(
                f"YAML frontmatter is missing {missing}. Keep all of `template`, "
                "`stage`, `author`, `date`, `company`, `ticker`, `exchange` filled in "
                "(memo routing fields `to`/`from`/`re` are optional but encouraged)."
            ))

    if "SECTIONS_MISSING" in g.flags:
        s.append(core(
            "Required section(s) appear to be missing: "
            + ", ".join(memo.sections_missing) + ". The Stage 2 brief lists all "
            "six numbered sections — each should be a separate `## N. Title` "
            "heading the grader can locate without searching."
        ))

    if "RATIONALE_THIN" in g.flags:
        s.append(core(
            "The **Selection Rationale** section is light on prose. The rubric "
            "weights this 25% — aim for 80+ words explaining *why this company* "
            "(industry relevance, your career angle, what makes its ratios "
            "analytically interesting). Specificity is the dimension that earns "
            "full marks; you can be terse if every sentence carries a distinct "
            "rationale point."
        ))

    h = memo.hypothesis_count
    if "NO_FALSIFIABLE_HYPOTHESES" in g.flags:
        s.append(core(
            "**Preliminary Observations** should contain 2–3 falsifiable, "
            "directional hypotheses in the form *\"I expect X because Y\"* — "
            "this is the rubric's bright-line test for full credit on the "
            "Analytical Framing criterion. Examples: *\"I expect rising leverage "
            "ratios from FY2023→FY2024 because Vinamilk financed its Indochina "
            "expansion with USD debt issued in early 2024.\"* Open-ended framings "
            "(*\"we'll see what the ratios show\"*) do not earn full credit."
        ))
    elif "HYPOTHESES_UNDER_TARGET" in g.flags:
        s.append(core(
            f"You have {h} clean *\"I expect X because Y\"* hypothesis(es); the "
            "target is 2–3. Add another in the same falsifiable, directional "
            "form — pick a different ratio category from the one(s) you've "
            "already covered."
        ))
    elif "HYPOTHESES_SOME_HEDGED" in g.flags:
        s.append(core(
            "One or more of your hypotheses hedges with *\"may\"* or *\"might\"* "
            "rather than the bright *\"I expect X because Y\"* form. Tighten the "
            "verb — *\"I expect X because Y\"* is a falsifiable claim; "
            "*\"X may happen because Y\"* isn't. Same idea, sharper voice."
        ))

    if "SOURCES_MISSING" in g.flags or "SOURCES_VAGUE" in g.flags:
        s.append(core(
            "The **Data Availability & Sources** section needs named, specific "
            "sources — SEC EDGAR, the company's IR page URL, the relevant "
            "exchange's disclosure portal, and any commercial database you'll "
            "use (Mergent / FactSet / Capital IQ). *\"The internet\"* and "
            "*\"Google\"* are not sources."
        ))
    elif "SOURCES_THIN" in g.flags:
        s.append(core(
            "Only one named source detected in your **Data Availability** "
            "section. Strong submissions name 3+ — typically the primary filing "
            "venue (SEC EDGAR / exchange portal), the company's IR page, and a "
            "secondary database for market data (Yahoo Finance, Mergent, etc.)."
        ))

    if "WORD_COUNT_LOW" in g.flags or "WORD_COUNT_BELOW_TARGET" in g.flags:
        s.append(core(
            f"Prose word count is around {memo.word_count_prose}; target is "
            "400–600. The brief is a tight format but it shouldn't read as "
            "skeletal — expand the Selection Rationale and Preliminary "
            "Observations sections in particular."
        ))
    elif "WORD_COUNT_HIGH" in g.flags:
        s.append(core(
            f"Prose word count is around {memo.word_count_prose}; target is "
            "400–600. Tighten any sections that read as repetitive — a senior "
            "analyst memo is short."
        ))

    if "MEMO_NOT_IN_REPO" in g.flags:
        s.append(core(
            f"The memo was submitted via Lamaku but doesn't appear at "
            f"`{STAGE_TARGET_DIR}/` in your repo. By Stage 5 the memo must live "
            "in the repo (the Lamaku fallback is a one-off, not the canonical "
            "path) — commit it now and you'll save yourself the cleanup later."
        ))

    if "INSTRUCTOR_NOT_COLLABORATOR" in g.flags:
        s.append(core(
            f"Stage 2's submission checklist asks you to add **`{INSTRUCTOR_GITHUB_HANDLE}`** "
            "as a Write collaborator on your repo so I can leave feedback "
            "directly on the document — concrete, tracked suggestions you can "
            "accept, push back on, or revise, the way a manager or auditor "
            "marks up a draft memo. Steps: Repo → Settings → Collaborators → "
            "Add people → choose **Write**. Without this I can't open a "
            "feedback document against your work, and Stage 5's "
            "feedback-incorporation rubric line gets harder."
        ))

    # Backward-bucket carry-forwards from Stage 0 / Stage 1.
    if "CARRY_OVER_DIRS" in g.flags:
        s.append(backward(
            "**Still open from Stage 0:** directory skeleton is still "
            "incomplete. Not re-deducted here; closing it before the deadline "
            "can bump your Stage 0 score at the post-deadline revision sweep."
        ))
    if "CARRY_OVER_READMES" in g.flags:
        s.append(backward(
            "**Still open from Stage 0:** placeholder README(s) remain in "
            "directory subfolders. Not re-deducted here; closing this before "
            "the deadline can bump your Stage 0 score at the post-deadline "
            "revision sweep."
        ))
    if "CARRY_OVER_COMMITS" in g.flags:
        s.append(backward(
            "**Still open from Stage 0:** commit-message hygiene still uneven. "
            "Not re-deducted here; tightening verb-lead messages (and closing "
            "the original Stage 0 gap) can bump that score at the sweep."
        ))
    if "CARRY_OVER_TEMPLATE_PATH" in g.flags:
        s.append(backward(
            "**Still open from Stage 1:** template not at the canonical "
            f"`{STAGE_TARGET_DIR}/{getattr(memo, 'template_filename', 'performance-ratios-template.xlsx')}` "
            "path. Not re-deducted; moving the file before Stage 3 unblocks "
            "the tooling and can bump your Stage 1 score at the sweep."
        ))

    core_count = sum(1 for x in s if x.bucket == "core")
    if "STRONG" in g.flags and core_count == 0:
        s.append(core(
            "Strong submission across all four criteria — falsifiable hypotheses, "
            "specific sources, on-spec structure. Keep the same voice for Stage 3's "
            "data-population pass."
        ))

    # Always end with forward guidance.
    fwd = next_stage_pointer(STAGE_N)
    if fwd is not None:
        s.append(fwd)

    return s


def _summary_tagline(g: Grade) -> str:
    flags = set(g.flags)
    if "MEMO_NOT_SUBMITTED" in flags:
        return "No memo submitted — floor does not apply."
    if "STRONG" in flags:
        return "Strong on merit; see suggestions for refinements."
    if "NO_FALSIFIABLE_HYPOTHESES" in flags:
        return "Floor likely — hypotheses need the \"I expect X because Y\" form."
    if "INSTRUCTOR_NOT_COLLABORATOR" in flags:
        return "Add instructor as Write collaborator; otherwise solid."
    if any(f.startswith("CARRY_OVER_") for f in flags):
        return "Strong submission; Stage-0/1 carry-over tips noted (no double-deduction)."
    return "See per-student suggestions for refinements."


def _student_section(n: int, g: Grade, floor_pct: int) -> str:
    sub = g.submission
    memo = g.memo
    submitted = sub.submitted_at.strftime("%Y-%m-%d %I:%M %p") if sub.submitted_at else "—"
    raw = g.raw_total
    floor_applied = _floor_was_applied(g, floor_pct)
    final = _final_score(g, floor_pct)
    letter = _letter_for(final)

    lines: list[str] = []
    suffix = ", floor applied" if floor_applied else ""
    lines.append(f"## {n}. {sub.student_name} — **{final} / 100** ({letter}{suffix})")
    lines.append("")
    if sub.repo_url:
        lines.append(f"**Repo:** {sub.repo_url}")
    lines.append(f"**Memo filename:** `{memo.filename}`")
    if memo.frontmatter.get("company"):
        ticker = memo.frontmatter.get("ticker", "")
        exchange = memo.frontmatter.get("exchange", "")
        suffix2 = f" ({ticker}:{exchange})" if ticker or exchange else ""
        lines.append(f"**Company:** {memo.frontmatter['company']}{suffix2}")
    lines.append(f"**Submitted:** {submitted}")
    lines.append("")
    lines.append("| Criterion | Earned | Notes |")
    lines.append("|-----------|--------|-------|")

    # Selection & Rationale
    sel_note = "Company named with rationale; both sections present."
    if "SELECTION_SECTIONS_MISSING" in g.flags:
        sel_note = "Required Selection-Rationale or Company-Overview section missing."
    elif "COMPANY_INFO_INCOMPLETE" in g.flags:
        sel_note = "Company info incomplete in frontmatter or Overview section."
    elif "RATIONALE_THIN" in g.flags:
        sel_note = "Sections present but Selection Rationale is thin (< 60 words)."
    lines.append(f"| Company Selection & Rationale | {g.score_selection} / 25 | {sel_note} |")

    # Hypotheses
    hyp_note = (
        f"{memo.hypothesis_count} hypothesis(es) in *\"I expect X because Y\"* form"
        f"{f'; {memo.soft_hypothesis_count} soft hedge(s) detected' if memo.soft_hypothesis_count else ''}."
    )
    lines.append(f"| Analytical Framing & Hypotheses | {g.score_hypotheses} / 25 | {hyp_note} |")

    # Sources
    src_note = (
        f"{len(memo.sources_named)} named source(s) detected"
        + (f" — {', '.join(memo.sources_named[:5])}." if memo.sources_named else ".")
    )
    if memo.sources_vague:
        src_note += " ⚠ vague-source language present."
    lines.append(f"| Data Source Identification | {g.score_sources} / 25 | {src_note} |")

    # Professionalism
    pro_bits = []
    if "FRONTMATTER_INCOMPLETE" in g.flags:
        pro_bits.append(f"frontmatter missing {len(memo.frontmatter_missing_fields)} field(s)")
    if "FILENAME_NONSTANDARD" in g.flags:
        pro_bits.append("filename off-convention")
    if "SECTIONS_MISSING" in g.flags:
        pro_bits.append(f"{len(memo.sections_missing)} section(s) missing")
    if memo.word_count_prose:
        pro_bits.append(f"prose ~{memo.word_count_prose} words")
    pro_note = "; ".join(pro_bits) if pro_bits else "On-spec structure, frontmatter, and filename."
    lines.append(f"| Professionalism & Communication | {g.score_professionalism} / 25 | {pro_note} |")

    if g.collab_penalty:
        lines.append(f"| **Rubric subtotal** | **{g.rubric_subtotal} / 100** | |")
        lines.append(
            f"| Collaborator-status penalty | **−{g.collab_penalty}** | "
            f"Instructor (`@{INSTRUCTOR_GITHUB_HANDLE}`) is not a Write "
            "collaborator on the repo, so tracked feedback can't be "
            "delivered against the work. See suggestions below for how to add. |"
        )

    if floor_applied:
        floor_value = round(TOTAL_POINTS * floor_pct / 100)
        lines.append(f"| **Raw total** | **{raw} / 100** | |")
        lines.append(
            f"| **Floor adjustment** | **+{floor_value - raw}** | "
            f"Working public repo present — floor of {floor_pct} applied. |"
        )
        lines.append(f"| **Final** | **{final} / 100** | |")
    else:
        merit_note = " Above the floor — earned on merit." if raw >= round(
            TOTAL_POINTS * floor_pct / 100
        ) else ""
        lines.append(f"| **Total** | **{final} / 100** |{merit_note} |")
    lines.append("")

    suggestions = _suggestions_for(g)
    lines.extend(render_suggestions(suggestions, stage_n=STAGE_N))
    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def _build_full_report(grades: list[Grade], floor_pct: int, today: datetime) -> str:
    lines = [
        "# BUS-629 Stage 2 — Grade Report",
        "",
        f"**Stage:** {STAGE_LABEL} (10% of project score)",
        f"**Graded:** {today.strftime('%Y-%m-%d')}",
        f"**Submissions reviewed:** {len(grades)}",
        f"**Floor policy:** {floor_pct}% floor for working-repo submissions.",
        "**Double-deduction policy:** Do not re-deduct for issues already "
        "flagged in Stage 0 or Stage 1. Useful suggestions may be repeated as "
        "forward-looking tips without point deductions.",
        "**Tracked-feedback policy:** When the instructor "
        f"(`{INSTRUCTOR_GITHUB_HANDLE}`) has Write access, individualized "
        "feedback is delivered as a tracked review document on the student's "
        "repo (the same workflow an auditor or supervising analyst would use "
        "to mark up a draft).",
        f"**Collaborator-status penalty:** 5 points off the raw total when the "
        f"instructor is **not** a Write collaborator on the student's repo "
        "(instructor direction, 2026-05-13). Stage 2's submission checklist "
        "requires this; the tracked-feedback workflow can't run otherwise.",
        "",
        "---",
        "",
        "## Rubric (recap)",
        "",
        "| Criterion | Weight |",
        "|-----------|--------|",
        "| Company Selection & Rationale | 25% |",
        "| Analytical Framing & Hypotheses (falsifiable; directional) | 25% |",
        "| Data Source Identification | 25% |",
        "| Professionalism & Communication | 25% |",
        "",
        "---",
        "",
    ]
    for i, g in enumerate(grades, 1):
        lines.append(_student_section(i, g, floor_pct))
    lines.append("## Class summary")
    lines.append("")
    lines.append("| Student | Score | Notes |")
    lines.append("|---------|-------|-------|")
    for g in grades:
        lines.append(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |"
        )
    lines.append("")
    finals = [_final_score(g, floor_pct) for g in grades if g.memo.found]
    submitted_n = len(finals)
    mean = sum(finals) / submitted_n if submitted_n else 0.0
    floors = sum(1 for g in grades if _floor_was_applied(g, floor_pct))
    lines.append(f"**Mean (submissions only):** {mean:.1f}")
    lines.append(f"**Submission rate:** {submitted_n} of {len(grades)}")
    lines.append(f"**Floor applied:** {floors} of {submitted_n} submissions")
    lines.append("")
    return "\n".join(lines)


def write_or_update_grade_report(
    grades: list[Grade], floor_pct: int, report_path: Path,
    today: datetime | None = None,
) -> list[Grade]:
    today = today or datetime.now()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    if not report_path.exists():
        report_path.write_text(_build_full_report(grades, floor_pct, today),
                               encoding="utf-8")
        return list(grades)

    text = report_path.read_text(encoding="utf-8")
    existing_urls = {
        u.lower().rstrip("/")
        for u in re.findall(r"https?://github\.com/[A-Za-z0-9_./-]+", text)
    }
    existing_names = {
        _normalize_name(n)
        for n in re.findall(r"^## \d+\. ([^—\n]+?) — \*\*\d+ / 100\*\*",
                            text, re.MULTILINE)
    }
    new_grades = []
    for g in grades:
        url_key = g.submission.repo_url.lower().rstrip("/")
        name_key = _normalize_name(g.submission.student_name)
        if url_key and url_key in existing_urls:
            continue
        if name_key and name_key in existing_names:
            continue
        new_grades.append(g)
    if not new_grades:
        return []

    nums = [int(n) for n in re.findall(r"^## (\d+)\.", text, re.MULTILINE)]
    next_n = (max(nums) + 1) if nums else 1
    chunks = []
    for g in new_grades:
        chunks.append(_student_section(next_n, g, floor_pct))
        next_n += 1
    new_block = "\n".join(chunks)

    if "## Class summary" in text:
        text = text.replace(
            "## Class summary", new_block + "\n## Class summary", 1
        )
    else:
        text = text.rstrip() + "\n\n" + new_block

    table_re = re.compile(
        r"(## Class summary\s*\n+"
        r"\| Student \| Score \| Notes \|\n"
        r"\| ?-+ ?\| ?-+ ?\| ?-+ ?\|\n"
        r"(?:\|[^\n]*\|\n)+)"
    )
    m = table_re.search(text)
    if m:
        new_rows = "".join(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |\n"
            for g in new_grades
        )
        text = text[: m.end()] + new_rows + text[m.end():]

    # Recompute stats (excluding non-submissions).
    student_re = re.compile(
        r"^## \d+\.[^—\n]*— \*\*(\d+) / 100\*\*([^\n]*)",
        re.MULTILINE,
    )
    submitted_finals = []
    total_students = 0
    for sm in student_re.finditer(text):
        total_students += 1
        if "no submission" in sm.group(2).lower():
            continue
        submitted_finals.append(int(sm.group(1)))
    floor_count = sum(
        1 for fs in submitted_finals if fs == round(TOTAL_POINTS * floor_pct / 100)
    )
    mean = sum(submitted_finals) / len(submitted_finals) if submitted_finals else 0.0

    text = re.sub(
        r"\*\*Mean \(submissions only\):\*\* [^\n]*",
        f"**Mean (submissions only):** {mean:.1f}", text, count=1,
    )
    text = re.sub(
        r"\*\*Submission rate:\*\* \d+ of \d+(?:\s*\([^)]*\))?",
        f"**Submission rate:** {len(submitted_finals)} of {total_students}",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Floor applied:\*\* \d+ of \d+ submissions(?:\s*\([^)]*\))?",
        f"**Floor applied:** {floor_count} of {len(submitted_finals)} submissions",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Submissions reviewed:\*\* [^\n]*",
        f"**Submissions reviewed:** {total_students}",
        text, count=1,
    )
    names_added = ", ".join(g.submission.student_name for g in new_grades)
    text = re.sub(
        r"\*\*Graded:\*\* [^\n]*",
        f"**Graded:** {today.strftime('%Y-%m-%d')} ({names_added} added)",
        text, count=1,
    )

    report_path.write_text(text, encoding="utf-8")
    return new_grades


# ------------------------------------------------------------------
# PR-feedback push (clones repo, creates feedback branch, opens PR)
# ------------------------------------------------------------------

def _build_feedback_md(g: Grade, floor_pct: int, today: datetime) -> str:
    sub = g.submission
    memo = g.memo
    final = _final_score(g, floor_pct)
    letter = _letter_for(final)
    suggestions = _suggestions_for(g)

    lines = [
        f"# Stage 2 Memo Feedback — {sub.student_name}",
        "",
        f"**Reviewer:** Adam Stauffer (`@{INSTRUCTOR_GITHUB_HANDLE}`)  ",
        f"**Date:** {today.strftime('%Y-%m-%d')}  ",
        f"**Memo:** [`{STAGE_TARGET_DIR}/{memo.filename}`](./{memo.filename})  ",
        f"**Score:** {final} / 100 ({letter})  ",
        "",
        "## Rubric scores",
        "",
        "| Criterion | Earned |",
        "|-----------|--------|",
        f"| Company Selection & Rationale | {g.score_selection} / 25 |",
        f"| Analytical Framing & Hypotheses | {g.score_hypotheses} / 25 |",
        f"| Data Source Identification | {g.score_sources} / 25 |",
        f"| Professionalism & Communication | {g.score_professionalism} / 25 |",
        f"| **Total** | **{g.raw_total} / 100** |",
        "",
    ]
    if suggestions:
        lines.extend(
            render_suggestions(
                suggestions,
                stage_n=STAGE_N,
                heading="## Suggestions for improvement",
            )
        )
    else:
        lines.extend([
            "## Suggestions for improvement",
            "",
            "- _No specific suggestions — this is a strong submission as-is._",
            "",
        ])
    lines.extend([
        "## How this feedback workflow works",
        "",
        "Stage 2 feedback is delivered as a tracked feedback document (a "
        "GitHub Pull Request) against your repo — the same idea as a "
        "marked-up draft a manager hands back, an auditor's review note, "
        "or an analyst peer-review on a deal memo: concrete, traceable, "
        "and either accepted or pushed back on. You can:",
        "",
        "1. **Read** the suggestions inline.",
        "2. **Reply** with questions or pushback in the conversation thread.",
        "3. **Address** the suggestions in your next commit(s).",
        "4. **Merge** the feedback document when you're done — that closes "
        "the loop and creates a historical record for Stage 5's *Stage 2 "
        "feedback incorporation* rubric line.",
        "",
        "_(Stage 5 grades how you incorporated this feedback — track your "
        "responses in `docs/decisions/{date}-stage2-feedback-response.md` "
        "or by commits referencing this review.)_",
        "",
    ])
    return "\n".join(lines)


_SAFE_OWNER_RE = re.compile(r"^[A-Za-z0-9](?:[A-Za-z0-9_.-]{0,38})$")
_SAFE_REPO_RE = re.compile(r"^[A-Za-z0-9_.][A-Za-z0-9_.-]{0,99}$")
_SAFE_LASTNAME_RE = re.compile(r"^[a-z][a-z0-9-]{0,39}$")
_SAFE_BRANCH_RE = re.compile(r"^[A-Za-z0-9_][A-Za-z0-9._/-]{0,99}$")


def push_feedback_pr(g: Grade, floor_pct: int, today: datetime) -> tuple[bool, str]:
    repo = g.repo
    sub = g.submission
    if not (repo.queried and repo.accessible):
        return False, f"repo not accessible ({repo.error or 'no metadata'})"
    if not (repo.collaborator_check_ran and repo.instructor_is_collaborator):
        return False, f"instructor `@{INSTRUCTOR_GITHUB_HANDLE}` is not a Write collaborator"

    # Validate every student-controlled value that's about to flow into
    # `gh`/`git` argv. argv (not shell=True) blocks classic command injection,
    # but `gh` and `git` parse leading-`-` values as flags, so a hostile owner
    # like `--upload-pack=evil` would otherwise reach the subprocess as a flag.
    if not _SAFE_OWNER_RE.fullmatch(sub.owner or ""):
        return False, f"refusing unsafe owner value: {sub.owner!r}"
    if not _SAFE_REPO_RE.fullmatch(sub.repo or ""):
        return False, f"refusing unsafe repo value: {sub.repo!r}"
    if not _SAFE_BRANCH_RE.fullmatch(repo.default_branch or ""):
        return False, f"refusing unsafe base-branch value: {repo.default_branch!r}"

    lastname = sub.student_name.split()[-1].lower()
    if not _SAFE_LASTNAME_RE.fullmatch(lastname):
        return False, f"refusing unsafe lastname value: {lastname!r}"
    branch = f"instructor/stage2-feedback-{today.strftime('%Y-%m-%d')}-{lastname}"
    fb_filename = (
        f"{today.strftime('%Y-%m-%d')}-instructor-stage2-feedback-{lastname}.md"
    )
    fb_md = _build_feedback_md(g, floor_pct, today)

    work_root = Path(tempfile.mkdtemp(prefix="bus629-stage2-feedback-"))
    try:
        clone_res = subprocess.run(
            ["gh", "repo", "clone", f"{sub.owner}/{sub.repo}", str(work_root / "repo")],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
        )
        if clone_res.returncode != 0:
            return False, f"clone failed: {clone_res.stderr.strip()}"
        repo_dir = work_root / "repo"

        # Branch
        subprocess.run(
            ["git", "-C", str(repo_dir), "checkout", "-b", branch],
            check=True, capture_output=True, text=True,
        )

        # Write feedback file
        fb_target_dir = repo_dir / STAGE_TARGET_DIR
        fb_target_dir.mkdir(parents=True, exist_ok=True)
        fb_path = fb_target_dir / fb_filename
        fb_path.write_text(fb_md, encoding="utf-8")

        # Commit
        subprocess.run(
            ["git", "-C", str(repo_dir), "add", str(fb_path)],
            check=True, capture_output=True, text=True,
        )
        subprocess.run(
            ["git", "-C", str(repo_dir), "commit",
             "-m", f"Stage 2 feedback for {sub.student_name}"],
            check=True, capture_output=True, text=True,
        )
        push_res = subprocess.run(
            ["git", "-C", str(repo_dir), "push", "-u", "origin", branch],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
        )
        if push_res.returncode != 0:
            return False, f"push failed: {push_res.stderr.strip()}"

        # Open PR
        pr_body = (
            f"Stage 2 review for **{sub.student_name}** — full feedback is in "
            f"`{STAGE_TARGET_DIR}/{fb_filename}`.\n\n"
            f"This PR is the canonical record for the Stage 5 *feedback "
            "incorporation* rubric line — please respond in commits or PR "
            "comments rather than re-doing the work silently.\n\n"
            f"_Generated by `grade_stage2.py` on {today.strftime('%Y-%m-%d')}._"
        )
        pr_res = subprocess.run(
            ["gh", "pr", "create",
             "-R", f"{sub.owner}/{sub.repo}",
             "--head", branch,
             "--base", repo.default_branch,
             "--title", f"Stage 2 feedback for {sub.student_name}",
             "--body", pr_body],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
        )
        if pr_res.returncode != 0:
            return False, f"PR create failed: {pr_res.stderr.strip()}"
        pr_url = pr_res.stdout.strip().splitlines()[-1] if pr_res.stdout else ""
        return True, pr_url
    finally:
        shutil.rmtree(work_root, ignore_errors=True)


# ------------------------------------------------------------------
# Worksheet
# ------------------------------------------------------------------

def build_worksheet(grades: list[Grade], floor_pct: int, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"
    headers = [
        "Student", "Submitted", "Repo URL", "Memo Filename",
        "Selection /25", "Hypotheses /25", "Sources /25", "Professionalism /25",
        "Raw /100", "Final /100", "Floored /100",
        "Hypothesis count", "Soft hedges", "Sources named",
        "Word count (prose)", "Sections missing",
        "Flags", "Comments",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F8CBAD")
    floor_fill = PatternFill("solid", fgColor="E2EFDA")

    final_col = headers.index("Final /100") + 1
    floor_col = headers.index("Floored /100") + 1
    flags_col = headers.index("Flags") + 1

    for g in grades:
        sub = g.submission
        memo = g.memo
        submitted = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else ""
        comments = g.repo.error or g.memo.error
        row = [
            sub.student_name, submitted, sub.repo_url, memo.filename,
            g.score_selection, g.score_hypotheses, g.score_sources, g.score_professionalism,
            g.raw_total, g.raw_total, None,
            memo.hypothesis_count, memo.soft_hypothesis_count, len(memo.sources_named),
            memo.word_count_prose, ", ".join(memo.sections_missing),
            ", ".join(g.flags), comments,
        ]
        ws.append(row)
        r = ws.max_row
        final_ref = f"{get_column_letter(final_col)}{r}"
        ws.cell(
            row=r, column=floor_col,
            value=f'=IF({final_ref}=0,0,MAX({final_ref},{floor_value}))',
        ).fill = floor_fill
        if comments and not memo.found:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = error_fill
        elif g.flags and g.flags != ["STRONG"]:
            ws.cell(row=r, column=flags_col).fill = flag_fill

    widths = [24, 17, 50, 40, 12, 14, 12, 16, 10, 11, 13, 14, 12, 14, 16, 24, 38, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(output_path)


# ------------------------------------------------------------------
# Sweep entry point — rescore a student against current repo state
# ------------------------------------------------------------------

GITHUB_URL_RE = re.compile(
    r"https?://github\.com/(?P<owner>[A-Za-z0-9_.-]+)/(?P<repo>[A-Za-z0-9_.-]+)"
)


def _download_text(owner: str, repo: str, path: str, branch: str) -> str | None:
    """Fetch a text file from a repo. Returns content or None on failure."""
    import base64
    raw = _gh(
        "api", f"repos/{owner}/{repo}/contents/{path}",
        "-X", "GET", "-f", f"ref={branch}",
    )
    if not raw:
        return None
    try:
        meta = json.loads(raw)
    except json.JSONDecodeError:
        return None
    content = meta.get("content")
    if not content:
        return None
    try:
        return base64.b64decode(content).decode("utf-8", errors="replace")
    except Exception:
        return None


def rescore_from_repo(
    student_name: str,
    repo_url: str,
    submitted_at: datetime | None = None,
    prior: "PriorGrade | None" = None,
    student_id: str = "",
) -> Grade | None:
    """Sweep entry point: rescore Stage 2 against current repo state.

    Scans `docs/decisions/` for any `*-selection.md`. If found, fetches and
    inspects it. If multiple exist, takes the one whose filename sorts last
    (i.e., latest date prefix wins).
    """
    m = GITHUB_URL_RE.search(repo_url)
    if not m:
        return None
    owner, repo = m.group("owner"), m.group("repo")
    repo_info = inspect_repo(owner, repo, "selection.md")

    import tempfile
    tmp = Path(tempfile.mkdtemp(prefix=f"sweep_s2_{owner}_"))
    # Write the temp file under the actual repo filename so inspect_memo()'s
    # filename validation sees the real name (not a generic "memo.md" placeholder).
    # If the repo doesn't have the file, fall back to "memo.md" so inspect_memo
    # still runs (and reports memo_not_found).
    real_filename = (
        Path(repo_info.memo_repo_path).name
        if repo_info.memo_repo_path else "memo.md"
    )
    memo_dest = tmp / real_filename
    if repo_info.memo_in_repo and repo_info.memo_repo_path and repo_info.accessible:
        text = _download_text(
            owner, repo, repo_info.memo_repo_path, repo_info.default_branch
        )
        if text is not None:
            memo_dest.write_text(text, encoding="utf-8")

    memo_info = inspect_memo(memo_dest)

    sub = Submission(
        student_id=student_id,
        student_name=student_name,
        submitted_at=submitted_at,
        memo_path=memo_dest,
        repo_url=repo_url,
        owner=owner,
        repo=repo,
    )
    return score(sub, memo_info, repo_info, prior)


# ------------------------------------------------------------------
# Workflow helper
# ------------------------------------------------------------------

def _detect_workflow(export_path: Path):
    if not (export_path.is_file() and export_path.suffix.lower() == ".zip"):
        return None
    parent = export_path.parent
    if parent.name.lower() != "ungraded":
        return None
    graded = parent.parent / "graded"
    scratch = parent / f"_{export_path.stem}_extracted"
    return graded, scratch


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("export", type=Path)
    p.add_argument("--floor", type=int, default=DEFAULT_FLOOR_PCT)
    p.add_argument("--prior-stage0", type=Path, default=None)
    p.add_argument("--prior-stage1", type=Path, default=None)
    p.add_argument("--out", type=Path, default=None)
    p.add_argument("--no-move", action="store_true")
    pr_group = p.add_mutually_exclusive_group()
    pr_group.add_argument("--push-pr", dest="push_pr", action="store_true",
                          help="Push individualized feedback PR to each student's repo "
                               "(requires instructor to be a Write collaborator)")
    pr_group.add_argument("--no-push-pr", dest="push_pr", action="store_false",
                          help="Skip the PR push (default).")
    p.set_defaults(push_pr=False)
    args = p.parse_args(argv)

    subs = discover_submissions(args.export)
    if not subs:
        print("No submissions discovered.")
        return 1
    print(f"Found {len(subs)} submission(s).")

    base_root = args.export.resolve().parents[2]
    s0 = args.prior_stage0 or (base_root / "stage0" / "graded" / "STAGE0_GRADES.md")
    s1 = args.prior_stage1 or (base_root / "stage1" / "graded" / "STAGE1_GRADES.md")
    prior0 = parse_prior_report(s0) if s0.exists() else {}
    prior1 = parse_prior_report(s1) if s1.exists() else {}
    if prior0:
        print(f"Loaded Stage 0 records: {len(prior0)} from {s0}")
    if prior1:
        print(f"Loaded Stage 1 records: {len(prior1)} from {s1}")
    prior = merge_prior(prior0, prior1)

    for s in subs:
        pg = lookup_prior(prior, s.student_name)
        if pg and pg.repo_url:
            m = GITHUB_URL_RE.search(pg.repo_url)
            if m:
                s.repo_url = pg.repo_url
                s.owner = m.group("owner")
                s.repo = m.group("repo")

    grades: list[Grade] = []
    for s in subs:
        print(f"  inspecting {s.student_name} ...", end=" ")
        memo = inspect_memo(s.memo_path)
        if s.owner:
            repo = inspect_repo(s.owner, s.repo, memo.filename)
        else:
            repo = RepoInspection()
        pg = lookup_prior(prior, s.student_name)
        g = score(s, memo, repo, pg)
        grades.append(g)
        if repo.queried and repo.accessible and repo.collaborator_check_ran:
            collab_str = (
                f"Y({repo.collaborator_permission})"
                if repo.instructor_is_collaborator
                else f"N({repo.collaborator_permission or 'none'})"
            )
        elif repo.queried and not repo.accessible:
            collab_str = "?(repo not accessible)"
        elif repo.queried:
            collab_str = "?(gh missing/failed)"
        else:
            collab_str = "?(no repo URL)"
        print(
            f"raw={g.raw_total}/100 "
            f"(hyp={memo.hypothesis_count}/{memo.hypothesis_count + memo.soft_hypothesis_count}, "
            f"src={len(memo.sources_named)}, words={memo.word_count_prose}, "
            f"collab={collab_str}, "
            f"flags={','.join(g.flags) or '-'})"
        )

    workflow = _detect_workflow(args.export)
    if args.out is not None:
        worksheet_path = args.out
    elif workflow is not None:
        graded_dir, _ = workflow
        graded_dir.mkdir(parents=True, exist_ok=True)
        ws_dir = graded_dir / "_worksheets"
        ws_dir.mkdir(exist_ok=True)
        worksheet_path = ws_dir / f"stage2-{args.export.stem}.xlsx"
    else:
        worksheet_path = (
            (args.export.parent if args.export.is_file() else args.export)
            / "_grading" / "stage2-grading-worksheet.xlsx"
        )

    build_worksheet(grades, args.floor, worksheet_path)
    print(f"\nWrote worksheet: {worksheet_path}")

    today = datetime.now()
    if workflow is not None:
        graded_dir, scratch = workflow
        report_path = graded_dir / "STAGE2_GRADES.md"
        new_entries = write_or_update_grade_report(grades, args.floor, report_path, today)
        if new_entries:
            names = ", ".join(g.submission.student_name for g in new_entries)
            print(f"Updated report: {report_path} (+{len(new_entries)} new: {names})")
        else:
            print(f"Report already up to date: {report_path}")

        if args.push_pr:
            print()
            print("--- Pushing per-student feedback PRs ---")
            for g in grades:
                if not g.memo.found:
                    continue
                ok, msg = push_feedback_pr(g, args.floor, today)
                if ok:
                    print(f"  ✓ {g.submission.student_name}: PR opened — {msg}")
                else:
                    print(f"  · {g.submission.student_name}: {msg}")

        if not args.no_move:
            dest = graded_dir / args.export.name
            if dest.exists():
                print(f"Source zip already at {dest} — leaving original in place.")
            else:
                shutil.move(str(args.export), str(dest))
                print(f"Moved source zip → {dest}")
            if scratch.exists():
                shutil.rmtree(scratch, ignore_errors=True)
                print(f"Cleaned up scratch extract at {scratch}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
