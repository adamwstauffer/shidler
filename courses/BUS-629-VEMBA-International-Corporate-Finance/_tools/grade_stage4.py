"""BUS-629 Stage 4 (LLM-drafted technical specification) grading scanner.

Stage 4 submissions are markdown: a spec file at
`docs/specs/YYYY-MM-DD-{lastname}-{company-slug}-spec.md` plus a prompt log
entry. Students may submit either by pushing to their repo + uploading the
URL pointer to Lamaku, or by uploading the `.md` file(s) directly to Lamaku
as a fallback.

This script:

  1. Reads the Lamaku export zip (or extracted directory). Two folder
     shapes are supported:
       (a) Direct upload — `.md` files in the folder. We read them in place.
       (b) URL pointer only — an HTML file with the repo URL. We pull the
           spec / prompt log from the repo via the `gh` CLI.
  2. For each student, inspects:
       - Spec sections present (Part A 1-7, Part B 8-11)
       - Coverage of named-range notation (`BAL_*`, `INC_*`, `CASH_*`, `RATIO_*`)
       - Ratio count and validation rules in Part A
       - Analysis-spec depth in Part B
       - Prompt log presence and visible HIL iteration evidence
       - (Optional repo check) spec file landed at `docs/specs/`
  3. Looks the student up in prior-stage reports (Stage 0/1/2/3) for repo
     URL and carry-over flags.
  4. Scores the 4-criterion rubric (25/25/25/25), 70% floor.
  5. Writes/updates `../graded/STAGE4_GRADES.md` (internal — has scores).
  6. Writes per-student PR-ready feedback files under
     `../graded/_pr_feedback/{lastname}/feedback-file.md` (NO scores —
     these are intended to be opened as PRs on student repos).
  7. Moves the source zip to `graded/` and cleans scratch.

The four-criterion rubric (per `stage4-technical-specification.md`):

    Model spec — Data & Structure (Part A 1–5)   /25
    Model spec — Ratios & Validation (Part A 6–7) /25
    Analysis spec (Part B 8–11)                   /25
    Spec craft + prompt log + HIL iteration       /25
                                                  = /100

Floor policy: 70% for any submission with a substantive spec file present.
Pass `--floor` to override.

Score-privacy: per course policy, score numbers live ONLY in the internal
`STAGE4_GRADES.md` and instructor email. The per-student feedback files
written under `_pr_feedback/` deliberately omit scores so they can be
pushed to public student repos as PRs.

ACCOUNTING STANDARDS POLICY (2026-05-24):
    Students analyzing non-US-GAAP companies (VAS, IFRS, K-IFRS, SFRS(I),
    CAS, etc.) should document standard-specific behaviors in their spec.
    A spec that correctly reflects VAS or IFRS conventions is NOT
    deficient — it is demonstrating accounting-standard literacy.

    POSITIVELY CREDIT students who:
    - Identify the reporting framework in Section 1 (Scope & Objective)
    - Add standard-aware validation rules (e.g., "INC_depreciation must
      be zero under VAS — use CASH_depreciation for cash coverage")
    - Note where template-default formulas need adaptation for their
      framework

    Do NOT deduct for:
    - INC_depreciation = 0 under VAS (correct; bundled in COGS/SGA)
    - Capitalized development costs under IFRS (correct; IAS 38)
    - Absent LIFO reserve for IFRS/VAS companies (LIFO prohibited)
    - Different revenue recognition timing under VAS (no IFRS 15 equiv)

    Reference: docs/decisions/2026-05-24-accounting-standards-conversion-framework.md

USAGE:
    python grade_stage4.py <export.zip> [--floor=70] [--no-move]
        [--prior-stage0=path/to/STAGE0_GRADES.md]
        [--prior-stage1=path/to/STAGE1_GRADES.md]
        [--prior-stage2=path/to/STAGE2_GRADES.md]
        [--prior-stage3=path/to/STAGE3_GRADES.md]

REQUIREMENTS:
    - `gh` CLI authenticated against github.com (optional; falls back to
      submission-file-only scoring without it)
"""
from __future__ import annotations

import argparse
import base64
import json
import re
import shutil
import subprocess
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _grading_comments import (
    Suggestion,
    backward,
    core,
    forward,
    next_stage_pointer,
    render_suggestions,
)
from _safe_zip import safe_extractall

STAGE_N = 4
DEFAULT_FLOOR_PCT = 70
TOTAL_POINTS = 100
STAGE_LABEL = "Stage 4 — LLM-Drafted Technical Specification"
STAGE_TARGET_DIR = "docs/specs"
INSTRUCTOR_GITHUB_HANDLE = "adamwstauffer"

# Filename conventions.
SPEC_FILENAME_RE = re.compile(
    r"^(?P<date>\d{4}-\d{2}-\d{2})-"
    r"(?P<lastname>[a-z]+(?:-[a-z]+)*)-"
    r"(?P<company>[a-z0-9]+(?:-[a-z0-9]+)*)-"
    r"spec\.md$",
    re.IGNORECASE,
)
PROMPT_LOG_FILENAME_HINTS = ("prompt-log", "promptlog", "prompts")

# Part A section keywords (1–5 = Data & Structure; 6–7 = Ratios & Validation).
# We match heading lines like "### 1. Scope & Objective" or "## 1 Scope".
SECTION_PATTERNS = {
    1: [r"scope\s*(&|and)?\s*objective", r"scope\b"],
    2: [r"model\s+architecture", r"\barchitecture\b"],
    3: [r"data\s+inputs?", r"\binputs?\b"],
    4: [r"named[- ]range", r"naming\s+convention"],
    5: [r"derived\s+inputs?", r"intermediates?"],
    6: [r"ratio\s+definitions?", r"ratio\s+formulas?", r"\bformulas?\b"],
    7: [r"validation\s+rules?", r"\bvalidation\b"],
    8: [r"analysis\s+requirements?"],
    9: [r"du\s*pont\s+decomposition", r"du\s*pont\s+system"],
    10: [r"strategic\s+recommendation", r"recommendation\s+requirements?"],
    11: [r"output\s+format"],
}

# Ratio categories we expect to see referenced somewhere in Part A.6.
RATIO_CATEGORIES = [
    "performance", "profitability", "efficiency",
    "leverage", "liquidity", "du pont",
]

# Named-range prefixes we expect to appear in a substantive spec.
NAMED_RANGE_PREFIXES = ("BAL_", "INC_", "CASH_", "RATIO_", "startYear_",
                        "currentYear_", "avg_")

# HIL evidence cues. The brief lists three acceptable forms (before/after note,
# round-2 prompt, annotated diff). We score by signal strength.
HIL_STRONG_PATTERNS = [
    re.compile(r"\bbefore\s*/?\s*after\b", re.IGNORECASE),
    re.compile(r"\b(round|prompt)\s*2\b", re.IGNORECASE),
    re.compile(r"\bsecond[- ]prompt\b", re.IGNORECASE),
    re.compile(r"\bre[- ]prompted?\b", re.IGNORECASE),
    re.compile(r"\bv1\s*[→\-]+\s*v2\b", re.IGNORECASE),
    re.compile(r"\bfirst draft\b", re.IGNORECASE),
    re.compile(r"\biteration\b", re.IGNORECASE),
    re.compile(r"\bannotated diff\b", re.IGNORECASE),
    re.compile(r"\brevised\s+(the\s+)?(prompt|spec)\b", re.IGNORECASE),
]
HIL_WEAK_PATTERNS = [
    re.compile(r"\bself[- ]?assessment\b", re.IGNORECASE),
    re.compile(r"\blimitation(s)? to note\b", re.IGNORECASE),
    re.compile(r"\bdesign\s+decisions?\b", re.IGNORECASE),
    re.compile(r"\bnotes? for next\b", re.IGNORECASE),
]

GITHUB_URL_RE = re.compile(
    r"https?://github\.com/(?P<owner>[A-Za-z0-9_.-]+)/(?P<repo>[A-Za-z0-9_.-]+)"
)
FOLDER_NAME_RE = re.compile(
    r"^(?P<sid>\d+)-\d+\s*-\s*(?P<name>.+?)\s*-\s*"
    r"(?P<month>[A-Za-z]+)\s+(?P<day>\d+),\s*(?P<year>\d{4})\s+"
    r"(?P<h>\d{1,4})\s*(?P<ampm>AM|PM)\s*$"
)

LETTER_GRADE_SCALE = [
    ("A+", 97, None), ("A", 93, 97), ("A-", 90, 93),
    ("B+", 87, 90), ("B", 83, 87), ("B-", 80, 83),
    ("C+", 77, 80), ("C", 73, 77), ("C-", 70, 73),
    ("D+", 67, 70), ("D", 65, 67), ("F", 0, 65),
]


# ------------------------------------------------------------------
# Validation regexes for subprocess argv safety
# ------------------------------------------------------------------
_SAFE_OWNER_RE = re.compile(r"^[A-Za-z0-9](?:[A-Za-z0-9_.-]{0,38})$")
_SAFE_REPO_RE = re.compile(r"^[A-Za-z0-9_.][A-Za-z0-9_.-]{0,99}$")
_SAFE_PATH_RE = re.compile(r"^[A-Za-z0-9._/-]{1,200}$")
_SAFE_BRANCH_RE = re.compile(r"^[A-Za-z0-9_][A-Za-z0-9._/-]{0,99}$")
_SAFE_HANDLE_RE = re.compile(r"^[A-Za-z0-9](?:[A-Za-z0-9-]{0,38})$")


# ------------------------------------------------------------------
# Submission discovery
# ------------------------------------------------------------------

@dataclass
class Submission:
    student_id: str
    student_name: str
    submitted_at: datetime | None
    spec_path: Path | None             # local path to spec .md (if uploaded directly)
    prompt_log_path: Path | None       # local path to prompt log .md (if uploaded)
    iteration_path: Path | None        # local path to optional iteration .md
    repo_url: str = ""
    owner: str = ""
    repo: str = ""


def _parse_folder_name(name: str):
    m = FOLDER_NAME_RE.match(name)
    if not m:
        return None, None, None
    h = m.group("h")
    if len(h) == 3:
        hour, minute = int(h[0]), int(h[1:])
    elif len(h) == 4:
        hour, minute = int(h[:2]), int(h[2:])
    else:
        hour, minute = int(h), 0
    ampm = m.group("ampm").upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    try:
        dt = datetime.strptime(
            f"{m.group('month')} {m.group('day')} {m.group('year')}", "%b %d %Y"
        ).replace(hour=hour, minute=minute)
    except ValueError:
        dt = None
    return m.group("sid"), m.group("name").strip(), dt


def _scan_html_for_url(folder: Path) -> str | None:
    for html in folder.glob("*.html"):
        text = html.read_text(encoding="utf-8", errors="ignore")
        m = GITHUB_URL_RE.search(text)
        if m:
            return m.group(0)
    return None


def discover_submissions(export_path: Path) -> list[Submission]:
    if export_path.is_file() and export_path.suffix.lower() == ".zip":
        scratch = export_path.parent / f"_{export_path.stem}_extracted"
        scratch.mkdir(exist_ok=True)
        with zipfile.ZipFile(export_path) as zf:
            safe_extractall(zf, scratch)
        root = scratch
    elif export_path.is_dir():
        root = export_path
    else:
        raise SystemExit(f"Cannot read submission export at {export_path}")

    subs: list[Submission] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        sid, name, dt = _parse_folder_name(child.name)
        if not name:
            continue
        spec_path: Path | None = None
        prompt_log_path: Path | None = None
        iteration_path: Path | None = None
        for md in child.glob("*.md"):
            low = md.name.lower()
            if "spec" in low and "retrospective" not in low:
                spec_path = md
            elif any(h in low for h in PROMPT_LOG_FILENAME_HINTS):
                prompt_log_path = md
            elif "iteration" in low or "hil" in low:
                iteration_path = md

        repo_url = _scan_html_for_url(child) or ""
        owner, repo = "", ""
        if repo_url:
            m = GITHUB_URL_RE.search(repo_url)
            if m:
                owner, repo = m.group("owner"), m.group("repo")

        subs.append(Submission(
            student_id=sid or "",
            student_name=name,
            submitted_at=dt,
            spec_path=spec_path,
            prompt_log_path=prompt_log_path,
            iteration_path=iteration_path,
            repo_url=repo_url,
            owner=owner,
            repo=repo,
        ))
    subs.sort(key=lambda s: s.student_name.lower())
    return subs


# ------------------------------------------------------------------
# Prior-stage parsing (carry-over awareness)
# ------------------------------------------------------------------

@dataclass
class PriorGrade:
    student_name: str
    repo_url: str
    final_score: int
    raw_section: str
    carry_over_tags: set[str] = field(default_factory=set)


_CARRY_OVER_PATTERNS = [
    ("INSTRUCTOR_NOT_COLLAB", re.compile(r"not a write collaborator|"
                                          r"instructor_not_collaborator",
                                          re.IGNORECASE)),
    ("FILENAME_CONVENTION",  re.compile(r"filename convention|canonical filename",
                                        re.IGNORECASE)),
    ("DIRS_INCOMPLETE",      re.compile(r"missing.*director|directory.*missing",
                                        re.IGNORECASE)),
    ("COMMIT_HYGIENE",       re.compile(r"commit message|descriptive commit",
                                        re.IGNORECASE)),
]


def _normalize_name(name: str) -> str:
    tokens = re.findall(r"[A-Za-z]+", name.lower())
    return " ".join(sorted(tokens))


def parse_prior_report(report_path: Path) -> dict[str, PriorGrade]:
    if not report_path.exists():
        return {}
    text = report_path.read_text(encoding="utf-8")
    out: dict[str, PriorGrade] = {}
    header_re = re.compile(r"^## (\d+)\. (.+?) — \*\*(\d+) / 100\*\*", re.MULTILINE)
    headers = [(m.start(), m) for m in header_re.finditer(text)]
    headers.append((len(text), None))
    for i in range(len(headers) - 1):
        start, m = headers[i]
        end, _ = headers[i + 1]
        section = text[start:end]
        name = m.group(2).strip()
        score = int(m.group(3))
        url_m = GITHUB_URL_RE.search(section)
        repo_url = url_m.group(0) if url_m else ""
        carry = set()
        for tag, pat in _CARRY_OVER_PATTERNS:
            if pat.search(section):
                carry.add(tag)
        out[_normalize_name(name)] = PriorGrade(
            student_name=name, repo_url=repo_url, final_score=score,
            raw_section=section, carry_over_tags=carry,
        )
    return out


def lookup_prior(prior: dict[str, PriorGrade], student_name: str):
    key = _normalize_name(student_name)
    if key in prior:
        return prior[key]
    submitted_tokens = set(key.split())
    for pkey, pg in prior.items():
        if not pkey:
            continue
        prior_tokens = set(pkey.split())
        if submitted_tokens.issubset(prior_tokens) or prior_tokens.issubset(submitted_tokens):
            return pg
    return None


def merge_prior(*priors: dict[str, PriorGrade]) -> dict[str, PriorGrade]:
    out: dict[str, PriorGrade] = {}
    for prior in priors:
        for key, pg in prior.items():
            if key not in out:
                out[key] = PriorGrade(
                    student_name=pg.student_name, repo_url=pg.repo_url,
                    final_score=pg.final_score, raw_section=pg.raw_section,
                    carry_over_tags=set(pg.carry_over_tags),
                )
            else:
                out[key].carry_over_tags |= pg.carry_over_tags
                if not out[key].repo_url and pg.repo_url:
                    out[key].repo_url = pg.repo_url
    return out


# ------------------------------------------------------------------
# Spec / prompt-log inspection
# ------------------------------------------------------------------

@dataclass
class SpecInspection:
    found: bool = False
    filename: str = ""
    filename_valid: bool = False
    filename_lastname: str = ""
    filename_company: str = ""

    word_count: int = 0
    sections_found: dict[int, bool] = field(default_factory=dict)  # 1..11
    section_word_counts: dict[int, int] = field(default_factory=dict)
    named_range_hits: dict[str, int] = field(default_factory=dict)
    ratio_categories_present: list[str] = field(default_factory=list)
    ratio_table_rows: int = 0          # rows in ratio tables (rough count)
    validation_rules_count: int = 0
    has_yaml_frontmatter: bool = False
    error: str = ""


@dataclass
class PromptLogInspection:
    found: bool = False
    filename: str = ""
    word_count: int = 0
    prompt_blocks: int = 0
    hil_strong_hits: int = 0
    hil_weak_hits: int = 0
    has_iteration_file: bool = False   # separate analysis/validation/...iteration.md
    error: str = ""


def _word_count(text: str) -> int:
    return len(re.findall(r"\b\w+\b", text or ""))


def _section_slice(text: str, target_section: int) -> tuple[int, int] | None:
    """Find the [start, end) text slice for a section number (1..11).

    Heading detection: a Markdown heading line of the form
    `## N. Title` or `### N. Title` (allowing optional whitespace, optional
    "Section" prefix). Section ends where the next numbered heading begins
    or where Part B / a horizontal rule introduces a new top-level region.
    """
    heading_re = re.compile(
        rf"^\s*#{{1,4}}\s*(?:section\s+)?{target_section}[\.\)]?\s+", re.MULTILINE
    )
    m = heading_re.search(text)
    if not m:
        # Fallback: keyword-based section detection. Useful when a student
        # numbers sections inconsistently (e.g., "Scope & Objective" with
        # no leading "1.").
        for pat in SECTION_PATTERNS.get(target_section, []):
            kw_re = re.compile(rf"^\s*#{{1,4}}[^\n]*{pat}", re.MULTILINE | re.IGNORECASE)
            m = kw_re.search(text)
            if m:
                break
        if not m:
            return None
    start = m.end()
    # Find the next section heading (any numbered heading 1..15) or `## Part`.
    next_re = re.compile(
        r"^\s*#{1,4}\s*(?:section\s+)?(?:\d+[\.\)]\s+|Part\s+[AB]\b|Appendix\b)",
        re.MULTILINE | re.IGNORECASE,
    )
    nm = next_re.search(text, m.end())
    end = nm.start() if nm else len(text)
    return (start, end)


def inspect_spec(path: Path | None, content: str | None = None,
                 source_label: str = "") -> SpecInspection:
    info = SpecInspection()
    if path is None and content is None:
        info.error = "spec file not provided"
        return info

    if content is None:
        try:
            content = path.read_text(encoding="utf-8")
        except Exception as e:
            info.error = f"could not read spec: {e}"
            return info
        info.filename = path.name
    else:
        info.filename = source_label or "(in-repo)"

    info.found = True
    info.word_count = _word_count(content)

    fm = SPEC_FILENAME_RE.match(info.filename)
    if fm:
        info.filename_valid = True
        info.filename_lastname = fm.group("lastname")
        info.filename_company = fm.group("company")

    info.has_yaml_frontmatter = content.lstrip().startswith("---")

    # Section detection + word counts per section
    for n in range(1, 12):
        sl = _section_slice(content, n)
        if sl:
            info.sections_found[n] = True
            seg = content[sl[0]:sl[1]]
            info.section_word_counts[n] = _word_count(seg)
        else:
            info.sections_found[n] = False
            info.section_word_counts[n] = 0

    # Named-range hits across the whole document
    for prefix in NAMED_RANGE_PREFIXES:
        info.named_range_hits[prefix] = len(re.findall(
            rf"`{re.escape(prefix)}", content
        )) + len(re.findall(rf"\b{re.escape(prefix)}[a-zA-Z_]+", content))

    # Ratio category coverage — search the Section 6 slice if present, else whole doc
    sl6 = _section_slice(content, 6) or (0, len(content))
    s6_text = content[sl6[0]:sl6[1]].lower()
    present = []
    for cat in RATIO_CATEGORIES:
        if cat in s6_text or cat.replace(" ", "-") in s6_text:
            present.append(cat)
    info.ratio_categories_present = present

    # Rough ratio table-row count: markdown table rows inside Section 6 that
    # contain at least one backtick block (a formula or named range) and are
    # not separator rows. Header rows ("| Metric | Formula |") would still
    # count except headers rarely include backticks — so this is a decent
    # proxy for "ratio rows with a formula stated."
    if sl6 != (0, len(content)):
        count = 0
        for line in s6_text.splitlines():
            stripped = line.strip()
            if not stripped.startswith("|") or "---" in stripped:
                continue
            if stripped.count("|") < 3:
                continue
            if "`" not in stripped:
                continue
            count += 1
        info.ratio_table_rows = count

    # Validation rules: count rows in Section 7 table or "Vn" markers
    sl7 = _section_slice(content, 7)
    if sl7:
        s7 = content[sl7[0]:sl7[1]]
        # Count V1..Vn markers
        v_markers = set(re.findall(r"\bV\d+\b", s7))
        # Or count table rows that look like rules
        body_rows = re.findall(r"^\|\s*(?:V\d+|\d+|[A-Z][a-z])", s7, re.MULTILINE)
        info.validation_rules_count = max(len(v_markers), len(body_rows))

    return info


def inspect_prompt_log(path: Path | None, content: str | None = None,
                       source_label: str = "") -> PromptLogInspection:
    info = PromptLogInspection()
    if path is None and content is None:
        info.error = "prompt log not provided"
        return info

    if content is None:
        try:
            content = path.read_text(encoding="utf-8")
        except Exception as e:
            info.error = f"could not read prompt log: {e}"
            return info
        info.filename = path.name
    else:
        info.filename = source_label or "(in-repo)"

    info.found = True
    info.word_count = _word_count(content)

    # Count distinct prompt blocks ("## Prompt N", "### Prompt N", or
    # "**Prompt N:**"). Generous matcher.
    prompt_heads = re.findall(
        r"^(?:#{2,4}\s*Prompt\s*\d+|^\*\*Prompt\s*\d+[\*:])",
        content, re.MULTILINE | re.IGNORECASE,
    )
    info.prompt_blocks = len(prompt_heads)

    for p in HIL_STRONG_PATTERNS:
        info.hil_strong_hits += len(p.findall(content))
    for p in HIL_WEAK_PATTERNS:
        info.hil_weak_hits += len(p.findall(content))

    return info


# ------------------------------------------------------------------
# Optional repo inspection
# ------------------------------------------------------------------

def _gh(*args: str) -> str:
    try:
        proc = subprocess.run(
            ["gh", *args], check=False, capture_output=True, text=True,
            encoding="utf-8", errors="replace",
        )
    except FileNotFoundError:
        return ""
    if proc.returncode != 0:
        return ""
    return proc.stdout or ""


@dataclass
class RepoInspection:
    queried: bool = False
    accessible: bool = False
    private: bool = False
    default_branch: str = "main"
    spec_in_repo: bool = False
    spec_repo_path: str = ""
    prompt_log_in_repo: bool = False
    prompt_log_repo_path: str = ""
    iteration_in_repo: bool = False
    iteration_repo_path: str = ""
    instructor_is_collaborator: bool = False
    collaborator_permission: str = ""
    collaborator_check_ran: bool = False
    error: str = ""


def _safe_inputs(owner: str, repo: str, branch: str | None = None) -> bool:
    if not _SAFE_OWNER_RE.match(owner or ""):
        return False
    if not _SAFE_REPO_RE.match(repo or ""):
        return False
    if branch is not None and not _SAFE_BRANCH_RE.match(branch):
        return False
    return True


def inspect_repo(owner: str, repo: str) -> RepoInspection:
    info = RepoInspection(queried=True)
    if not owner or not repo:
        info.error = "no repo URL known"
        return info
    if not _safe_inputs(owner, repo):
        info.error = "unsafe owner/repo characters"
        return info
    meta_raw = _gh("api", f"repos/{owner}/{repo}")
    if not meta_raw:
        info.error = "repo not accessible"
        return info
    try:
        meta = json.loads(meta_raw)
    except json.JSONDecodeError:
        info.error = "could not parse repo metadata"
        return info
    info.accessible = True
    info.private = bool(meta.get("private"))
    info.default_branch = meta.get("default_branch") or "main"
    if not _SAFE_BRANCH_RE.match(info.default_branch):
        info.default_branch = "main"

    tree_raw = _gh(
        "api", f"repos/{owner}/{repo}/git/trees/{info.default_branch}",
        "-X", "GET", "-f", "recursive=1",
    )
    if tree_raw:
        try:
            tree = json.loads(tree_raw).get("tree", [])
        except json.JSONDecodeError:
            tree = []
        for e in tree:
            if e.get("type") != "blob":
                continue
            path = e.get("path", "")
            low = path.lower()
            if low.endswith("-spec.md") and "docs/specs" in low:
                info.spec_in_repo = True
                info.spec_repo_path = path
            elif low.endswith("prompt-log.md") or low.endswith("/prompt_log.md"):
                info.prompt_log_in_repo = True
                info.prompt_log_repo_path = path
            elif "stage4-iteration" in low or "stage4_iteration" in low:
                info.iteration_in_repo = True
                info.iteration_repo_path = path

    if _SAFE_HANDLE_RE.match(INSTRUCTOR_GITHUB_HANDLE):
        perm_raw = _gh(
            "api",
            f"repos/{owner}/{repo}/collaborators/{INSTRUCTOR_GITHUB_HANDLE}/permission",
        )
        if perm_raw:
            info.collaborator_check_ran = True
            try:
                perm = json.loads(perm_raw).get("permission", "")
            except json.JSONDecodeError:
                perm = ""
            info.collaborator_permission = perm
            info.instructor_is_collaborator = perm in {"admin", "write", "maintain"}
    return info


def _fetch_repo_file(owner: str, repo: str, path: str, branch: str) -> str:
    if not _safe_inputs(owner, repo, branch):
        return ""
    if not _SAFE_PATH_RE.match(path):
        return ""
    raw = _gh(
        "api", f"repos/{owner}/{repo}/contents/{path}",
        "-X", "GET", "-f", f"ref={branch}",
        "-H", "Accept: application/vnd.github.raw",
    )
    return raw


# ------------------------------------------------------------------
# Scoring
# ------------------------------------------------------------------

@dataclass
class Grade:
    submission: Submission
    spec: SpecInspection
    plog: PromptLogInspection
    repo: RepoInspection
    prior: PriorGrade | None

    score_data_structure: int = 0    # /25 (Part A 1-5)
    score_ratios_validation: int = 0 # /25 (Part A 6-7)
    score_analysis: int = 0          # /25 (Part B 8-11)
    score_craft_hil: int = 0         # /25 (spec craft + prompt log + HIL)

    flags: list[str] = field(default_factory=list)

    @property
    def raw_total(self) -> int:
        return (self.score_data_structure + self.score_ratios_validation
                + self.score_analysis + self.score_craft_hil)


def _section_present_pts(spec: SpecInspection, n: int, weight: int,
                         min_words: int = 30) -> int:
    """Award `weight` if the section is present and substantive, half if
    present but thin, zero if missing."""
    if not spec.sections_found.get(n):
        return 0
    if spec.section_word_counts.get(n, 0) < min_words:
        return weight // 2
    return weight


def score(sub: Submission, spec: SpecInspection, plog: PromptLogInspection,
          repo: RepoInspection, prior: PriorGrade | None) -> Grade:
    g = Grade(submission=sub, spec=spec, plog=plog, repo=repo, prior=prior)
    if not spec.found:
        g.flags.append("SPEC_NOT_SUBMITTED")
        return g

    # ----- Criterion 1: Data & Structure (Part A 1-5), 25 pts -----
    c1 = 0
    c1 += _section_present_pts(spec, 1, 4, min_words=60)      # Scope
    c1 += _section_present_pts(spec, 2, 4, min_words=60)      # Architecture
    c1 += _section_present_pts(spec, 3, 8, min_words=120)     # Data Inputs
    c1 += _section_present_pts(spec, 4, 4, min_words=40)      # Named-range conv
    c1 += _section_present_pts(spec, 5, 5, min_words=80)      # Derived Inputs
    g.score_data_structure = min(25, c1)

    if not spec.sections_found.get(3):
        g.flags.append("DATA_INPUTS_MISSING")
    if not spec.sections_found.get(4):
        g.flags.append("NAMED_RANGES_MISSING")

    # ----- Criterion 2: Ratios & Validation (Part A 6-7), 25 pts -----
    c2 = 0
    # Section 6 — base 5pts for presence; then up to 12 pts based on ratio
    # category coverage + 3 pts for ratio table rows (>=20 rows = full).
    if spec.sections_found.get(6):
        c2 += 5
        cats = len(spec.ratio_categories_present)
        c2 += round(12 * min(cats, len(RATIO_CATEGORIES)) / len(RATIO_CATEGORIES))
        rows = spec.ratio_table_rows
        if rows >= 20:
            c2 += 3
        elif rows >= 10:
            c2 += 2
        elif rows >= 1:
            c2 += 1
        else:
            g.flags.append("RATIO_TABLE_THIN")
    else:
        g.flags.append("RATIOS_MISSING")
    # Section 7 — base 2 pts for presence + up to 3 pts on count
    if spec.sections_found.get(7):
        c2 += 2
        if spec.validation_rules_count >= 5:
            c2 += 3
        elif spec.validation_rules_count >= 2:
            c2 += 2
        elif spec.validation_rules_count >= 1:
            c2 += 1
    else:
        g.flags.append("VALIDATION_MISSING")
    g.score_ratios_validation = min(25, c2)

    # ----- Criterion 3: Analysis Spec (Part B 8-11), 25 pts -----
    c3 = 0
    c3 += _section_present_pts(spec, 8, 8, min_words=120)    # Analysis Reqs
    c3 += _section_present_pts(spec, 9, 5, min_words=60)     # Du Pont
    c3 += _section_present_pts(spec, 10, 6, min_words=80)    # Strategic recs
    c3 += _section_present_pts(spec, 11, 6, min_words=60)    # Output Format
    g.score_analysis = min(25, c3)
    if not any(spec.sections_found.get(n) for n in (8, 9, 10, 11)):
        g.flags.append("PART_B_MISSING")

    # ----- Criterion 4: Spec craft + prompt log + HIL, 25 pts -----
    c4 = 0
    # Spec craft: word count + named-range usage + YAML frontmatter
    if spec.word_count >= 1500:
        c4 += 4
    elif spec.word_count >= 800:
        c4 += 2
    elif spec.word_count >= 300:
        c4 += 1
    else:
        g.flags.append("SPEC_THIN")
    total_nr_hits = sum(spec.named_range_hits.values())
    if total_nr_hits >= 40:
        c4 += 4
    elif total_nr_hits >= 15:
        c4 += 2
    elif total_nr_hits >= 5:
        c4 += 1
    else:
        g.flags.append("NAMED_RANGES_THIN")
    if spec.has_yaml_frontmatter:
        c4 += 1

    # Prompt log presence + depth
    if plog.found:
        if plog.word_count >= 400:
            c4 += 4
        elif plog.word_count >= 150:
            c4 += 2
        elif plog.word_count >= 50:
            c4 += 1
    else:
        g.flags.append("PROMPT_LOG_MISSING")

    # HIL evidence: strongest signal wins.
    # - Annotated iteration file present → 12 pts (max)
    # - prompt_blocks >= 2 OR hil_strong_hits >= 2 → 10 pts (round-2 or before/after)
    # - hil_strong_hits >= 1 → 7 pts (single strong cue)
    # - hil_weak_hits >= 2 → 4 pts (weak self-assessment)
    # - else → 0
    has_iter_file = (sub.iteration_path is not None) or repo.iteration_in_repo or plog.has_iteration_file
    if has_iter_file:
        c4 += 12
    elif plog.prompt_blocks >= 2 or plog.hil_strong_hits >= 2:
        c4 += 10
    elif plog.hil_strong_hits >= 1:
        c4 += 7
    elif plog.hil_weak_hits >= 2:
        c4 += 4
    else:
        g.flags.append("HIL_MISSING")

    g.score_craft_hil = min(25, c4)

    # ----- Repo presence & carry-forward (informational) -----
    if repo.queried and repo.accessible:
        if not repo.spec_in_repo:
            g.flags.append("SPEC_NOT_IN_REPO")
        if not repo.prompt_log_in_repo:
            g.flags.append("PROMPT_LOG_NOT_IN_REPO")
        if repo.collaborator_check_ran and not repo.instructor_is_collaborator:
            g.flags.append("INSTRUCTOR_NOT_COLLABORATOR")

    if not spec.filename_valid:
        g.flags.append("FILENAME_NONSTANDARD")

    if not g.flags and g.raw_total >= 95:
        g.flags = ["STRONG"]
    return g


# ------------------------------------------------------------------
# Suggestions
# ------------------------------------------------------------------

def _suggestions_for(g: Grade) -> list[Suggestion]:
    s: list[Suggestion] = []
    spec = g.spec
    plog = g.plog

    if "SPEC_NOT_SUBMITTED" in g.flags:
        s.append(core(
            "No spec was found in the Lamaku submission. The Stage 4 deliverable "
            "is a markdown spec at "
            "`docs/specs/YYYY-MM-DD-{lastname}-{company-slug}-spec.md` plus a "
            "prompt-log entry. If you submitted the URL pointer alone, please "
            "confirm the spec lives at that path in your repo."
        ))
        return s

    if "FILENAME_NONSTANDARD" in g.flags:
        s.append(core(
            f"The spec filename `{spec.filename}` doesn't match the convention "
            "`YYYY-MM-DD-{lastname}-{company-slug}-spec.md` (lowercase, "
            "hyphen-separated, ISO date prefix, `-spec.md` suffix). Stage 5 "
            "tooling indexes specs by this name — rename in-repo and re-commit."
        ))

    # Part A — Data & Structure
    if "DATA_INPUTS_MISSING" in g.flags:
        s.append(core(
            "**Section 3 (Data Inputs) is missing or sparse.** This is the section "
            "the Stage 5 LLM reads to populate the model — every named range "
            "(`BAL_*`, `INC_*`, `CASH_*`) needs an explicit numeric value drawn "
            "from your Stage 3 workbook. \"Total assets\" is not a spec; "
            "\"`BAL_assets_total_curr = 104,816 USD millions`\" is."
        ))
    if "NAMED_RANGES_MISSING" in g.flags:
        s.append(core(
            "**Section 4 (Named Range Conventions) is missing or sparse.** Map "
            "out the `BAL_*` / `INC_*` / `CASH_*` / `RATIO_*` / `startYear_*` / "
            "`avg_*` prefixes with an example each — the Stage 5 LLM has zero "
            "context and needs the naming grammar laid out before it can produce "
            "correct formulas."
        ))

    # Part A — Ratios & Validation
    if "RATIOS_MISSING" in g.flags:
        s.append(core(
            "**Section 6 (Ratio Definitions & Formulas) is missing.** This is the "
            "core of the model spec — every ratio (Performance, Profitability, "
            "Efficiency, Leverage, Liquidity, Du Pont) with formula in named-range "
            "notation and expected unit. Aim for 25+ ratios."
        ))
    elif "RATIO_TABLE_THIN" in g.flags:
        present = ", ".join(spec.ratio_categories_present) or "none detected"
        s.append(core(
            f"**Section 6 covers fewer ratios than expected.** Categories detected: "
            f"{present}. The full template carries 25+ ratios across six categories "
            f"(Performance, Profitability, Efficiency, Leverage, Liquidity, Du Pont) "
            f"— each one needs a formula in named-range notation so the Stage 5 LLM "
            f"can execute without guessing."
        ))
    if "VALIDATION_MISSING" in g.flags:
        s.append(core(
            "**Section 7 (Validation Rules) is missing.** Internal consistency "
            "checks (Balance Sheet balance, Du Pont vs direct ROE, no #DIV/0! on "
            "the Ratios tab) tell the Stage 5 LLM what to verify before writing "
            "narrative. Without them, the LLM has no way to know its output is "
            "self-consistent."
        ))

    # Part B
    if "PART_B_MISSING" in g.flags:
        s.append(core(
            "**Part B (Analysis Specification, sections 8–11) is missing.** Part A "
            "defines the model; Part B defines what to *do* with the model — "
            "Analysis Requirements (8), Du Pont Decomposition (9), Strategic "
            "Recommendation Requirements (10), Output Format (11). Without Part B "
            "the Stage 5 LLM gets a model but no marching orders."
        ))
    else:
        for n, label, suggestion in (
            (8, "Analysis Requirements", "what to interpret per category, "
                                          "benchmarks, cross-category connections"),
            (9, "Du Pont Decomposition", "the four-factor breakdown and which "
                                          "driver matters most"),
            (10, "Strategic Recommendation Requirements", "how many recs (3–5), "
                                                            "evidence standard, "
                                                            "level of specificity"),
            (11, "Output Format", "deliverable structure, sections in order, "
                                    "length targets, tone, audience"),
        ):
            if not spec.sections_found.get(n):
                s.append(core(
                    f"**Section {n} ({label}) appears missing.** The brief asks for "
                    f"{suggestion}."
                ))

    # Spec craft
    if "SPEC_THIN" in g.flags:
        s.append(core(
            f"The spec is on the short side ({spec.word_count} words). The brief "
            f"targets 3–5 pages — most students land at 1,500–2,500 words once "
            f"every section is fleshed out. Sparse sections at Stage 4 turn into "
            f"hallucinations at Stage 5."
        ))
    if "NAMED_RANGES_THIN" in g.flags:
        s.append(core(
            "Named-range notation (`BAL_*`, `INC_*`, `CASH_*`, `RATIO_*`) "
            "appears infrequently in the spec. The Stage 5 LLM uses these names "
            "as anchors — replace generic phrasing (\"net income\") with the "
            "exact range name (`INC_net`) wherever a formula or value is cited."
        ))

    # Prompt log + HIL
    if "PROMPT_LOG_MISSING" in g.flags:
        s.append(core(
            "**Prompt log not found.** Add a `deliverables/prompt-log.md` entry "
            "for each meaningful spec-drafting session — at minimum: intent, "
            "exact prompt(s) submitted, LLM used, and a note on what changed "
            "between rounds. The prompt log is half of the \"spec craft\" "
            "rubric line."
        ))
    if "HIL_MISSING" in g.flags:
        s.append(core(
            "**No visible HIL iteration evidence.** The brief asks for one of: "
            "(a) a 150–250 word before/after note in the prompt log describing "
            "a gap you found in the LLM's first draft and what you changed, "
            "(b) a clearly-labeled round-2 prompt that responded to a specific "
            "round-1 gap, or (c) an annotated diff at "
            "`analysis/validation/YYYY-MM-DD-{lastname}-{company}-stage4-iteration.md`. "
            "Self-assessment of the final draft is not the same — the iteration "
            "must show *the gap you caught* and *what you changed*."
        ))

    if "SPEC_NOT_IN_REPO" in g.flags:
        s.append(core(
            f"The spec was submitted via Lamaku but the canonical path "
            f"`docs/specs/{spec.filename}` doesn't appear in your repo. Stage 5 "
            f"requires every prior artifact to live in the repo (Lamaku fallback "
            f"does not apply at Stage 5). Commit the spec to `docs/specs/` now "
            f"to avoid the cleanup later."
        ))
    if "PROMPT_LOG_NOT_IN_REPO" in g.flags:
        s.append(core(
            "The prompt log should live at `deliverables/prompt-log.md` in your "
            "repo. The file accumulates across stages — append your Stage 4 "
            "session to it rather than creating a separate file per stage."
        ))

    if "INSTRUCTOR_NOT_COLLABORATOR" in g.flags:
        s.append(backward(
            f"**Still open from Stage 2:** `@{INSTRUCTOR_GITHUB_HANDLE}` is not a "
            "Write collaborator on your repo, which blocks the tracked-feedback "
            "PR workflow. Settings → Collaborators → Add people → **Write**. "
            "Not re-deducted here; closing this before the deadline can bump "
            "your Stage 2 score at the revision sweep — and Stage 5's "
            "feedback-incorporation rubric line depends on it (5% of S5)."
        ))

    core_count = sum(1 for x in s if x.bucket == "core")
    if "STRONG" in g.flags and core_count == 0:
        s.append(core(
            "Strong submission — Parts A and B both fully developed, named-range "
            "notation used consistently, and visible HIL iteration on the prompt "
            "log. Stage 5 builds directly on this — feed *only* the spec to the "
            "LLM at Stage 5, verify five ratios by hand, and the deliverable falls "
            "out the other side."
        ))

    fwd = next_stage_pointer(STAGE_N)
    if fwd is not None:
        s.append(fwd)
    return s


def _summary_tagline(g: Grade) -> str:
    flags = set(g.flags)
    if "SPEC_NOT_SUBMITTED" in flags:
        return "No spec submitted — floor does not apply."
    if "STRONG" in flags:
        return "Strong on merit; see suggestions for refinements."
    if "PART_B_MISSING" in flags:
        return "Part B missing — Stage 5 LLM will have no marching orders."
    if "RATIOS_MISSING" in flags or "RATIO_TABLE_THIN" in flags:
        return "Section 6 (Ratio Definitions) needs expansion."
    if "HIL_MISSING" in flags:
        return "Spec drafted but no visible HIL iteration on the prompt log."
    if "DATA_INPUTS_MISSING" in flags:
        return "Section 3 (Data Inputs) needs numerical values from Stage 3."
    return "See per-student suggestions for refinements."


# ------------------------------------------------------------------
# Report generation
# ------------------------------------------------------------------

def _letter_for(score_pts: int) -> str:
    for letter, lo, hi in LETTER_GRADE_SCALE:
        lo_pts = round(lo * TOTAL_POINTS / 100)
        if score_pts >= lo_pts and (hi is None or score_pts < round(hi * TOTAL_POINTS / 100)):
            return letter
    return "F"


def _final_score(g: Grade, floor_pct: int) -> int:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    if g.spec.found and g.raw_total < floor_value:
        return floor_value
    return g.raw_total


def _floor_was_applied(g: Grade, floor_pct: int) -> bool:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    return g.spec.found and g.raw_total < floor_value


def _student_section(n: int, g: Grade, floor_pct: int) -> str:
    """Internal grade report section. Has scores."""
    sub = g.submission
    spec = g.spec
    submitted = sub.submitted_at.strftime("%Y-%m-%d %I:%M %p") if sub.submitted_at else "—"
    raw = g.raw_total
    floor_applied = _floor_was_applied(g, floor_pct)
    final = _final_score(g, floor_pct)
    letter = _letter_for(final)

    lines: list[str] = []
    suffix = ", floor applied" if floor_applied else ""
    lines.append(f"## {n}. {sub.student_name} — **{final} / 100** ({letter}{suffix})")
    lines.append("")
    if sub.repo_url:
        lines.append(f"**Repo:** {sub.repo_url}")
    lines.append(f"**Spec:** `{spec.filename}`" if spec.found else "**Spec:** *(not submitted)*")
    lines.append(f"**Submitted:** {submitted}")
    lines.append("")
    lines.append("| Criterion | Earned | Notes |")
    lines.append("|-----------|--------|-------|")

    # Criterion 1
    secs_a = ", ".join(str(i) for i in (1, 2, 3, 4, 5) if spec.sections_found.get(i))
    lines.append(
        f"| Model spec — Data & Structure (Part A 1–5) | "
        f"{g.score_data_structure} / 25 | Sections present: {secs_a or 'none'}. "
        f"Spec word count: {spec.word_count}. |"
    )

    # Criterion 2
    cats = ", ".join(g.spec.ratio_categories_present) or "none detected"
    lines.append(
        f"| Model spec — Ratios & Validation (Part A 6–7) | "
        f"{g.score_ratios_validation} / 25 | Categories: {cats}; ratio rows ~"
        f"{spec.ratio_table_rows}; validation rules: {spec.validation_rules_count}. |"
    )

    # Criterion 3
    secs_b = ", ".join(str(i) for i in (8, 9, 10, 11) if spec.sections_found.get(i))
    lines.append(
        f"| Analysis spec (Part B 8–11) | {g.score_analysis} / 25 | "
        f"Sections present: {secs_b or 'none'}. |"
    )

    # Criterion 4
    nr_total = sum(g.spec.named_range_hits.values())
    plog_state = (
        f"prompt log {g.plog.word_count}w / {g.plog.prompt_blocks} block(s); "
        f"HIL strong={g.plog.hil_strong_hits} weak={g.plog.hil_weak_hits}"
        if g.plog.found else "no prompt log"
    )
    lines.append(
        f"| Spec craft + prompt log + HIL | {g.score_craft_hil} / 25 | "
        f"Named-range hits: {nr_total}; {plog_state}. |"
    )

    if floor_applied:
        floor_value = round(TOTAL_POINTS * floor_pct / 100)
        lines.append(f"| **Raw total** | **{raw} / 100** | |")
        lines.append(
            f"| **Floor adjustment** | **+{floor_value - raw}** | "
            f"Spec present — floor of {floor_pct} applied. |"
        )
        lines.append(f"| **Final** | **{final} / 100** | |")
    else:
        merit_note = " Above the floor — earned on merit." if raw >= round(
            TOTAL_POINTS * floor_pct / 100
        ) else ""
        lines.append(f"| **Total** | **{final} / 100** |{merit_note} |")
    lines.append("")

    suggestions = _suggestions_for(g)
    lines.extend(render_suggestions(suggestions, stage_n=STAGE_N))
    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def _build_full_report(grades: list[Grade], floor_pct: int, today: datetime) -> str:
    lines = [
        "# BUS-629 Stage 4 — Grade Report",
        "",
        f"**Stage:** {STAGE_LABEL} (20% of project score)",
        f"**Graded:** {today.strftime('%Y-%m-%d')}",
        f"**Submissions reviewed:** {len(grades)}",
        f"**Floor policy:** {floor_pct}% floor for submissions with a "
        "substantive spec file.",
        "**Score privacy:** scores live in this internal file only. The "
        "per-student suggestion blocks are also rendered under "
        "`_pr_feedback/<lastname>/feedback-file.md` *without scores* for "
        "PR delivery to student repos.",
        "",
        "---",
        "",
        "## Rubric (recap)",
        "",
        "| Criterion | Weight |",
        "|-----------|--------|",
        "| Model spec — Data & Structure (Part A 1–5) | 25% |",
        "| Model spec — Ratios & Validation (Part A 6–7) | 25% |",
        "| Analysis spec (Part B 8–11) | 25% |",
        "| Spec craft + prompt log + HIL iteration | 25% |",
        "",
        "---",
        "",
    ]
    for i, g in enumerate(grades, 1):
        lines.append(_student_section(i, g, floor_pct))
    lines.append("## Class summary")
    lines.append("")
    lines.append("| Student | Score | Notes |")
    lines.append("|---------|-------|-------|")
    for g in grades:
        lines.append(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |"
        )
    lines.append("")
    finals = [_final_score(g, floor_pct) for g in grades if g.spec.found]
    submitted_n = len(finals)
    mean = sum(finals) / submitted_n if submitted_n else 0.0
    floors = sum(1 for g in grades if _floor_was_applied(g, floor_pct))
    lines.append(f"**Mean (submissions only):** {mean:.1f}")
    lines.append(f"**Submission rate:** {submitted_n} of {len(grades)}")
    lines.append(f"**Floor applied:** {floors} of {submitted_n} submissions")
    lines.append("")
    return "\n".join(lines)


def write_or_update_grade_report(
    grades: list[Grade], floor_pct: int, report_path: Path,
    today: datetime | None = None,
) -> list[Grade]:
    today = today or datetime.now()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    if not report_path.exists():
        report_path.write_text(_build_full_report(grades, floor_pct, today),
                               encoding="utf-8")
        return list(grades)

    text = report_path.read_text(encoding="utf-8")
    existing_urls = {
        u.lower().rstrip("/")
        for u in re.findall(r"https?://github\.com/[A-Za-z0-9_./-]+", text)
    }
    existing_names = {
        _normalize_name(n)
        for n in re.findall(r"^## \d+\. ([^—\n]+?) — \*\*\d+ / 100\*\*",
                            text, re.MULTILINE)
    }
    new_grades = []
    for g in grades:
        url_key = (g.submission.repo_url or "").lower().rstrip("/")
        name_key = _normalize_name(g.submission.student_name)
        if url_key and url_key in existing_urls:
            continue
        if name_key and name_key in existing_names:
            continue
        new_grades.append(g)
    if not new_grades:
        return []

    nums = [int(n) for n in re.findall(r"^## (\d+)\.", text, re.MULTILINE)]
    next_n = (max(nums) + 1) if nums else 1
    chunks = []
    for g in new_grades:
        chunks.append(_student_section(next_n, g, floor_pct))
        next_n += 1
    new_block = "\n".join(chunks)

    if "## Class summary" in text:
        text = text.replace("## Class summary", new_block + "\n## Class summary", 1)
    else:
        text = text.rstrip() + "\n\n" + new_block

    table_re = re.compile(
        r"(## Class summary\s*\n+"
        r"\| Student \| Score \| Notes \|\n"
        r"\| ?-+ ?\| ?-+ ?\| ?-+ ?\|\n"
        r"(?:\|[^\n]*\|\n)+)"
    )
    m = table_re.search(text)
    if m:
        new_rows = "".join(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |\n"
            for g in new_grades
        )
        text = text[: m.end()] + new_rows + text[m.end():]

    text = re.sub(
        r"\*\*Graded:\*\* [^\n]*",
        f"**Graded:** {today.strftime('%Y-%m-%d')} "
        f"({', '.join(g.submission.student_name for g in new_grades)} added)",
        text, count=1,
    )

    report_path.write_text(text, encoding="utf-8")
    return new_grades


# ------------------------------------------------------------------
# PR-ready feedback file (no scores)
# ------------------------------------------------------------------

def _lastname_slug(name: str) -> str:
    """First reasonable last-name token, lowercased, [a-z0-9-] only.

    Heuristic: Vietnamese names often appear as `Family Middle Given` and the
    given name is what graders use. Western names follow `First Last`. We
    fall back to the last whitespace-separated token, which works for both
    when the export normalizes to "First Last" / "Family Given" forms.
    """
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9'-]+", name)
    if not tokens:
        return "student"
    last = tokens[-1].lower()
    return re.sub(r"[^a-z0-9-]", "", last) or "student"


def write_pr_feedback(g: Grade, today: datetime, feedback_dir: Path) -> Path:
    """Write a scoreless feedback file ready to push as a PR to the student's repo.

    File lands at `<feedback_dir>/<lastname>/feedback-file.md`. The student-
    facing structure mirrors the format we'd commit at
    `_feedback/YYYY-MM-DD-stage4-review.md` on their repo.
    """
    slug = _lastname_slug(g.submission.student_name)
    out_dir = feedback_dir / slug
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "feedback-file.md"

    date_str = today.strftime("%Y-%m-%d")
    spec = g.spec
    plog = g.plog

    lines: list[str] = []
    lines.append(f"# Stage 4 review — {date_str}")
    lines.append("")
    lines.append(
        f"Reviewing the Stage 4 spec at "
        f"`docs/specs/{spec.filename}`" if spec.found and spec.filename
        else "Reviewing your Stage 4 submission."
    )
    lines.append("")

    # Section-coverage table — no scores, just presence/absence.
    lines.append("## Section coverage")
    lines.append("")
    lines.append("| Section | Present | Word count |")
    lines.append("|---|---|---|")
    sect_labels = {
        1: "1. Scope & Objective",
        2: "2. Model Architecture",
        3: "3. Data Inputs",
        4: "4. Named Range Conventions",
        5: "5. Derived Inputs",
        6: "6. Ratio Definitions & Formulas",
        7: "7. Validation Rules",
        8: "8. Analysis Requirements (Part B)",
        9: "9. Du Pont Decomposition (Part B)",
        10: "10. Strategic Recommendations (Part B)",
        11: "11. Output Format (Part B)",
    }
    for n in range(1, 12):
        present = "✓" if spec.sections_found.get(n) else "—"
        wc = spec.section_word_counts.get(n, 0)
        lines.append(f"| {sect_labels[n]} | {present} | {wc} |")
    lines.append("")

    # Quick observations
    lines.append("## Observations")
    lines.append("")
    nr_total = sum(spec.named_range_hits.values())
    lines.append(
        f"- Spec length: **{spec.word_count} words** (brief targets 3–5 pages, "
        f"~1,500–2,500 words)."
    )
    lines.append(
        f"- Named-range notation usage: **{nr_total} hit(s)** across "
        f"`BAL_*`, `INC_*`, `CASH_*`, `RATIO_*`, `startYear_*`, `currentYear_*`, "
        f"`avg_*`."
    )
    cats = g.spec.ratio_categories_present
    if cats:
        lines.append(
            f"- Ratio categories detected in Section 6: "
            f"**{', '.join(cats)}** ({len(cats)}/6)."
        )
    if spec.ratio_table_rows:
        lines.append(f"- Ratio table rows in Section 6: **{spec.ratio_table_rows}**.")
    if spec.validation_rules_count:
        lines.append(
            f"- Validation rules counted in Section 7: "
            f"**{spec.validation_rules_count}**."
        )
    if plog.found:
        lines.append(
            f"- Prompt log: **{plog.word_count} words**, "
            f"{plog.prompt_blocks} explicit prompt block(s); "
            f"HIL signals: {plog.hil_strong_hits} strong, {plog.hil_weak_hits} weak."
        )
    else:
        lines.append("- Prompt log: **not detected** in the submission.")
    lines.append("")

    # Suggestions — same as the internal report, just without scores.
    suggestions = _suggestions_for(g)
    lines.extend(render_suggestions(suggestions, stage_n=STAGE_N))
    lines.append("")
    lines.append(
        "*This review is feedback-only — no scores included.* Score numbers "
        "live in the internal grade report and the instructor's email; this "
        "file is intended for review against your repo state."
    )
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


# ------------------------------------------------------------------
# Worksheet
# ------------------------------------------------------------------

def build_worksheet(grades: list[Grade], floor_pct: int, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"
    headers = [
        "Student", "Submitted", "Repo URL", "Spec filename",
        "Data & Structure /25", "Ratios & Validation /25",
        "Analysis Spec /25", "Craft + HIL /25",
        "Raw /100", "Final /100", "Floored /100",
        "Sections present (A)", "Sections present (B)",
        "Spec words", "Named-range hits", "Ratio rows",
        "Validation rules", "Prompt log words", "Prompt blocks",
        "HIL strong", "HIL weak", "Flags",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F8CBAD")
    floor_fill = PatternFill("solid", fgColor="E2EFDA")
    final_col = headers.index("Final /100") + 1
    floor_col = headers.index("Floored /100") + 1
    flags_col = headers.index("Flags") + 1

    for g in grades:
        sub = g.submission
        spec = g.spec
        plog = g.plog
        submitted = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else ""
        secs_a = ",".join(str(i) for i in (1, 2, 3, 4, 5) if spec.sections_found.get(i))
        secs_b = ",".join(str(i) for i in (8, 9, 10, 11) if spec.sections_found.get(i))
        nr_total = sum(spec.named_range_hits.values())
        row = [
            sub.student_name, submitted, sub.repo_url, spec.filename,
            g.score_data_structure, g.score_ratios_validation,
            g.score_analysis, g.score_craft_hil,
            g.raw_total, g.raw_total, None,
            secs_a, secs_b, spec.word_count, nr_total,
            spec.ratio_table_rows, spec.validation_rules_count,
            plog.word_count, plog.prompt_blocks,
            plog.hil_strong_hits, plog.hil_weak_hits,
            ", ".join(g.flags),
        ]
        ws.append(row)
        r = ws.max_row
        final_ref = f"{get_column_letter(final_col)}{r}"
        ws.cell(
            row=r, column=floor_col,
            value=f'=IF({final_ref}=0,0,MAX({final_ref},{floor_value}))',
        ).fill = floor_fill
        if not spec.found:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = error_fill
        elif g.flags and g.flags != ["STRONG"]:
            ws.cell(row=r, column=flags_col).fill = flag_fill

    widths = [24, 17, 50, 40, 12, 14, 12, 12, 10, 11, 13,
              18, 18, 12, 14, 12, 14, 14, 12, 11, 11, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(output_path)


# ------------------------------------------------------------------
# Sweep entry point — rescore a student against current repo state
# ------------------------------------------------------------------

def _download_blob(owner: str, repo: str, path: str, branch: str) -> str | None:
    if not _safe_inputs(owner, repo, branch):
        return None
    if not _SAFE_PATH_RE.match(path):
        return None
    raw = _gh(
        "api", f"repos/{owner}/{repo}/contents/{path}",
        "-X", "GET", "-f", f"ref={branch}",
    )
    if not raw:
        return None
    try:
        meta = json.loads(raw)
    except json.JSONDecodeError:
        return None
    content = meta.get("content")
    if not content:
        return None
    try:
        return base64.b64decode(content).decode("utf-8", errors="replace")
    except Exception:
        return None


def rescore_from_repo(
    student_name: str,
    repo_url: str,
    submitted_at: datetime | None = None,
    prior: PriorGrade | None = None,
    student_id: str = "",
) -> Grade | None:
    """Sweep entry point: rescore Stage 4 against current repo state.

    Pulls the spec file (any `docs/specs/*-spec.md`), prompt log
    (`deliverables/prompt-log.md`), and optional iteration file
    (`analysis/validation/*stage4-iteration.md`) from the repo and rescores.
    """
    m = GITHUB_URL_RE.search(repo_url)
    if not m:
        return None
    owner, repo = m.group("owner"), m.group("repo")
    repo_info = inspect_repo(owner, repo)
    spec_text: str | None = None
    plog_text: str | None = None
    if repo_info.accessible:
        if repo_info.spec_in_repo:
            spec_text = _download_blob(
                owner, repo, repo_info.spec_repo_path, repo_info.default_branch
            )
        if repo_info.prompt_log_in_repo:
            plog_text = _download_blob(
                owner, repo, repo_info.prompt_log_repo_path, repo_info.default_branch
            )

    spec_filename = (
        Path(repo_info.spec_repo_path).name
        if repo_info.spec_repo_path else "(in-repo spec)"
    )
    spec_info = inspect_spec(None, content=spec_text or "",
                              source_label=spec_filename) if spec_text else SpecInspection()
    plog_info = inspect_prompt_log(None, content=plog_text or "",
                                    source_label="prompt-log.md") if plog_text else PromptLogInspection()

    sub = Submission(
        student_id=student_id, student_name=student_name,
        submitted_at=submitted_at, spec_path=None, prompt_log_path=None,
        iteration_path=None, repo_url=repo_url, owner=owner, repo=repo,
    )
    return score(sub, spec_info, plog_info, repo_info, prior)


# ------------------------------------------------------------------
# Workflow helper
# ------------------------------------------------------------------

def _detect_workflow(export_path: Path):
    if not (export_path.is_file() and export_path.suffix.lower() == ".zip"):
        return None
    parent = export_path.parent
    if parent.name.lower() != "ungraded":
        return None
    graded = parent.parent / "graded"
    scratch = parent / f"_{export_path.stem}_extracted"
    return graded, scratch


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("export", type=Path)
    p.add_argument("--floor", type=int, default=DEFAULT_FLOOR_PCT)
    p.add_argument("--prior-stage0", type=Path, default=None)
    p.add_argument("--prior-stage1", type=Path, default=None)
    p.add_argument("--prior-stage2", type=Path, default=None)
    p.add_argument("--prior-stage3", type=Path, default=None)
    p.add_argument("--out", type=Path, default=None)
    p.add_argument("--no-move", action="store_true")
    args = p.parse_args(argv)

    subs = discover_submissions(args.export)
    if not subs:
        print("No submissions discovered.")
        return 1
    print(f"Found {len(subs)} submission(s).")

    base_root = args.export.resolve().parents[2]
    s0 = args.prior_stage0 or (base_root / "stage0" / "graded" / "STAGE0_GRADES.md")
    s1 = args.prior_stage1 or (base_root / "stage1" / "graded" / "STAGE1_GRADES.md")
    s2 = args.prior_stage2 or (base_root / "stage2" / "graded" / "STAGE2_GRADES.md")
    s3 = args.prior_stage3 or (base_root / "stage3" / "graded" / "STAGE3_GRADES.md")
    prior0 = parse_prior_report(s0) if s0.exists() else {}
    prior1 = parse_prior_report(s1) if s1.exists() else {}
    prior2 = parse_prior_report(s2) if s2.exists() else {}
    prior3 = parse_prior_report(s3) if s3.exists() else {}
    for label, p_path, p_map in (
        ("Stage 0", s0, prior0), ("Stage 1", s1, prior1),
        ("Stage 2", s2, prior2), ("Stage 3", s3, prior3),
    ):
        if p_map:
            print(f"Loaded {label} records: {len(p_map)} from {p_path}")
    prior = merge_prior(prior0, prior1, prior2, prior3)

    # Repo-URL fill-in from priors.
    for s in subs:
        if s.repo_url:
            continue
        pg = lookup_prior(prior, s.student_name)
        if pg and pg.repo_url:
            m = GITHUB_URL_RE.search(pg.repo_url)
            if m:
                s.repo_url = pg.repo_url
                s.owner = m.group("owner")
                s.repo = m.group("repo")

    grades: list[Grade] = []
    for s in subs:
        print(f"  inspecting {s.student_name} ...", end=" ", flush=True)
        # Spec from upload, or fall back to repo if zip had only an HTML pointer.
        spec_info = SpecInspection()
        plog_info = PromptLogInspection()
        repo = RepoInspection()
        if s.owner:
            repo = inspect_repo(s.owner, s.repo)
        if s.spec_path:
            spec_info = inspect_spec(s.spec_path)
        elif repo.accessible and repo.spec_in_repo:
            spec_text = _download_blob(
                s.owner, s.repo, repo.spec_repo_path, repo.default_branch
            )
            if spec_text:
                spec_info = inspect_spec(
                    None, content=spec_text,
                    source_label=Path(repo.spec_repo_path).name,
                )
        if s.prompt_log_path:
            plog_info = inspect_prompt_log(s.prompt_log_path)
        elif repo.accessible and repo.prompt_log_in_repo:
            plog_text = _download_blob(
                s.owner, s.repo, repo.prompt_log_repo_path, repo.default_branch
            )
            if plog_text:
                plog_info = inspect_prompt_log(
                    None, content=plog_text, source_label="prompt-log.md"
                )

        pg = lookup_prior(prior, s.student_name)
        g = score(s, spec_info, plog_info, repo, pg)
        grades.append(g)
        if not spec_info.found:
            print("NO SPEC FOUND")
        else:
            if repo.queried and repo.accessible and repo.collaborator_check_ran:
                collab = (
                    f"collab=Y({repo.collaborator_permission})"
                    if repo.instructor_is_collaborator
                    else f"collab=N({repo.collaborator_permission or 'none'})"
                )
            elif repo.queried and not repo.accessible:
                collab = "collab=?(repo not accessible)"
            elif repo.queried:
                collab = "collab=?(gh missing/failed)"
            else:
                collab = "collab=?(no repo URL)"
            print(
                f"raw={g.raw_total}/100 "
                f"a={g.score_data_structure} b={g.score_ratios_validation} "
                f"c={g.score_analysis} d={g.score_craft_hil} "
                f"sections_A={sum(1 for i in (1,2,3,4,5) if spec_info.sections_found.get(i))}/5 "
                f"sections_B={sum(1 for i in (8,9,10,11) if spec_info.sections_found.get(i))}/4 "
                f"{collab} flags={','.join(g.flags) or '-'}"
            )

    workflow = _detect_workflow(args.export)
    if args.out is not None:
        worksheet_path = args.out
    elif workflow is not None:
        graded_dir, _ = workflow
        graded_dir.mkdir(parents=True, exist_ok=True)
        ws_dir = graded_dir / "_worksheets"
        ws_dir.mkdir(exist_ok=True)
        worksheet_path = ws_dir / f"stage4-{args.export.stem}.xlsx"
    else:
        worksheet_path = (
            (args.export.parent if args.export.is_file() else args.export)
            / "_grading" / "stage4-grading-worksheet.xlsx"
        )

    build_worksheet(grades, args.floor, worksheet_path)
    print(f"\nWrote worksheet: {worksheet_path}")

    today = datetime.now()
    if workflow is not None:
        graded_dir, scratch = workflow
        report_path = graded_dir / "STAGE4_GRADES.md"
        new_entries = write_or_update_grade_report(
            grades, args.floor, report_path, today
        )
        if new_entries:
            names = ", ".join(g.submission.student_name for g in new_entries)
            print(f"Updated report: {report_path} (+{len(new_entries)} new: {names})")
        else:
            print(f"Report already up to date: {report_path}")

        # Per-student PR feedback files (no scores).
        pr_dir = graded_dir / "_pr_feedback"
        pr_dir.mkdir(exist_ok=True)
        for g in grades:
            if not g.spec.found:
                continue
            out = write_pr_feedback(g, today, pr_dir)
            print(f"Wrote PR feedback: {out}")

        if not args.no_move:
            dest = graded_dir / args.export.name
            if dest.exists():
                print(f"Source zip already exists at {dest} — leaving original in place.")
            else:
                shutil.move(str(args.export), str(dest))
                print(f"Moved source zip → {dest}")
            if scratch.exists():
                shutil.rmtree(scratch, ignore_errors=True)
                print(f"Cleaned up scratch extract at {scratch}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
