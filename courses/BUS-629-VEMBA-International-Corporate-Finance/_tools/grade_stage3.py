"""BUS-629 Stage 3 (populated financials workbook) grading scanner.

Stage 3 submissions are an `.xlsx` workbook named per the convention
`YYYY-MM-DD-{lastname}-{company-slug}-financials.xlsx` containing Income
Statement, Balance Sheet, and Cash Flow data for the company selected at
Stage 2. The workbook is typically committed to `models/builds/` in the
student's repo; Lamaku upload is an accepted fallback.

This script:

  1. Reads the Lamaku export zip (or extracted directory).
  2. For each student, inspects the submitted `.xlsx`:
       - filename convention
       - sheet presence (Cover/Notes, Balance Sheet, Income Statement,
         Cash Flow, Ratios)
       - named-range completeness (% of BAL_*, INC_*, CASH_*, startYear_*
         cells populated)
       - balance-sheet identity (Assets = Liabilities + Equity, both years)
       - source-documentation completeness on Cover OR Notes tab
       - Ratios-tab error scan (#REF!, #DIV/0!, #NAME?, #VALUE!)
  3. Looks the student up in prior-stage reports (Stage 0/1/2) for repo
     URL and carry-over flags (no double-deduction policy).
  4. Optional repo check via `gh`: does the workbook exist at
     `models/builds/`? Is the instructor still a Write collaborator?
  5. Scores the 4-criterion rubric (40/25/20/15).
  6. Writes/updates `../graded/STAGE3_GRADES.md` (append mode, dedupes).
  7. Moves the source zip to `graded/` and cleans scratch.

The four-criterion rubric (per `stage3-model-population-validation.md`):

    Data accuracy (ties to source 10-K)        /40
    Completeness (both years populated)        /25
    Source documentation                        /20
    Auto-computed ratios resolve cleanly        /15
                                               = /100

Floor policy: 75% for working-repo submissions with a populated workbook
(set via instructor direction, 2026-05-14). Pass `--floor` to override.

Double-deduction policy: issues already flagged in Stage 0/1/2 are not
re-deducted here. They appear as forward-looking tips with no point loss.

USAGE:
    python grade_stage3.py <export.zip> [--floor=75] [--no-move]
        [--prior-stage0=path/to/STAGE0_GRADES.md]
        [--prior-stage1=path/to/STAGE1_GRADES.md]
        [--prior-stage2=path/to/STAGE2_GRADES.md]

REQUIREMENTS:
    - openpyxl (`pip install openpyxl`)
    - `gh` CLI authenticated against github.com (optional; falls back to
      file-only scoring without it)
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _grading_comments import (
    Suggestion,
    backward,
    core,
    forward,
    next_stage_pointer,
    render_suggestions,
)

STAGE_N = 3
DEFAULT_FLOOR_PCT = 75
TOTAL_POINTS = 100
STAGE_LABEL = "Stage 3 — Populated Financials Workbook"
STAGE_TARGET_DIR = "models/builds"
INSTRUCTOR_GITHUB_HANDLE = "adamwstauffer"

# Named-range buckets we expect to find populated. Each bucket is scored
# separately; missing names in a bucket reduce that bucket's contribution to
# the completeness score.
REQUIRED_BAL_CURR = [
    "BAL_assets_total_curr", "BAL_liabilities_total_curr",
    "BAL_equity_shareholders_curr", "BAL_assets_current_curr",
    "BAL_liabilities_current_curr", "BAL_cash_marketable_securities_curr",
    "BAL_receivables_curr", "BAL_inventories_curr", "BAL_debt_long_term_curr",
    "BAL_debt_short_term_curr", "BAL_accounts_payable_curr",
    "BAL_fixed_assets_net_curr", "BAL_retained_earnings_curr",
    "BAL_common_stock_curr",
]
REQUIRED_BAL_PRIOR = [
    "BAL_assets_total_prior", "BAL_liabilities_total_prior",
    "BAL_equity_shareholders_prior", "BAL_assets_current_prior",
    "BAL_liabilities_current_prior", "BAL_cash_marketable_securities_prior",
    "BAL_receivables_prior", "BAL_inventories_prior",
    "BAL_debt_long_term_prior", "BAL_debt_short_term_prior",
    "BAL_accounts_payable_prior", "BAL_fixed_assets_net_prior",
    "BAL_retained_earnings_prior", "BAL_common_stock_prior",
]
REQUIRED_INC = [
    "INC_sales", "INC_cost_goods_sold", "INC_sga", "INC_ebit",
    "INC_interest_expense", "INC_taxable_income", "INC_taxes", "INC_net",
]
REQUIRED_CASH = ["CASH_operating", "CASH_investments", "CF_depreciation_amortization"]
REQUIRED_ASSUMPTIONS = ["share_price", "shares_outstanding", "cost_capital", "tax_rate"]

# Cover/Notes tab metadata we want to see filled in. Each field gets 5 pts.
DOC_FIELDS = ["company_or_ticker", "fiscal_year", "source_url", "reporting_standard",
              "currency_or_units", "tax_rate_doc"]

# Filename convention: YYYY-MM-DD-{lastname}-{company-slug}-financials.xlsx
FILENAME_RE = re.compile(
    r"^(?P<date>\d{4}-\d{2}-\d{2})-"
    r"(?P<lastname>[a-z]+(?:-[a-z]+)*)-"
    r"(?P<company>[a-z0-9]+(?:-[a-z0-9]+)*)-"
    r"financials\.xlsx$"
)

GITHUB_URL_RE = re.compile(
    r"https?://github\.com/(?P<owner>[A-Za-z0-9_.-]+)/(?P<repo>[A-Za-z0-9_.-]+)"
)
FOLDER_NAME_RE = re.compile(
    r"^(?P<sid>\d+)-\d+\s*-\s*(?P<name>.+?)\s*-\s*"
    r"(?P<month>[A-Za-z]+)\s+(?P<day>\d+),\s*(?P<year>\d{4})\s+"
    r"(?P<h>\d{1,4})\s*(?P<ampm>AM|PM)\s*$"
)

LETTER_GRADE_SCALE = [
    ("A+", 97, None), ("A", 93, 97), ("A-", 90, 93),
    ("B+", 87, 90), ("B", 83, 87), ("B-", 80, 83),
    ("C+", 77, 80), ("C", 73, 77), ("C-", 70, 73),
    ("D+", 67, 70), ("D", 65, 67), ("F", 0, 65),
]

ERROR_VALUES = {"#REF!", "#DIV/0!", "#NAME?", "#VALUE!", "#NULL!", "#N/A", "#NUM!"}


# ------------------------------------------------------------------
# Submission discovery
# ------------------------------------------------------------------

@dataclass
class Submission:
    student_id: str
    student_name: str
    submitted_at: datetime | None
    workbook_path: Path
    repo_url: str = ""
    owner: str = ""
    repo: str = ""


def _parse_folder_name(name: str):
    m = FOLDER_NAME_RE.match(name)
    if not m:
        return None, None, None
    h = m.group("h")
    if len(h) == 3:
        hour, minute = int(h[0]), int(h[1:])
    elif len(h) == 4:
        hour, minute = int(h[:2]), int(h[2:])
    else:
        hour, minute = int(h), 0
    ampm = m.group("ampm").upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0
    try:
        dt = datetime.strptime(
            f"{m.group('month')} {m.group('day')} {m.group('year')}", "%b %d %Y"
        ).replace(hour=hour, minute=minute)
    except ValueError:
        dt = None
    return m.group("sid"), m.group("name").strip(), dt


def discover_submissions(export_path: Path) -> list[Submission]:
    if export_path.is_file() and export_path.suffix.lower() == ".zip":
        scratch = export_path.parent / f"_{export_path.stem}_extracted"
        scratch.mkdir(exist_ok=True)
        with zipfile.ZipFile(export_path) as zf:
            zf.extractall(scratch)
        root = scratch
    elif export_path.is_dir():
        root = export_path
    else:
        raise SystemExit(f"Cannot read submission export at {export_path}")

    subs: list[Submission] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        sid, name, dt = _parse_folder_name(child.name)
        if not name:
            continue
        # Find the most-recent-looking xlsx (ignore Excel temp files)
        candidates = [
            p for p in child.glob("*.xlsx")
            if not p.name.startswith("~$")
        ]
        if not candidates:
            print(f"  WARN: no .xlsx in '{child.name}' — skipping {name}")
            continue
        wb_path = candidates[0]
        subs.append(Submission(
            student_id=sid or "",
            student_name=name,
            submitted_at=dt,
            workbook_path=wb_path,
        ))
    subs.sort(key=lambda s: s.student_name.lower())
    return subs


# ------------------------------------------------------------------
# Prior-stage parsing (carry-over awareness)
# ------------------------------------------------------------------

@dataclass
class PriorGrade:
    student_name: str
    repo_url: str
    final_score: int
    raw_section: str
    carry_over_tags: set[str] = field(default_factory=set)


_CARRY_OVER_PATTERNS = [
    ("READMES_PLACEHOLDER", re.compile(r"placeholder readme|empty readme")),
    ("DIRS_INCOMPLETE",     re.compile(r"missing.*director|directory.*missing")),
    ("BIO_PLACEHOLDER",     re.compile(r"(fill in|expand).*bio")),
    ("COMMIT_HYGIENE",      re.compile(r"commit message|tighten commit|descriptive commit")),
    ("GITIGNORE",           re.compile(r"\.ds_store|gitignore|~\$.*xlsx")),
    ("CASING",              re.compile(r"casing|capitaliz")),
    ("LICENSE",             re.compile(r"add a (license|`license`)")),
    ("FILENAME_CONVENTION", re.compile(r"phuong|first name|last name|"
                                       r"filename convention|canonical filename")),
    ("BIO_LENGTH",          re.compile(r"\d{2,3}\s*words.*target|bio.*length")),
]


def _normalize_name(name: str) -> str:
    tokens = re.findall(r"[A-Za-z]+", name.lower())
    return " ".join(sorted(tokens))


def parse_prior_report(report_path: Path) -> dict[str, PriorGrade]:
    if not report_path.exists():
        return {}
    text = report_path.read_text(encoding="utf-8")
    out: dict[str, PriorGrade] = {}
    header_re = re.compile(r"^## (\d+)\. (.+?) — \*\*(\d+) / 100\*\*", re.MULTILINE)
    headers = [(m.start(), m) for m in header_re.finditer(text)]
    headers.append((len(text), None))
    for i in range(len(headers) - 1):
        start, m = headers[i]
        end, _ = headers[i + 1]
        section = text[start:end]
        name = m.group(2).strip()
        score = int(m.group(3))
        url_m = GITHUB_URL_RE.search(section)
        repo_url = url_m.group(0) if url_m else ""
        carry = set()
        for tag, pat in _CARRY_OVER_PATTERNS:
            if pat.search(section.lower()):
                carry.add(tag)
        out[_normalize_name(name)] = PriorGrade(
            student_name=name, repo_url=repo_url, final_score=score,
            raw_section=section, carry_over_tags=carry,
        )
    return out


def lookup_prior(prior: dict[str, PriorGrade], student_name: str):
    key = _normalize_name(student_name)
    if key in prior:
        return prior[key]
    submitted_tokens = set(key.split())
    for pkey, pg in prior.items():
        if not pkey:
            continue
        prior_tokens = set(pkey.split())
        if submitted_tokens.issubset(prior_tokens) or prior_tokens.issubset(submitted_tokens):
            return pg
    return None


def merge_prior(*priors: dict[str, PriorGrade]) -> dict[str, PriorGrade]:
    out: dict[str, PriorGrade] = {}
    for prior in priors:
        for key, pg in prior.items():
            if key not in out:
                out[key] = PriorGrade(
                    student_name=pg.student_name, repo_url=pg.repo_url,
                    final_score=pg.final_score, raw_section=pg.raw_section,
                    carry_over_tags=set(pg.carry_over_tags),
                )
            else:
                out[key].carry_over_tags |= pg.carry_over_tags
                if not out[key].repo_url and pg.repo_url:
                    out[key].repo_url = pg.repo_url
    return out


# ------------------------------------------------------------------
# Workbook inspection
# ------------------------------------------------------------------

@dataclass
class WorkbookInspection:
    found: bool = False
    filename: str = ""
    filename_valid: bool = False
    filename_lastname: str = ""
    filename_company: str = ""

    sheets: list[str] = field(default_factory=list)
    named_ranges: dict[str, float | int | str | None] = field(default_factory=dict)
    error_cells: list[tuple[str, str, str]] = field(default_factory=list)  # (sheet, coord, value)
    error_cells_on_ratios: int = 0
    ratios_tab_numeric_cells: int = 0

    bs_balances_curr: bool = False
    bs_balances_prior: bool = False
    bs_diff_curr: float = 0.0
    bs_diff_prior: float = 0.0

    bal_curr_filled: int = 0
    bal_prior_filled: int = 0
    inc_filled: int = 0
    cash_filled: int = 0
    assumptions_filled: int = 0

    doc_fields_filled: dict[str, str] = field(default_factory=dict)
    doc_evidence: dict[str, str] = field(default_factory=dict)  # field -> snippet
    company_named: str = ""
    error: str = ""


def _gather_named_ranges(wb) -> dict[str, object]:
    """Pull current (computed) value for every defined name, looking it up
    on the sheet+coord it points at."""
    out: dict[str, object] = {}
    for name in wb.defined_names:
        try:
            dn = wb.defined_names[name]
        except KeyError:
            continue
        for sn, coord in dn.destinations:
            if sn not in wb.sheetnames:
                continue
            try:
                v = wb[sn][coord].value
            except Exception:
                v = None
            out[name] = v
            break  # only first destination
    return out


def _is_filled(v) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return False
        if s in ERROR_VALUES:
            return False
        # Tolerate purposeful "0" entries (e.g. dividends = 0) — they count as filled.
        return True
    return True


def _to_number(v) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _scan_for_errors(wb) -> list[tuple[str, str, str]]:
    """Return list of (sheet, coord, value) tuples for any cell whose
    computed value is an Excel error string."""
    errs = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip() in ERROR_VALUES:
                    errs.append((sn, cell.coordinate, v.strip()))
    return errs


def _scan_metadata_tab(wb) -> tuple[dict[str, str], dict[str, str], str]:
    """Look at the Cover / Notes / Cover & Instructions tab(s) for the
    Stage 3 documentation fields. Returns (filled, evidence, company_name).

    The scan is forgiving: we look for line/label-style content anywhere
    on the tabs whose name contains 'cover', 'notes', or 'instructions'.
    """
    filled: dict[str, str] = {}
    evidence: dict[str, str] = {}
    company = ""

    keywords = {
        "company_or_ticker": [r"\b(company|ticker)\b"],
        "fiscal_year": [r"\bfiscal year\b", r"\bFY\s*\d{4}\b", r"year[- ]end"],
        "source_url": [r"\bsec[- ]edgar\b", r"\b10[- ]?k\b", r"\b20[- ]?f\b",
                       r"\bannual report\b", r"hose|hnx|sgx|set|psx|bursa|idx",
                       r"\bIR\b|investor relations", r"https?://"],
        "reporting_standard": [r"\bUS\s*GAAP\b", r"\bIFRS\b", r"\bVAS\b",
                               r"reporting standard"],
        "currency_or_units": [r"\b(USD|EUR|VND|JPY|RMB|CNY|HKD|SGD|GBP|KRW|INR)\b",
                              r"\b\$?\s*(thousands?|millions?|billions?|'000)\b",
                              r"figures in"],
        "tax_rate_doc": [r"tax rate", r"\b\d{1,2}\s*%\s*(statutory|effective)?"],
    }

    target_sheets = [
        sn for sn in wb.sheetnames
        if any(k in sn.lower() for k in ("cover", "notes", "instructions",
                                          "metadata", "assumptions"))
    ]
    if not target_sheets:
        return filled, evidence, company

    for sn in target_sheets:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                txt = v.strip()
                if not txt or len(txt) > 500:
                    continue
                low = txt.lower()
                # Capture company name early if we hit a "Company: X" pattern.
                if not company:
                    m = re.match(r"(?i)^\s*company\s*[:\-]\s*(.+)$", txt)
                    if m and len(m.group(1)) > 1:
                        company = m.group(1).strip()
                # On a label cell, look at the next cell to the right for the
                # value. We capture either-or — any cell with the keyword that
                # also has a non-placeholder neighbor counts as evidence.
                for field_name, pats in keywords.items():
                    if field_name in filled:
                        continue
                    for pat in pats:
                        if re.search(pat, txt, re.IGNORECASE):
                            neighbor = None
                            try:
                                neighbor = ws.cell(cell.row, cell.column + 1).value
                            except Exception:
                                neighbor = None
                            evidence_text = (
                                str(neighbor).strip() if _is_filled(neighbor) else txt
                            )
                            # Reject placeholder neighbors like '[your company name]'
                            if (isinstance(neighbor, str)
                                and re.match(r"^\s*\[", neighbor.strip())):
                                continue
                            # If the keyword *is* the value (e.g. row contains
                            # 'iQIYI Inc' next to 'Company:'), accept either.
                            if _is_filled(neighbor) or len(txt) > 6:
                                filled[field_name] = evidence_text[:200]
                                evidence[field_name] = (
                                    f"{sn}!{cell.coordinate}: {txt[:60]}"
                                )
                                break
    return filled, evidence, company


def inspect_workbook(wb_path: Path) -> WorkbookInspection:
    info = WorkbookInspection(filename=wb_path.name)
    if not wb_path.exists():
        info.error = "workbook file not found"
        return info

    m = FILENAME_RE.match(wb_path.name)
    if m:
        info.filename_valid = True
        info.filename_lastname = m.group("lastname")
        info.filename_company = m.group("company")

    try:
        wb = load_workbook(wb_path, data_only=True)
    except Exception as e:
        info.error = f"could not open workbook: {e}"
        return info

    info.found = True
    info.sheets = list(wb.sheetnames)
    info.named_ranges = _gather_named_ranges(wb)
    info.error_cells = _scan_for_errors(wb)
    info.error_cells_on_ratios = sum(
        1 for sn, _, _ in info.error_cells if "ratio" in sn.lower()
    )

    # Count numeric cells on the Ratios tab (used as denominator for the
    # "ratios resolve cleanly" criterion).
    ratios_sheet = next(
        (sn for sn in wb.sheetnames if "ratio" in sn.lower()), None
    )
    if ratios_sheet:
        ws = wb[ratios_sheet]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, (int, float)) and v != 0:
                    info.ratios_tab_numeric_cells += 1

    # Balance-sheet identity checks
    a_curr = _to_number(info.named_ranges.get("BAL_assets_total_curr"))
    le_curr = (_to_number(info.named_ranges.get("BAL_liabilities_total_curr")) or 0) + \
              (_to_number(info.named_ranges.get("BAL_equity_shareholders_curr")) or 0)
    if a_curr is not None and le_curr is not None:
        info.bs_diff_curr = a_curr - le_curr
        info.bs_balances_curr = abs(info.bs_diff_curr) < max(1.0, abs(a_curr) * 0.001)
    a_prior = _to_number(info.named_ranges.get("BAL_assets_total_prior"))
    le_prior = (_to_number(info.named_ranges.get("BAL_liabilities_total_prior")) or 0) + \
               (_to_number(info.named_ranges.get("BAL_equity_shareholders_prior")) or 0)
    if a_prior is not None and le_prior is not None:
        info.bs_diff_prior = a_prior - le_prior
        info.bs_balances_prior = abs(info.bs_diff_prior) < max(1.0, abs(a_prior) * 0.001)

    # Completeness tallies
    info.bal_curr_filled = sum(
        1 for n in REQUIRED_BAL_CURR if _is_filled(info.named_ranges.get(n))
    )
    info.bal_prior_filled = sum(
        1 for n in REQUIRED_BAL_PRIOR if _is_filled(info.named_ranges.get(n))
    )
    info.inc_filled = sum(
        1 for n in REQUIRED_INC if _is_filled(info.named_ranges.get(n))
    )
    info.cash_filled = sum(
        1 for n in REQUIRED_CASH if _is_filled(info.named_ranges.get(n))
    )
    info.assumptions_filled = sum(
        1 for n in REQUIRED_ASSUMPTIONS if _is_filled(info.named_ranges.get(n))
    )

    # Source documentation
    filled, evidence, company = _scan_metadata_tab(wb)
    info.doc_fields_filled = filled
    info.doc_evidence = evidence
    info.company_named = company
    return info


# ------------------------------------------------------------------
# Optional repo inspection
# ------------------------------------------------------------------

def _gh(*args: str) -> str:
    try:
        proc = subprocess.run(
            ["gh", *args], check=False, capture_output=True, text=True,
            encoding="utf-8", errors="replace",
        )
    except FileNotFoundError:
        return ""
    if proc.returncode != 0:
        return ""
    return proc.stdout or ""


@dataclass
class RepoInspection:
    queried: bool = False
    accessible: bool = False
    private: bool = False
    default_branch: str = "main"
    file_in_repo: bool = False
    file_repo_path: str = ""
    instructor_is_collaborator: bool = False
    error: str = ""


def inspect_repo(owner: str, repo: str, expected_filename: str) -> RepoInspection:
    info = RepoInspection(queried=True)
    if not owner or not repo:
        info.error = "no repo URL known (lookup from prior stages failed)"
        return info
    meta_raw = _gh("api", f"repos/{owner}/{repo}")
    if not meta_raw:
        info.error = "repo not accessible"
        return info
    try:
        meta = json.loads(meta_raw)
    except json.JSONDecodeError:
        info.error = "could not parse repo metadata"
        return info
    info.accessible = True
    info.private = bool(meta.get("private"))
    info.default_branch = meta.get("default_branch") or "main"

    tree_raw = _gh(
        "api", f"repos/{owner}/{repo}/git/trees/{info.default_branch}",
        "-X", "GET", "-f", "recursive=1",
    )
    if tree_raw:
        try:
            tree = json.loads(tree_raw).get("tree", [])
        except json.JSONDecodeError:
            tree = []
        for e in tree:
            if e.get("type") != "blob":
                continue
            path = e.get("path", "")
            if path.lower().endswith(expected_filename.lower()):
                info.file_in_repo = True
                info.file_repo_path = path
                break

    perm_raw = _gh(
        "api",
        f"repos/{owner}/{repo}/collaborators/{INSTRUCTOR_GITHUB_HANDLE}/permission",
    )
    if perm_raw:
        try:
            perm = json.loads(perm_raw).get("permission", "")
        except json.JSONDecodeError:
            perm = ""
        info.instructor_is_collaborator = perm in {"admin", "write", "maintain"}
    return info


# ------------------------------------------------------------------
# Scoring
# ------------------------------------------------------------------

@dataclass
class Grade:
    submission: Submission
    wb: WorkbookInspection
    repo: RepoInspection
    prior: PriorGrade | None

    score_accuracy: int = 0       # /40
    score_completeness: int = 0   # /25
    score_sources: int = 0        # /20
    score_ratios: int = 0         # /15

    flags: list[str] = field(default_factory=list)

    @property
    def raw_total(self) -> int:
        return (
            self.score_accuracy + self.score_completeness
            + self.score_sources + self.score_ratios
        )


def score(sub: Submission, wb: WorkbookInspection, repo: RepoInspection,
          prior: PriorGrade | None) -> Grade:
    g = Grade(submission=sub, wb=wb, repo=repo, prior=prior)
    if not wb.found:
        g.flags.append("WORKBOOK_NOT_SUBMITTED")
        return g

    # ----- Criterion 1: Data accuracy (40 pts) -----
    # 20 pts: BS balances both years (10 each)
    # 10 pts: zero workbook-wide error cells outside the Ratios tab
    # 10 pts: sanity checks (Total Assets > 0, INC_sales > 0, signs ok)
    acc = 0
    if wb.bs_balances_curr:
        acc += 10
    else:
        g.flags.append("BS_UNBALANCED_CURR")
    if wb.bs_balances_prior:
        acc += 10
    else:
        if _to_number(wb.named_ranges.get("BAL_assets_total_prior")) is None:
            g.flags.append("BS_PRIOR_BLANK")
        else:
            g.flags.append("BS_UNBALANCED_PRIOR")

    # No error cells outside the Ratios tab (Ratios-tab errors fold into criterion 4)
    non_ratios_errors = [
        (s, c, v) for s, c, v in wb.error_cells if "ratio" not in s.lower()
    ]
    if not non_ratios_errors:
        acc += 10
    elif len(non_ratios_errors) <= 2:
        acc += 6
        g.flags.append("ERROR_CELLS_FEW")
    else:
        acc += 2
        g.flags.append("ERROR_CELLS_MANY")

    # Sanity: revenue positive, total assets positive, signs plausible
    sanity = 0
    sales = _to_number(wb.named_ranges.get("INC_sales"))
    ta = _to_number(wb.named_ranges.get("BAL_assets_total_curr"))
    inv = _to_number(wb.named_ranges.get("BAL_inventories_curr"))
    cash = _to_number(wb.named_ranges.get("BAL_cash_marketable_securities_curr"))
    if sales is not None and sales > 0:
        sanity += 4
    elif sales is None:
        g.flags.append("REVENUE_MISSING")
    else:
        g.flags.append("REVENUE_NONPOSITIVE")
    if ta is not None and ta > 0:
        sanity += 3
    else:
        g.flags.append("TOTAL_ASSETS_MISSING")
    # Inventory and cash should be >= 0
    if (inv is None or inv >= 0) and (cash is None or cash >= 0):
        sanity += 3
    else:
        g.flags.append("SIGN_ANOMALY")
    acc += sanity
    g.score_accuracy = min(40, acc)

    # ----- Criterion 2: Completeness (25 pts) -----
    # 8 pts: BAL current-year (proportional)
    # 8 pts: BAL prior-year (proportional)
    # 5 pts: INC current-year
    # 2 pts: CASH current-year
    # 2 pts: 4 assumptions filled
    comp = 0
    comp += round(8 * wb.bal_curr_filled / len(REQUIRED_BAL_CURR))
    comp += round(8 * wb.bal_prior_filled / len(REQUIRED_BAL_PRIOR))
    comp += round(5 * wb.inc_filled / len(REQUIRED_INC))
    comp += round(2 * wb.cash_filled / max(1, len(REQUIRED_CASH)))
    comp += round(2 * wb.assumptions_filled / len(REQUIRED_ASSUMPTIONS))
    g.score_completeness = min(25, comp)
    if wb.bal_prior_filled < len(REQUIRED_BAL_PRIOR) * 0.6:
        g.flags.append("PRIOR_YEAR_THIN")
    if wb.inc_filled < len(REQUIRED_INC) * 0.7:
        g.flags.append("INC_THIN")

    # ----- Criterion 3: Source documentation (20 pts) -----
    # 6 fields × ~3.3 pts each; round to nearest int.
    filled_count = len(wb.doc_fields_filled)
    g.score_sources = round(20 * filled_count / len(DOC_FIELDS))
    if filled_count < len(DOC_FIELDS) // 2:
        g.flags.append("DOC_FIELDS_THIN")

    # ----- Criterion 4: Auto-computed ratios resolve cleanly (15 pts) -----
    # 15 if Ratios tab has zero error cells and at least 10 numeric outputs,
    # 12 if Ratios tab has 1–2 errors,
    # 8 if many errors,
    # 0 if the tab is blank or unreadable.
    if wb.ratios_tab_numeric_cells == 0:
        g.score_ratios = 0
        g.flags.append("RATIOS_BLANK")
    elif wb.error_cells_on_ratios == 0:
        if wb.ratios_tab_numeric_cells >= 10:
            g.score_ratios = 15
        else:
            g.score_ratios = 12
            g.flags.append("RATIOS_SPARSE")
    elif wb.error_cells_on_ratios <= 2:
        g.score_ratios = 11
        g.flags.append("RATIOS_FEW_ERRORS")
    else:
        g.score_ratios = 6
        g.flags.append("RATIOS_MANY_ERRORS")

    # Repo presence + collaborator state (informational; do not deduct here
    # beyond the carry-over from Stage 2).
    if repo.queried and repo.accessible:
        if not repo.file_in_repo:
            g.flags.append("FILE_NOT_IN_REPO")
        if not repo.instructor_is_collaborator:
            g.flags.append("INSTRUCTOR_NOT_COLLABORATOR")

    if not wb.filename_valid:
        g.flags.append("FILENAME_NONSTANDARD")

    if not g.flags and g.raw_total >= 95:
        g.flags = ["STRONG"]
    return g


# ------------------------------------------------------------------
# Suggestions
# ------------------------------------------------------------------

def _suggestions_for(g: Grade) -> list[Suggestion]:
    """Auto-generated, kind-worded suggestions for Stage 3.

    Returns a list of `Suggestion` objects tagged by bucket:
      - CORE     — observations about Stage 3's rubric performance
      - BACKWARD — carry-forwards from Stage 0 / 1 / 2 (no points lost;
                   can bump the prior stage's score at the post-deadline
                   revision sweep)
      - FORWARD  — looking ahead to Stage 4
    """
    s: list[Suggestion] = []
    wb = g.wb

    if "WORKBOOK_NOT_SUBMITTED" in g.flags:
        s.append(core(
            "No populated workbook was found in the Lamaku submission. The Stage 3 "
            "deliverable is an `.xlsx` at "
            "`models/builds/YYYY-MM-DD-{lastname}-{company-slug}-financials.xlsx` "
            "with Income Statement, Balance Sheet, and Cash Flow filled in from "
            "the 10-K (or VAS/IFRS equivalent) of the company you selected at Stage 2."
        ))
        return s

    if "FILENAME_NONSTANDARD" in g.flags:
        s.append(core(
            f"The filename `{wb.filename}` doesn't quite match the convention "
            "`YYYY-MM-DD-{lastname}-{company-slug}-financials.xlsx` "
            "(all lowercase, hyphen-separated, ISO date prefix, no extra suffix). "
            "Stage 4/5 tooling indexes the workbook by this name — please rename "
            "in-repo and re-commit."
        ))

    if "BS_UNBALANCED_CURR" in g.flags:
        s.append(core(
            f"**Balance Sheet doesn't balance for the current year.** "
            f"`BAL_assets_total_curr − (BAL_liabilities_total_curr + BAL_equity_shareholders_curr)` "
            f"= {wb.bs_diff_curr:,.0f}. Most common cause: a missing line item "
            "(intangibles, deferred tax, accumulated OCI) or a sign error on one "
            "of the prior-year vs current-year inputs. Trace by category — "
            "current assets / fixed assets / other assets on the left vs current "
            "liabilities / long-term debt / equity on the right — and the gap "
            "usually points at one line."
        ))
    if "BS_UNBALANCED_PRIOR" in g.flags:
        s.append(core(
            f"**Balance Sheet doesn't balance for the prior year.** "
            f"Gap = {wb.bs_diff_prior:,.0f}. Prior-year balances feed every "
            "start-of-year ratio (ROA, asset turnover, inventory turnover); "
            "if these don't tie, the Stage 5 ratio analysis inherits the gap. "
            "Tie the prior-year column to the *prior* fiscal year column in "
            "the 10-K — many filings show 2-year comparatives on the same page."
        ))
    if "BS_PRIOR_BLANK" in g.flags:
        s.append(core(
            "Prior-year balance-sheet column appears blank or sparsely populated. "
            "Every `startYear_*` named range pulls from the BAL_*_prior cells, so "
            "without these, ROA/ROE/asset-turnover ratios cannot compute correctly. "
            "Most 10-Ks present two years side-by-side on each financial statement — "
            "fill in the prior-year column from the same filing."
        ))

    if "ERROR_CELLS_MANY" in g.flags or "ERROR_CELLS_FEW" in g.flags:
        outside_ratios = [
            f"{sn}!{coord} = {val}" for sn, coord, val in wb.error_cells
            if "ratio" not in sn.lower()
        ][:5]
        if outside_ratios:
            s.append(core(
                "Error cells detected outside the Ratios tab — likely caused by "
                "broken formulas or missing inputs. Examples: "
                + "; ".join(outside_ratios)
                + ". Resolve these first — they usually cascade into the Ratios tab."
            ))

    if "REVENUE_MISSING" in g.flags or "REVENUE_NONPOSITIVE" in g.flags:
        s.append(core(
            "`INC_sales` looks empty or non-positive. Double-check that Net Sales / "
            "Revenue is entered on the Income Statement at the line bound to "
            "`INC_sales`. (Right-click the cell → *Define Name* to confirm.)"
        ))
    if "TOTAL_ASSETS_MISSING" in g.flags:
        s.append(core(
            "`BAL_assets_total_curr` looks empty. Total Assets should be at the "
            "bottom of the Assets column on the Balance Sheet — confirm the cell "
            "is named `BAL_assets_total_curr` and contains the sum."
        ))
    if "SIGN_ANOMALY" in g.flags:
        s.append(core(
            "One or more cells where a non-negative value is expected (inventory, "
            "cash) appears to be negative. Cross-check the source 10-K — the most "
            "common cause is entering a contra-account from a different section."
        ))

    if "PRIOR_YEAR_THIN" in g.flags:
        s.append(core(
            f"Only {wb.bal_prior_filled}/{len(REQUIRED_BAL_PRIOR)} prior-year "
            "balance-sheet line items are populated. The 10-K's comparative balance "
            "sheet shows current + prior in the same statement — they're free data "
            "and they unlock start-of-year ratios."
        ))
    if "INC_THIN" in g.flags:
        s.append(core(
            f"Only {wb.inc_filled}/{len(REQUIRED_INC)} income-statement line items "
            "are populated. Each one (Sales, COGS, SG&A, EBIT, Interest, Tax, Net "
            "Income) drives at least one ratio — the rubric expects all of them filled."
        ))

    if "DOC_FIELDS_THIN" in g.flags:
        missing_doc = [
            f for f in DOC_FIELDS if f not in wb.doc_fields_filled
        ]
        s.append(core(
            "Cover / Notes tab needs more company-context detail. Missing fields "
            "detected: " + ", ".join(missing_doc) + ". The Stage 3 spec asks for "
            "source URL, reporting standard (US GAAP / IFRS / VAS), currency and "
            "units, fiscal year-end, and the tax-rate basis. Even a one-line "
            "answer per field counts."
        ))

    if "RATIOS_MANY_ERRORS" in g.flags:
        ratios_errs = [
            f"{coord} = {val}" for sn, coord, val in wb.error_cells
            if "ratio" in sn.lower()
        ][:5]
        s.append(core(
            f"The Ratios tab has {wb.error_cells_on_ratios} error cell(s) "
            f"({', '.join(ratios_errs)}). A `#DIV/0!` usually means a denominator "
            "is empty (often a `startYear_*` value); `#NAME?` means a named range "
            "isn't defined; `#REF!` means a referenced cell was deleted. The "
            "Legend / Cover tab lists every named range — cross-check each errored "
            "row against it."
        ))
    elif "RATIOS_FEW_ERRORS" in g.flags:
        ratios_errs = [
            f"{coord} = {val}" for sn, coord, val in wb.error_cells
            if "ratio" in sn.lower()
        ][:3]
        s.append(core(
            f"The Ratios tab has a small number of error cells "
            f"({', '.join(ratios_errs)}). Each one usually traces back to a single "
            "empty input cell — fixing the input clears the dependent ratio."
        ))
    elif "RATIOS_SPARSE" in g.flags:
        s.append(core(
            "The Ratios tab is populated but sparse — fewer than ~10 ratios "
            "resolved to numeric values. The full template computes 20+ ratios "
            "across Performance, Profitability, Efficiency, Leverage, Liquidity, "
            "and Du Pont — most of them should populate once the statement tabs "
            "are complete."
        ))

    if "FILE_NOT_IN_REPO" in g.flags:
        s.append(core(
            f"The workbook was submitted via Lamaku but doesn't appear at "
            f"`{STAGE_TARGET_DIR}/` in your repo. The Stage 3 spec accepts the "
            "Lamaku fallback, but by Stage 5 the file must live in the repo — "
            "commit it now and you'll save yourself the cleanup later."
        ))

    if "INSTRUCTOR_NOT_COLLABORATOR" in g.flags:
        # This was first flagged at Stage 2 — treat as backward carry-forward.
        s.append(backward(
            f"**Still open from Stage 2:** `@{INSTRUCTOR_GITHUB_HANDLE}` is not a "
            "Write collaborator on your repo, which continues to block the "
            "tracked-feedback workflow (the way an auditor or supervising "
            "analyst would mark up a draft for you to revise). Repo → Settings "
            "→ Collaborators → Add people → **Write**. Not re-deducted here, "
            "but closing this before the deadline can bump your Stage 2 score "
            "at the sweep — and Stage 5's feedback-incorporation rubric line "
            "depends on it."
        ))

    # Generic backward bucket — any other CARRY_OVER_* flags.
    if "CARRY_OVER_DIRS" in g.flags:
        s.append(backward(
            "**Still open from Stage 0:** directory skeleton still incomplete. "
            "Not re-deducted here; closing it before the deadline can bump your "
            "Stage 0 score at the post-deadline revision sweep."
        ))
    if "CARRY_OVER_READMES" in g.flags:
        s.append(backward(
            "**Still open from Stage 0:** placeholder README(s) in directory "
            "subfolders. Not re-deducted here; a sentence of purpose in each "
            "closes the carry-forward and can bump your Stage 0 score at the sweep."
        ))

    core_count = sum(1 for x in s if x.bucket == "core")
    if "STRONG" in g.flags and core_count == 0:
        s.append(core(
            "Strong submission across all four criteria — balance sheet balances "
            "both years, named ranges populated, documentation complete, and "
            "ratios resolve cleanly. Stage 4 (the spec) is largely a writing task "
            "on top of what you've already built — keep the same discipline."
        ))

    fwd = next_stage_pointer(STAGE_N)
    if fwd is not None:
        s.append(fwd)

    return s


def _summary_tagline(g: Grade) -> str:
    flags = set(g.flags)
    if "WORKBOOK_NOT_SUBMITTED" in flags:
        return "No workbook submitted — floor does not apply."
    if "STRONG" in flags:
        return "Strong on merit; see suggestions for refinements."
    if "BS_UNBALANCED_CURR" in flags or "BS_UNBALANCED_PRIOR" in flags:
        return "Balance Sheet doesn't balance — trace by category."
    if "RATIOS_MANY_ERRORS" in flags or "RATIOS_BLANK" in flags:
        return "Ratios tab not resolving cleanly — fix upstream input gaps."
    if "PRIOR_YEAR_THIN" in flags or "BS_PRIOR_BLANK" in flags:
        return "Prior-year column needs more data — unlocks start-of-year ratios."
    if "DOC_FIELDS_THIN" in flags:
        return "Workbook populated; Cover/Notes documentation needs more detail."
    return "See per-student suggestions for refinements."


# ------------------------------------------------------------------
# Report generation
# ------------------------------------------------------------------

def _letter_for(score: int) -> str:
    for letter, lo, hi in LETTER_GRADE_SCALE:
        lo_pts = round(lo * TOTAL_POINTS / 100)
        if score >= lo_pts and (hi is None or score < round(hi * TOTAL_POINTS / 100)):
            return letter
    return "F"


def _final_score(g: Grade, floor_pct: int) -> int:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    if g.wb.found and g.raw_total < floor_value:
        return floor_value
    return g.raw_total


def _floor_was_applied(g: Grade, floor_pct: int) -> bool:
    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    return g.wb.found and g.raw_total < floor_value


def _student_section(n: int, g: Grade, floor_pct: int) -> str:
    sub = g.submission
    wb = g.wb
    submitted = sub.submitted_at.strftime("%Y-%m-%d %I:%M %p") if sub.submitted_at else "—"
    raw = g.raw_total
    floor_applied = _floor_was_applied(g, floor_pct)
    final = _final_score(g, floor_pct)
    letter = _letter_for(final)

    lines: list[str] = []
    suffix = ", floor applied" if floor_applied else ""
    lines.append(f"## {n}. {sub.student_name} — **{final} / 100** ({letter}{suffix})")
    lines.append("")
    if sub.repo_url:
        lines.append(f"**Repo:** {sub.repo_url}")
    lines.append(f"**Workbook:** `{wb.filename}`")
    if wb.company_named:
        lines.append(f"**Company:** {wb.company_named}")
    lines.append(f"**Submitted:** {submitted}")
    lines.append("")
    lines.append("| Criterion | Earned | Notes |")
    lines.append("|-----------|--------|-------|")

    # Accuracy notes
    acc_bits = []
    if wb.bs_balances_curr:
        acc_bits.append("BS balances (curr ✓)")
    else:
        acc_bits.append(f"BS curr off by {wb.bs_diff_curr:,.0f}")
    if wb.bs_balances_prior:
        acc_bits.append("BS balances (prior ✓)")
    elif "BS_PRIOR_BLANK" in g.flags:
        acc_bits.append("BS prior blank")
    else:
        acc_bits.append(f"BS prior off by {wb.bs_diff_prior:,.0f}")
    non_ratios_errs = [
        e for e in wb.error_cells if "ratio" not in e[0].lower()
    ]
    if non_ratios_errs:
        acc_bits.append(f"{len(non_ratios_errs)} error cell(s) off-ratios")
    else:
        acc_bits.append("no error cells outside Ratios")
    lines.append(f"| Data accuracy | {g.score_accuracy} / 40 | {'; '.join(acc_bits)}. |")

    # Completeness notes
    comp_note = (
        f"BAL curr {wb.bal_curr_filled}/{len(REQUIRED_BAL_CURR)}; "
        f"BAL prior {wb.bal_prior_filled}/{len(REQUIRED_BAL_PRIOR)}; "
        f"INC {wb.inc_filled}/{len(REQUIRED_INC)}; "
        f"CASH {wb.cash_filled}/{len(REQUIRED_CASH)}; "
        f"assumptions {wb.assumptions_filled}/{len(REQUIRED_ASSUMPTIONS)}."
    )
    lines.append(f"| Completeness | {g.score_completeness} / 25 | {comp_note} |")

    # Source documentation notes
    if wb.doc_fields_filled:
        src_note = f"{len(wb.doc_fields_filled)}/{len(DOC_FIELDS)} doc fields detected on Cover/Notes."
    else:
        src_note = "Cover/Notes tab metadata not detected."
    lines.append(f"| Source documentation | {g.score_sources} / 20 | {src_note} |")

    # Ratios notes
    if wb.ratios_tab_numeric_cells == 0:
        rat_note = "Ratios tab appears blank or formulas not computed."
    elif wb.error_cells_on_ratios == 0:
        rat_note = f"{wb.ratios_tab_numeric_cells} ratio outputs resolve to numbers; no errors."
    else:
        rat_note = (
            f"{wb.ratios_tab_numeric_cells} ratio output(s) resolve; "
            f"{wb.error_cells_on_ratios} error cell(s) on the Ratios tab."
        )
    lines.append(f"| Auto-computed ratios | {g.score_ratios} / 15 | {rat_note} |")

    if floor_applied:
        floor_value = round(TOTAL_POINTS * floor_pct / 100)
        lines.append(f"| **Raw total** | **{raw} / 100** | |")
        lines.append(
            f"| **Floor adjustment** | **+{floor_value - raw}** | "
            f"Working public repo present — floor of {floor_pct} applied. |"
        )
        lines.append(f"| **Final** | **{final} / 100** | |")
    else:
        merit_note = " Above the floor — earned on merit." if raw >= round(
            TOTAL_POINTS * floor_pct / 100
        ) else ""
        lines.append(f"| **Total** | **{final} / 100** |{merit_note} |")
    lines.append("")

    suggestions = _suggestions_for(g)
    lines.extend(render_suggestions(suggestions, stage_n=STAGE_N))
    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def _build_full_report(grades: list[Grade], floor_pct: int, today: datetime) -> str:
    lines = [
        "# BUS-629 Stage 3 — Grade Report",
        "",
        f"**Stage:** {STAGE_LABEL} (20% of project score)",
        f"**Graded:** {today.strftime('%Y-%m-%d')}",
        f"**Submissions reviewed:** {len(grades)}",
        f"**Floor policy:** {floor_pct}% floor for submissions with a populated workbook.",
        "**Double-deduction policy:** Do not re-deduct for issues already flagged "
        "in Stage 0/1/2. Useful suggestions may be repeated as forward-looking "
        "tips without point loss.",
        "",
        "---",
        "",
        "## Rubric (recap)",
        "",
        "| Criterion | Weight |",
        "|-----------|--------|",
        "| Data accuracy (ties to source 10-K) | 40% |",
        "| Completeness (both years populated) | 25% |",
        "| Source documentation | 20% |",
        "| Auto-computed ratios resolve cleanly | 15% |",
        "",
        "---",
        "",
    ]
    for i, g in enumerate(grades, 1):
        lines.append(_student_section(i, g, floor_pct))
    lines.append("## Class summary")
    lines.append("")
    lines.append("| Student | Score | Notes |")
    lines.append("|---------|-------|-------|")
    for g in grades:
        lines.append(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |"
        )
    lines.append("")
    finals = [_final_score(g, floor_pct) for g in grades if g.wb.found]
    submitted_n = len(finals)
    mean = sum(finals) / submitted_n if submitted_n else 0.0
    floors = sum(1 for g in grades if _floor_was_applied(g, floor_pct))
    lines.append(f"**Mean (submissions only):** {mean:.1f}")
    lines.append(f"**Submission rate:** {submitted_n} of {len(grades)}")
    lines.append(f"**Floor applied:** {floors} of {submitted_n} submissions")
    lines.append("")
    return "\n".join(lines)


def write_or_update_grade_report(
    grades: list[Grade], floor_pct: int, report_path: Path,
    today: datetime | None = None,
) -> list[Grade]:
    today = today or datetime.now()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    if not report_path.exists():
        report_path.write_text(_build_full_report(grades, floor_pct, today),
                               encoding="utf-8")
        return list(grades)

    text = report_path.read_text(encoding="utf-8")
    existing_urls = {
        u.lower().rstrip("/")
        for u in re.findall(r"https?://github\.com/[A-Za-z0-9_./-]+", text)
    }
    existing_names = {
        _normalize_name(n)
        for n in re.findall(r"^## \d+\. ([^—\n]+?) — \*\*\d+ / 100\*\*",
                            text, re.MULTILINE)
    }
    new_grades = []
    for g in grades:
        url_key = g.submission.repo_url.lower().rstrip("/")
        name_key = _normalize_name(g.submission.student_name)
        if url_key and url_key in existing_urls:
            continue
        if name_key and name_key in existing_names:
            continue
        new_grades.append(g)
    if not new_grades:
        return []

    nums = [int(n) for n in re.findall(r"^## (\d+)\.", text, re.MULTILINE)]
    next_n = (max(nums) + 1) if nums else 1
    chunks = []
    for g in new_grades:
        chunks.append(_student_section(next_n, g, floor_pct))
        next_n += 1
    new_block = "\n".join(chunks)

    if "## Class summary" in text:
        text = text.replace(
            "## Class summary", new_block + "\n## Class summary", 1
        )
    else:
        text = text.rstrip() + "\n\n" + new_block

    table_re = re.compile(
        r"(## Class summary\s*\n+"
        r"\| Student \| Score \| Notes \|\n"
        r"\| ?-+ ?\| ?-+ ?\| ?-+ ?\|\n"
        r"(?:\|[^\n]*\|\n)+)"
    )
    m = table_re.search(text)
    if m:
        new_rows = "".join(
            f"| {g.submission.student_name} | "
            f"{_final_score(g, floor_pct)} / 100 | {_summary_tagline(g)} |\n"
            for g in new_grades
        )
        text = text[: m.end()] + new_rows + text[m.end():]

    student_re = re.compile(
        r"^## \d+\.[^—\n]*— \*\*(\d+) / 100\*\*([^\n]*)", re.MULTILINE
    )
    submitted_finals = []
    total_students = 0
    for sm in student_re.finditer(text):
        total_students += 1
        if "no workbook" in sm.group(2).lower() or "no submission" in sm.group(2).lower():
            continue
        submitted_finals.append(int(sm.group(1)))
    floor_count = sum(
        1 for fs in submitted_finals if fs == round(TOTAL_POINTS * floor_pct / 100)
    )
    mean = sum(submitted_finals) / len(submitted_finals) if submitted_finals else 0.0

    text = re.sub(
        r"\*\*Mean \(submissions only\):\*\* [^\n]*",
        f"**Mean (submissions only):** {mean:.1f}", text, count=1,
    )
    text = re.sub(
        r"\*\*Submission rate:\*\* \d+ of \d+(?:\s*\([^)]*\))?",
        f"**Submission rate:** {len(submitted_finals)} of {total_students}",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Floor applied:\*\* \d+ of \d+ submissions(?:\s*\([^)]*\))?",
        f"**Floor applied:** {floor_count} of {len(submitted_finals)} submissions",
        text, count=1,
    )
    text = re.sub(
        r"\*\*Submissions reviewed:\*\* [^\n]*",
        f"**Submissions reviewed:** {total_students}",
        text, count=1,
    )
    names_added = ", ".join(g.submission.student_name for g in new_grades)
    text = re.sub(
        r"\*\*Graded:\*\* [^\n]*",
        f"**Graded:** {today.strftime('%Y-%m-%d')} ({names_added} added)",
        text, count=1,
    )

    report_path.write_text(text, encoding="utf-8")
    return new_grades


# ------------------------------------------------------------------
# Worksheet
# ------------------------------------------------------------------

def build_worksheet(grades: list[Grade], floor_pct: int, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading"
    headers = [
        "Student", "Submitted", "Repo URL", "Workbook", "Company",
        "Accuracy /40", "Completeness /25", "Sources /20", "Ratios /15",
        "Raw /100", "Final /100", "Floored /100",
        "BS curr ✓", "BS prior ✓", "BAL curr filled", "BAL prior filled",
        "INC filled", "CASH filled", "Doc fields", "Ratios numeric", "Ratios errors",
        "Flags", "Comments",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="024731")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

    floor_value = round(TOTAL_POINTS * floor_pct / 100)
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F8CBAD")
    floor_fill = PatternFill("solid", fgColor="E2EFDA")
    final_col = headers.index("Final /100") + 1
    floor_col = headers.index("Floored /100") + 1
    flags_col = headers.index("Flags") + 1

    for g in grades:
        sub = g.submission
        wbi = g.wb
        submitted = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else ""
        comments = g.repo.error or g.wb.error
        row = [
            sub.student_name, submitted, sub.repo_url, wbi.filename, wbi.company_named,
            g.score_accuracy, g.score_completeness, g.score_sources, g.score_ratios,
            g.raw_total, g.raw_total, None,
            "Y" if wbi.bs_balances_curr else "N",
            "Y" if wbi.bs_balances_prior else "N",
            wbi.bal_curr_filled, wbi.bal_prior_filled, wbi.inc_filled, wbi.cash_filled,
            len(wbi.doc_fields_filled), wbi.ratios_tab_numeric_cells,
            wbi.error_cells_on_ratios,
            ", ".join(g.flags), comments,
        ]
        ws.append(row)
        r = ws.max_row
        final_ref = f"{get_column_letter(final_col)}{r}"
        ws.cell(
            row=r, column=floor_col,
            value=f'=IF({final_ref}=0,0,MAX({final_ref},{floor_value}))',
        ).fill = floor_fill
        if comments and not wbi.found:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = error_fill
        elif g.flags and g.flags != ["STRONG"]:
            ws.cell(row=r, column=flags_col).fill = flag_fill

    widths = [24, 17, 50, 40, 22, 12, 14, 12, 11, 10, 11, 13,
              10, 10, 13, 13, 11, 11, 11, 13, 12, 38, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(output_path)


# ------------------------------------------------------------------
# Sweep entry point — rescore a student against current repo state
# ------------------------------------------------------------------

GITHUB_URL_RE = re.compile(
    r"https?://github\.com/(?P<owner>[A-Za-z0-9_.-]+)/(?P<repo>[A-Za-z0-9_.-]+)"
)


def _download_blob(owner: str, repo: str, path: str, branch: str, dest: Path) -> bool:
    """Fetch a binary blob from a repo to `dest`. Returns True on success."""
    import base64
    raw = _gh(
        "api", f"repos/{owner}/{repo}/contents/{path}",
        "-X", "GET", "-f", f"ref={branch}",
    )
    if not raw:
        return False
    try:
        meta = json.loads(raw)
    except json.JSONDecodeError:
        return False
    content = meta.get("content")
    if not content:
        return False
    try:
        dest.write_bytes(base64.b64decode(content))
    except Exception:
        return False
    return True


def rescore_from_repo(
    student_name: str,
    repo_url: str,
    submitted_at: datetime | None = None,
    prior: "PriorGrade | None" = None,
    student_id: str = "",
) -> Grade | None:
    """Sweep entry point: rescore Stage 3 against current repo state.

    Scans `models/builds/` for any `*-financials.xlsx`. If found, downloads
    and inspects it. If multiple exist, takes the lexicographically last
    match (latest ISO-date prefix wins).
    """
    m = GITHUB_URL_RE.search(repo_url)
    if not m:
        return None
    owner, repo = m.group("owner"), m.group("repo")
    repo_info = inspect_repo(owner, repo, "financials.xlsx")

    import tempfile
    tmp = Path(tempfile.mkdtemp(prefix=f"sweep_s3_{owner}_"))
    wb_dest = tmp / "financials.xlsx"
    if repo_info.file_in_repo and repo_info.file_repo_path and repo_info.accessible:
        _download_blob(
            owner, repo, repo_info.file_repo_path, repo_info.default_branch, wb_dest
        )

    wb_info = inspect_workbook(wb_dest)
    if repo_info.file_repo_path:
        wb_info.filename = Path(repo_info.file_repo_path).name

    sub = Submission(
        student_id=student_id,
        student_name=student_name,
        submitted_at=submitted_at,
        workbook_path=wb_dest,
        repo_url=repo_url,
        owner=owner,
        repo=repo,
    )
    return score(sub, wb_info, repo_info, prior)


# ------------------------------------------------------------------
# Workflow helper
# ------------------------------------------------------------------

def _detect_workflow(export_path: Path):
    if not (export_path.is_file() and export_path.suffix.lower() == ".zip"):
        return None
    parent = export_path.parent
    if parent.name.lower() != "ungraded":
        return None
    graded = parent.parent / "graded"
    scratch = parent / f"_{export_path.stem}_extracted"
    return graded, scratch


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("export", type=Path)
    p.add_argument("--floor", type=int, default=DEFAULT_FLOOR_PCT)
    p.add_argument("--prior-stage0", type=Path, default=None)
    p.add_argument("--prior-stage1", type=Path, default=None)
    p.add_argument("--prior-stage2", type=Path, default=None)
    p.add_argument("--out", type=Path, default=None)
    p.add_argument("--no-move", action="store_true")
    args = p.parse_args(argv)

    subs = discover_submissions(args.export)
    if not subs:
        print("No submissions discovered.")
        return 1
    print(f"Found {len(subs)} submission(s).")

    base_root = args.export.resolve().parents[2]
    s0 = args.prior_stage0 or (base_root / "stage0" / "graded" / "STAGE0_GRADES.md")
    s1 = args.prior_stage1 or (base_root / "stage1" / "graded" / "STAGE1_GRADES.md")
    s2 = args.prior_stage2 or (base_root / "stage2" / "graded" / "STAGE2_GRADES.md")
    prior0 = parse_prior_report(s0) if s0.exists() else {}
    prior1 = parse_prior_report(s1) if s1.exists() else {}
    prior2 = parse_prior_report(s2) if s2.exists() else {}
    if prior0:
        print(f"Loaded Stage 0 records: {len(prior0)} from {s0}")
    if prior1:
        print(f"Loaded Stage 1 records: {len(prior1)} from {s1}")
    if prior2:
        print(f"Loaded Stage 2 records: {len(prior2)} from {s2}")
    prior = merge_prior(prior0, prior1, prior2)

    for s in subs:
        pg = lookup_prior(prior, s.student_name)
        if pg and pg.repo_url:
            m = GITHUB_URL_RE.search(pg.repo_url)
            if m:
                s.repo_url = pg.repo_url
                s.owner = m.group("owner")
                s.repo = m.group("repo")

    grades: list[Grade] = []
    for s in subs:
        print(f"  inspecting {s.student_name} ...", end=" ")
        wbi = inspect_workbook(s.workbook_path)
        if s.owner:
            repo = inspect_repo(s.owner, s.repo, wbi.filename)
        else:
            repo = RepoInspection()
        pg = lookup_prior(prior, s.student_name)
        g = score(s, wbi, repo, pg)
        grades.append(g)
        if wbi.error:
            print(f"ERROR: {wbi.error}")
        else:
            print(
                f"raw={g.raw_total}/100 acc={g.score_accuracy} comp={g.score_completeness} "
                f"src={g.score_sources} rat={g.score_ratios} "
                f"BS=({'Y' if wbi.bs_balances_curr else 'N'}/"
                f"{'Y' if wbi.bs_balances_prior else 'N'}) "
                f"flags={','.join(g.flags) or '-'}"
            )

    workflow = _detect_workflow(args.export)
    if args.out is not None:
        worksheet_path = args.out
    elif workflow is not None:
        graded_dir, _ = workflow
        graded_dir.mkdir(parents=True, exist_ok=True)
        ws_dir = graded_dir / "_worksheets"
        ws_dir.mkdir(exist_ok=True)
        worksheet_path = ws_dir / f"stage3-{args.export.stem}.xlsx"
    else:
        worksheet_path = (
            (args.export.parent if args.export.is_file() else args.export)
            / "_grading" / "stage3-grading-worksheet.xlsx"
        )

    build_worksheet(grades, args.floor, worksheet_path)
    print(f"\nWrote worksheet: {worksheet_path}")

    today = datetime.now()
    if workflow is not None:
        graded_dir, scratch = workflow
        report_path = graded_dir / "STAGE3_GRADES.md"
        new_entries = write_or_update_grade_report(grades, args.floor, report_path, today)
        if new_entries:
            names = ", ".join(g.submission.student_name for g in new_entries)
            print(f"Updated report: {report_path} (+{len(new_entries)} new: {names})")
        else:
            print(f"Report already up to date: {report_path}")

        if not args.no_move:
            dest = graded_dir / args.export.name
            if dest.exists():
                print(f"Source zip already exists at {dest} -- leaving original in place.")
            else:
                shutil.move(str(args.export), str(dest))
                print(f"Moved source zip -> {dest}")
            if scratch.exists():
                shutil.rmtree(scratch, ignore_errors=True)
                print(f"Cleaned up scratch extract at {scratch}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
