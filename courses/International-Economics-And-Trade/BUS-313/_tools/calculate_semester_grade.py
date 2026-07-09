"""
Semester grade calculator template for BUS-313 (D2L "GradesExport" workbooks).

Builds a "Semester Grades" sheet with live formulas (no hardcoded values) on
top of the raw "Grades" export sheet. Reusable across semesters: assignment
names/counts can change (e.g. chapter list, EC items) and the header-detection
logic below will re-discover them by pattern rather than fixed column letters.

Usage:
    python calculate_semester_grade.py <path_to_GradesExport.xlsx> [--drop USERNAME ...]

Grading logic (confirmed with instructor, 2026-07 BUS-313-601 SU26):
  - Each chapter homework: MAX(regular score, makeup score) -- makeup is a
    lower-max-point alternate for a missed regular assignment.
  - "Professional Profile" one-pager assignment is informational only and is
    NOT included in the semester point total.
  - Points Earned (ex EC) = HW total + Midterm + Final + Attendance + Team Project.
  - Denominator is fixed at 100 (per course syllabus weights: Attendance 10% +
    Homework 20% + Group Project 20% + Midterm 25% + Final 25%).
  - Extra credit (Team Project EC) is added on top and capped so the total
    never exceeds 100.
  - Final grade = ROUNDUP to nearest whole number.
  - Letter grade formula (instructor-specified, use verbatim for future terms):
    =IFS(X>=97,"A+",X>=93,"A",X>=90,"A-",X>=87,"B+",X>=83,"B",X>=80,"B-",
         X>=77,"C+",X>=73,"C",X>=70,"C-",X>=67,"D+",X>=65,"D",TRUE,"F")

To reuse for a future semester's export: update SKIP_FROM_TOTAL_MATCH below if
the non-counted assignment's name changes, then run this script against the
new export file.
"""
import argparse
import re
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment

SKIP_FROM_TOTAL_MATCH = 'Professional Profile'  # header substring excluded from point total

LETTER_FORMULA_TEMPLATE = (
    '=IFS({fg}>=97,"A+",{fg}>=93,"A",{fg}>=90,"A-",{fg}>=87,"B+",{fg}>=83,"B",'
    '{fg}>=80,"B-",{fg}>=77,"C+",{fg}>=73,"C",{fg}>=70,"C-",{fg}>=67,"D+",'
    '{fg}>=65,"D",TRUE,"F")'
)


def key_of(h):
    h = re.sub(r'\s*Points Grade.*$', '', h)
    h = re.sub(r'\s*\(Makeup\)\s*$', '', h)
    h = re.sub(r'[^a-z0-9]+', ' ', h.lower()).strip()
    return h


def short_ch_label(h):
    h = re.sub(r'\s*Points Grade.*$', '', h)
    m = re.match(r'Ch\.?\s*(\d+)', h) or re.match(r'Chapter\s*(\d+)', h)
    return f'Ch{m.group(1)}' if m else h[:8]


def build(path, drop_usernames=None):
    drop_usernames = set(drop_usernames or [])
    wb = load_workbook(path)
    ws = wb['Grades']
    last_row = ws.max_row
    headers = {c: ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}

    orgid_col = username_col = last_col = first_col = None
    assignment_skip_col = midterm_col = final_col = attendance_col = team_col = teamec_col = None
    reg_chapters, mk_chapters, reg_order = {}, {}, []

    for c in range(1, ws.max_column + 1):
        h = headers[c]
        if h == 'OrgDefinedId':
            orgid_col = c
        elif h == 'Username':
            username_col = c
        elif h == 'Last Name':
            last_col = c
        elif h == 'First Name':
            first_col = c
        if not h or 'Points Grade' not in str(h):
            continue
        if SKIP_FROM_TOTAL_MATCH in h:
            assignment_skip_col = c
        elif 'Midterm' in h:
            midterm_col = c
        elif 'Final' in h:
            final_col = c
        elif 'Attendance' in h:
            attendance_col = c
        elif 'Team Project (EC)' in h:
            teamec_col = c
        elif 'Team Project' in h:
            team_col = c
        elif '(Makeup)' in h:
            mk_chapters[key_of(h)] = c
        else:
            reg_chapters[key_of(h)] = c
            reg_order.append((key_of(h), c, h))

    missing = [k for k in reg_chapters if k not in mk_chapters]
    if missing:
        print(f'WARNING: no makeup pair found for: {missing}', file=sys.stderr)

    if 'Semester Grades' in wb.sheetnames:
        del wb['Semester Grades']
    out = wb.create_sheet('Semester Grades')

    bold = Font(name='Arial', bold=True, size=10)
    normal = Font(name='Arial', size=10)
    header_fill = PatternFill('solid', start_color='024731', end_color='024731')
    header_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    c_orgid, c_username, c_last, c_first = 1, 2, 3, 4
    c_assign_skip = 5
    first_ch_col = 6
    n_chapters = len(reg_order)
    c_hw_total = first_ch_col + n_chapters
    c_midterm = c_hw_total + 1
    c_final = c_midterm + 1
    c_attendance = c_final + 1
    c_team = c_attendance + 1
    c_points_ex_ec = c_team + 1
    c_teamec = c_points_ex_ec + 1
    c_total_incl_ec = c_teamec + 1
    c_final_pct = c_total_incl_ec + 1
    c_letter = c_final_pct + 1

    hdr_row = 1
    out.cell(row=hdr_row, column=c_orgid, value='OrgDefinedId')
    out.cell(row=hdr_row, column=c_username, value='Username')
    out.cell(row=hdr_row, column=c_last, value='Last Name')
    out.cell(row=hdr_row, column=c_first, value='First Name')
    out.cell(row=hdr_row, column=c_assign_skip, value='Assignment 1 (info only, not counted)')
    for i, (k, regc, h) in enumerate(reg_order):
        out.cell(row=hdr_row, column=first_ch_col + i, value=f'{short_ch_label(h)} Best (1.25)')
    hw_max = round(n_chapters * 1.25, 2)
    out.cell(row=hdr_row, column=c_hw_total, value=f'HW Total ({hw_max})')
    out.cell(row=hdr_row, column=c_midterm, value='Midterm (25)')
    out.cell(row=hdr_row, column=c_final, value='Final (25)')
    out.cell(row=hdr_row, column=c_attendance, value='Attendance (10)')
    out.cell(row=hdr_row, column=c_team, value='Team Project (15)')
    out.cell(row=hdr_row, column=c_points_ex_ec, value='Points Earned (ex EC)')
    out.cell(row=hdr_row, column=c_teamec, value='Team Project EC (3)')
    out.cell(row=hdr_row, column=c_total_incl_ec, value='Total incl. EC (capped /100)')
    out.cell(row=hdr_row, column=c_final_pct, value='Final Grade (roundup)')
    out.cell(row=hdr_row, column=c_letter, value='Letter Grade')

    for c in range(1, c_letter + 1):
        cell = out.cell(row=hdr_row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    out_row = 2
    for src_row in range(2, last_row + 1):
        username_val = ws.cell(row=src_row, column=username_col).value
        if username_val in drop_usernames:
            continue

        out.cell(row=out_row, column=c_orgid, value=f"='Grades'!{get_column_letter(orgid_col)}{src_row}")
        out.cell(row=out_row, column=c_username, value=f"='Grades'!{get_column_letter(username_col)}{src_row}")
        out.cell(row=out_row, column=c_last, value=f"='Grades'!{get_column_letter(last_col)}{src_row}")
        out.cell(row=out_row, column=c_first, value=f"='Grades'!{get_column_letter(first_col)}{src_row}")
        out.cell(row=out_row, column=c_assign_skip, value=f"='Grades'!{get_column_letter(assignment_skip_col)}{src_row}")

        ch_cells = []
        for j, (k, regc, h) in enumerate(reg_order):
            mkc = mk_chapters[k]
            col_idx = first_ch_col + j
            formula = f"=MAX('Grades'!{get_column_letter(regc)}{src_row},'Grades'!{get_column_letter(mkc)}{src_row})"
            out.cell(row=out_row, column=col_idx, value=formula)
            ch_cells.append(get_column_letter(col_idx) + str(out_row))

        out.cell(row=out_row, column=c_hw_total, value=f"=SUM({ch_cells[0]}:{ch_cells[-1]})")
        out.cell(row=out_row, column=c_midterm, value=f"='Grades'!{get_column_letter(midterm_col)}{src_row}")
        out.cell(row=out_row, column=c_final, value=f"='Grades'!{get_column_letter(final_col)}{src_row}")
        out.cell(row=out_row, column=c_attendance, value=f"='Grades'!{get_column_letter(attendance_col)}{src_row}")
        out.cell(row=out_row, column=c_team, value=f"='Grades'!{get_column_letter(team_col)}{src_row}")

        hw = get_column_letter(c_hw_total) + str(out_row)
        mt = get_column_letter(c_midterm) + str(out_row)
        fn = get_column_letter(c_final) + str(out_row)
        at = get_column_letter(c_attendance) + str(out_row)
        tm = get_column_letter(c_team) + str(out_row)
        out.cell(row=out_row, column=c_points_ex_ec, value=f"=SUM({hw},{mt},{fn},{at},{tm})")

        out.cell(row=out_row, column=c_teamec, value=f"='Grades'!{get_column_letter(teamec_col)}{src_row}")

        pe = get_column_letter(c_points_ex_ec) + str(out_row)
        ec = get_column_letter(c_teamec) + str(out_row)
        out.cell(row=out_row, column=c_total_incl_ec, value=f"=MIN(100,{pe}+{ec})")

        ti = get_column_letter(c_total_incl_ec) + str(out_row)
        out.cell(row=out_row, column=c_final_pct, value=f"=ROUNDUP({ti},0)")

        fg = get_column_letter(c_final_pct) + str(out_row)
        out.cell(row=out_row, column=c_letter, value=LETTER_FORMULA_TEMPLATE.format(fg=fg))

        for c in range(1, c_letter + 1):
            cell = out.cell(row=out_row, column=c)
            cell.font = bold if c == c_final_pct else normal
            cell.border = border

        out_row += 1

    out.column_dimensions[get_column_letter(c_orgid)].width = 12
    out.column_dimensions[get_column_letter(c_username)].width = 12
    out.column_dimensions[get_column_letter(c_last)].width = 12
    out.column_dimensions[get_column_letter(c_first)].width = 12
    out.column_dimensions[get_column_letter(c_assign_skip)].width = 14
    for j in range(n_chapters):
        out.column_dimensions[get_column_letter(first_ch_col + j)].width = 10
    for c in [c_hw_total, c_midterm, c_final, c_attendance, c_team, c_points_ex_ec,
              c_teamec, c_total_incl_ec, c_final_pct, c_letter]:
        out.column_dimensions[get_column_letter(c)].width = 14

    out.freeze_panes = get_column_letter(c_assign_skip + 1) + '2'

    out.cell(row=hdr_row, column=c_final_pct).comment = Comment(
        "Denominator fixed at 100 per course syllabus weights (Attendance 10 + HW 20 + "
        "Group Project 20 + Midterm 25 + Final 25 = 100). Assignment 1 (Professional Profile) "
        "is informational only and excluded from the point total per instructor direction. "
        "Team Project EC is the only extra credit; Total incl. EC is capped at 100. "
        "Letter grade formula and this workbook structure are the standing template for "
        "future-semester grade calculations -- see script docstring for the exact formula.",
        "Claude"
    )

    wb.save(path)
    return out_row - 2, n_chapters


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('path', help='Path to the GradesExport .xlsx (edited in place)')
    parser.add_argument('--drop', nargs='*', default=[], help='Usernames to exclude (e.g. withdrawn students)')
    args = parser.parse_args()
    n_students, n_chapters = build(args.path, drop_usernames=args.drop)
    print(f'Semester Grades sheet rebuilt: {n_students} students, {n_chapters} chapters.')
