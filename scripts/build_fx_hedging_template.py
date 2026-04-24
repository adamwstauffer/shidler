"""
Build FIN-321 FX Transaction Hedging Template workbook.

Generates a UH-branded Excel template with Cover, Receivable, Payable, and Notes tabs,
following the spec at courses/FIN-321-International-Finance-And-Securities/_templates/example-spec-template.md.

Output: courses/FIN-321-International-Finance-And-Securities/_templates/excel/FIN321_FX_Hedging_Template.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import ColorChoice

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "courses" / "FIN-321-International-Finance-And-Securities" / "_templates" / "excel" / "FIN321_FX_Hedging_Template.xlsx"

# Brand tokens (from docs/_branding/design.json v1.0.0)
UH_GREEN = "024731"
UH_GREEN_700 = "013D26"
UH_GREEN_50 = "E6F2EF"
BLACK = "000000"
WHITE = "FFFFFF"
SILVER = "B2B2B2"
NEUTRAL_600 = "525252"
NEUTRAL_200 = "E5E5E5"
YELLOW_INPUT = "FFFF00"
BLUE_ASSUMP = "0000FF"
RED_EXT = "B43232"

FONT = "Open Sans"

# Reusable styles
def f(bold=False, size=10, color=BLACK, italic=False):
    return Font(name=FONT, bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

thin = Side(border_style="thin", color=SILVER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

# Number formats
FMT_USD = '"$"#,##0;("$"#,##0);-'
FMT_FX = "0.0000"
FMT_PCT = "0.00%"
FMT_INT = "#,##0"
FMT_FACTOR = "0.0000"


def style_banner(ws, row, title, subtitle, sub2=None, ncols=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = f(bold=True, size=9, color=WHITE)
    c.fill = fill(UH_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=ncols)
    c = ws.cell(row=row+1, column=1, value=subtitle)
    c.font = f(bold=True, size=14, color=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row+1].height = 24

    if sub2:
        ws.merge_cells(start_row=row+2, start_column=1, end_row=row+2, end_column=ncols)
        c = ws.cell(row=row+2, column=1, value=sub2)
        c.font = f(size=10, color=NEUTRAL_600, italic=True)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row+2].height = 18


def section_heading(ws, row, text, ncols=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = f(bold=True, size=12, color=UH_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    # Bottom border in silver
    c.border = Border(bottom=Side(border_style="thin", color=UH_GREEN))
    ws.row_dimensions[row].height = 20


def header_row(ws, row, headers, start_col=2):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = f(bold=True, size=10, color=WHITE)
        c.fill = fill(UH_GREEN)
        c.alignment = CENTER
        c.border = box
    ws.row_dimensions[row].height = 20


def input_cell(ws, row, col, value, number_format=None, assumption=False, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(YELLOW_INPUT)
    c.font = f(color=(BLUE_ASSUMP if assumption else BLACK), bold=bold)
    c.alignment = RIGHT
    c.border = box
    if number_format:
        c.number_format = number_format
    return c


def formula_cell(ws, row, col, formula, number_format=None, cross_sheet=False, bold=False):
    c = ws.cell(row=row, column=col, value=formula)
    color = UH_GREEN if cross_sheet else BLACK
    c.font = f(color=color, bold=bold)
    c.alignment = RIGHT
    c.border = box
    if number_format:
        c.number_format = number_format
    return c


def label_cell(ws, row, col, text, bold=False, italic=False, indent=0, color=BLACK):
    c = ws.cell(row=row, column=col, value=text)
    c.font = f(bold=bold, italic=italic, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent, wrap_text=True)
    c.border = box
    return c


def set_column_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def add_named(wb, name, ref):
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn


# ---------------------------------------------------------------------------
# Cover tab
# ---------------------------------------------------------------------------
def build_cover(wb):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False

    set_column_widths(ws, {"A": 2, "B": 26, "C": 42, "D": 18, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 2})

    style_banner(
        ws, 1,
        "UNIVERSITY OF HAWAIʻI AT MĀNOA · SHIDLER COLLEGE OF BUSINESS",
        "FIN 321 — International Finance & Securities",
        sub2="FX Transaction Hedging Template · Built to spec v0.1 · UH Mānoa brand v1.0.0",
    )

    # Purpose
    section_heading(ws, 5, "Purpose")
    ws.merge_cells("B6:K7")
    c = ws.cell(row=6, column=2, value=(
        "This workbook implements the four-strategy FX transaction-hedging framework specified in "
        "_templates/example-spec-template.md. Given a foreign-currency exposure (receivable or payable), "
        "it compares No Hedge, Forward, Money-Market, and Option strategies across a ±5% spot-rate "
        "sensitivity grid, using a consistent set of named ranges and standardized calculation flow."
    ))
    c.font = f(size=10)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[6].height = 36
    ws.row_dimensions[7].height = 36

    # How to use
    section_heading(ws, 9, "How to Use")
    steps = [
        ("1.", "Open the Receivable or Payable tab (whichever matches your exposure)."),
        ("2.", "Enter scenario inputs in the YELLOW cells: notional, spot, forward, interest rates, tenor, strike, and premium."),
        ("3.", "Review the DERIVED block (DF_USD, DF_FC, option premium FV) — these update automatically."),
        ("4.", "Read the STRATEGY SUMMARY panel for base-case USD outcome per strategy."),
        ("5.", "Inspect the SENSITIVITY grid and chart to see how each strategy performs across S_T ± 5%."),
        ("6.", "Record data sources and access dates on the NOTES tab (required for reproducibility)."),
        ("7.", "Confirm the parity check: USD_MM should match USD_FWD within rounding."),
    ]
    for i, (n, txt) in enumerate(steps):
        r = 10 + i
        ws.cell(row=r, column=2, value=n).font = f(bold=True, color=UH_GREEN)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
        c = ws.cell(row=r, column=3, value=txt)
        c.font = f(size=10)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 18

    # Legend
    legend_row = 10 + len(steps) + 2
    section_heading(ws, legend_row, "Legend — Cell Color Coding")
    header_row(ws, legend_row + 1, ["Swatch", "Role", "Description"], start_col=2)

    legend_items = [
        (YELLOW_INPUT, BLACK, False, "Input", "Editable scenario values: notional, spot, forward, rates, strikes, premia."),
        (WHITE, BLUE_ASSUMP, True, "Assumption / Hardcode", "Blue text on white — analyst judgment knobs (e.g., strike override, step size)."),
        (WHITE, BLACK, False, "Formula", "Black text on white — all calculations and derived values. Do not edit."),
        (WHITE, UH_GREEN, False, "Cross-sheet link", "UH Green text — pulls from another tab in this workbook."),
        (WHITE, RED_EXT, False, "External link", "Dark red text — external data pull (Bloomberg, FRED, etc.)."),
        (UH_GREEN, WHITE, True, "Header / Banner", "UH Green fill with white text — section headers and table headers."),
    ]
    for i, (bg, fg, bold, role, desc) in enumerate(legend_items):
        r = legend_row + 2 + i
        swatch = ws.cell(row=r, column=2, value="■ sample")
        swatch.fill = fill(bg)
        swatch.font = f(color=fg, bold=bold, size=10)
        swatch.alignment = CENTER
        swatch.border = box
        label_cell(ws, r, 3, role, bold=True)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=11)
        c = ws.cell(row=r, column=4, value=desc)
        c.font = f(size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = box
        ws.row_dimensions[r].height = 18

    # Named range index
    idx_row = legend_row + 2 + len(legend_items) + 2
    section_heading(ws, idx_row, "Named Range Index")
    header_row(ws, idx_row + 1, ["Standardized Name (spec)", "Workbook Name", "Description", "Unit", "Sheet", "Cell"], start_col=2)

    name_index = [
        # (spec_name, workbook_name, desc, unit, sheet, cell)
        ("FC_AMT", "FC_AMT_rec", "Foreign-currency receivable", "FC units", "Receivable", "$D$5"),
        ("S0_in", "S0_in_rec", "Spot exchange rate", "USD/FC", "Receivable", "$D$6"),
        ("F0_in", "F0_in_rec", "Forward rate", "USD/FC", "Receivable", "$D$7"),
        ("R_USD", "R_USD_rec", "USD interest rate (annual)", "%", "Receivable", "$D$8"),
        ("R_FC", "R_FC_rec", "FC interest rate (annual)", "%", "Receivable", "$D$9"),
        ("T_DAYS", "T_DAYS_rec", "Days to settlement", "Days", "Receivable", "$D$10"),
        ("BASIS", "BASIS_rec", "Day-count denominator", "Days", "Receivable", "$D$11"),
        ("K_PUT", "K_PUT_rec", "Put option strike", "USD/FC", "Receivable", "$D$12"),
        ("PREM_PUT", "PREM_PUT_rec", "Put premium (USD per 1 FC)", "USD", "Receivable", "$D$13"),
        ("STEP_FRAC", "STEP_FRAC_rec", "Sensitivity step size", "%", "Receivable", "$D$14"),
        ("DF_USD", "DF_USD_rec", "USD accumulation factor", "×", "Receivable", "$D$17"),
        ("DF_FC", "DF_FC_rec", "FC accumulation factor", "×", "Receivable", "$D$18"),
        ("FV_PREM_PUT", "FV_PREM_PUT_rec", "FV of put premium (neg.)", "USD", "Receivable", "$D$19"),
        ("USD_FWD", "USD_FWD_rec", "Forward hedge USD proceeds", "USD", "Receivable", "$C$24"),
        ("USD_MM", "USD_MM_rec", "Money-market hedge USD proceeds", "USD", "Receivable", "$C$25"),
        ("USD_BASE_PUT", "USD_BASE_PUT_rec", "Put hedge USD proceeds @ baseline", "USD", "Receivable", "$C$26"),
        ("USD_FLOOR_PUT", "USD_FLOOR_PUT_rec", "Worst-case put outcome on grid", "USD", "Receivable", "$C$27"),
        # Payable side
        ("FC_AMT", "FC_AMT_pay", "Foreign-currency payable", "FC units", "Payable", "$D$5"),
        ("S0_in", "S0_in_pay", "Spot exchange rate", "USD/FC", "Payable", "$D$6"),
        ("F0_in", "F0_in_pay", "Forward rate", "USD/FC", "Payable", "$D$7"),
        ("R_USD", "R_USD_pay", "USD interest rate (annual)", "%", "Payable", "$D$8"),
        ("R_FC", "R_FC_pay", "FC interest rate (annual)", "%", "Payable", "$D$9"),
        ("T_DAYS", "T_DAYS_pay", "Days to settlement", "Days", "Payable", "$D$10"),
        ("BASIS", "BASIS_pay", "Day-count denominator", "Days", "Payable", "$D$11"),
        ("K_CALL", "K_CALL_pay", "Call option strike", "USD/FC", "Payable", "$D$12"),
        ("PREM_CALL", "PREM_CALL_pay", "Call premium (USD per 1 FC)", "USD", "Payable", "$D$13"),
        ("STEP_FRAC", "STEP_FRAC_pay", "Sensitivity step size", "%", "Payable", "$D$14"),
        ("DF_USD", "DF_USD_pay", "USD accumulation factor", "×", "Payable", "$D$17"),
        ("DF_FC", "DF_FC_pay", "FC accumulation factor", "×", "Payable", "$D$18"),
        ("FV_PREM_CALL", "FV_PREM_CALL_pay", "FV of call premium (neg.)", "USD", "Payable", "$D$19"),
        ("USD_FWD", "USD_FWD_pay", "Forward hedge USD outlay", "USD", "Payable", "$C$24"),
        ("USD_MM", "USD_MM_pay", "Money-market hedge USD outlay", "USD", "Payable", "$C$25"),
        ("USD_BASE_CALL", "USD_BASE_CALL_pay", "Call hedge USD outlay @ baseline", "USD", "Payable", "$C$26"),
        ("USD_CEILING_CALL", "USD_CEILING_CALL_pay", "Worst-case call outcome on grid", "USD", "Payable", "$C$27"),
    ]
    for i, (spec, wb_name, desc, unit, sh, cell) in enumerate(name_index):
        r = idx_row + 2 + i
        label_cell(ws, r, 2, spec).font = f(bold=True, color=UH_GREEN, size=10)
        label_cell(ws, r, 3, wb_name).font = f(color=BLACK, size=10)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center")
        label_cell(ws, r, 4, desc).font = f(size=10)
        label_cell(ws, r, 5, unit).alignment = CENTER
        label_cell(ws, r, 6, sh).alignment = CENTER
        label_cell(ws, r, 7, cell).alignment = CENTER
        ws.row_dimensions[r].height = 16
    ws.merge_cells(start_row=idx_row + 1, start_column=4, end_row=idx_row + 1, end_column=4)

    # Footer
    foot_row = idx_row + 2 + len(name_index) + 2
    ws.merge_cells(start_row=foot_row, start_column=2, end_row=foot_row, end_column=11)
    c = ws.cell(row=foot_row, column=2, value=(
        "Prepared per UH Mānoa brand standards (docs/_branding/design.json v1.0.0). "
        "Primary green #024731 · Black #000000 · Silver #B2B2B2 · Body Open Sans, 10–12 pt · "
        "ADA-compliant contrast · Flush-left alignment · No custom palettes or gradients."
    ))
    c.font = f(size=9, color=NEUTRAL_600, italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = Border(top=Side(border_style="thin", color=SILVER))
    ws.row_dimensions[foot_row].height = 30

    return ws


# ---------------------------------------------------------------------------
# Exposure tab builder (receivable / payable share structure)
# ---------------------------------------------------------------------------
def build_exposure(wb, mode):
    """mode: 'rec' (receivable) or 'pay' (payable)."""
    is_rec = (mode == "rec")
    sheet_name = "Receivable" if is_rec else "Payable"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    set_column_widths(ws, {
        "A": 2, "B": 4, "C": 38, "D": 15, "E": 12, "F": 14,
        "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 16, "M": 2
    })

    # Banner
    title = "FX Hedging — Foreign-Currency Receivable" if is_rec else "FX Hedging — Foreign-Currency Payable"
    sub = ("Example: U.S. exporter expecting FC-denominated revenue in T_DAYS days."
           if is_rec else
           "Example: U.S. importer owing FC-denominated invoice in T_DAYS days.")
    style_banner(ws, 1,
        "UNIVERSITY OF HAWAIʻI AT MĀNOA · SHIDLER COLLEGE OF BUSINESS · FIN 321",
        title, sub2=sub)

    # --- Inputs section ---
    section_heading(ws, 4, "Inputs (edit YELLOW cells)")

    # Labels and starting values
    # Strike and premium rows differ between rec (put) and pay (call)
    # Defaults chosen so F0 approximately satisfies covered interest-rate parity
    # given R_USD = 5.5%, R_FC = 4.5%, T=365, BASIS=360: F0_parity ≈ S0 × (1.05576/1.04563) ≈ 1.2621
    rows = [
        ("FC_AMT_" + mode, "Foreign-currency notional", "FC units", 10_000_000 if is_rec else 5_000_000, FMT_INT, False),
        ("S0_in_" + mode, "Spot exchange rate (S₀)", "USD/FC", 1.2500, FMT_FX, False),
        ("F0_in_" + mode, "Forward rate to settlement (F₀)", "USD/FC", 1.2620, FMT_FX, False),
        ("R_USD_" + mode, "USD interest rate (annual)", "%", 0.055, FMT_PCT, False),
        ("R_FC_" + mode, "FC interest rate (annual)", "%", 0.045, FMT_PCT, False),
        ("T_DAYS_" + mode, "Days to settlement", "Days", 365, FMT_INT, False),
        ("BASIS_" + mode, "Day-count denominator (assumption)", "Days", 360, FMT_INT, True),
        ("K_PUT_" + mode if is_rec else "K_CALL_" + mode,
            ("Put option strike (assumption — default = S₀)" if is_rec else "Call option strike (assumption — default = S₀)"),
            "USD/FC", "=S0_in_" + mode, FMT_FX, True),
        ("PREM_PUT_" + mode if is_rec else "PREM_CALL_" + mode,
            "Option premium (USD per 1 FC)", "USD", 0.017 if is_rec else 0.021, FMT_FX, False),
        ("STEP_FRAC_" + mode, "Sensitivity step size (assumption)", "%", 0.01, FMT_PCT, True),
    ]

    for i, (nm, desc, unit, val, fmt, is_assump) in enumerate(rows):
        r = 5 + i
        label_cell(ws, r, 2, nm).font = f(bold=True, color=UH_GREEN, size=10)
        label_cell(ws, r, 3, desc)
        input_cell(ws, r, 4, val, number_format=fmt, assumption=is_assump)
        label_cell(ws, r, 5, unit).alignment = CENTER
        # Register as global named range
        ref = f"'{sheet_name}'!$D${r}"
        add_named(wb, nm, ref)

    # --- Derived block ---
    section_heading(ws, 16, "Derived Values")
    derived = [
        ("DF_USD_" + mode, "USD accumulation factor = 1 + R_USD × T_DAYS/BASIS",
         f"=1+R_USD_{mode}*T_DAYS_{mode}/BASIS_{mode}", FMT_FACTOR),
        ("DF_FC_" + mode, "FC accumulation factor = 1 + R_FC × T_DAYS/BASIS",
         f"=1+R_FC_{mode}*T_DAYS_{mode}/BASIS_{mode}", FMT_FACTOR),
    ]
    if is_rec:
        derived.append(("FV_PREM_PUT_rec",
            "FV of put premium (negative cash flow at t₀, grown to maturity)",
            "=-PREM_PUT_rec*FC_AMT_rec*DF_USD_rec", FMT_USD))
    else:
        derived.append(("FV_PREM_CALL_pay",
            "FV of call premium (negative cash flow at t₀, grown to maturity)",
            "=-PREM_CALL_pay*FC_AMT_pay*DF_USD_pay", FMT_USD))

    for i, (nm, desc, formula, fmt) in enumerate(derived):
        r = 17 + i
        label_cell(ws, r, 2, nm).font = f(bold=True, color=UH_GREEN, size=10)
        label_cell(ws, r, 3, desc)
        formula_cell(ws, r, 4, formula, number_format=fmt)
        label_cell(ws, r, 5, "").alignment = CENTER
        add_named(wb, nm, f"'{sheet_name}'!$D${r}")

    # --- Strategy summary ---
    section_heading(ws, 21, "Strategy Summary (Baseline at S_T = S₀)")
    verb_up = "Proceeds" if is_rec else "Outlay"
    header_row(ws, 22, ["Strategy", f"USD {verb_up} @ Baseline", "Locked-in across grid?", "Hedge Profit vs. No Hedge"], start_col=2)

    # Strategy values live in column C (index 3). No Hedge = row 23, Forward = 24, MM = 25, Option = 26, Floor/Ceiling = 27.
    if is_rec:
        strat_rows = [
            ("No Hedge",    "=S0_in_rec*FC_AMT_rec",                    "No",  "—"),
            ("Forward",     "=FC_AMT_rec*F0_in_rec",                    "Yes", "=C24-C23"),
            ("Money Market","=(FC_AMT_rec/DF_FC_rec)*S0_in_rec*DF_USD_rec", "Yes", "=C25-C23"),
            ("Option (Put)","=MAX(S0_in_rec,K_PUT_rec)*FC_AMT_rec+FV_PREM_PUT_rec", "No", "=C26-C23"),
            ("—  Put floor across grid",
                           "=MIN(G33:G43)",                             "—",   "—"),
        ]
    else:
        strat_rows = [
            ("No Hedge",    "=S0_in_pay*FC_AMT_pay",                    "No",  "—"),
            ("Forward",     "=FC_AMT_pay*F0_in_pay",                    "Yes", "=C23-C24"),
            ("Money Market","=(FC_AMT_pay/DF_FC_pay)*S0_in_pay*DF_USD_pay", "Yes", "=C23-C25"),
            ("Option (Call)","=MIN(S0_in_pay,K_CALL_pay)*FC_AMT_pay-FV_PREM_CALL_pay", "No", "=C23-C26"),
            ("—  Call ceiling across grid",
                           "=MAX(G33:G43)",                             "—",   "—"),
        ]

    for i, (strat, formula, locked, profit) in enumerate(strat_rows):
        r = 22 + i + 1
        label_cell(ws, r, 2, strat).font = f(bold=(i in (4,)), size=10)
        formula_cell(ws, r, 3, formula, number_format=FMT_USD, bold=(i == 4))
        label_cell(ws, r, 4, locked).alignment = CENTER
        if profit == "—":
            label_cell(ws, r, 5, "—").alignment = CENTER
        else:
            formula_cell(ws, r, 5, profit, number_format=FMT_USD)

    # Register scalar named ranges (strategy rows 23–27)
    if is_rec:
        add_named(wb, "USD_FWD_rec",         f"'{sheet_name}'!$C$24")
        add_named(wb, "USD_MM_rec",          f"'{sheet_name}'!$C$25")
        add_named(wb, "USD_BASE_PUT_rec",    f"'{sheet_name}'!$C$26")
        add_named(wb, "USD_FLOOR_PUT_rec",   f"'{sheet_name}'!$C$27")
    else:
        add_named(wb, "USD_FWD_pay",         f"'{sheet_name}'!$C$24")
        add_named(wb, "USD_MM_pay",          f"'{sheet_name}'!$C$25")
        add_named(wb, "USD_BASE_CALL_pay",   f"'{sheet_name}'!$C$26")
        add_named(wb, "USD_CEILING_CALL_pay",f"'{sheet_name}'!$C$27")

    # Parity check
    r = 28
    label_cell(ws, r, 2, "Parity check").font = f(bold=True, color=UH_GREEN, size=10)
    if is_rec:
        formula_cell(ws, r, 3, "=USD_MM_rec-USD_FWD_rec", number_format=FMT_USD)
    else:
        formula_cell(ws, r, 3, "=USD_MM_pay-USD_FWD_pay", number_format=FMT_USD)
    label_cell(ws, r, 4, "Should be near zero under covered interest-rate parity.").font = f(italic=True, color=NEUTRAL_600, size=9)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    # --- Sensitivity grid ---
    section_heading(ws, 30, f"Sensitivity: USD {verb_up} by Strategy vs. S_T")

    strat_label_no   = "No Hedge"
    strat_label_fwd  = "Forward"
    strat_label_mm   = "Money Market"
    strat_label_opt  = "Option (Put)" if is_rec else "Option (Call)"

    headers = ["n", "S_T", strat_label_no, strat_label_fwd, strat_label_mm, strat_label_opt,
               "Profit: Fwd", "Profit: MM", "Profit: Opt", "Overall Winner", "Best Active Hedge"]
    header_row(ws, 32, headers, start_col=2)
    ws.row_dimensions[32].height = 28

    # 11 rows, n = -5..+5
    first_row = 33
    last_row = 43
    for i, n in enumerate(range(-5, 6)):
        r = first_row + i
        # Column B: n
        ws.cell(row=r, column=2, value=n).font = f(color=NEUTRAL_600, size=10)
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=2).border = box

        # Column C: S_T = S0 × (1 + n × STEP_FRAC)
        s0 = f"S0_in_{mode}"
        sf = f"STEP_FRAC_{mode}"
        formula_cell(ws, r, 3, f"=$B{r}*{sf}*{s0}+{s0}", number_format=FMT_FX)

        # Column D: No hedge
        if is_rec:
            formula_cell(ws, r, 4, f"=$C{r}*FC_AMT_rec", number_format=FMT_USD)
        else:
            formula_cell(ws, r, 4, f"=$C{r}*FC_AMT_pay", number_format=FMT_USD)

        # Column E: Forward (constant)
        if is_rec:
            formula_cell(ws, r, 5, "=USD_FWD_rec", number_format=FMT_USD)
        else:
            formula_cell(ws, r, 5, "=USD_FWD_pay", number_format=FMT_USD)

        # Column F: Money market (constant)
        if is_rec:
            formula_cell(ws, r, 6, "=USD_MM_rec", number_format=FMT_USD)
        else:
            formula_cell(ws, r, 6, "=USD_MM_pay", number_format=FMT_USD)

        # Column G: Option
        if is_rec:
            # Put: proceeds = MAX(S_T, K_PUT) × FC_AMT + FV_PREM_PUT
            formula_cell(ws, r, 7, f"=MAX($C{r},K_PUT_rec)*FC_AMT_rec+FV_PREM_PUT_rec", number_format=FMT_USD)
        else:
            # Call: outlay = MIN(S_T, K_CALL) × FC_AMT - FV_PREM_CALL
            formula_cell(ws, r, 7, f"=MIN($C{r},K_CALL_pay)*FC_AMT_pay-FV_PREM_CALL_pay", number_format=FMT_USD)

        # Hedge profit columns (vs. no hedge)
        if is_rec:
            formula_cell(ws, r, 8, f"=$E{r}-$D{r}", number_format=FMT_USD)  # forward
            formula_cell(ws, r, 9, f"=$F{r}-$D{r}", number_format=FMT_USD)  # MM
            formula_cell(ws, r, 10, f"=$G{r}-$D{r}", number_format=FMT_USD) # option
            # Overall winner
            formula_cell(ws, r, 11,
                f'=INDEX({{"No Hedge","Forward","Money Market","Option"}},MATCH(MAX($D{r}:$G{r}),$D{r}:$G{r},0))')
            # Best active hedge
            formula_cell(ws, r, 12,
                f'=INDEX({{"Forward","Money Market","Option"}},MATCH(MAX($E{r}:$G{r}),$E{r}:$G{r},0))')
        else:
            formula_cell(ws, r, 8, f"=$D{r}-$E{r}", number_format=FMT_USD)
            formula_cell(ws, r, 9, f"=$D{r}-$F{r}", number_format=FMT_USD)
            formula_cell(ws, r, 10, f"=$D{r}-$G{r}", number_format=FMT_USD)
            formula_cell(ws, r, 11,
                f'=INDEX({{"No Hedge","Forward","Money Market","Option"}},MATCH(MIN($D{r}:$G{r}),$D{r}:$G{r},0))')
            formula_cell(ws, r, 12,
                f'=INDEX({{"Forward","Money Market","Option"}},MATCH(MIN($E{r}:$G{r}),$E{r}:$G{r},0))')

        # Highlight baseline row (n == 0) in light green
        if n == 0:
            for col in range(2, 13):
                cc = ws.cell(row=r, column=col)
                cc.fill = fill(UH_GREEN_50)

    # Winner label column styling
    for r in range(first_row, last_row + 1):
        ws.cell(row=r, column=11).font = f(bold=True, color=UH_GREEN)
        ws.cell(row=r, column=11).alignment = CENTER
        ws.cell(row=r, column=12).font = f(color=NEUTRAL_600, italic=True)
        ws.cell(row=r, column=12).alignment = CENTER

    # --- Chart ---
    chart_row = last_row + 3
    chart = LineChart()
    chart.title = f"USD {verb_up} by Strategy vs. Future Spot S_T"
    chart.style = 2
    chart.y_axis.title = f"USD {verb_up}"
    chart.x_axis.title = "S_T (USD per FC)"
    chart.height = 10
    chart.width = 20
    chart.legend.position = "b"

    # Series: No Hedge, Forward, MM, Option
    for col_offset, series_name, hex_color, dashed in [
        (4, strat_label_no,  BLACK,        False),
        (5, strat_label_fwd, UH_GREEN,     False),
        (6, strat_label_mm,  UH_GREEN_700, True),
        (7, strat_label_opt, NEUTRAL_600,  False),
    ]:
        data_ref = Reference(ws, min_col=col_offset, min_row=32, max_col=col_offset, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        series = chart.series[-1]
        series.graphicalProperties = GraphicalProperties(
            solidFill=hex_color,
            ln=LineProperties(solidFill=hex_color, w=22000 if col_offset in (5, 7) else 18000,
                              prstDash="dash" if dashed else "solid")
        )

    cats = Reference(ws, min_col=3, min_row=first_row, max_col=3, max_row=last_row)
    chart.set_categories(cats)

    ws.add_chart(chart, f"B{chart_row}")

    # Freeze header rows
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Notes tab
# ---------------------------------------------------------------------------
def build_notes(wb):
    ws = wb.create_sheet("Notes")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, {"A": 2, "B": 28, "C": 24, "D": 18, "E": 36, "F": 18, "G": 18, "H": 2})

    style_banner(ws, 1,
        "UNIVERSITY OF HAWAIʻI AT MĀNOA · SHIDLER COLLEGE OF BUSINESS · FIN 321",
        "Data Sources & Access Dates",
        sub2="Record sources for every market input. Required for Stage 3 reproducibility.")

    section_heading(ws, 4, "Market Data Inputs")
    header_row(ws, 5, ["Input", "Tab", "Cell", "Source (vendor / URL / dataset)", "Access Date", "Analyst"], start_col=2)

    inputs = [
        ("Spot rate (S₀)", "Receivable", "$D$6"),
        ("Forward rate (F₀)", "Receivable", "$D$7"),
        ("USD interest rate", "Receivable", "$D$8"),
        ("FC interest rate", "Receivable", "$D$9"),
        ("Put premium", "Receivable", "$D$13"),
        ("Spot rate (S₀)", "Payable", "$D$6"),
        ("Forward rate (F₀)", "Payable", "$D$7"),
        ("USD interest rate", "Payable", "$D$8"),
        ("FC interest rate", "Payable", "$D$9"),
        ("Call premium", "Payable", "$D$13"),
    ]
    for i, (inp, sh, cell) in enumerate(inputs):
        r = 6 + i
        label_cell(ws, r, 2, inp)
        label_cell(ws, r, 3, sh).alignment = CENTER
        label_cell(ws, r, 4, cell).alignment = CENTER
        # Yellow input cells for analyst to fill in
        input_cell(ws, r, 5, "", number_format="@").alignment = LEFT
        input_cell(ws, r, 6, "", number_format="yyyy-mm-dd").alignment = CENTER
        input_cell(ws, r, 7, "", number_format="@").alignment = LEFT

    # Analytical assumptions narrative
    sec2 = 6 + len(inputs) + 2
    section_heading(ws, sec2, "Analytical Assumptions & Judgment Calls")
    ws.merge_cells(start_row=sec2+1, start_column=2, end_row=sec2+6, end_column=7)
    c = ws.cell(row=sec2+1, column=2, value=(
        "Record any deviations from the default assumptions in §3 of the spec. Examples:\n"
        "  • BASIS override (360 vs. 365) and why.\n"
        "  • Strike override away from S₀ (e.g., out-of-the-money put for cheaper premium).\n"
        "  • Transaction-cost adjustments or bid-ask spread assumptions.\n"
        "  • Rationale for scenario construction (step size, range)."
    ))
    c.font = f(size=10)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.fill = fill(UH_GREEN_50)
    c.border = box


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_cover(wb)
    build_exposure(wb, "rec")
    build_exposure(wb, "pay")
    build_notes(wb)

    # Set workbook default font (applied to all cells without explicit font via new styles)
    # Note: openpyxl doesn't natively support setting default font; individual cells set explicitly above.

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
