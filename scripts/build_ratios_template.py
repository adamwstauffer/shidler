"""Build BUS-314 Accounting & Performance Ratios Excel template.

Generates a blank-input workbook with named ranges, color-coded cells,
a branded cover sheet, and pre-wired ratio formulas that return "—"
until inputs are populated (via IFERROR wrappers).

Output (legacy, archived): _archive/bus314/accounting-ratios/_templates/excel/BUS314_Ratios_Template.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT = (
    "_archive/bus314/"
    "accounting-ratios/_templates/excel/BUS314_Ratios_Template.xlsx"
)

UH_GREEN = "024731"
UH_GREEN_LIGHT = "E6F2EF"
BLACK = "000000"
WHITE = "FFFFFF"
SILVER = "B2B2B2"
NEUTRAL_600 = "525252"

INPUT_BG = "FFFFCC"          # light yellow — data-entry cells (10-K numbers)
ASSUMPTION_BG = "DCE6F1"     # light blue bg — analyst assumptions
ASSUMPTION_TEXT = "0000FF"   # blue text for hardcoded assumption values
FORMULA_TEXT = "008000"      # green text — cross-sheet / derived formulas
OUTPUT_BG = "F2F2F2"         # light gray — computed ratio outputs
HEADER_BG = UH_GREEN
HEADER_TEXT = WHITE

FONT = "Open Sans"
FONT_FALLBACK = "Arial"

wb = Workbook()
wb.remove(wb.active)


def f_header(size=11):
    return Font(name=FONT, size=size, bold=True, color=HEADER_TEXT)


def f_section(size=11):
    return Font(name=FONT, size=size, bold=True, color=UH_GREEN)


def f_label(bold=False, size=10):
    return Font(name=FONT, size=size, bold=bold, color=BLACK)


def f_input():
    return Font(name=FONT, size=10, color=BLACK)


def f_assumption():
    return Font(name=FONT, size=10, bold=True, color=ASSUMPTION_TEXT)


def f_formula():
    return Font(name=FONT, size=10, color=FORMULA_TEXT)


def f_output():
    return Font(name=FONT, size=10, color=BLACK)


def f_mono(size=9):
    return Font(name="Consolas", size=size, color=NEUTRAL_600)


def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


thin = Side(border_style="thin", color=SILVER)
box = Border(top=thin, bottom=thin, left=thin, right=thin)


def style_header_row(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = f_header()
        cell.fill = fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = box


def style_section_row(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = f_section()
        cell.fill = fill(UH_GREEN_LIGHT)


def add_name(name, ref):
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn


# =============================================================================
# COVER SHEET
# =============================================================================
cover = wb.create_sheet("Cover")
cover.sheet_view.showGridLines = False
cover.column_dimensions["A"].width = 2
cover.column_dimensions["B"].width = 28
cover.column_dimensions["C"].width = 70
cover.column_dimensions["D"].width = 2

# Brand banner
cover.row_dimensions[2].height = 8
cover.cell(row=2, column=2).fill = fill(UH_GREEN)
cover.cell(row=2, column=3).fill = fill(UH_GREEN)

cover["B4"] = "UNIVERSITY OF HAWAIʻI AT MĀNOA · SHIDLER COLLEGE OF BUSINESS"
cover["B4"].font = Font(name=FONT, size=9, bold=True, color=UH_GREEN)
cover.merge_cells("B4:C4")

cover["B5"] = "BUS-314 International Corporate Finance"
cover["B5"].font = Font(name=FONT, size=18, bold=True, color=BLACK)
cover.merge_cells("B5:C5")

cover["B6"] = "Accounting & Performance Ratios — Excel Template"
cover["B6"].font = Font(name=FONT, size=12, color=NEUTRAL_600)
cover.merge_cells("B6:C6")

cover.row_dimensions[7].height = 4
cover.cell(row=7, column=2).fill = fill(SILVER)
cover.cell(row=7, column=3).fill = fill(SILVER)

# How to use
cover["B9"] = "How to Use This Template"
cover["B9"].font = f_section(size=13)
cover.merge_cells("B9:C9")

how_to_use = [
    ("1.", "Enter prior-year and current-year balance-sheet figures on the Balance Sheet tab."),
    ("2.", "Enter current-year income-statement figures on the Income Statement tab."),
    ("3.", "Enter current-year cash-flow items on the Cash Flow Statement tab."),
    ("4.", "On the Ratios tab, enter the four analyst assumptions at top: share price, shares outstanding, cost of capital, tax rate. Also enter the current/prior fiscal years."),
    ("5.", "On the Ratios tab Outputs section, compute each ratio in column C using the named-range formulas shown in column D. The Du Pont ROA self-check on column F confirms your work."),
    ("6.", "Document company metadata, data source, and AI usage on the Notes tab."),
]
r = 10
for num, text in how_to_use:
    cover.cell(row=r, column=2, value=num).font = Font(name=FONT, size=10, bold=True, color=UH_GREEN)
    cover.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="top")
    cover.cell(row=r, column=3, value=text).font = f_label(size=10)
    cover.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    cover.row_dimensions[r].height = 18
    r += 1

r += 1
# Color key
cover.cell(row=r, column=2, value="Color Key").font = f_section(size=13)
cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1

color_key_header_row = r
cover.cell(row=r, column=2, value="Style").font = f_header()
cover.cell(row=r, column=2).fill = fill(HEADER_BG)
cover.cell(row=r, column=3, value="Meaning").font = f_header()
cover.cell(row=r, column=3).fill = fill(HEADER_BG)
r += 1

color_rows = [
    ("Sample", INPUT_BG, BLACK, False, "Yellow background — DATA INPUTS. Numbers pulled from 10-K / financial statements. Overwrite with your company's figures."),
    ("12.3%", ASSUMPTION_BG, ASSUMPTION_TEXT, True, "Light-blue background + blue text — ASSUMPTIONS. Analyst inputs: share price, shares outstanding, WACC, tax rate, fiscal years."),
    ("f(x)", None, FORMULA_TEXT, False, "Green text — FORMULAS. Cross-sheet references and derived calculations. Do not overwrite."),
    ("1.23x", OUTPUT_BG, BLACK, True, "Gray background — RATIO OUTPUTS. Compute each ratio in column C using the named-range formulas shown alongside in column D."),
]
for sample, bg, txt_color, bold, desc in color_rows:
    c = cover.cell(row=r, column=2, value=sample)
    c.font = Font(name=FONT, size=10, bold=bold, color=txt_color)
    c.alignment = Alignment(horizontal="center")
    if bg:
        c.fill = fill(bg)
    c.border = box
    d = cover.cell(row=r, column=3, value=desc)
    d.font = f_label(size=10)
    d.alignment = Alignment(wrap_text=True, vertical="center")
    d.border = box
    cover.row_dimensions[r].height = 30
    r += 1

r += 1
cover.cell(row=r, column=2, value="Named-Range Convention").font = f_section(size=13)
cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
conv_rows = [
    ("BAL_[item]_[yr]", "Balance-sheet line items, suffixed with fiscal year (e.g., BAL_assets_total_2024)."),
    ("INC_[item]", "Income-statement items (e.g., INC_sales, INC_ebit, INC_net)."),
    ("CASH_[item]", "Cash-flow items (e.g., CASH_operating)."),
    ("share_price, shares_outstanding, cost_capital, tax_rate", "Analyst assumptions (no prefix)."),
    ("startYear_[item]", "Alias for prior-year balance (e.g., startYear_equity ≡ BAL_equity_shareholders_prior)."),
    ("currentYear_[item]", "Alias for current-year balance or derived current-year figure."),
    ("avg_[item]", "Mean of start-of-year and current-year (e.g., avg_equity)."),
    ("RATIO_[name]", "Key ratios reused in Du Pont decomposition (asset_turnover, operating_profit_margin, leverage, debt_burden)."),
]
c = cover.cell(row=r, column=2, value="Prefix / Pattern")
c.font = f_header(); c.fill = fill(HEADER_BG)
c = cover.cell(row=r, column=3, value="Meaning")
c.font = f_header(); c.fill = fill(HEADER_BG)
r += 1
for pattern, meaning in conv_rows:
    c = cover.cell(row=r, column=2, value=pattern)
    c.font = f_mono(10); c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = box
    d = cover.cell(row=r, column=3, value=meaning)
    d.font = f_label(size=10); d.alignment = Alignment(wrap_text=True, vertical="center")
    d.border = box
    cover.row_dimensions[r].height = 24
    r += 1

r += 2
cover.cell(row=r, column=2, value="Companion Specification").font = f_section(size=13)
cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
cover.cell(row=r, column=2, value="See:").font = f_label(size=10)
cover.cell(row=r, column=3, value="accounting-ratios/_templates/example-spec-template.md").font = f_mono(10)
r += 1
cover.cell(row=r, column=2, value="Skill:").font = f_label(size=10)
cover.cell(row=r, column=3, value=".claude/skills/accounting-ratios/SKILL.md").font = f_mono(10)

r += 3
cover.cell(row=r, column=2, value="Prepared per UH Mānoa brand standards (docs/_branding/design.json). Primary green #024731 · Open Sans typography · ADA-compliant contrast.").font = Font(name=FONT, size=8, italic=True, color=NEUTRAL_600)
cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
cover.cell(row=r, column=2).alignment = Alignment(wrap_text=True)


# =============================================================================
# BALANCE SHEET
# =============================================================================
bs = wb.create_sheet("Balance Sheet")
bs.sheet_view.showGridLines = False
bs.column_dimensions["A"].width = 2
bs.column_dimensions["B"].width = 38
bs.column_dimensions["C"].width = 16
bs.column_dimensions["D"].width = 16
bs.column_dimensions["E"].width = 2
bs.column_dimensions["F"].width = 38
bs.column_dimensions["G"].width = 16
bs.column_dimensions["H"].width = 16

bs["B2"] = "Balance Sheet"
bs["B2"].font = Font(name=FONT, size=16, bold=True, color=UH_GREEN)
bs["B3"] = "All figures in $millions unless otherwise noted"
bs["B3"].font = Font(name=FONT, size=9, italic=True, color=NEUTRAL_600)

# Header row (row 5)
bs["B5"] = "ASSETS"; bs["C5"] = "FY Current"; bs["D5"] = "FY Prior"
bs["F5"] = "LIABILITIES & EQUITY"; bs["G5"] = "FY Current"; bs["H5"] = "FY Prior"
for col in [2, 3, 4, 6, 7, 8]:
    c = bs.cell(row=5, column=col)
    c.font = f_header(); c.fill = fill(HEADER_BG); c.border = box
    c.alignment = Alignment(horizontal="center" if col in [3, 4, 7, 8] else "left")

# Balance-sheet structure:
# Left column: assets | Right column: liabilities & equity
bs_assets = [
    # (label, named_range_base, is_subtotal)
    ("Current Assets", None, "section"),
    ("Cash and marketable securities", "cash_marketable_securities", "input"),
    ("Receivables", "receivables", "input"),
    ("Inventories", "inventories", "input"),
    ("Other current assets", "other_current_assets", "input"),
    ("   Total current assets", "assets_current", "subtotal"),
    ("", None, "spacer"),
    ("Fixed Assets", None, "section"),
    ("Property, plant, and equipment", "ppe_gross", "input"),
    ("Less accumulated depreciation", "accumulated_depreciation", "input"),
    ("   Net tangible fixed assets", "fixed_assets_net", "subtotal"),
    ("", None, "spacer"),
    ("Other Assets", None, "section"),
    ("Intangible assets (goodwill)", "intangibles", "input"),
    ("Other assets", "other_assets", "input"),
    ("", None, "spacer"),
    ("Total Assets", "assets_total", "total"),
]

bs_liab = [
    ("Current Liabilities", None, "section"),
    ("Debt due for repayment", "debt_short_term", "input"),
    ("Accounts payable", "accounts_payable", "input"),
    ("Other current liabilities", "other_current_liabilities", "input"),
    ("   Total current liabilities", "liabilities_current", "subtotal"),
    ("", None, "spacer"),
    ("Long-term debt", "debt_long_term", "input"),
    ("Other long-term liabilities", "other_long_term_liabilities", "input"),
    ("", None, "spacer"),
    ("Total Liabilities", "liabilities_total", "subtotal"),
    ("", None, "spacer"),
    ("SHAREHOLDERS' EQUITY", None, "section"),
    ("Common stock and paid-in capital", "common_stock", "input"),
    ("Retained earnings", "retained_earnings", "input"),
    ("   Total shareholders' equity", "equity_shareholders", "subtotal"),
    ("", None, "spacer"),
    ("", None, "spacer"),
    ("Total Liabilities + Equity", "total_liab_equity", "total"),
]


def write_bs_column(items, start_col_label, start_col_curr, start_col_prior, start_row, side):
    row = start_row
    # Track which named ranges we've emitted cells for
    for label, base, kind in items:
        cell_label = bs.cell(row=row, column=start_col_label, value=label)
        cell_curr = bs.cell(row=row, column=start_col_curr)
        cell_prior = bs.cell(row=row, column=start_col_prior)

        if kind == "section":
            cell_label.font = Font(name=FONT, size=10, bold=True, color=UH_GREEN)
        elif kind == "input":
            cell_label.font = f_label(size=10)
            for c in (cell_curr, cell_prior):
                c.fill = fill(INPUT_BG)
                c.font = f_input()
                c.number_format = '#,##0;(#,##0);"—"'
                c.border = box
        elif kind == "subtotal":
            cell_label.font = f_label(bold=True, size=10)
            cell_curr.font = f_label(bold=True, size=10)
            cell_prior.font = f_label(bold=True, size=10)
            cell_curr.number_format = '#,##0;(#,##0);"—"'
            cell_prior.number_format = '#,##0;(#,##0);"—"'
            cell_curr.fill = fill(OUTPUT_BG)
            cell_prior.fill = fill(OUTPUT_BG)
            cell_curr.border = box; cell_prior.border = box
        elif kind == "total":
            cell_label.font = Font(name=FONT, size=11, bold=True, color=WHITE)
            cell_label.fill = fill(UH_GREEN)
            for c in (cell_curr, cell_prior):
                c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
                c.fill = fill(UH_GREEN)
                c.number_format = '#,##0;(#,##0);"—"'
        row += 1
    return row


end_left = write_bs_column(bs_assets, 2, 3, 4, 6, "assets")
end_right = write_bs_column(bs_liab, 6, 7, 8, 6, "liab")

# Wire the subtotal / total formulas
# Left side
bs["C11"] = "=SUM(C7:C10)"   # Total current assets
bs["D11"] = "=SUM(D7:D10)"
bs["C16"] = "=C14-C15"       # Net tangible fixed assets
bs["D16"] = "=D14-D15"
bs["C22"] = "=C11+C16+C19+C20"  # Total Assets
bs["D22"] = "=D11+D16+D19+D20"

# Right side
bs["G10"] = "=SUM(G7:G9)"    # Total current liabilities
bs["H10"] = "=SUM(H7:H9)"
bs["G15"] = "=G10+G12+G13"   # Total Liabilities
bs["H15"] = "=H10+H12+H13"
bs["G20"] = "=G18+G19"       # Total shareholders equity
bs["H20"] = "=H18+H19"
bs["G23"] = "=G15+G20"       # Total liab + equity
bs["H23"] = "=H15+H20"

# Named ranges for balance sheet items.
# Curr-year maps to column C (assets) / G (liab); prior-year maps to column D / H.
# Row-tracking helper — rebuild the row map for named ranges.
# LEFT (assets) rows 6–22
bs_name_map_assets = {
    "cash_marketable_securities": 7,
    "receivables": 8,
    "inventories": 9,
    "other_current_assets": 10,
    "assets_current": 11,
    "ppe_gross": 14,
    "accumulated_depreciation": 15,
    "fixed_assets_net": 16,
    "intangibles": 19,
    "other_assets": 20,
    "assets_total": 22,
}
# RIGHT (liab/equity)
bs_name_map_liab = {
    "debt_short_term": 7,
    "accounts_payable": 8,
    "other_current_liabilities": 9,
    "liabilities_current": 10,
    "debt_long_term": 12,
    "other_long_term_liabilities": 13,
    "liabilities_total": 15,
    "common_stock": 18,
    "retained_earnings": 19,
    "equity_shareholders": 20,
}

for base, row in bs_name_map_assets.items():
    add_name(f"BAL_{base}_curr", f"'Balance Sheet'!$C${row}")
    add_name(f"BAL_{base}_prior", f"'Balance Sheet'!$D${row}")
for base, row in bs_name_map_liab.items():
    add_name(f"BAL_{base}_curr", f"'Balance Sheet'!$G${row}")
    add_name(f"BAL_{base}_prior", f"'Balance Sheet'!$H${row}")

bs.freeze_panes = "B6"


# =============================================================================
# INCOME STATEMENT
# =============================================================================
ist = wb.create_sheet("Income Statement")
ist.sheet_view.showGridLines = False
ist.column_dimensions["A"].width = 2
ist.column_dimensions["B"].width = 45
ist.column_dimensions["C"].width = 16
ist.column_dimensions["D"].width = 14

ist["B2"] = "Income Statement"
ist["B2"].font = Font(name=FONT, size=16, bold=True, color=UH_GREEN)
ist["B3"] = "Current fiscal year · figures in $millions"
ist["B3"].font = Font(name=FONT, size=9, italic=True, color=NEUTRAL_600)

ist["B5"] = "Line Item"; ist["C5"] = "Amount"; ist["D5"] = "% of Sales"
for col in [2, 3, 4]:
    c = ist.cell(row=5, column=col)
    c.font = f_header(); c.fill = fill(HEADER_BG); c.border = box

is_items = [
    # (label, named_range, kind, formula)
    ("Net sales", "sales", "input", None),
    ("less Cost of goods sold", "cost_goods_sold", "input", None),
    ("less Selling, general & administrative expenses", "sga", "input", None),
    ("less Depreciation", "depreciation", "input", None),
    ("   Earnings before interest and taxes (EBIT)", "ebit", "subtotal", "=C6-C7-C8-C9"),
    ("plus Other income", "other_income", "input", None),
    ("less Interest expense", "interest_expense", "input", None),
    ("   Taxable income", "taxable_income", "subtotal", "=C10+C11-C12"),
    ("less Taxes", "taxes", "input", None),
    ("   Net income", "net", "total", "=C13-C14"),
    ("", None, "spacer", None),
    ("Allocation of net income", None, "section", None),
    ("   Dividends", "dividends", "input", None),
    ("   Addition to retained earnings", "addition_retained_earnings", "subtotal", "=C15-C18"),
]

row = 6
is_row_map = {}
for label, name, kind, formula in is_items:
    ist.cell(row=row, column=2, value=label)
    val_cell = ist.cell(row=row, column=3)
    pct_cell = ist.cell(row=row, column=4)

    if kind == "input":
        ist.cell(row=row, column=2).font = f_label(size=10)
        val_cell.fill = fill(INPUT_BG); val_cell.font = f_input()
        val_cell.number_format = '#,##0;(#,##0);"—"'; val_cell.border = box
        pct_cell.font = f_formula()
        pct_cell.number_format = '0.0%;(0.0%);"—"'
        pct_cell.value = f'=IFERROR({val_cell.coordinate}/$C$6,"")'
    elif kind == "subtotal":
        ist.cell(row=row, column=2).font = f_label(bold=True, size=10)
        val_cell.font = f_label(bold=True, size=10); val_cell.fill = fill(OUTPUT_BG)
        val_cell.number_format = '#,##0;(#,##0);"—"'; val_cell.border = box
        val_cell.value = formula
        pct_cell.font = f_formula()
        pct_cell.number_format = '0.0%;(0.0%);"—"'
        pct_cell.value = f'=IFERROR({val_cell.coordinate}/$C$6,"")'
    elif kind == "total":
        ist.cell(row=row, column=2).font = Font(name=FONT, size=11, bold=True, color=WHITE)
        ist.cell(row=row, column=2).fill = fill(UH_GREEN)
        val_cell.font = Font(name=FONT, size=11, bold=True, color=WHITE); val_cell.fill = fill(UH_GREEN)
        val_cell.number_format = '#,##0;(#,##0);"—"'
        val_cell.value = formula
        pct_cell.font = Font(name=FONT, size=11, bold=True, color=WHITE); pct_cell.fill = fill(UH_GREEN)
        pct_cell.number_format = '0.0%;(0.0%);"—"'
        pct_cell.value = f'=IFERROR({val_cell.coordinate}/$C$6,"")'
    elif kind == "section":
        ist.cell(row=row, column=2).font = Font(name=FONT, size=10, bold=True, color=UH_GREEN)

    if name:
        is_row_map[name] = row
    row += 1

# 100% row for Net sales is_row_map should now show sales at row 6
ist["D6"].value = 1  # 100% of sales
ist["D6"].number_format = '0.0%'
ist["D6"].font = f_formula()

# Named ranges — INC_*
for name, row in is_row_map.items():
    add_name(f"INC_{name}", f"'Income Statement'!$C${row}")

ist.freeze_panes = "B6"


# =============================================================================
# CASH FLOW STATEMENT
# =============================================================================
cfs = wb.create_sheet("Cash Flow Statement")
cfs.sheet_view.showGridLines = False
cfs.column_dimensions["A"].width = 2
cfs.column_dimensions["B"].width = 50
cfs.column_dimensions["C"].width = 16
cfs.column_dimensions["D"].width = 24

cfs["B2"] = "Cash Flow Statement"
cfs["B2"].font = Font(name=FONT, size=16, bold=True, color=UH_GREEN)
cfs["B3"] = "Current fiscal year · figures in $millions"
cfs["B3"].font = Font(name=FONT, size=9, italic=True, color=NEUTRAL_600)

cfs["B5"] = "Line Item"; cfs["C5"] = "Amount"; cfs["D5"] = "Source / Notes"
for col in [2, 3, 4]:
    c = cfs.cell(row=5, column=col)
    c.font = f_header(); c.fill = fill(HEADER_BG); c.border = box

cfs_items = [
    ("Operations:", None, "section", None, ""),
    ("Net income", None, "link", "=INC_net", "[Income Statement]"),
    ("plus Depreciation", None, "link", "=INC_depreciation", "[Income Statement]"),
    ("   plus Changes in working capital items", None, "subsection", None, ""),
    ("      Decrease (increase) in accounts receivable", "chg_receivables", "input", None, "[Balance Sheet]"),
    ("      Decrease (increase) in inventories", "chg_inventories", "input", None, "[Balance Sheet]"),
    ("      Decrease (increase) in other current assets", "chg_other_current_assets", "input", None, "[Balance Sheet]"),
    ("      Increase (decrease) in accounts payable", "chg_accounts_payable", "input", None, "[Balance Sheet]"),
    ("      Increase (decrease) in other current liabilities", "chg_other_current_liabilities", "input", None, "[Balance Sheet]"),
    ("         Total change in working capital", "change_wc", "subtotal", "=SUM(C10:C14)", ""),
    ("   Cash provided by operations", "operating", "total", "=C7+C8+C15", ""),
    ("", None, "spacer", None, ""),
    ("Investments:", None, "section", None, ""),
    ("Capital expenditures", "capex", "input", None, ""),
    ("plus Sales (acquisitions) of long-term assets", "acq_long_term", "input", None, ""),
    ("plus Other investing activities", "other_investing", "input", None, ""),
    ("   Cash provided by (used for) investments", "investments", "total", "=SUM(C19:C21)", ""),
    ("", None, "spacer", None, ""),
    ("Financing activities:", None, "section", None, ""),
    ("Increase (decrease) in short-term debt", "chg_short_term_debt", "input", None, ""),
    ("plus Increase (decrease) in long-term debt", "chg_long_term_debt", "input", None, ""),
    ("plus Dividends paid", "dividends_paid", "input", None, ""),
    ("plus Issues (repurchases) of stock", "stock_issues_repurchases", "input", None, ""),
    ("plus Other", "other_financing", "input", None, ""),
    ("   Cash provided by (used for) financing", "financing", "total", "=SUM(C25:C29)", ""),
    ("", None, "spacer", None, ""),
    ("Net increase (decrease) in cash", "net_change_cash", "total", "=C16+C22+C30", ""),
]

row = 6
cfs_row_map = {}
for label, name, kind, formula, source in cfs_items:
    cfs.cell(row=row, column=2, value=label)
    val_cell = cfs.cell(row=row, column=3)
    src_cell = cfs.cell(row=row, column=4, value=source)
    src_cell.font = Font(name=FONT, size=9, italic=True, color=NEUTRAL_600)

    if kind == "input":
        cfs.cell(row=row, column=2).font = f_label(size=10)
        val_cell.fill = fill(INPUT_BG); val_cell.font = f_input()
        val_cell.number_format = '#,##0;(#,##0);"—"'; val_cell.border = box
    elif kind == "link":
        cfs.cell(row=row, column=2).font = f_label(size=10)
        val_cell.font = f_formula()
        val_cell.number_format = '#,##0;(#,##0);"—"'
        val_cell.value = formula
    elif kind == "subtotal":
        cfs.cell(row=row, column=2).font = f_label(bold=True, size=10)
        val_cell.font = f_label(bold=True, size=10); val_cell.fill = fill(OUTPUT_BG)
        val_cell.number_format = '#,##0;(#,##0);"—"'; val_cell.border = box
        val_cell.value = formula
    elif kind == "total":
        cfs.cell(row=row, column=2).font = Font(name=FONT, size=11, bold=True, color=WHITE)
        cfs.cell(row=row, column=2).fill = fill(UH_GREEN)
        val_cell.font = Font(name=FONT, size=11, bold=True, color=WHITE); val_cell.fill = fill(UH_GREEN)
        val_cell.number_format = '#,##0;(#,##0);"—"'
        val_cell.value = formula
    elif kind == "section":
        cfs.cell(row=row, column=2).font = Font(name=FONT, size=10, bold=True, color=UH_GREEN)
    elif kind == "subsection":
        cfs.cell(row=row, column=2).font = Font(name=FONT, size=10, italic=True, color=NEUTRAL_600)

    if name:
        cfs_row_map[name] = row
    row += 1

# CASH_ named ranges — only for the two required outputs per spec
add_name("CASH_operating", f"'Cash Flow Statement'!$C${cfs_row_map['operating']}")
add_name("CASH_investments", f"'Cash Flow Statement'!$C${cfs_row_map['investments']}")

cfs.freeze_panes = "B6"


# =============================================================================
# RATIOS TAB
# =============================================================================
rt = wb.create_sheet("Ratios")
rt.sheet_view.showGridLines = False
rt.column_dimensions["A"].width = 2
rt.column_dimensions["B"].width = 44
rt.column_dimensions["C"].width = 16
rt.column_dimensions["D"].width = 62
rt.column_dimensions["E"].width = 36

rt["B2"] = "Ratios & Derived Inputs"
rt["B2"].font = Font(name=FONT, size=16, bold=True, color=UH_GREEN)
rt["B3"] = "Enter the four assumptions at top; all ratios populate automatically."
rt["B3"].font = Font(name=FONT, size=9, italic=True, color=NEUTRAL_600)

# Section: Assumptions (rows 5-11)
rt["B5"] = "Metric"; rt["C5"] = "Input"; rt["D5"] = "Named Range Formula"; rt["E5"] = "Named Range"
for col in [2, 3, 4, 5]:
    c = rt.cell(row=5, column=col)
    c.font = f_header(); c.fill = fill(HEADER_BG); c.border = box

assumptions = [
    # (label, value, formula_desc, name, format)
    ("Current fiscal year", 2024, None, "yearCurrent", "0"),
    ("Start of year (prior)", "=yearCurrent-1", "yearCurrent - 1", "yearStart", "0"),
    ("Share price", None, None, "share_price", '"$"#,##0.00'),
    ("Shares outstanding (M)", None, None, "shares_outstanding", "#,##0"),
    ("Cost of capital", 0.09, None, "cost_capital", "0.0%"),
    ("Tax rate", 0.21, None, "tax_rate", "0.0%"),
    ("Market capitalization", "=share_price*shares_outstanding", "share_price × shares_outstanding", "market_capitalization", "#,##0"),
]

row = 6
for label, value, formula_desc, name, fmt in assumptions:
    rt.cell(row=row, column=2, value=label).font = f_label(size=10)
    vc = rt.cell(row=row, column=3, value=value)
    rt.cell(row=row, column=4, value=formula_desc).font = f_mono(9)
    rt.cell(row=row, column=5, value=name).font = f_mono(9)
    vc.number_format = fmt
    vc.border = box

    if isinstance(value, str) and value.startswith("="):
        vc.font = f_formula()
    else:
        vc.fill = fill(ASSUMPTION_BG)
        vc.font = f_assumption()

    add_name(name, f"Ratios!$C${row}")
    row += 1

# Spacer + section header: Start of Year
row += 1
rt.cell(row=row, column=2, value="── Start of Year (Prior-Year Balance Sheet) ──")
style_section_row(rt, row, range(2, 6))
rt.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
row += 1

start_year_items = [
    ("Equity (start of year)", "=BAL_equity_shareholders_prior", "BAL_equity_shareholders_prior", "startYear_equity"),
    ("Inventories (start of year)", "=BAL_inventories_prior", "BAL_inventories_prior", "startYear_inventory"),
    ("Receivables (start of year)", "=BAL_receivables_prior", "BAL_receivables_prior", "startYear_receivables"),
    ("Total assets (start of year)", "=BAL_assets_total_prior", "BAL_assets_total_prior", "startYear_total_assets"),
    ("Total capitalization (start of year)", "=BAL_debt_long_term_prior+BAL_equity_shareholders_prior", "BAL_debt_long_term_prior + BAL_equity_shareholders_prior", "startYear_total_capitalization"),
]
for label, formula, desc, name in start_year_items:
    rt.cell(row=row, column=2, value=label).font = f_label(size=10)
    vc = rt.cell(row=row, column=3, value=formula)
    vc.font = f_formula(); vc.number_format = '#,##0;(#,##0);"—"'; vc.border = box
    rt.cell(row=row, column=4, value=desc).font = f_mono(9)
    rt.cell(row=row, column=5, value=name).font = f_mono(9)
    add_name(name, f"Ratios!$C${row}")
    row += 1

# Current year derived
row += 1
rt.cell(row=row, column=2, value="── Current Year (Derived) ──")
style_section_row(rt, row, range(2, 6))
rt.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
row += 1

current_items = [
    ("After-tax operating income", "=INC_net+(1-tax_rate)*INC_interest_expense", "INC_net + (1 − tax_rate) × INC_interest_expense", "currentYear_after_tax_operating_income", '#,##0;(#,##0);"—"'),
    ("Average daily sales", "=IFERROR(INC_sales/365,\"\")", "INC_sales / 365", "currentYear_daily_sales_average", '#,##0.00;(#,##0.00);"—"'),
    ("Book value of equity", "=BAL_equity_shareholders_curr", "BAL_equity_shareholders_curr", "currentYear_equity", '#,##0;(#,##0);"—"'),
    ("Cash + Marketable securities", "=BAL_cash_marketable_securities_curr", "BAL_cash_marketable_securities_curr", "currentYear_cash_marketable_securities", '#,##0;(#,##0);"—"'),
    ("Current assets", "=BAL_assets_current_curr", "BAL_assets_current_curr", "currentYear_assets_current", '#,##0;(#,##0);"—"'),
    ("Current liabilities", "=BAL_liabilities_current_curr", "BAL_liabilities_current_curr", "currentYear_liabilities_current", '#,##0;(#,##0);"—"'),
    ("Daily COGS", "=IFERROR(INC_cost_goods_sold/365,\"\")", "INC_cost_goods_sold / 365", "currentYear_cost_goods_sold_daily", '#,##0.00;(#,##0.00);"—"'),
    ("Long-term debt", "=BAL_debt_long_term_curr", "BAL_debt_long_term_curr", "currentYear_debt_long_term", '#,##0;(#,##0);"—"'),
    ("Net working capital", "=BAL_assets_current_curr-BAL_liabilities_current_curr", "BAL_assets_current_curr − BAL_liabilities_current_curr", "currentYear_working_capital_net", '#,##0;(#,##0);"—"'),
    ("Total assets", "=BAL_assets_total_curr", "BAL_assets_total_curr", "currentYear_assets_total", '#,##0;(#,##0);"—"'),
    ("Total capitalization", "=BAL_debt_long_term_curr+BAL_equity_shareholders_curr", "currentYear_debt_long_term + currentYear_equity", "currentYear_total_capitalization", '#,##0;(#,##0);"—"'),
    ("Total liabilities", "=BAL_liabilities_total_curr", "BAL_liabilities_total_curr", "currentYear_liabilities_total", '#,##0;(#,##0);"—"'),
]
for label, formula, desc, name, fmt in current_items:
    rt.cell(row=row, column=2, value=label).font = f_label(size=10)
    vc = rt.cell(row=row, column=3, value=formula)
    vc.font = f_formula(); vc.number_format = fmt; vc.border = box
    rt.cell(row=row, column=4, value=desc).font = f_mono(9)
    rt.cell(row=row, column=5, value=name).font = f_mono(9)
    add_name(name, f"Ratios!$C${row}")
    row += 1

# Averages
row += 1
rt.cell(row=row, column=2, value="── Mixed Year (Averages) ──")
style_section_row(rt, row, range(2, 6))
rt.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
row += 1

avg_items = [
    ("Average equity", "=AVERAGE(startYear_equity,currentYear_equity)", "AVERAGE(startYear_equity, currentYear_equity)", "avg_equity"),
    ("Average total assets", "=AVERAGE(startYear_total_assets,currentYear_assets_total)", "AVERAGE(startYear_total_assets, currentYear_assets_total)", "avg_total_assets"),
    ("Average total capitalization", "=AVERAGE(startYear_total_capitalization,currentYear_total_capitalization)", "AVERAGE(startYear_total_capitalization, currentYear_total_capitalization)", "avg_total_capitalization"),
]
for label, formula, desc, name in avg_items:
    rt.cell(row=row, column=2, value=label).font = f_label(size=10)
    vc = rt.cell(row=row, column=3, value=formula)
    vc.font = f_formula(); vc.number_format = '#,##0.00;(#,##0.00);"—"'; vc.border = box
    rt.cell(row=row, column=4, value=desc).font = f_mono(9)
    rt.cell(row=row, column=5, value=name).font = f_mono(9)
    add_name(name, f"Ratios!$C${row}")
    row += 1

# Ratio outputs header
row += 2
rt.cell(row=row, column=2, value="Ratio"); rt.cell(row=row, column=3, value="Output")
rt.cell(row=row, column=4, value="Named Range Formula (reference)"); rt.cell(row=row, column=5, value="Named Range")
rt.cell(row=row, column=6, value="Check")
for col in [2, 3, 4, 5, 6]:
    c = rt.cell(row=row, column=col)
    c.font = f_header(); c.fill = fill(HEADER_BG); c.border = box
row += 1

ratio_sections = [
    ("Performance", [
        ("Market value added (MVA)", "=market_capitalization-currentYear_equity", "market_capitalization − currentYear_equity", None, '#,##0;(#,##0);"—"'),
        ("Market-to-book ratio", "=IFERROR(market_capitalization/currentYear_equity,\"\")", "market_capitalization / currentYear_equity", None, '0.00"x"'),
        ("Economic value added (EVA)", "=currentYear_after_tax_operating_income-(cost_capital*startYear_total_capitalization)", "currentYear_after_tax_operating_income − (cost_capital × startYear_total_capitalization)", None, '#,##0;(#,##0);"—"'),
    ]),
    ("Profitability", [
        ("Return on assets (ROA)", "=IFERROR(currentYear_after_tax_operating_income/startYear_total_assets,\"\")", "currentYear_after_tax_operating_income / startYear_total_assets", None, '0.0%;(0.0%);"—"'),
        ("Return on capital (ROC)", "=IFERROR(currentYear_after_tax_operating_income/startYear_total_capitalization,\"\")", "currentYear_after_tax_operating_income / startYear_total_capitalization", None, '0.0%;(0.0%);"—"'),
        ("Return on equity (ROE)", "=IFERROR(INC_net/startYear_equity,\"\")", "INC_net / startYear_equity", None, '0.0%;(0.0%);"—"'),
        ("Return on assets (ROA) [avg]", "=IFERROR(currentYear_after_tax_operating_income/avg_total_assets,\"\")", "currentYear_after_tax_operating_income / avg_total_assets", None, '0.0%;(0.0%);"—"'),
        ("Return on capital (ROC) [avg]", "=IFERROR(currentYear_after_tax_operating_income/avg_total_capitalization,\"\")", "currentYear_after_tax_operating_income / avg_total_capitalization", None, '0.0%;(0.0%);"—"'),
        ("Return on equity (ROE) [avg]", "=IFERROR(INC_net/avg_equity,\"\")", "INC_net / avg_equity", None, '0.0%;(0.0%);"—"'),
    ]),
    ("Efficiency", [
        ("Asset turnover", "=IFERROR(INC_sales/startYear_total_assets,\"\")", "INC_sales / startYear_total_assets", "RATIO_asset_turnover", '0.00"x"'),
        ("Receivables turnover", "=IFERROR(INC_sales/startYear_receivables,\"\")", "INC_sales / startYear_receivables", None, '0.00"x"'),
        ("Average collection period (days)", "=IFERROR(startYear_receivables/currentYear_daily_sales_average,\"\")", "startYear_receivables / currentYear_daily_sales_average", None, '0.0" days"'),
        ("Inventory turnover", "=IFERROR(INC_cost_goods_sold/startYear_inventory,\"\")", "INC_cost_goods_sold / startYear_inventory", None, '0.00"x"'),
        ("Days in inventory", "=IFERROR(startYear_inventory/currentYear_cost_goods_sold_daily,\"\")", "startYear_inventory / currentYear_cost_goods_sold_daily", None, '0.0" days"'),
        ("Profit margin", "=IFERROR(INC_net/INC_sales,\"\")", "INC_net / INC_sales", None, '0.0%;(0.0%);"—"'),
        ("Operating profit margin", "=IFERROR(currentYear_after_tax_operating_income/INC_sales,\"\")", "currentYear_after_tax_operating_income / INC_sales", "RATIO_operating_profit_margin", '0.0%;(0.0%);"—"'),
    ]),
    ("Leverage", [
        ("Long-term debt ratio", "=IFERROR(currentYear_debt_long_term/(currentYear_debt_long_term+currentYear_equity),\"\")", "currentYear_debt_long_term / (currentYear_debt_long_term + currentYear_equity)", None, '0.0%;(0.0%);"—"'),
        ("Long-term debt-equity ratio", "=IFERROR(currentYear_debt_long_term/currentYear_equity,\"\")", "currentYear_debt_long_term / currentYear_equity", None, '0.00"x"'),
        ("Total debt ratio", "=IFERROR(currentYear_liabilities_total/currentYear_assets_total,\"\")", "currentYear_liabilities_total / currentYear_assets_total", None, '0.0%;(0.0%);"—"'),
        ("Times interest earned", "=IFERROR(INC_ebit/INC_interest_expense,\"\")", "INC_ebit / INC_interest_expense", None, '0.00"x"'),
        ("Cash coverage ratio", "=IFERROR((INC_ebit+INC_depreciation)/INC_interest_expense,\"\")", "(INC_ebit + INC_depreciation) / INC_interest_expense", None, '0.00"x"'),
        ("Debt burden", "=IFERROR(INC_net/currentYear_after_tax_operating_income,\"\")", "INC_net / currentYear_after_tax_operating_income", "RATIO_debt_burden", '0.000'),
        ("Leverage ratio", "=IFERROR(currentYear_assets_total/currentYear_equity,\"\")", "currentYear_assets_total / currentYear_equity", "RATIO_leverage", '0.00"x"'),
    ]),
    ("Liquidity", [
        ("Net working capital to assets", "=IFERROR(currentYear_working_capital_net/currentYear_assets_total,\"\")", "currentYear_working_capital_net / currentYear_assets_total", None, '0.0%;(0.0%);"—"'),
        ("Current ratio", "=IFERROR(currentYear_assets_current/currentYear_liabilities_current,\"\")", "currentYear_assets_current / currentYear_liabilities_current", None, '0.00"x"'),
        ("Quick ratio", "=IFERROR((currentYear_cash_marketable_securities+BAL_receivables_curr)/currentYear_liabilities_current,\"\")", "(currentYear_cash_marketable_securities + BAL_receivables_curr) / currentYear_liabilities_current", None, '0.00"x"'),
        ("Cash ratio", "=IFERROR(currentYear_cash_marketable_securities/currentYear_liabilities_current,\"\")", "currentYear_cash_marketable_securities / currentYear_liabilities_current", None, '0.00"x"'),
    ]),
    ("Du Pont System", [
        ("Return on assets (Du Pont)", "=IFERROR(RATIO_asset_turnover*RATIO_operating_profit_margin,\"\")", "RATIO_asset_turnover × RATIO_operating_profit_margin", None, '0.0%;(0.0%);"—"'),
        ("Return on equity (Du Pont)", "=IFERROR(RATIO_leverage*RATIO_asset_turnover*RATIO_operating_profit_margin*RATIO_debt_burden,\"\")", "RATIO_leverage × RATIO_asset_turnover × RATIO_operating_profit_margin × RATIO_debt_burden", None, '0.0%;(0.0%);"—"'),
    ]),
]

# Capture rows of key ratios for the Du Pont self-check and Notes references.
roa_direct_row = roe_direct_row = roa_dupont_row = roe_dupont_row = None

for section_name, items in ratio_sections:
    # Section header
    sc = rt.cell(row=row, column=2, value=section_name)
    sc.font = Font(name=FONT, size=11, bold=True, color=WHITE)
    sc.fill = fill(UH_GREEN)
    for col in [3, 4, 5, 6]:
        rt.cell(row=row, column=col).fill = fill(UH_GREEN)
    row += 1
    for label, formula, desc, name, fmt in items:
        rt.cell(row=row, column=2, value=label).font = f_label(size=10)
        # Column C left blank — students compute each ratio themselves using
        # the named-range formula shown in column D as a reference.
        vc = rt.cell(row=row, column=3)
        vc.font = f_output(); vc.fill = fill(OUTPUT_BG); vc.number_format = fmt
        vc.border = box
        vc.alignment = Alignment(horizontal="right")
        rt.cell(row=row, column=4, value=desc).font = f_mono(9)
        rt.cell(row=row, column=5, value=name if name else "").font = f_mono(9)
        if name:
            add_name(name, f"Ratios!$C${row}")
        if label == "Return on assets (ROA)":
            roa_direct_row = row
        elif label == "Return on equity (ROE)":
            roe_direct_row = row
        elif label == "Return on assets (Du Pont)":
            roa_dupont_row = row
        elif label == "Return on equity (Du Pont)":
            roe_dupont_row = row
        row += 1

# Du Pont ROA self-check (exact identity): Du Pont ROA == direct ROA when
# both use start-of-year assets. Stays blank until the student fills in
# both source cells; then displays match or signed delta.
if roa_direct_row and roa_dupont_row:
    chk = rt.cell(row=roa_dupont_row, column=6)
    chk.value = (
        f'=IF(OR(C{roa_dupont_row}="",C{roa_direct_row}=""),"",'
        f'IF(ROUND(C{roa_dupont_row},6)=ROUND(C{roa_direct_row},6),'
        f'"✓ matches direct ROA",'
        f'"✗ differs by "&TEXT(ABS(C{roa_dupont_row}-C{roa_direct_row}),"0.00%")))'
    )
    chk.font = Font(name=FONT, size=10, color=UH_GREEN)
    chk.fill = fill(OUTPUT_BG)
    chk.alignment = Alignment(horizontal="left", vertical="center")
    chk.number_format = "@"

rt.column_dimensions["F"].width = 28

rt.freeze_panes = "B6"


# =============================================================================
# NOTES TAB
# =============================================================================
notes = wb.create_sheet("Notes")
notes.sheet_view.showGridLines = False
notes.column_dimensions["A"].width = 2
notes.column_dimensions["B"].width = 28
notes.column_dimensions["C"].width = 70

notes["B2"] = "Notes & Assumptions"
notes["B2"].font = Font(name=FONT, size=16, bold=True, color=UH_GREEN)

notes_rows = [
    ("Company", "[Your Company Name]"),
    ("Ticker", "[TICKER]"),
    ("Industry", "[Industry]"),
    ("Fiscal Year (current)", "[FYYYYY]"),
    ("Fiscal Year (prior)", "[FYYYYY]"),
    ("Data Source", "[SEC EDGAR 10-K URL / Yahoo Finance / etc.]"),
    ("Units", "All figures in $millions unless otherwise noted"),
    ("Tax Rate", "[Statutory 21% / Effective rate from financials — and why]"),
    ("Cost of Capital", "[X.X% — source/method, e.g., CAPM, class working WACC]"),
    ("Share Price", "[FY-end closing price]"),
    ("Shares Outstanding", "[M shares — diluted weighted avg]"),
    ("", ""),
    ("BUS-314 Project Use", "Template for the Accounting & Performance Ratios project"),
    ("Stage 2 Task", "Populate Balance Sheet, Income Statement, and Cash Flow with your selected company's figures; enter the four assumptions on the Ratios tab; then compute all ratios in column C using the named-range formulas shown in column D."),
    ("Stage 3 Task", "Document the model using example-spec-template.md"),
    ("Stage 4 Task", "Interpret ratios and write CFO recommendation memo"),
    ("AI Usage", "[Document any AI tools used and how they were applied]"),
    ("", ""),
    ("Self-checks", "Cross-checks built into the Ratios sheet to verify your work."),
    ("Du Pont ROA", f"Du Pont ROA (Ratios!C{roa_dupont_row}) is mathematically identical to direct ROA (Ratios!C{roa_direct_row}). The inline check on Ratios!F{roa_dupont_row} confirms the match — investigate any discrepancy before submitting."),
    ("Du Pont ROE", f"Du Pont ROE (Ratios!C{roe_dupont_row}) will not exactly equal direct ROE (Ratios!C{roe_direct_row}) in this model: the leverage component uses current-year balances while asset turnover uses prior-year assets. Explain this time-mismatch in your Stage 4 memo."),
]
row = 4
for label, val in notes_rows:
    if label == "":
        row += 1
        continue
    lc = notes.cell(row=row, column=2, value=label)
    lc.font = f_label(bold=True, size=10); lc.alignment = Alignment(vertical="top")
    vc = notes.cell(row=row, column=3, value=val)
    vc.font = f_input(); vc.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

# Reorder sheets: Cover first
wb.move_sheet("Cover", offset=-5)

wb.save(OUTPUT)
print(f"Wrote {OUTPUT}")
print(f"Defined names: {len(list(wb.defined_names))}")
