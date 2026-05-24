import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl.worksheet.properties

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GAAP Bridge"

OPEN_SANS = "Open Sans"
header_font = Font(name=OPEN_SANS, bold=True, size=14, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")
section_font = Font(name=OPEN_SANS, bold=True, size=11, color="000000")
section_fill = PatternFill("solid", fgColor="D9E1F2")
label_font = Font(name=OPEN_SANS, size=10, color="000000")
input_font = Font(name=OPEN_SANS, size=10, color="0000FF")
input_fill = PatternFill("solid", fgColor="F2F2F2")
note_font = Font(name=OPEN_SANS, size=9, italic=True, color="666666")
col_head_font = Font(name=OPEN_SANS, bold=True, size=10, color="FFFFFF")
col_head_fill = PatternFill("solid", fgColor="4472C4")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 40
ws.column_dimensions["F"].width = 35

# --- HEADER ---
ws.merge_cells("A1:F1")
ws["A1"] = "GAAP BRIDGE — Accounting Standards Conversion Disclosure"
ws["A1"].font = header_font
ws["A1"].fill = header_fill
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

# --- COMPANY INFO ---
r = 3
ws.merge_cells(f"A{r}:F{r}")
ws[f"A{r}"] = "Company & Reporting Basis"
ws[f"A{r}"].font = section_font
ws[f"A{r}"].fill = section_fill
ws[f"A{r}"].alignment = center

info_labels = [
    ("Company Name", ""),
    ("Reporting Standard", "e.g., VAS, IFRS, CAS, JGAAP, Ind AS, US GAAP"),
    ("IFRS Convergence Status", "Identical / Converged / Partially converged / Distinct"),
    ("Conversion Tier Applied", "Tier 1 (none) / Tier 2 (disclosure) / Tier 3 (quantitative)"),
    ("Reporting Currency", "e.g., VND, CNY, JPY, INR, USD"),
    ("Fiscal Year End", "e.g., December 31, 2024"),
    ("Source Document", "e.g., Annual Report FY2024, 10-K, 20-F"),
    ("Source URL", ""),
]

for i, (label, hint) in enumerate(info_labels):
    row = r + 1 + i
    ws[f"B{row}"] = label
    ws[f"B{row}"].font = label_font
    ws[f"B{row}"].border = thin_border
    ws[f"C{row}"].font = input_font
    ws[f"C{row}"].fill = input_fill
    ws[f"C{row}"].border = thin_border
    ws.merge_cells(f"C{row}:D{row}")
    if hint:
        ws[f"E{row}"] = hint
        ws[f"E{row}"].font = note_font
        ws[f"E{row}"].alignment = wrap

# --- ADJUSTMENT ANALYSIS ---
r = 14
ws.merge_cells(f"A{r}:F{r}")
ws[f"A{r}"] = "Cross-Standard Adjustment Analysis"
ws[f"A{r}"].font = section_font
ws[f"A{r}"].fill = section_fill
ws[f"A{r}"].alignment = center

r = 15
col_headers = ["#", "Area", "Applies?", "Direction of Impact", "Magnitude / Notes", "Annual Report Source"]
for j, h in enumerate(col_headers):
    col = get_column_letter(j + 1)
    ws[f"{col}{r}"] = h
    ws[f"{col}{r}"].font = col_head_font
    ws[f"{col}{r}"].fill = col_head_fill
    ws[f"{col}{r}"].alignment = center
    ws[f"{col}{r}"].border = thin_border

adjustments = [
    (
        1,
        "Inventory Method (LIFO/FIFO)",
        "IFRS/VAS prohibit LIFO. If comparing to US GAAP LIFO companies, add LIFO reserve to their inventory. "
        "If your company uses weighted-average vs. FIFO, note the method.",
    ),
    (
        2,
        "Leases (IFRS 16 / VAS off-BS)",
        "IFRS 16: all leases on balance sheet (ROU asset + lease liability); EBITDA inflated vs. US GAAP "
        "operating-lease model. VAS: operating leases are OFF balance sheet — assets and liabilities understated "
        "vs. IFRS/US GAAP peers.",
    ),
    (
        3,
        "R&D / Development Costs",
        "IFRS (IAS 38): qualifying development costs capitalized as intangible assets. US GAAP: nearly all R&D "
        "expensed. VAS: practice varies by MoF circular. Note capitalized amount and annual amortization.",
    ),
    (
        4,
        "Impairment Reversals",
        "IFRS allows reversal of prior impairment losses (except goodwill) — inflates operating income in "
        "recovery years. US GAAP prohibits reversals. Check P&L for any impairment reversal gain.",
    ),
    (
        5,
        "Revenue Recognition Timing",
        "IFRS 15 and ASC 606 are largely converged (5-step model). VAS follows legacy IAS 18 — may differ for "
        "licensing, construction, multi-element arrangements. Note any disclosed policy differences.",
    ),
]

for i, (num, area, guidance) in enumerate(adjustments):
    row = r + 1 + (i * 3)

    ws[f"A{row}"] = num
    ws[f"A{row}"].font = Font(name=OPEN_SANS, bold=True, size=10)
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="top")
    ws[f"A{row}"].border = thin_border

    ws[f"B{row}"] = area
    ws[f"B{row}"].font = Font(name=OPEN_SANS, bold=True, size=10)
    ws[f"B{row}"].alignment = wrap
    ws[f"B{row}"].border = thin_border

    for col in ["C", "D", "E", "F"]:
        ws[f"{col}{row}"].font = input_font
        ws[f"{col}{row}"].fill = input_fill
        ws[f"{col}{row}"].alignment = wrap
        ws[f"{col}{row}"].border = thin_border

    ws.merge_cells(f"B{row+1}:F{row+1}")
    ws[f"B{row+1}"] = guidance
    ws[f"B{row+1}"].font = note_font
    ws[f"B{row+1}"].alignment = wrap
    ws.row_dimensions[row + 1].height = 45

# --- ADDITIONAL ADJUSTMENTS ---
r_extra = 32
ws.merge_cells(f"A{r_extra}:F{r_extra}")
ws[f"A{r_extra}"] = "Additional Adjustments (if applicable)"
ws[f"A{r_extra}"].font = section_font
ws[f"A{r_extra}"].fill = section_fill
ws[f"A{r_extra}"].alignment = center

r_extra += 1
for j, h in enumerate(col_headers):
    col = get_column_letter(j + 1)
    ws[f"{col}{r_extra}"] = h
    ws[f"{col}{r_extra}"].font = col_head_font
    ws[f"{col}{r_extra}"].fill = col_head_fill
    ws[f"{col}{r_extra}"].alignment = center
    ws[f"{col}{r_extra}"].border = thin_border

for i in range(3):
    row = r_extra + 1 + i
    ws[f"A{row}"] = 6 + i
    ws[f"A{row}"].font = label_font
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws[f"{col}{row}"].font = input_font
        ws[f"{col}{row}"].fill = input_fill
        ws[f"{col}{row}"].alignment = wrap
        ws[f"{col}{row}"].border = thin_border

# --- AFFECTED RATIOS ---
r_ratios = 39
ws.merge_cells(f"A{r_ratios}:F{r_ratios}")
ws[f"A{r_ratios}"] = "Ratios Materially Affected by Cross-Standard Differences"
ws[f"A{r_ratios}"].font = section_font
ws[f"A{r_ratios}"].fill = section_fill
ws[f"A{r_ratios}"].alignment = center

ratio_headers = ["#", "Ratio Name", "GAAP Bridge Row #", "Direction", "Estimated Impact", "Notes"]
r_ratios += 1
for j, h in enumerate(ratio_headers):
    col = get_column_letter(j + 1)
    ws[f"{col}{r_ratios}"] = h
    ws[f"{col}{r_ratios}"].font = col_head_font
    ws[f"{col}{r_ratios}"].fill = col_head_fill
    ws[f"{col}{r_ratios}"].alignment = center
    ws[f"{col}{r_ratios}"].border = thin_border

affected_ratios = [
    "EBITDA / EV/EBITDA",
    "Current Ratio",
    "Debt-to-Equity",
    "Return on Assets (ROA)",
    "Asset Turnover",
    "Gross Margin",
    "Inventory Turnover",
    "Operating Income / EBIT",
]

for i, ratio in enumerate(affected_ratios):
    row = r_ratios + 1 + i
    ws[f"A{row}"] = i + 1
    ws[f"A{row}"].font = label_font
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    ws[f"B{row}"] = ratio
    ws[f"B{row}"].font = label_font
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws[f"{col}{row}"].border = thin_border
    for col in ["C", "D", "E", "F"]:
        ws[f"{col}{row}"].font = input_font
        ws[f"{col}{row}"].fill = input_fill
        ws[f"{col}{row}"].alignment = wrap

# --- SUMMARY ---
r_summary = 51
ws.merge_cells(f"A{r_summary}:F{r_summary}")
ws[f"A{r_summary}"] = "Summary Assessment"
ws[f"A{r_summary}"].font = section_font
ws[f"A{r_summary}"].fill = section_fill
ws[f"A{r_summary}"].alignment = center

summary_labels = [
    ("Total adjustments identified", ""),
    ("Adjustments quantified", ""),
    ("Adjustments noted but not quantified", ""),
    ("Overall impact on comparability", "Low / Medium / High"),
    ("Analyst name", ""),
    ("Date completed", ""),
]

for i, (label, hint) in enumerate(summary_labels):
    row = r_summary + 1 + i
    ws[f"B{row}"] = label
    ws[f"B{row}"].font = label_font
    ws[f"B{row}"].border = thin_border
    ws.merge_cells(f"C{row}:D{row}")
    ws[f"C{row}"].font = input_font
    ws[f"C{row}"].fill = input_fill
    ws[f"C{row}"].border = thin_border
    if hint:
        ws[f"E{row}"] = hint
        ws[f"E{row}"].font = note_font

# --- REFERENCE FOOTER ---
r_footer = 60
ws.merge_cells(f"A{r_footer}:F{r_footer}")
ws[f"A{r_footer}"] = (
    "Reference: docs/decisions/2026-05-24-accounting-standards-conversion-framework.md | "
    "Quick-ref: docs/templates/gaap-conversion-quickref.md"
)
ws[f"A{r_footer}"].font = Font(name=OPEN_SANS, size=8, italic=True, color="999999")

ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = "landscape"
ws.freeze_panes = "A3"

out_path = "courses/BUS-629-VEMBA-International-Corporate-Finance/models/templates/gaap-bridge-template.xlsx"
wb.save(out_path)
print(f"Created: {out_path}")
