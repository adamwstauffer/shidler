"""Verify a 'Final Grades' workbook produced by compute_final_grades.py.

Checks performed against each input file:
  1. Letter-grade formula consistency across all student rows.
  2. Cross-tab row alignment: every Grades!XN reference inside a Final Grades
     formula points to row N == student's row in Final Grades.
  3. Identity columns (OrgDefinedId, Username, Last/First/Email) match between
     Grades and Final Grades for every row.
  4. Cached computed values (after Excel has saved the workbook) match an
     independent recomputation from raw inputs in the Grades sheet.
  5. Distribution preview.

Usage:
    python scripts/grading/verify_final_grades.py <course_key> <final_grades.xlsx> [<course_key2> <file2> ...]

Example:
    python scripts/grading/verify_final_grades.py \\
        BUS-313 "courses/.../BUS-313-001-SP26_FinalGrades.xlsx" \\
        FIN-321 "courses/.../FIN-321-002-SP26_FinalGrades.xlsx"
"""
import argparse
import math
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from configs import (
    COURSES,
    DEFAULT_WEIGHTS,
    FINAL_CAP,
    LETTER_SCALE,
    LETTER_FALLBACK,
)
from compute_final_grades import detect_chapters, find_col


GRADES_REF = re.compile(r"Grades!\$?([A-Z]+)\$?(\d+)")


def expected_letter(p):
    for thresh, letter in LETTER_SCALE:
        if p >= thresh:
            return letter
    return LETTER_FALLBACK


def safe_num(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def formula_text(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v
    if hasattr(v, 'text'):
        return v.text
    return str(v)


def verify(course_key, path, weights=None, final_cap=FINAL_CAP):
    weights = weights or DEFAULT_WEIGHTS
    cfg = COURSES[course_key]

    print(f"\n{'=' * 78}")
    print(f"VERIFYING [{course_key}] {Path(path).name}")
    print('=' * 78)

    if not Path(path).exists():
        print(f"  MISSING FILE: {path}")
        return False

    wb = load_workbook(path)
    if 'Grades' not in wb.sheetnames or 'Final Grades' not in wb.sheetnames:
        print(f"  MISSING SHEETS: have {wb.sheetnames}")
        return False
    grades = wb['Grades']
    fg = wb['Final Grades']

    # Read source headers from the Grades sheet
    headers = [grades.cell(row=1, column=i).value for i in range(1, grades.max_column + 1)]
    n_students = grades.max_row - 1

    # Resolve source columns (mirrors compute logic)
    att_col = find_col(headers, cfg['attendance'])
    mid_col = find_col(headers, cfg['midterm'])
    final_col = find_col(headers, cfg['final_exam'])
    project_cols = [find_col(headers, p) for p in cfg['project']]
    ec_cols = [find_col(headers, name) for name, _max in cfg['extra_credit']]
    chapters = detect_chapters(headers, cfg['chapter_prefix_re'], cfg['makeup_marker'])

    # Locate Final Grades columns by header text
    fg_headers = [fg.cell(row=1, column=i).value for i in range(1, fg.max_column + 1)]
    def fg_col(needle):
        for i, h in enumerate(fg_headers, start=1):
            if h and needle.lower() in str(h).lower():
                return i
        raise KeyError(f"Final Grades sheet missing column matching {needle!r}")

    cols = {
        'att':    fg_col('Attendance ('),
        'hw':     fg_col('HW ('),
        'proj':   fg_col('Project ('),
        'mid':    fg_col('Midterm ('),
        'fin':    fg_col('Final ('),
        'sub':    fg_col('Subtotal'),
        'total':  fg_col('Total w/ EC'),
        'final':  fg_col('Final Grade'),
        'letter': fg_col('Letter Grade'),
    }

    # --- 1) Letter formula consistency -----------------------------------
    f2 = formula_text(fg.cell(row=2, column=cols['letter']).value)
    print(f"  Letter formula sample (row 2): {f2}")
    norm = re.sub(r'\d+', 'N', f2 or '')
    odd = []
    for r in range(3, n_students + 2):
        f = formula_text(fg.cell(row=r, column=cols['letter']).value)
        if f is None:
            continue
        if re.sub(r'\d+', 'N', f) != norm:
            odd.append((r, f))
    if odd:
        print(f"  WARN: {len(odd)} letter formulas differ in structure")
        for r, f in odd[:3]:
            print(f"    row {r}: {f}")
    else:
        print(f"  Letter formulas: consistent across {n_students} rows.")

    # --- 2) Cross-tab row alignment --------------------------------------
    misaligned = []
    for r in range(2, n_students + 2):
        for fcol in cols.values():
            f = formula_text(fg.cell(row=r, column=fcol).value)
            if not isinstance(f, str) or not f.startswith('='):
                continue
            for col_letters, row_str in GRADES_REF.findall(f):
                ref_row = int(row_str)
                if ref_row != r:
                    misaligned.append((r, fcol, col_letters, ref_row, f[:80]))
    if misaligned:
        print(f"  FAIL cross-tab alignment: {len(misaligned)} mismatches")
        for x in misaligned[:5]:
            print(f"    {x}")
        return False
    print(f"  Cross-tab alignment OK across {n_students} students x all formulas.")

    # --- 3) Identity column alignment ------------------------------------
    id_mismatches = []
    for r in range(2, n_students + 2):
        for col_idx in range(1, 6):
            g = grades.cell(row=r, column=col_idx).value
            f = fg.cell(row=r, column=col_idx).value
            if str(g if g is not None else '') != str(f if f is not None else ''):
                id_mismatches.append((r, col_idx, g, f))
    if id_mismatches:
        print(f"  FAIL identity: {len(id_mismatches)} mismatches")
        for x in id_mismatches[:5]:
            print(f"    {x}")
        return False
    print(f"  Identity columns aligned across all rows.")

    # --- 4) Independent value recomputation ------------------------------
    summary = []
    for r in range(2, n_students + 2):
        def gv(idx_zero):
            return safe_num(grades.cell(row=r, column=idx_zero + 1).value)

        att = min(weights['attendance'], gv(att_col))
        hw_raw = sum(max(gv(o), gv(m)) for _ch, o, m in chapters)
        hw = min(weights['homework'], hw_raw)
        proj = min(weights['project'], sum(gv(c) for c in project_cols))
        mid = min(weights['midterm'], gv(mid_col))
        fin = min(weights['final_exam'], gv(final_col))
        sub = att + hw + proj + mid + fin
        ec = sum(gv(c) for c in ec_cols)
        total = sub + ec
        final = min(final_cap, math.ceil(total))
        summary.append({
            'r': r,
            'last': fg.cell(row=r, column=3).value,
            'first': fg.cell(row=r, column=4).value,
            'att': att, 'hw': hw, 'proj': proj, 'mid': mid, 'fin': fin,
            'sub': sub, 'ec': ec, 'total': total, 'final': final,
            'letter': expected_letter(final),
        })

    wb2 = load_workbook(path, data_only=True)
    fg2 = wb2['Final Grades']
    cached_present = False
    diffs = []
    for s in summary:
        r = s['r']
        rowvals = {k: fg2.cell(row=r, column=cols[k]).value for k in cols}
        if any(isinstance(v, (int, float)) for v in rowvals.values() if v is not None):
            cached_present = True
        for k in ('att', 'hw', 'proj', 'mid', 'fin', 'sub', 'total', 'final'):
            cv = rowvals.get(k)
            if cv is None or isinstance(cv, str):
                continue
            if abs(float(cv) - s[k]) > 0.005:
                diffs.append((r, s['last'], s['first'], k, float(cv), s[k]))
        cv_letter = rowvals.get('letter')
        if isinstance(cv_letter, str) and cv_letter and cv_letter != s['letter']:
            diffs.append((r, s['last'], s['first'], 'letter', cv_letter, s['letter']))

    if not cached_present:
        print("  No cached values present — open in Excel once, save, and re-run for value-level check.")
    elif diffs:
        print(f"  FAIL value check: {len(diffs)} differences")
        for d in diffs[:10]:
            print(f"    {d}")
        return False
    else:
        print(f"  All cached values match independent recomputation.")

    # --- 5) Distribution preview -----------------------------------------
    c = Counter(s['letter'] for s in summary)
    order = [lg for _, lg in LETTER_SCALE] + [LETTER_FALLBACK]
    print(f"  Distribution: " + " ".join(f"{k}:{c.get(k, 0)}" for k in order))
    return True


def main():
    p = argparse.ArgumentParser(description=__doc__.split('\n\n')[0])
    p.add_argument('args', nargs='+',
                   help='Pairs of (course_key, file_path) to verify')
    a = p.parse_args()
    if len(a.args) % 2 != 0:
        p.error('Expected an even number of args (course_key file ... pairs).')

    pairs = list(zip(a.args[0::2], a.args[1::2]))
    all_ok = True
    for course_key, path in pairs:
        ok = verify(course_key, path)
        all_ok = all_ok and ok
    sys.exit(0 if all_ok else 1)


if __name__ == '__main__':
    main()
