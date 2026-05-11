"""Compute final grades from a Brightspace export and write a 'Final Grades'
sheet alongside the original 'Grades' sheet.

Method (per student, capped within formulas; weights == max points per component):
  Attendance       = MIN(weight, raw)
  Homework         = MIN(weight, sum of MAX(original, makeup) per chapter)
  Project          = MIN(weight, sum of project-stage columns)
  Midterm / Final  = MIN(weight, raw)
  Subtotal         = sum of the five components above   (out of 100)
  + Extra credit   = sum of EC columns (added on top, no cap on EC itself)
  Final Grade      = MIN(100, ROUNDUP(Subtotal + EC, 0))
  Letter Grade     = IFS over the configured scale

Cross-tab references: every formula points to the SAME row number in 'Grades'
as the student's row in 'Final Grades' (row alignment preserved 1:1).

Usage:
    python scripts/grading/compute_final_grades.py <course_key> <input.xlsx> <output.xlsx>

Example:
    python scripts/grading/compute_final_grades.py BUS-313 \\
        "courses/BUS-313.../ignore/2026 Spring/grades/SP26_GradesExport.xlsx" \\
        "courses/BUS-313.../ignore/2026 Spring/grades/SP26_FinalGrades.xlsx"

The course_key indexes into configs.COURSES.
"""
import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from configs import (
    COURSES,
    DEFAULT_WEIGHTS,
    FINAL_CAP,
    LETTER_SCALE,
    LETTER_FALLBACK,
)


UH_GREEN = '024731'
HEADER_FONT = Font(name='Open Sans', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color=UH_GREEN)
BODY_FONT = Font(name='Open Sans', size=10)
BOLD_FONT = Font(name='Open Sans', size=10, bold=True)
CENTER = Alignment(horizontal='center', vertical='center')
THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def find_col(headers, needle):
    """Find the 0-based column index whose header contains needle (case-insensitive).
    Raises if 0 or >1 matches."""
    needle_l = needle.lower()
    matches = [i for i, h in enumerate(headers) if needle_l in str(h).lower()]
    if not matches:
        raise KeyError(f"No column matched substring {needle!r}")
    if len(matches) > 1:
        sample = ', '.join(repr(headers[i]) for i in matches[:3])
        raise KeyError(f"Ambiguous: {len(matches)} cols matched {needle!r}: {sample}")
    return matches[0]


def detect_chapters(headers, prefix_re, makeup_marker):
    """Detect (orig_idx, makeup_idx) pairs for chapter HW columns.

    Returns a list of pairs ordered by chapter number. A chapter without
    a makeup falls back to (orig_idx, orig_idx) so MAX(orig, orig) == orig.
    """
    rx = re.compile(prefix_re, re.IGNORECASE)
    marker_l = makeup_marker.lower()

    originals, makeups = {}, {}  # chapter_num -> col_idx
    for idx, h in enumerate(headers):
        s = str(h)
        m = rx.match(s)
        if not m:
            continue
        ch_num = int(m.group(1))
        is_makeup = marker_l in s.lower()
        bucket = makeups if is_makeup else originals
        if ch_num in bucket:
            raise ValueError(
                f"Duplicate chapter {ch_num} ({'makeup' if is_makeup else 'original'}): "
                f"cols {bucket[ch_num]} and {idx} ({s!r})"
            )
        bucket[ch_num] = idx

    pairs = []
    for ch_num in sorted(originals):
        o = originals[ch_num]
        m = makeups.get(ch_num, o)  # fallback to original if no makeup
        pairs.append((ch_num, o, m))

    # Warn about any orphan makeups (makeup with no original)
    orphans = sorted(set(makeups) - set(originals))
    if orphans:
        print(f"  WARNING: makeup columns with no matching original: chapters {orphans}",
              file=sys.stderr)
    return pairs


def build_letter_formula(target_cell, scale, fallback):
    """Return =IFS(...) string evaluating to a letter grade based on target_cell."""
    parts = ['IFS(']
    for thresh, letter in scale:
        parts.append(f'{target_cell}>={thresh},"{letter}",')
    parts.append(f'TRUE,"{fallback}")')
    return '=' + ''.join(parts)


def src_col_letter(zero_idx):
    return get_column_letter(zero_idx + 1)


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER


def compute_final_grades(course_key, src_path, out_path,
                         weights=None, final_cap=FINAL_CAP,
                         letter_scale=LETTER_SCALE,
                         letter_fallback=LETTER_FALLBACK):
    if course_key not in COURSES:
        raise KeyError(f"Unknown course {course_key!r}; configured: {list(COURSES)}")
    cfg = COURSES[course_key]
    weights = weights or DEFAULT_WEIGHTS

    df = pd.read_excel(src_path)
    headers = list(df.columns)
    n_students = len(df)

    # Resolve column indices via header substring matching
    att_col = find_col(headers, cfg['attendance'])
    mid_col = find_col(headers, cfg['midterm'])
    final_col = find_col(headers, cfg['final_exam'])
    project_cols = [find_col(headers, p) for p in cfg['project']]
    ec_specs = [(find_col(headers, name), name, max_pts) for name, max_pts in cfg['extra_credit']]
    chapters = detect_chapters(headers, cfg['chapter_prefix_re'], cfg['makeup_marker'])

    print(f"[{course_key}] resolved {len(chapters)} chapters, "
          f"{len(project_cols)} project col(s), {len(ec_specs)} EC col(s).")

    # Open source workbook (preserves Grades sheet untouched)
    wb = load_workbook(src_path)
    if 'Final Grades' in wb.sheetnames:
        del wb['Final Grades']
    fg = wb.create_sheet('Final Grades', 0)

    # Headers
    out_headers = [
        'OrgDefinedId', 'Username', 'Last Name', 'First Name', 'Email',
        f'Attendance ({weights["attendance"]})',
        f'HW ({weights["homework"]})',
        f'Project ({weights["project"]})',
        f'Midterm ({weights["midterm"]})',
        f'Final ({weights["final_exam"]})',
        'Subtotal (100)',
    ]
    ec_start_col_1based = len(out_headers) + 1
    for _, name, max_pts in ec_specs:
        out_headers.append(f'EC: {name} ({max_pts})')
    out_headers += [
        'Total w/ EC',
        f'Final Grade (max {final_cap})',
        'Letter Grade',
    ]
    final_grade_col_1based = len(out_headers) - 1
    letter_col_1based = len(out_headers)
    for j, h in enumerate(out_headers, start=1):
        cell = fg.cell(row=1, column=j, value=h)
        style_header(cell)

    GR = 'Grades'

    for i in range(n_students):
        r = i + 2  # excel row (header is row 1)
        # Identity columns — copy values directly from source
        for col_zero in range(5):
            fg.cell(row=r, column=col_zero + 1, value=df.iloc[i, col_zero])

        # Attendance (col F)
        fg.cell(row=r, column=6,
                value=f"=MIN({weights['attendance']},IFERROR({GR}!{src_col_letter(att_col)}{r},0))")

        # HW (col G): sum of MAX(orig, makeup) per chapter, capped at weight
        hw_terms = []
        for _ch, o, m in chapters:
            oc = f"{GR}!{src_col_letter(o)}{r}"
            mc = f"{GR}!{src_col_letter(m)}{r}"
            hw_terms.append(f"MAX(IFERROR({oc},0),IFERROR({mc},0))")
        fg.cell(row=r, column=7,
                value=f"=MIN({weights['homework']},{'+'.join(hw_terms)})")

        # Project (col H): sum of project columns, capped at weight
        proj_terms = [f"IFERROR({GR}!{src_col_letter(c)}{r},0)" for c in project_cols]
        fg.cell(row=r, column=8,
                value=f"=MIN({weights['project']},{'+'.join(proj_terms)})")

        # Midterm (col I), Final (col J)
        fg.cell(row=r, column=9,
                value=f"=MIN({weights['midterm']},IFERROR({GR}!{src_col_letter(mid_col)}{r},0))")
        fg.cell(row=r, column=10,
                value=f"=MIN({weights['final_exam']},IFERROR({GR}!{src_col_letter(final_col)}{r},0))")

        # Subtotal (col K)
        fg.cell(row=r, column=11, value=f"=SUM(F{r}:J{r})")

        # Extra-credit columns
        ec_cells = []
        for k, (col_zero, _name, _maxp) in enumerate(ec_specs):
            cell_col = ec_start_col_1based + k
            fg.cell(row=r, column=cell_col,
                    value=f"=IFERROR({GR}!{src_col_letter(col_zero)}{r},0)")
            ec_cells.append(get_column_letter(cell_col) + str(r))

        # Total w/ EC
        total_col = ec_start_col_1based + len(ec_specs)
        if ec_cells:
            fg.cell(row=r, column=total_col, value=f"=K{r}+" + "+".join(ec_cells))
        else:
            fg.cell(row=r, column=total_col, value=f"=K{r}")

        # Final Grade (capped at 100)
        total_letter = get_column_letter(total_col)
        final_letter = get_column_letter(final_grade_col_1based)
        fg.cell(row=r, column=final_grade_col_1based,
                value=f"=MIN({final_cap},ROUNDUP({total_letter}{r},0))")

        # Letter Grade
        fg.cell(row=r, column=letter_col_1based,
                value=build_letter_formula(f'{final_letter}{r}', letter_scale, letter_fallback))

        # Styling
        for j in range(1, letter_col_1based + 1):
            cell = fg.cell(row=r, column=j)
            cell.font = BODY_FONT
            cell.border = BORDER
            if j >= 6:
                cell.alignment = CENTER
                if j in (final_grade_col_1based, letter_col_1based):
                    cell.font = BOLD_FONT
                cell.number_format = ('0' if j in (final_grade_col_1based, letter_col_1based)
                                       else '0.00')

    # Column widths (a sensible default profile)
    widths = [12, 18, 18, 15, 30] + [12, 12, 14, 12, 12, 14] + [16] * len(ec_specs) + [14, 16, 14]
    for j, w in enumerate(widths[: letter_col_1based], start=1):
        fg.column_dimensions[get_column_letter(j)].width = w
    fg.row_dimensions[1].height = 30
    fg.freeze_panes = 'F2'

    # ---- Grade Distribution table to the right of letter grade column ----
    last_data_row = n_students + 1
    dist_col = letter_col_1based + 2  # leave one blank column

    fg.merge_cells(start_row=1, start_column=dist_col, end_row=1, end_column=dist_col + 2)
    hc = fg.cell(row=1, column=dist_col, value='Grade Distribution')
    style_header(hc)
    for c in range(dist_col, dist_col + 3):
        cell = fg.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for k, h in enumerate(['Letter', 'Count', '% of Class']):
        cell = fg.cell(row=2, column=dist_col + k, value=h)
        style_header(cell)

    LC = get_column_letter(letter_col_1based)
    letters = [lg for _, lg in letter_scale] + [letter_fallback]
    for i_row, lg in enumerate(letters, start=3):
        fg.cell(row=i_row, column=dist_col, value=lg).font = BOLD_FONT
        fg.cell(row=i_row, column=dist_col + 1,
                value=f'=COUNTIF({LC}2:{LC}{last_data_row},"{lg}")')
        fg.cell(row=i_row, column=dist_col + 2,
                value=f'=IF(COUNTA({LC}2:{LC}{last_data_row})=0,0,'
                      f'COUNTIF({LC}2:{LC}{last_data_row},"{lg}")/COUNTA({LC}2:{LC}{last_data_row}))')
        for k in range(3):
            cell = fg.cell(row=i_row, column=dist_col + k)
            cell.alignment = CENTER
            cell.border = BORDER
            cell.font = BODY_FONT if k != 0 else BOLD_FONT
        fg.cell(row=i_row, column=dist_col + 2).number_format = '0.0%'

    total_row = 3 + len(letters)
    fg.cell(row=total_row, column=dist_col, value='Total').font = BOLD_FONT
    cnt_col_letter = get_column_letter(dist_col + 1)
    pct_col_letter = get_column_letter(dist_col + 2)
    fg.cell(row=total_row, column=dist_col + 1,
            value=f'=SUM({cnt_col_letter}3:{cnt_col_letter}{total_row - 1})').font = BOLD_FONT
    fg.cell(row=total_row, column=dist_col + 2,
            value=f'=SUM({pct_col_letter}3:{pct_col_letter}{total_row - 1})').font = BOLD_FONT
    fg.cell(row=total_row, column=dist_col + 2).number_format = '0.0%'
    for k in range(3):
        cell = fg.cell(row=total_row, column=dist_col + k)
        cell.alignment = CENTER
        cell.border = BORDER
        cell.fill = PatternFill('solid', start_color='E8F0EC')

    for k, w in enumerate([10, 10, 12]):
        fg.column_dimensions[get_column_letter(dist_col + k)].width = w

    wb.save(out_path)
    print(f"[{course_key}] wrote {out_path} ({n_students} students).")


def main():
    p = argparse.ArgumentParser(description=__doc__.split('\n\n')[0])
    p.add_argument('course_key', help='Course key in configs.COURSES (e.g. BUS-313)')
    p.add_argument('input_xlsx', help='Brightspace gradebook export (.xlsx)')
    p.add_argument('output_xlsx', help='Output workbook path (.xlsx)')
    args = p.parse_args()
    compute_final_grades(args.course_key, args.input_xlsx, args.output_xlsx)


if __name__ == '__main__':
    main()
